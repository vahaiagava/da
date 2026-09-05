#!/usr/bin/env python3
"""test_core_f1_f2_omzet.py — CORE TEST F1 (impor pesanan marketplace) + F2 (rekap turunan).

SATU berkas uji untuk seluruh rantai "satu sumber omzet":
  BAGIAN A — MESIN (tanpa DB): baca berkas nyata 65 kolom, lewati baris deskripsi,
             kamus nilai, pengelompokan 601 baris → 559 pesanan, angka uang.
  BAGIAN B — API impor (HTTP nyata): upload → pratinjau → commit → dedupe →
             rollback, termasuk penolakan berkas platform yang salah.
  BAGIAN C — F2 rekap harian turunan: satu angka omzet di semua pintu
             (rekap harian, performa, target bulanan, dashboard).

Berkas contoh: samples/TikTok_UntukDikirim_2026-07-19.xlsx (toko TikTok Outfit Boutique)

Angka yang WAJIB sama (dihitung ulang langsung dari berkas, bukan disalin):
  601 baris data · 559 pesanan · Σ omzet produk Rp 59.783.811 ·
  Σ order amount Rp 62.805.113 · Σ diskon penjual Rp 48.020.983 ·
  Σ harga coret Rp 109.179.000 · Σ qty 603 pcs · 514 baris item pre-order

Pakai:  python3 /app/test_core_f1_f2_omzet.py
"""
from __future__ import annotations

import os
import sys
import time
from collections import Counter, OrderedDict

import requests

sys.path.insert(0, "/app/backend")

BASE = "http://localhost:8001"
ADMIN = {"email": "admin@garment.com", "password": "Admin@123"}
SAMPLE = "/app/samples/TikTok_UntukDikirim_2026-07-19.xlsx"
ACCOUNT_CODE = "TIKTOK-OUTFIT"

RESULTS: list[tuple[str, bool, str]] = []


def ok(name, detail=""):
    RESULTS.append((name, True, detail))
    print(f"  \033[92mPASS\033[0m  {name}" + (f" — {detail}" if detail else ""))


def fail(name, detail=""):
    RESULTS.append((name, False, detail))
    print(f"  \033[91mFAIL\033[0m  {name}" + (f" — {detail}" if detail else ""))


def check(name, cond, detail=""):
    (ok if cond else fail)(name, detail)
    return bool(cond)


def money(v) -> int:
    return int(round(float(v or 0)))


# ══════════════════════════════════════════════════════════════════════════════
# Kebenaran dasar: dihitung LANGSUNG dari berkas dengan openpyxl (pembanding)
# ══════════════════════════════════════════════════════════════════════════════
def truth_from_file() -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(SAMPLE)
    ws = wb.active
    hdr = [c.value for c in ws[1]]
    i = {h: n for n, h in enumerate(hdr)}
    data = list(ws.iter_rows(min_row=3, values_only=True))   # baris 2 = deskripsi

    def num(v):
        if v is None or v == "":
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace("Rp", "").replace(" ", "").replace(",", "")
        try:
            return float(s)
        except ValueError:
            return 0.0

    orders: "OrderedDict[str, list]" = OrderedDict()
    for r in data:
        orders.setdefault(str(r[i["Order ID"]]), []).append(r)

    return {
        "headers": [h for h in hdr if h],
        "rows": len(data),
        "orders": len(orders),
        "revenue_product": money(sum(num(r[i["SKU Subtotal After Discount"]]) for r in data)),
        "revenue_gross": money(sum(num(r[i["SKU Subtotal Before Discount"]]) for r in data)),
        "seller_discount": money(sum(num(r[i["SKU Seller Discount"]]) for r in data)),
        "platform_discount": money(sum(num(r[i["SKU Platform Discount"]]) for r in data)),
        "order_amount": money(sum(num(rs[0][i["Order Amount"]]) for rs in orders.values())),
        "quantity": int(sum(num(r[i["Quantity"]]) for r in data)),
        "preorder_items": sum(1 for r in data if str(r[i["Normal or Pre-order"]]) == "Pre-order"),
        "dates": sorted({str(r[i["Created Time"]])[:10] for r in data}),
        "order_ids": list(orders.keys()),
    }


# ══════════════════════════════════════════════════════════════════════════════
# BAGIAN A — MESIN (tanpa database)
# ══════════════════════════════════════════════════════════════════════════════
def part_a_engine(truth: dict) -> dict:
    print("\n[A] MESIN IMPOR — berkas nyata 65 kolom → 1 dokumen per pesanan")
    from core import marketing_import_engine as eng
    from core.marketing_import_schema import get_source_type

    st = get_source_type("marketplace_orders")
    check("A1 jenis impor `marketplace_orders` terdaftar", st.collection == "marketing_orders",
          f"koleksi={st.collection} group_by={st.group_by}")

    with open(SAMPLE, "rb") as fh:
        raw = fh.read()
    headers, rows = eng.parse_table(raw, os.path.basename(SAMPLE))
    check("A2 65 kolom terbaca", len(headers) == 65, f"{len(headers)} kolom")

    mapping = eng.auto_map(headers, st)
    report = eng.mapping_report(mapping, st)
    mapped = [m for m in mapping if m.get("field")]
    check("A3 pemetaan otomatis TANPA AI menutup semua kolom wajib", report["ready"],
          f"belum terpetakan wajib={report.get('missing_required')}")
    check("A4 hampir semua kolom dikenali (≥60 dari 65)", len(mapped) >= 60,
          f"{len(mapped)}/65 dikenali; tak dikenali="
          f"{[m['column'] for m in mapping if not m.get('field')][:6]}")

    rows2, n_skipped = eng.strip_description_rows(rows, mapping, st)
    check("A5 baris deskripsi kolom dilewati (bukan dihitung data)",
          n_skipped == 1 and len(rows2) == truth["rows"],
          f"dilewati={n_skipped} sisa={len(rows2)} (harap {truth['rows']})")

    built = eng.build_rows(rows2, mapping, st, limit=eng.MAX_ROWS)
    check("A6 601 baris SKU dikelompokkan menjadi 559 pesanan",
          len(built) == truth["orders"], f"{len(built)} entri (harap {truth['orders']})")
    errs = [b for b in built if b["status"] == "error"]
    check("A7 0 pesanan ditolak", not errs,
          f"{len(errs)} ditolak; contoh={errs[0]['errors'][:2] if errs else []}")

    tot_items = sum(len(b["data"]["items"]) for b in built)
    check("A8 601 item masuk items[]", tot_items == truth["rows"], f"{tot_items} item")

    rev_prod = money(sum(sum(money(it.get("sku_subtotal_after_discount"))
                             for it in b["data"]["items"]) for b in built))
    rev_gross = money(sum(sum(money(it.get("sku_subtotal_before_discount"))
                              for it in b["data"]["items"]) for b in built))
    sel_disc = money(sum(sum(money(it.get("sku_seller_discount"))
                             for it in b["data"]["items"]) for b in built))
    ord_amt = money(sum(money(b["data"].get("order_amount")) for b in built))
    qty = int(sum(sum(int(it.get("quantity") or 0) for it in b["data"]["items"]) for b in built))
    pre = sum(1 for b in built for it in b["data"]["items"] if it.get("is_preorder") is True)

    check("A9 Σ omzet produk sama dengan berkas", rev_prod == truth["revenue_product"],
          f"Rp {rev_prod:,} (harap Rp {truth['revenue_product']:,})")
    check("A10 Σ harga coret sama dengan berkas", rev_gross == truth["revenue_gross"],
          f"Rp {rev_gross:,}")
    check("A11 Σ diskon penjual sama dengan berkas", sel_disc == truth["seller_discount"],
          f"Rp {sel_disc:,}")
    check("A12 Σ Order Amount TIDAK dijumlah antar baris (uang per pesanan)",
          ord_amt == truth["order_amount"], f"Rp {ord_amt:,} (harap Rp {truth['order_amount']:,})")
    check("A13 Σ qty sama dengan berkas", qty == truth["quantity"], f"{qty} pcs")
    check("A14 514 baris item pre-order terbaca", pre == truth["preorder_items"], f"{pre} item")

    statuses = Counter(b["data"].get("status") for b in built)
    check("A15 kamus status: 'Perlu dikirim' → paid",
          set(statuses) == {"paid"}, str(dict(statuses)))
    raws = Counter(b["data"].get("status_raw") for b in built)
    check("A16 nilai asli platform tetap disimpan (status_raw)",
          set(raws) == {"Perlu dikirim"}, str(dict(raws)))
    couriers = Counter(b["data"].get("courier") for b in built)
    check("A17 kamus kurir: J&T Express → jnt (kurir kosong tidak menghapus pesanan)",
          couriers.get("jnt", 0) > 500 and couriers.get("jne", 0) >= 1, str(dict(couriers)))
    channels = Counter(b["data"].get("order_channel") for b in built)
    check("A18 kamus kanal: LIVE→live · Videos→video · Product cards→product_card",
          set(channels) <= {"live", "video", "product_card", "search", "ads",
                            "affiliate", "campaign", "other"}, str(dict(channels)))
    mixed = [b for b in built if any("Kanal Pesanan berbeda" in w for w in b["warnings"])]
    check("A19 pesanan dengan kanal campur diberi PERINGATAN (bukan diam-diam)",
          len(mixed) >= 1, f"{len(mixed)} pesanan berkanal campur")

    fp1 = eng.format_fingerprint(headers)
    fp2 = eng.format_fingerprint(list(reversed(headers)))
    check("A20 sidik format stabil & peka susunan kolom",
          len(fp1) == 40 and fp1 != fp2, fp1[:12])

    guard = st.platform_guard
    pc = Counter(b["data"].get("purchase_channel") for b in built)
    check("A21 kolom penjaga platform terbaca (Purchase Channel)",
          guard == "purchase_channel" and set(pc) == {"TikTok"}, str(dict(pc)))
    return {"built": built, "mapping": mapping, "headers": headers}


# ══════════════════════════════════════════════════════════════════════════════
# BAGIAN B — API impor (HTTP nyata)
# ══════════════════════════════════════════════════════════════════════════════
def login() -> str:
    for attempt in range(3):
        r = requests.post(f"{BASE}/api/auth/login", json=ADMIN, timeout=30)
        if r.status_code == 200:
            j = r.json()
            return j.get("token") or j.get("access_token")
        time.sleep(6)
    raise SystemExit(f"login gagal: {r.status_code} {r.text[:200]}")


def part_b_api(truth: dict, token: str) -> dict:
    print("\n[B] API IMPOR — upload → pratinjau → commit → dedupe → rollback")
    H = {"Authorization": f"Bearer {token}"}
    HJ = {**H, "Content-Type": "application/json"}
    out: dict = {}

    accounts = requests.get(f"{BASE}/api/marketing/accounts", headers=HJ, timeout=60).json()
    acc = next((a for a in accounts if a.get("account_code") == ACCOUNT_CODE), None)
    shopee = next((a for a in accounts if (a.get("platform") == "shopee"
                                          and a.get("status") == "active")), None)
    if not check("B1 toko tujuan ada (TIKTOK-OUTFIT)", bool(acc), ACCOUNT_CODE):
        return out
    out["account"] = acc

    types = requests.get(f"{BASE}/api/marketing/data-import/source-types",
                         headers=HJ, timeout=60).json()
    lst = types.get("source_types") or types.get("types") or types
    keys = [t.get("key") for t in (lst if isinstance(lst, list) else [])]
    check("B2 jenis `marketplace_orders` muncul di daftar pilihan layar",
          "marketplace_orders" in keys, f"{len(keys)} jenis")

    def upload(account_id):
        with open(SAMPLE, "rb") as fh:
            return requests.post(
                f"{BASE}/api/marketing/data-import/upload", headers=H,
                files={"file": (os.path.basename(SAMPLE), fh,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                data={"source_type": "marketplace_orders", "account_id": account_id},
                timeout=180)

    r = upload(acc["id"])
    body = r.json() if r.headers.get("content-type", "").startswith("application/json") else {}
    if not check("B3 upload 200", r.status_code == 200, f"status={r.status_code} {str(body)[:200]}"):
        return out
    sess = body.get("session") or {}
    out["session_id"] = sess.get("id")
    check("B4 sesi mencatat 601 baris data (baris deskripsi tidak dihitung)",
          sess.get("total_rows") == truth["rows"], f"total_rows={sess.get('total_rows')}")
    check("B5 pratinjau melaporkan 559 pesanan", body.get("preview_total") == truth["orders"],
          f"preview_total={body.get('preview_total')}")
    summ = body.get("summary") or {}
    check("B6 0 baris galat di pratinjau", summ.get("error") == 0, str(summ))
    check("B7 sesi menyimpan sidik format (pemetaan bisa diingat)",
          bool(sess.get("format_fingerprint")), str(sess.get("format_fingerprint"))[:12])
    check("B8 sesi siap commit tanpa AI", sess.get("status") == "ready",
          f"status={sess.get('status')} ai_used={sess.get('ai_used')}")

    # ── berkas TikTok ke toko Shopee ⇒ DITOLAK ───────────────────────────────
    if shopee:
        r2 = upload(shopee["id"])
        b2 = r2.json() if r2.content else {}
        detail = str(b2.get("detail") or "")
        wrong_ok = (r2.status_code == 400 and "tiktok" in detail.lower())
        check("B9 berkas TikTok diunggah ke toko Shopee ⇒ ditolak dengan alasan platform",
              wrong_ok, f"status={r2.status_code} detail={detail[:140]}")
        if r2.status_code == 200:
            sid = ((b2.get("session") or {}).get("id"))
            if sid:
                requests.delete(f"{BASE}/api/marketing/data-import/sessions/{sid}",
                                headers=HJ, timeout=60)

    sid = out["session_id"]
    r = requests.post(f"{BASE}/api/marketing/data-import/sessions/{sid}/commit",
                      headers=HJ, json={"on_duplicate": "skip"}, timeout=300)
    cb = r.json() if r.content else {}
    if not check("B10 commit 200", r.status_code == 200, f"status={r.status_code} {str(cb)[:200]}"):
        return out
    check("B11 559 pesanan masuk · 0 ditolak",
          cb.get("inserted") == truth["orders"] and cb.get("rejected") == 0,
          f"inserted={cb.get('inserted')} rejected={cb.get('rejected')} "
          f"skipped={cb.get('skipped_duplicates')}")

    # ── angka di database ────────────────────────────────────────────────────
    agg = requests.get(f"{BASE}/api/marketing/orders?account_id={acc['id']}&limit=1",
                       headers=HJ, timeout=60)
    check("B12 endpoint daftar pesanan hidup", agg.status_code == 200, f"status={agg.status_code}")

    db = _db()
    q = {"account_id": acc["id"], "_import_session_id": sid}
    n_docs = db.marketing_orders.count_documents(q)
    check("B13 1 dokumen = 1 pesanan (559 dokumen, bukan 601)",
          n_docs == truth["orders"], f"{n_docs} dokumen")
    docs = list(db.marketing_orders.find(q, {"_id": 0}))
    s_rev = money(sum(d.get("revenue_product") or 0 for d in docs))
    s_ord = money(sum(d.get("order_amount") or 0 for d in docs))
    s_gross = money(sum(d.get("revenue_gross") or 0 for d in docs))
    s_disc = money(sum(d.get("seller_discount_total") or 0 for d in docs))
    s_qty = int(sum(d.get("quantity") or 0 for d in docs))
    n_items = sum(len(d.get("items") or []) for d in docs)
    check("B14 Σ revenue_product = Rp 59.783.811", s_rev == truth["revenue_product"],
          f"Rp {s_rev:,}")
    check("B15 Σ order_amount = Rp 62.805.113", s_ord == truth["order_amount"], f"Rp {s_ord:,}")
    check("B16 Σ revenue_gross = Rp 109.179.000", s_gross == truth["revenue_gross"],
          f"Rp {s_gross:,}")
    check("B17 Σ seller_discount_total = Rp 48.020.983", s_disc == truth["seller_discount"],
          f"Rp {s_disc:,}")
    check("B18 Σ quantity = 603 pcs", s_qty == truth["quantity"], f"{s_qty} pcs")
    check("B19 601 item tersimpan di items[]", n_items == truth["rows"], f"{n_items} item")
    check("B20 kompatibilitas pembaca lama: revenue = revenue_product & total_payment = order_amount",
          all(money(d.get("revenue")) == money(d.get("revenue_product"))
              and money(d.get("total_payment")) == money(d.get("order_amount")) for d in docs))
    check("B21 komisi platform TIDAK dikarang (platform_fee null, fee_known false)",
          all(d.get("platform_fee") is None and d.get("fee_known") is False for d in docs))
    linked = sum(1 for d in docs for it in (d.get("items") or []) if it.get("catalog_item_id"))
    out["linked_items"] = linked
    print(f"       (item tertaut katalog: {linked}/{n_items} — sisanya menunggu Pemetaan SKU)")

    # ── impor ulang berkas yang sama ⇒ 0 dokumen tambahan ────────────────────
    r = upload(acc["id"])
    sid2 = ((r.json() or {}).get("session") or {}).get("id")
    if sid2:
        r = requests.post(f"{BASE}/api/marketing/data-import/sessions/{sid2}/commit",
                          headers=HJ, json={"on_duplicate": "skip"}, timeout=300)
        cb2 = r.json() if r.content else {}
        total_now = db.marketing_orders.count_documents({"account_id": acc["id"]})
        check("B22 impor ulang berkas sama ⇒ 0 dokumen tambahan (dedupe order_id)",
              cb2.get("inserted") == 0 and cb2.get("skipped_duplicates") == truth["orders"],
              f"inserted={cb2.get('inserted')} skipped={cb2.get('skipped_duplicates')} "
              f"total_dokumen={total_now}")
        out["session_id_2"] = sid2
    return out


def _db():
    from dotenv import load_dotenv
    from pymongo import MongoClient
    load_dotenv("/app/backend/.env")
    cli = MongoClient(os.environ["MONGO_URL"], serverSelectionTimeoutMS=8000)
    return cli[os.environ.get("DB_NAME", "test_database")]


# ══════════════════════════════════════════════════════════════════════════════
# BAGIAN C — F2 rekap harian turunan (satu angka omzet di semua pintu)
# ══════════════════════════════════════════════════════════════════════════════
def part_c_rollup(truth: dict, token: str, ctx: dict):
    print("\n[C] F2 — REKAP HARIAN TURUNAN: satu angka omzet di semua pintu")
    acc = ctx.get("account")
    if not acc:
        fail("C0 toko tujuan tidak ada — bagian C dilewati")
        return
    HJ = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    db = _db()

    daily = list(db.marketing_sales_data.find(
        {"account_id": acc["id"], "revenue_type": "total"}, {"_id": 0}))
    derived = [d for d in daily if d.get("source") == "orders_auto"]
    check("C1 rekap harian TURUNAN dibuat otomatis sesudah impor",
          len(derived) == len(truth["dates"]),
          f"{len(derived)} dokumen (tanggal di berkas: {len(truth['dates'])})")
    s_rev = money(sum((d.get("metrics") or {}).get("revenue_product") or 0 for d in derived))
    check("C2 Σ metrics.revenue_product rekap harian = Rp 59.783.811",
          s_rev == truth["revenue_product"], f"Rp {s_rev:,}")
    s_orders = int(sum((d.get("metrics") or {}).get("orders") or 0 for d in derived))
    check("C3 Σ metrics.orders = 559 pesanan", s_orders == truth["orders"], f"{s_orders}")
    check("C4 dokumen turunan terkunci (locked_source) sehingga tidak bisa diketik ulang",
          all(d.get("locked_source") is True for d in derived))
    traffic_live = money(sum((d.get("traffic") or {}).get("live") or 0 for d in derived))
    check("C5 pecahan trafik per kanal ikut terisi (live > 0)", traffic_live > 0,
          f"live Rp {traffic_live:,}")

    year, month = 2026, 7
    r = requests.get(f"{BASE}/api/marketing/performance/overview"
                     f"?account_id={acc['id']}&date_from=2026-07-01&date_to=2026-07-31",
                     headers=HJ, timeout=90)
    perf = r.json() if r.status_code == 200 else {}
    pv = perf.get("data") or perf
    perf_rev = money(pv.get("total_revenue") or pv.get("revenue") or 0)
    perf_ord = int(pv.get("total_orders") or pv.get("orders") or 0)
    check("C6 /performance/overview memakai omzet produk yang sama",
          perf_rev == truth["revenue_product"],
          f"Rp {perf_rev:,} (harap Rp {truth['revenue_product']:,})")
    check("C7 /performance/overview menghitung 559 pesanan (bukan 601 baris)",
          perf_ord == truth["orders"], f"{perf_ord}")

    r = requests.get(f"{BASE}/api/marketing/targets/monthly-summary?year={year}&month={month}",
                     headers=HJ, timeout=90)
    ms = r.json() if r.status_code == 200 else {}
    msv = ms.get("data") or ms
    rows_acc = msv.get("accounts") or []
    mine = next((a for a in rows_acc if a.get("account_id") == acc["id"]), {})
    actual = money(((mine.get("actual") or {}).get("revenue")) or 0)
    check("C8 target bulanan Juli 2026 memakai angka yang sama",
          actual == truth["revenue_product"],
          f"actual.revenue=Rp {actual:,} (summary.rev_actual="
          f"Rp {money((msv.get('summary') or {}).get('rev_actual')):,})")

    r = requests.get(f"{BASE}/api/marketing/dashboard/overview", headers=HJ, timeout=90)
    check("C9 dashboard marketing tetap 200 (bukan 500)", r.status_code == 200,
          f"status={r.status_code}")

    # ── kunci entri manual + jalur override SPV ──────────────────────────────
    date0 = "2026-07-19"
    r = requests.post(f"{BASE}/api/marketing/sales-data", headers=HJ, timeout=60, json={
        "account_id": acc["id"], "date": date0, "revenue_type": "total",
        "revenue": 123456, "orders": 7})
    check("C10 entri manual ke tanggal turunan DITOLAK 409 dengan alasan jelas",
          r.status_code == 409, f"status={r.status_code} {str(r.text)[:140]}")
    r = requests.post(f"{BASE}/api/marketing/sales-data?override=true", headers=HJ, timeout=60, json={
        "account_id": acc["id"], "date": date0, "revenue_type": "total",
        "revenue": 123456, "orders": 7,
        "override_reason": "uji: koreksi manual oleh SPV"})
    ok_override = r.status_code in (200, 201)
    check("C11 SPV boleh MENGGANTI angka (override) dengan alasan tercatat",
          ok_override, f"status={r.status_code} {str(r.text)[:140]}")
    if ok_override:
        doc = db.marketing_sales_data.find_one(
            {"account_id": acc["id"], "date": date0, "revenue_type": "total"}, {"_id": 0})
        check("C12 dokumen override menyimpan jejak (source, alasan, oleh siapa)",
              (doc or {}).get("source") == "manual_override"
              and bool((doc or {}).get("override_reason")), str((doc or {}).get("source")))
        # kembalikan ke turunan
        r = requests.post(f"{BASE}/api/marketing/sales/recompute"
                          f"?account_id={acc['id']}&date_from={date0}&date_to={date0}"
                          f"&force=true", headers=HJ, timeout=90)
        doc = db.marketing_sales_data.find_one(
            {"account_id": acc["id"], "date": date0, "revenue_type": "total"}, {"_id": 0})
        check("C13 'Hitung Ulang (paksa)' memulihkan angka turunan dari pesanan",
              r.status_code == 200 and (doc or {}).get("source") == "orders_auto",
              f"status={r.status_code} source={(doc or {}).get('source')}")

    # ── rollback: angka harus ikut turun ────────────────────────────────────
    sid = ctx.get("session_id_2") or ctx.get("session_id")
    if sid:
        r = requests.post(f"{BASE}/api/marketing/data-import/sessions/{sid}/rollback",
                          headers=HJ, timeout=120)
        print(f"       (rollback sesi kedua: status={r.status_code})")
    sid1 = ctx.get("session_id")
    if sid1:
        r = requests.post(f"{BASE}/api/marketing/data-import/sessions/{sid1}/rollback",
                          headers=HJ, timeout=180)
        rb = r.json() if r.content else {}
        check("C14 rollback impor menghapus 559 pesanan sesi itu",
              r.status_code == 200 and rb.get("deleted") == truth["orders"],
              f"status={r.status_code} deleted={rb.get('deleted')}")
        left = db.marketing_orders.count_documents({"account_id": acc["id"]})
        check("C15 tidak ada pesanan sisa dari sesi yang di-rollback", left == 0,
              f"{left} pesanan tersisa")
        after = list(db.marketing_sales_data.find(
            {"account_id": acc["id"], "source": "orders_auto"}, {"_id": 0}))
        check("C16 rekap harian turunan ikut hilang (angka tidak 'nyangkut')",
              not after, f"{len(after)} dokumen tersisa")


def cleanup(acc_id: str | None):
    """Bersihkan jejak uji supaya data seed kembali seperti semula."""
    if not acc_id:
        return
    db = _db()
    db.marketing_orders.delete_many({"account_id": acc_id, "_import_source_type": "marketplace_orders"})
    db.marketing_sales_data.delete_many({"account_id": acc_id,
                                         "source": {"$in": ["orders_auto", "manual_override"]}})
    db.marketing_data_import_sessions.delete_many({"source_type": "marketplace_orders"})
    print("\n  (bersih-bersih: pesanan uji, rekap turunan, dan sesi impor uji dihapus)")


def main() -> int:
    print("=" * 90)
    print("CORE TEST F1+F2 — satu sumber omzet: impor pesanan → rekap harian turunan")
    print("=" * 90)
    if not os.path.exists(SAMPLE):
        print(f"Berkas contoh tidak ada: {SAMPLE}")
        return 1
    truth = truth_from_file()
    print(f"  kebenaran dari berkas: {truth['rows']} baris · {truth['orders']} pesanan · "
          f"Rp {truth['revenue_product']:,} · {len(truth['dates'])} tanggal")

    part_a_engine(truth)
    token = login()
    ctx = part_b_api(truth, token)
    part_c_rollup(truth, token, ctx)
    cleanup((ctx.get("account") or {}).get("id"))

    passed = sum(1 for _, p, _ in RESULTS if p)
    failed = [(n, d) for n, p, d in RESULTS if not p]
    print("\n" + "=" * 90)
    print(f"RINGKAS: {passed}/{len(RESULTS)} PASS")
    if failed:
        print("GAGAL:")
        for n, d in failed:
            print(f"  · {n} — {d}")
    print("=" * 90)
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
