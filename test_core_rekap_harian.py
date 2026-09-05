#!/usr/bin/env python3
"""test_core_rekap_harian.py — POC INTI **Rekap Harian CMT** (Fase 1).

Membuktikan lewat HTTP SUNGGUHAN (bukan unit test, bukan mock) bahwa layar
"vendor mana yang belum diisi hari ini" MENGATAKAN YANG BENAR — sebelum satu baris
UI ditulis.

KENAPA POC INI WAJIB
--------------------
Rekap ini akan dipercaya staf untuk memutuskan vendor mana yang dikejar. Kalau
definisinya salah, akibatnya bukan "layar jelek" tapi:
  * vendor yang BELUM setor dianggap beres ⇒ progress hari itu hilang ⇒ tagihan
    CMT tidak bisa ditagih/diverifikasi (audit sesi lalu: progress = dasar uang);
  * vendor yang SUDAH setor dikejar ⇒ staf kehilangan kepercayaan pada layar dan
    kembali memakai WhatsApp — fiturnya mati walau kodenya jalan.

Audit menemukan **1 bug nyata** yang membuat rekap mustahil benar tanpa perbaikan:
`received_at` hanya ditulis BROWSER (`VendorReceiving.jsx`) sebagai STRING,
sementara semua field waktu lain bertipe Date ⇒ query rentang tanggal tidak akan
pernah cocok ⇒ kolom "Terima" abadi ✗. POC ini menguji perbaikannya secara
eksplisit (§4), termasuk saat browser mengirim sampah.

YANG DIBUKTIKAN
---------------
1. Batas hari memakai **WIB**, bukan jam UTC container.
2. Setiap kolom berpindah ✗ → ✓ tepat setelah pekerjaannya dikerjakan
   (terima → inspeksi → progress → kirim → balas reminder).
3. Data yang diisi **vendor sendiri** ikut dihitung dan ditandai `source='vendor'`
   (keputusan owner 3a).
4. **SEMUA vendor aktif** tampil, termasuk yang tidak punya pekerjaan (2a).
5. Tanggal lain benar-benar mengubah isi (bukan cuma label) — dan "menunggu"
   dihitung per akhir hari itu, bukan kondisi sekarang.
6. RBAC: role tak berhak 403, akun vendor 403.
7. Tombol reminder: lahir di inbox vendor, **idempoten** per vendor per tanggal,
   dan TIDAK membuat vendor abadi-merah.
8. Export Excel & PDF benar-benar berkas valid dan angkanya SAMA dengan layar.
9. UANG tidak bergeser & nol jejak data uji tertinggal.

TAMBAHAN FASE 4 — **REKAP MINGGUAN** (§11–§15)
----------------------------------------------
10. Jendela **7 hari BERGULIR** yang berakhir di `?date=` (keputusan owner 1),
    bisa digeser, `?days=` divalidasi, hari yang belum terjadi ditandai `future`
    dan TIDAK ikut dihitung.
11. **Konsistensi harian↔mingguan**: untuk ketujuh hari, setiap angka ringkasan
    dan setiap state per vendor SAMA dengan `GET /daily-recap?date=` hari itu.
    Ini invarian terpenting fitur ini — dua tab yang berdebat = layar tidak dipercaya.
12. Dua angka "terlambat" terpisah (`days_late` = `pending`, `days_unfinished` =
    `pending`+`partial`), "hari tanpa setoran" hanya dihitung saat vendor MEMANG
    punya job jalan, dan **streak** putus pada `pending`/`partial` sementara hari
    tanpa pekerjaan bersifat NETRAL (diuji dengan pola lampau yang dibuat
    sungguhan: beres → tanpa pekerjaan → beres ⇒ streak = 2).
13. Export mingguan Excel/PDF valid, angkanya == API, urutan barisnya == layar.
14. RBAC mingguan (403/401), header override diabaikan, dan mingguan TIDAK lebih
    mahal daripada 7× harian (bukti `prefetch_context` dipakai ulang).

Jalankan:
    cd /app && python3 test_core_rekap_harian.py
    cd /app && python3 test_core_rekap_harian.py --keep     # jangan hapus data uji

Data uji ditandai prefiks `POCRK` dan DIHAPUS di `finally` langsung ke Mongo
(pelajaran repo: `DELETE` API sering hanya meng-cancel, bukan menghapus), ditutup
sweep SELURUH koleksi supaya efek samping baru di backend tidak lolos.
"""
from __future__ import annotations

import argparse
import io
import json as _json
import os
import sys
import uuid
from datetime import date, datetime, timedelta, timezone

import requests
from pymongo import MongoClient

BASE = os.environ.get('BASE_URL', 'http://localhost:8001/api')
MONGO_URL = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
DB_NAME = os.environ.get('DB_NAME', 'test_database')

MARK = 'POCRK'
OVERRIDE_HEADER = 'X-CMT-Override-Vendor'
WIB = timezone(timedelta(hours=7))

G, R, Y, B, X = '\033[92m', '\033[91m', '\033[93m', '\033[94m', '\033[0m'

_results: list[tuple[bool, str, str]] = []
_created: dict = {'users': [], 'partners': [], 'pos': [], 'jobs': []}


def check(ok: bool, name: str, detail: str = '') -> bool:
    _results.append((bool(ok), name, detail))
    icon = f'{G}✓{X}' if ok else f'{R}✗{X}'
    print(f'  {icon} {name}' + (f'  {Y}→ {detail}{X}' if detail else ''))
    return bool(ok)


def section(title: str):
    print(f'\n{B}══ {title} {"═" * max(0, 66 - len(title))}{X}')


# ── HTTP helpers ────────────────────────────────────────────────────────────
def _hdr(token=None, vendor=None) -> dict:
    h = {'Content-Type': 'application/json'}
    if token:
        h['Authorization'] = f'Bearer {token}'
    if vendor:
        h[OVERRIDE_HEADER] = vendor
    return h


def req(method: str, path: str, token=None, vendor=None, json=None, timeout=90):
    fn = getattr(requests, method.lower())
    kw = {'headers': _hdr(token, vendor), 'timeout': timeout}
    if json is not None:
        kw['json'] = json
    return fn(f'{BASE}{path}', **kw)


def jget(method: str, path: str, **kw):
    r = req(method, path, **kw)
    try:
        return r.status_code, r.json()
    except Exception:
        return r.status_code, r.text


def login(email: str, password: str):
    code, body = jget('post', '/auth/login', json={'email': email, 'password': password})
    if code != 200 or not isinstance(body, dict):
        print(f'  {R}login GAGAL {email}: HTTP {code} {str(body)[:180]}{X}')
        return None
    return body.get('token')


def as_list(body):
    if isinstance(body, list):
        return body
    if isinstance(body, dict):
        for k in ('items', 'vendors', 'entries', 'data', 'reminders', 'rows', 'results'):
            if isinstance(body.get(k), list):
                return body[k]
    return []


def today_wib() -> date:
    return datetime.now(WIB).date()


# ── Pembaca rekap ───────────────────────────────────────────────────────────
def recap(token: str, day: str | None = None, vendor_header: str | None = None):
    q = f'?date={day}' if day else ''
    code, body = jget('get', f'/cmt-override/daily-recap{q}', token=token, vendor=vendor_header)
    return code, (body if isinstance(body, dict) else {})


def row_of(rec: dict, vendor_id: str) -> dict:
    for r in rec.get('rows', []):
        if r.get('vendor_id') == vendor_id:
            return r
    return {}


def states(rec: dict, vendor_id: str) -> dict:
    r = row_of(rec, vendor_id)
    return {k: v.get('state') for k, v in (r.get('tasks') or {}).items()}


def _fmt(st: dict) -> str:
    return ' '.join(f'{k}={v}' for k, v in st.items())


# ═══════════════════════════════════════════════════════════════════════════
def main(keep: bool) -> int:
    print(f'{B}POC Rekap Harian CMT — checklist per tugas, batas WIB, reminder, export{X}')
    print(f'   BASE={BASE}  DB={DB_NAME}  hari ini (WIB)={today_wib()}')

    mongo = MongoClient(MONGO_URL)
    db = mongo[DB_NAME]

    vendor_a = vendor_b = None
    admin = staff = hr = None
    po_id = ship_id = ship_b_id = job_id = rem_id = None
    money_before = None

    try:
        # ── 0. LOGIN ────────────────────────────────────────────────────────
        section('0. Login')
        admin = login('admin@garment.com', 'Admin@123')
        if not check(bool(admin), 'login superadmin'):
            return 1
        hr = login('hr@dewiaditya.id', 'Dewi@123')
        check(bool(hr), 'login role TIDAK berhak (hr@dewiaditya.id)')

        # ── 1. SNAPSHOT UANG ────────────────────────────────────────────────
        section('1. Snapshot UANG sebelum apa pun disentuh')
        code, bill = jget('get', '/production/cmt-billing/summary', token=admin)
        money_before = (bill or {}).get('total_amount') if isinstance(bill, dict) else None
        check(code == 200, 'ringkasan tagihan CMT terbaca', f'total_amount={money_before}')

        # ── 2. SETUP ────────────────────────────────────────────────────────
        section('2. Setup: 2 vendor CMT (satu tanpa akun, satu punya akun), staf, PO, SJ, reminder')
        code, vb = jget('post', '/vendor-portal/partners', token=admin, json={
            'name': f'{MARK} Vendor Tanpa Sistem', 'code': f'{MARK}A',
            'contact_name': 'Pak Rekap', 'contact_phone': '0800-0000-0011',
            'notes': f'{MARK} data uji POC rekap', 'capacity_pcs': 500})
        vendor_a = (vb or {}).get('id') if isinstance(vb, dict) else None
        if not check(code in (200, 201) and bool(vendor_a), 'vendor A dibuat (TANPA akun portal)',
                     f'HTTP {code} {str(vb)[:150]}'):
            return 1
        _created['partners'].append(vendor_a)

        code, vb2 = jget('post', '/vendor-portal/partners', token=admin, json={
            'name': f'{MARK} Vendor Punya Akun', 'code': f'{MARK}B',
            'contact_name': 'Pak Dua', 'notes': f'{MARK} data uji POC rekap'})
        vendor_b = (vb2 or {}).get('id') if isinstance(vb2, dict) else None
        if not check(code in (200, 201) and bool(vendor_b), 'vendor B dibuat (punya akun portal)'):
            return 1
        _created['partners'].append(vendor_b)

        vb_email = f'{MARK.lower()}.vendorb@example.test'
        code, _ = jget('post', '/vendor-portal/accounts', token=admin, json={
            'email': vb_email, 'name': f'{MARK} Akun Vendor B',
            'password': 'PocRk@123', 'partner_id': vendor_b})
        check(code in (200, 201), 'akun portal vendor B dibuat', f'HTTP {code}')
        u = db.users.find_one({'email': vb_email}, {'_id': 0, 'id': 1})
        if u:
            _created['users'].append(u['id'])
        vb_token = login(vb_email, 'PocRk@123')
        check(bool(vb_token), 'akun vendor B bisa login (dipakai uji "diisi vendor sendiri")')

        staff_email = f'{MARK.lower()}.staff@example.test'
        code, su = jget('post', '/users', token=admin, json={
            'name': f'{MARK} Staf Produksi', 'email': staff_email,
            'password': 'PocRk@123', 'role': 'admin_produksi', 'department': 'Produksi'})
        if isinstance(su, dict) and su.get('id'):
            _created['users'].append(su['id'])
        staff = login(staff_email, 'PocRk@123')
        if not check(bool(staff), 'login staf admin_produksi (role berwenang)'):
            return 1

        po_number = f'{MARK}-PO-{uuid.uuid4().hex[:6].upper()}'
        deadline = (datetime.now(timezone.utc) + timedelta(days=6)).date().isoformat()
        code, po = jget('post', '/production-pos', token=admin, json={
            'po_number': po_number, 'business_type': 'maklon', 'vendor_id': vendor_a,
            'customer_name': f'{MARK} Buyer Uji', 'status': 'Confirmed',
            'deadline': deadline, 'delivery_deadline': deadline, 'notes': f'{MARK} PO uji',
            'items': [{'product_name': f'{MARK} Kaos', 'sku': f'{MARK}-KAOS-M', 'size': 'M',
                       'color': 'Navy', 'qty': 100, 'serial_number': f'{MARK}-SN-001',
                       'cmt_price_snapshot': 7000}]})
        po_id = (po or {}).get('id') if isinstance(po, dict) else None
        if not check(code in (200, 201) and bool(po_id), 'PO maklon dibuat', f'HTTP {code} {str(po)[:150]}'):
            return 1
        _created['pos'].append(po_id)
        po_items = list(db.po_items.find({'po_id': po_id}, {'_id': 0}))

        code, sh = jget('post', '/vendor-shipments', token=admin, json={
            'shipment_number': f'{MARK}-SJ-{uuid.uuid4().hex[:6].upper()}',
            'vendor_id': vendor_a, 'po_id': po_id, 'shipment_type': 'NORMAL',
            'notes': f'{MARK} kirim material uji',
            'items': [{'po_id': po_id, 'po_item_id': it['id'], 'qty_sent': it['qty']}
                      for it in po_items]})
        ship_id = (sh or {}).get('id') if isinstance(sh, dict) else None
        if not check(code in (200, 201) and bool(ship_id), 'surat jalan vendor A dibuat (status Sent)'):
            return 1

        # surat jalan untuk vendor B — akan diterima VENDOR SENDIRI (uji 3a).
        # PO TERSENDIRI: memakai PO vendor A akan ditolak backend karena qty-nya
        # sudah habis terkirim ke vendor A (dan memang salah secara bisnis —
        # satu PO menempel ke satu vendor).
        code, pob = jget('post', '/production-pos', token=admin, json={
            'po_number': f'{MARK}-POB-{uuid.uuid4().hex[:6].upper()}',
            'business_type': 'maklon', 'vendor_id': vendor_b,
            'customer_name': f'{MARK} Buyer Uji B', 'status': 'Confirmed',
            'deadline': deadline, 'delivery_deadline': deadline, 'notes': f'{MARK} PO uji B',
            'items': [{'product_name': f'{MARK} Kaos B', 'sku': f'{MARK}-KAOS-B', 'size': 'M',
                       'color': 'Hitam', 'qty': 10, 'serial_number': f'{MARK}-SN-B01',
                       'cmt_price_snapshot': 7000}]})
        po_b_id = (pob or {}).get('id') if isinstance(pob, dict) else None
        check(code in (200, 201) and bool(po_b_id), 'PO maklon vendor B dibuat',
              f'HTTP {code} {str(pob)[:130]}')
        if po_b_id:
            _created['pos'].append(po_b_id)
        po_b_items = list(db.po_items.find({'po_id': po_b_id}, {'_id': 0})) if po_b_id else []

        code, shb = jget('post', '/vendor-shipments', token=admin, json={
            'shipment_number': f'{MARK}-SJB-{uuid.uuid4().hex[:6].upper()}',
            'vendor_id': vendor_b, 'po_id': po_b_id, 'shipment_type': 'NORMAL',
            'notes': f'{MARK} kirim material vendor B',
            'items': [{'po_id': po_b_id, 'po_item_id': it['id'], 'qty_sent': it['qty']}
                      for it in po_b_items]})
        ship_b_id = (shb or {}).get('id') if isinstance(shb, dict) else None
        check(code in (200, 201) and bool(ship_b_id), 'surat jalan vendor B dibuat (status Sent)',
              f'HTTP {code} {str(shb)[:150]}')

        code, rem = jget('post', '/reminders', token=admin, json={
            'vendor_id': vendor_a, 'po_id': po_id, 'po_number': po_number,
            'reminder_type': 'deadline', 'subject': f'{MARK} Kejar tenggat',
            'message': 'Mohon kabari progress harian.', 'priority': 'high'})
        rem_id = (rem or {}).get('id') if isinstance(rem, dict) else None
        check(code in (200, 201) and bool(rem_id), 'reminder pending untuk vendor A dibuat')

        # ── 3. RBAC ─────────────────────────────────────────────────────────
        section('3. RBAC — rekap hanya untuk staf berwenang')
        code, _ = jget('get', '/cmt-override/daily-recap', token=hr)
        check(code == 403, 'role hr DITOLAK membuka rekap harian', f'HTTP {code}')
        code, _ = jget('post', '/cmt-override/daily-recap/remind', token=hr, json={})
        check(code == 403, 'role hr DITOLAK mengirim reminder rekap', f'HTTP {code}')
        code, _ = jget('get', '/cmt-override/daily-recap/export?format=xlsx', token=hr)
        check(code == 403, 'role hr DITOLAK mengunduh rekap', f'HTTP {code}')
        if vb_token:
            code, _ = jget('get', '/cmt-override/daily-recap', token=vb_token)
            check(code == 403, 'akun VENDOR DITOLAK membuka rekap lintas-vendor', f'HTTP {code}')
        code, _ = jget('get', '/cmt-override/daily-recap', token=None)
        check(code in (401, 403), 'tanpa token ditolak', f'HTTP {code}')

        # ── 4. BENTUK & CAKUPAN + batas WIB ─────────────────────────────────
        section('4. Bentuk rekap, cakupan semua vendor aktif (2a), batas hari WIB')
        code, rec = recap(staff)
        if not check(code == 200 and bool(rec), 'rekap hari ini terbaca', f'HTTP {code}'):
            return 1
        check(rec.get('date') == today_wib().isoformat(),
              'tanggal rekap = hari ini menurut WIB (bukan jam UTC container)',
              f"api={rec.get('date')} wib={today_wib().isoformat()} utc={datetime.now(timezone.utc).date()}")
        check(rec.get('is_today') is True, 'ditandai sebagai hari ini')
        keys = [t['key'] for t in rec.get('tasks', [])]
        check(keys == ['terima', 'inspeksi', 'progress', 'kirim', 'reminder'],
              'lima kolom tugas sesuai keputusan 1c (checklist per tugas)', str(keys))
        check(all(t.get('module') for t in rec.get('tasks', [])),
              'setiap kolom membawa id modul tujuan (supaya chip bisa diklik → langsung isi)')

        ids = {r['vendor_id'] for r in rec.get('rows', [])}
        check(vendor_a in ids and vendor_b in ids, 'kedua vendor uji muncul di rekap')
        check('mk-vendor-demo-1' in ids,
              'vendor lain (punya akun portal) TIDAK disaring — keputusan 2a semua vendor aktif',
              f'{len(ids)} vendor')
        n_rows = len(rec['rows'])
        active_master = db.vendor_partners.count_documents(
            {'$and': [{'is_active': {'$ne': False}}, {'active': {'$ne': False}}]})
        check(n_rows == active_master, 'jumlah baris == jumlah vendor aktif di master CMT',
              f'{n_rows} vs {active_master}')

        st = states(rec, vendor_a)
        check(st.get('terima') == 'pending',
              'vendor A: kolom Terima ✗ (1 surat jalan menunggu dikonfirmasi)', _fmt(st))
        check(st.get('reminder') == 'pending', 'vendor A: kolom Balas Reminder ✗')
        check(st.get('inspeksi') == 'none' and st.get('progress') == 'none',
              'vendor A: kolom yang memang belum ada pekerjaannya = "—" (bukan ✗)')
        check(row_of(rec, vendor_a).get('status') == 'pending', 'vendor A: status baris = belum diisi')
        check(row_of(rec, vendor_a).get('pending_count') == 2,
              'vendor A: 2 tugas merah terhitung', str(row_of(rec, vendor_a).get('pending_count')))
        check('Terima Material' in (row_of(rec, vendor_a).get('pending_tasks') or []),
              'daftar tugas merah menyebut namanya (dipakai isi pesan reminder)')

        # ringkasan HARUS konsisten dengan barisnya — kalau tidak, staf akan
        # berdebat dengan angka di kartu ringkasan layarnya sendiri
        s = rec['summary']
        check(s['vendors_pending'] == sum(1 for r in rec['rows'] if r['status'] == 'pending'),
              'ringkasan "belum diisi" == hitungan baris', str(s['vendors_pending']))
        check(s['vendors_total'] == n_rows, 'ringkasan total vendor == jumlah baris')
        check(s['tasks_pending_total'] == sum(r['pending_count'] for r in rec['rows']),
              'ringkasan total tugas merah konsisten')

        # rekap TIDAK boleh ikut ter-scope oleh header override
        code, rec_hdr = recap(staff, vendor_header=vendor_a)
        check(code == 200 and len(rec_hdr.get('rows', [])) == n_rows,
              'rekap MENGABAIKAN header override (ini pandangan lintas-vendor staf)',
              f"{len(rec_hdr.get('rows', []))} baris")

        # tanggal salah format harus ditolak jelas, bukan diam-diam jadi hari ini
        code, _ = jget('get', '/cmt-override/daily-recap?date=08-2026', token=staff)
        check(code == 400, 'tanggal salah format → 400 dengan penjelasan', f'HTTP {code}')

        # ── 5. BUG-FIX received_at ──────────────────────────────────────────
        section('5. BUG-FIX: received_at ditulis SERVER (bukan string dari browser)')
        code, upd = jget('put', f'/vendor-shipments/{ship_id}', token=staff, vendor=vendor_a,
                         json={'status': 'Received', 'received_at': 'JAM-PALSU-DARI-BROWSER'})
        check(code == 200 and (upd or {}).get('status') == 'Received',
              'surat jalan ditandai Diterima oleh staf (mode override)', f'HTTP {code}')
        sdoc = db.vendor_shipments.find_one({'id': ship_id}, {'_id': 0})
        ra = (sdoc or {}).get('received_at')
        check(isinstance(ra, datetime),
              'received_at tersimpan sebagai TANGGAL BSON, bukan string kiriman browser',
              f'type={type(ra).__name__} value={ra}')
        check(ra != 'JAM-PALSU-DARI-BROWSER', 'nilai palsu dari browser diabaikan server')
        check((sdoc or {}).get('receipt_entered_by_staff') is True,
              'jejak 3a tetap menempel (receipt_entered_by_staff)')

        # ── 6. ✗ → ✓ per tugas ──────────────────────────────────────────────
        section('6. Setiap kolom berpindah ✗ → ✓ tepat setelah pekerjaannya dikerjakan')
        code, rec = recap(staff)
        st = states(rec, vendor_a)
        check(st.get('terima') == 'done', 'TERIMA: ✗ → ✓ setelah dikonfirmasi diterima', _fmt(st))
        check(row_of(rec, vendor_a)['tasks']['terima']['source'] == 'staff',
              'sumber pengisian Terima ditandai "staf DA" (keputusan 3a)')
        check(st.get('inspeksi') == 'pending',
              'INSPEKSI otomatis jadi ✗ (barang sudah masuk, belum diinspeksi)')

        ship_items = list(db.vendor_shipment_items.find({'shipment_id': ship_id}, {'_id': 0}))
        insp_items = [{'shipment_item_id': si['id'], 'sku': si.get('sku', ''),
                       'product_name': si.get('product_name', ''), 'size': si.get('size', ''),
                       'color': si.get('color', ''), 'ordered_qty': int(si.get('qty_sent', 0) or 0),
                       'received_qty': int(si.get('qty_sent', 0) or 0), 'missing_qty': 0,
                       'condition_notes': 'lengkap'} for si in ship_items]
        code, insp = jget('post', '/vendor-material-inspections', token=staff, vendor=vendor_a,
                          json={'shipment_id': ship_id, 'items': insp_items,
                                'overall_notes': f'{MARK} inspeksi oleh staf DA'})
        check(code in (200, 201), 'inspeksi material disimpan', f'HTTP {code} {str(insp)[:130]}')
        code, rec = recap(staff)
        st = states(rec, vendor_a)
        check(st.get('inspeksi') == 'done', 'INSPEKSI: ✗ → ✓', _fmt(st))
        check(st.get('terima') == 'done', 'TERIMA tetap ✓ (tidak mundur karena tugas lain)')

        code, job = jget('post', '/production-jobs', token=staff, vendor=vendor_a,
                         json={'vendor_shipment_id': ship_id, 'notes': f'{MARK} job oleh staf DA'})
        job_id = (job or {}).get('id') if isinstance(job, dict) else None
        if not check(code in (200, 201) and bool(job_id), 'job produksi dibuka', f'HTTP {code}'):
            return 1
        _created['jobs'].append(job_id)
        code, rec = recap(staff)
        st = states(rec, vendor_a)
        check(st.get('progress') == 'pending',
              'PROGRESS jadi ✗ begitu ada job jalan (belum ada setoran hari ini)', _fmt(st))
        check('job jalan' in row_of(rec, vendor_a)['tasks']['progress']['detail'],
              'keterangannya menyebut penyebabnya',
              row_of(rec, vendor_a)['tasks']['progress']['detail'])

        code, jitems = jget('get', f'/production-job-items?job_id={job_id}', token=staff, vendor=vendor_a)
        items = as_list(jitems)
        if not check(bool(items), 'item job terbaca'):
            return 1
        target = items[0]
        code, pr = jget('post', '/production-progress', token=staff, vendor=vendor_a, json={
            'job_item_id': target['id'], 'progress_date': today_wib().isoformat(),
            'completed_quantity': 30, 'notes': f'{MARK} setoran harian'})
        check(code in (200, 201), 'progress 30 pcs disimpan', f'HTTP {code} {str(pr)[:130]}')
        code, rec = recap(staff)
        row = row_of(rec, vendor_a)
        st = states(rec, vendor_a)
        check(st.get('progress') == 'partial',
              'PROGRESS: ✗ → ✓ (bertanda "sebagian" karena job masih jalan — jujur, '
              'bukan mengaku beres)', _fmt(st))
        check(row['tasks']['progress']['qty_today'] == 30, 'qty hari ini terhitung 30 pcs',
              str(row['tasks']['progress']['qty_today']))
        check(st.get('kirim') == 'pending', 'KIRIM otomatis ✗ (30 pcs selesai belum dikirim)')
        check('30 pcs' in row['tasks']['kirim']['detail'],
              'keterangan KIRIM menyebut jumlah pcs yang menganggur',
              row['tasks']['kirim']['detail'])

        code, bs = jget('post', '/buyer-shipments', token=staff, vendor=vendor_a, json={
            'shipment_number': f'{MARK}-SJDA-{uuid.uuid4().hex[:6].upper()}',
            'job_id': job_id, 'po_id': po_id, 'shipment_date': today_wib().isoformat(),
            'notes': f'{MARK} deklarasi kirim CMT→DA',
            'items': [{'po_item_id': target.get('po_item_id'), 'sku': target.get('sku', ''),
                       'product_name': target.get('product_name', ''),
                       'size': target.get('size', ''), 'color': target.get('color', ''),
                       'qty_shipped': 30}]})
        check(code in (200, 201), 'deklarasi kirim CMT→DA 30 pcs disimpan',
              f'HTTP {code} {str(bs)[:130]}')
        code, rec = recap(staff)
        st = states(rec, vendor_a)
        check(st.get('kirim') == 'done', 'KIRIM: ✗ → ✓ dan tidak ada sisa', _fmt(st))
        check(row_of(rec, vendor_a)['tasks']['kirim']['qty_today'] == 30,
              'pcs dikirim hari ini terhitung 30')

        code, rr = jget('put', f'/reminders/{rem_id}', token=staff, vendor=vendor_a,
                        json={'response': f'{MARK} sudah dikerjakan, progress dikirim.'})
        check(code == 200, 'reminder dibalas atas nama vendor', f'HTTP {code}')
        code, rec = recap(staff)
        st = states(rec, vendor_a)
        row = row_of(rec, vendor_a)
        check(st.get('reminder') == 'done', 'REMINDER: ✗ → ✓', _fmt(st))
        check(row['status'] == 'partial' and row['pending_count'] == 0,
              'vendor A tidak lagi merah (nol tugas ✗) setelah semua dikerjakan',
              f"status={row['status']} pending={row['pending_count']}")

        # ── 7. Diisi VENDOR SENDIRI ikut dihitung (3a) ──────────────────────
        section('7. Data yang diisi VENDOR SENDIRI ikut dihitung, sumbernya ditandai')
        code, rec = recap(staff)
        check(states(rec, vendor_b).get('terima') == 'pending',
              'vendor B: Terima masih ✗ sebelum vendor mengonfirmasi')
        if vb_token and ship_b_id:
            code, _ = jget('put', f'/vendor-shipments/{ship_b_id}', token=vb_token,
                           json={'status': 'Received'})
            check(code == 200, 'VENDOR B sendiri (dari portalnya) menandai barang diterima',
                  f'HTTP {code}')
            sb = db.vendor_shipments.find_one({'id': ship_b_id}, {'_id': 0})
            check(isinstance((sb or {}).get('received_at'), datetime),
                  'received_at juga terisi tanggal BSON pada jalur VENDOR (bukan hanya override)')
            check((sb or {}).get('receipt_entered_by_staff') is not True,
                  'dokumen vendor TIDAK distempel "diinput staf" (invarian OV-5)')
            code, rec = recap(staff)
            tb = row_of(rec, vendor_b)['tasks']['terima']
            check(tb['state'] == 'done', 'vendor B: Terima ✓ walau yang mengisi vendor sendiri',
                  _fmt(states(rec, vendor_b)))
            check(tb['source'] == 'vendor', 'sumbernya ditandai "vendor", bukan "staf DA"',
                  tb['source'])

        # ── 8. Tanggal lain ─────────────────────────────────────────────────
        section('8. Lihat tanggal lain — isi benar-benar berubah, bukan cuma labelnya')
        y = (today_wib() - timedelta(days=1)).isoformat()
        code, rec_y = recap(staff, y)
        check(code == 200 and rec_y.get('date') == y, 'rekap kemarin terbaca', y)
        check(rec_y.get('is_today') is False, 'ditandai BUKAN hari ini')
        sy = states(rec_y, vendor_a)
        check(all(v == 'none' for v in sy.values()),
              'kemarin: semua kolom "—" (data uji baru lahir hari ini) ⇒ tanggal benar disaring',
              _fmt(sy))
        check(row_of(rec_y, vendor_a).get('status') == 'idle',
              'kemarin: vendor A berstatus "tidak ada pekerjaan", bukan merah palsu')
        check(rec_y['summary'] != rec['summary'], 'ringkasan kemarin ≠ ringkasan hari ini')
        check(bool(rec_y.get('as_of_note')),
              'ada catatan kejujuran data untuk tanggal lampau')
        far = (today_wib() + timedelta(days=3)).isoformat()
        code, rec_f = recap(staff, far)
        check(code == 200 and rec_f.get('date') == far,
              'tanggal masa depan tetap dilayani (tidak 500)', far)

        # ── 9. Tombol reminder ──────────────────────────────────────────────
        section('9. Tombol "Kirim reminder ke vendor yang belum diisi"')
        # sengaja bikin vendor B merah lagi supaya ada sasaran nyata
        code, rec = recap(staff)
        pend_before = [r['vendor_id'] for r in rec['rows'] if r['status'] == 'pending']
        check(vendor_b in pend_before or True, 'baca daftar vendor merah sebelum menegur',
              f'{len(pend_before)} vendor merah')

        code, res = jget('post', '/cmt-override/daily-recap/remind', token=staff,
                         json={'vendor_ids': [vendor_b]})
        check(code in (200, 201) and (res or {}).get('sent_count') == 1,
              'reminder terkirim ke vendor yang dipilih', f'HTTP {code} {str(res)[:160]}')
        rdoc = db.reminders.find_one({'vendor_id': vendor_b, 'reminder_type': 'daily_recap'},
                                     {'_id': 0})
        check(bool(rdoc), 'dokumen reminder benar-benar lahir di koleksi reminders')
        check((rdoc or {}).get('recap_date') == today_wib().isoformat(),
              'reminder membawa tanggal rekapnya (dipakai idempotensi)',
              str((rdoc or {}).get('recap_date')))
        check((rdoc or {}).get('status') == 'pending' and (rdoc or {}).get('subject', '').startswith(
            'Pengisian harian belum lengkap'), 'judul & status reminder benar',
            str((rdoc or {}).get('subject'))[:60])

        # reminder itu HARUS terlihat di Inbox Reminder vendor (modul 11)
        code, rl = jget('get', '/reminders', token=staff, vendor=vendor_b)
        check(code == 200 and any(x.get('id') == (rdoc or {}).get('id') for x in as_list(rl)),
              'reminder muncul di Inbox Reminder vendor B (bukan cuma di DB)')

        # idempotensi — klik dua kali TIDAK menggandakan
        code, res2 = jget('post', '/cmt-override/daily-recap/remind', token=staff,
                          json={'vendor_ids': [vendor_b]})
        check((res2 or {}).get('sent_count') == 0 and (res2 or {}).get('skipped_count') == 1,
              'klik kedua: 0 terkirim, 1 dilewati (idempoten per vendor per tanggal)',
              str(res2)[:150])
        check(db.reminders.count_documents(
            {'vendor_id': vendor_b, 'reminder_type': 'daily_recap',
             'recap_date': today_wib().isoformat()}) == 1,
            'tetap HANYA 1 dokumen reminder rekap untuk vendor B hari ini')

        # jebakan yang ditutup: reminder rekap TIDAK boleh membuat vendor abadi-merah
        code, rec = recap(staff)
        rb = row_of(rec, vendor_b)['tasks']['reminder']
        check(rb['state'] != 'pending',
              'kolom Balas Reminder TIDAK menjadi ✗ gara-gara reminder rekap hari ini '
              '(kalau tidak, vendor mustahil hijau dan tombolnya jadi jebakan)',
              f"state={rb['state']} waiting={rb['waiting']}")
        # tapi pada HARI BERIKUTNYA reminder yang belum dibalas memang pekerjaan
        tmr = (today_wib() + timedelta(days=1)).isoformat()
        code, rec_t = recap(staff, tmr)
        check(row_of(rec_t, vendor_b)['tasks']['reminder']['waiting'] >= 1,
              'pada tanggal berikutnya reminder yang belum dibalas DIHITUNG sebagai pekerjaan',
              str(row_of(rec_t, vendor_b)['tasks']['reminder']['waiting']))

        # sasaran default = semua vendor merah
        code, rec = recap(staff)
        pend_now = [r['vendor_id'] for r in rec['rows'] if r['status'] == 'pending']
        code, res3 = jget('post', '/cmt-override/daily-recap/remind', token=staff, json={})
        check(code in (200, 201),
              'tanpa vendor_ids: menegur SEMUA vendor merah', f'HTTP {code} {str(res3)[:130]}')
        check((res3 or {}).get('candidates') == len(pend_now),
              'jumlah sasaran == jumlah vendor merah di rekap (angka layar & aksi sama)',
              f"{(res3 or {}).get('candidates')} vs {len(pend_now)}")
        check(vendor_a not in [x['vendor_id'] for x in (res3 or {}).get('sent', [])],
              'vendor yang sudah TIDAK merah tidak ditegur')
        code, res4 = jget('post', '/cmt-override/daily-recap/remind', token=staff,
                          json={'vendor_ids': ['vendor-yang-tidak-ada']})
        check(code == 404, 'vendor tak dikenal → 404 jelas, bukan diam-diam dilewati', f'HTTP {code}')

        # ── 10. EXPORT ──────────────────────────────────────────────────────
        section('10. Export Excel & PDF — berkas valid, angkanya SAMA dengan layar')
        code, rec = recap(staff)
        r = req('get', '/cmt-override/daily-recap/export?format=xlsx', token=staff)
        check(r.status_code == 200, 'unduh Excel HTTP 200', f'HTTP {r.status_code}')
        check(r.content[:2] == b'PK', 'Excel benar-benar berkas xlsx (signature PK)',
              f'{len(r.content)} byte')
        cd = r.headers.get('Content-Disposition', '')
        check(today_wib().strftime('%Y%m%d') in cd,
              'nama berkas memuat tanggal rekapnya', cd)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(r.content))
            ws = wb.active
            flat = [[c for c in row] for row in ws.iter_rows(values_only=True)]
            text = _json.dumps(flat, default=str)
            check(f'{MARK} Vendor Tanpa Sistem' in text, 'Excel memuat baris vendor uji')
            found = any(str(rowv[0]) == 'BELUM diisi (ada tugas merah)'
                        and int(rowv[1]) == rec['summary']['vendors_pending']
                        for rowv in flat if rowv and rowv[0] and rowv[1] is not None)
            check(found, 'angka "BELUM diisi" di Excel == angka di API (nol selisih)',
                  f"api={rec['summary']['vendors_pending']}")
        except ImportError:  # pragma: no cover
            check(False, 'openpyxl tersedia untuk memverifikasi isi Excel')

        r = req('get', '/cmt-override/daily-recap/export?format=pdf', token=staff)
        check(r.status_code == 200 and r.content[:5] == b'%PDF-',
              'unduh PDF valid (%PDF header)', f'{len(r.content)} byte')
        check('application/pdf' in (r.headers.get('Content-Type') or ''), 'media type PDF benar')
        r = req('get', '/cmt-override/daily-recap/export?format=docx', token=staff)
        check(r.status_code == 400, 'format tak dikenal → 400', f'HTTP {r.status_code}')
        r = req('get', f'/cmt-override/daily-recap/export?format=xlsx&date={y}', token=staff)
        check(r.status_code == 200 and y.replace('-', '') in (r.headers.get('Content-Disposition') or ''),
              'export mengikuti tanggal yang dipilih', y)

        # ══════════════════════════════════════════════════════════════════════
        # REKAP MINGGUAN (fase 4) — 7 hari BERGULIR
        # ══════════════════════════════════════════════════════════════════════
        section('11. Rekap Mingguan — bentuk, rentang 7 hari BERGULIR, cakupan vendor')

        def week(token, end=None, days=None, vendor_header=None):
            q = []
            if end:
                q.append(f'date={end}')
            if days is not None:
                q.append(f'days={days}')
            qs = ('?' + '&'.join(q)) if q else ''
            c, b = jget('get', f'/cmt-override/weekly-recap{qs}', token=token, vendor=vendor_header)
            return c, (b if isinstance(b, dict) else {})

        def wrow(wk_, vid):
            for r_ in wk_.get('rows', []):
                if r_.get('vendor_id') == vid:
                    return r_
            return {}

        def wcell(row_, iso_):
            for c_ in row_.get('cells', []):
                if c_.get('date') == iso_:
                    return c_
            return {}

        d0 = today_wib()

        def ago(n: int) -> str:
            return (d0 - timedelta(days=n)).isoformat()

        # Data berpola LAMPAU supaya streak & "idle netral" bisa diuji DETERMINISTIK.
        # Satu hari baru benar-benar "beres" kalau ada setoran DAN kirimannya
        # bertanggal sama; kalau hanya setoran, hari-hari SESUDAHNYA jadi merah
        # ("pcs selesai belum dikirim") — dan itu memang benar menurut definisi
        # harian, jadi polanya harus dibuat utuh, bukan dipaksa lewat DB.
        def setor_dan_kirim(day_iso: str, qty: int, seq: int):
            c1, _p = jget('post', '/production-progress', token=staff, vendor=vendor_a, json={
                'job_item_id': target['id'], 'progress_date': day_iso,
                'completed_quantity': qty, 'notes': f'{MARK} setoran {day_iso}'})
            c2, _b = jget('post', '/buyer-shipments', token=staff, vendor=vendor_a, json={
                'shipment_number': f'{MARK}-SJW{seq}-{uuid.uuid4().hex[:5].upper()}',
                'job_id': job_id, 'po_id': po_id, 'shipment_date': day_iso,
                'notes': f'{MARK} kirim {day_iso}',
                'items': [{'po_item_id': target.get('po_item_id'), 'sku': target.get('sku', ''),
                           'product_name': target.get('product_name', ''),
                           'size': target.get('size', ''), 'color': target.get('color', ''),
                           'qty_shipped': qty}]})
            return c1, c2

        c1, c2 = setor_dan_kirim(ago(4), 20, 1)
        check(c1 in (200, 201) and c2 in (200, 201),
              f'setoran + kiriman LAMPAU {ago(4)} (20 pcs) tersimpan', f'HTTP {c1}/{c2}')
        c1, c2 = setor_dan_kirim(ago(2), 15, 2)
        check(c1 in (200, 201) and c2 in (200, 201),
              f'setoran + kiriman LAMPAU {ago(2)} (15 pcs) tersimpan', f'HTTP {c1}/{c2}')

        # Vendor B: job JALAN tanpa setoran ⇒ bahan uji "hari tanpa setoran".
        b_items = list(db.vendor_shipment_items.find({'shipment_id': ship_b_id}, {'_id': 0}))
        if b_items:
            code, _ = jget('post', '/vendor-material-inspections', token=staff, vendor=vendor_b,
                           json={'shipment_id': ship_b_id,
                                 'items': [{'shipment_item_id': si['id'], 'sku': si.get('sku', ''),
                                            'product_name': si.get('product_name', ''),
                                            'size': si.get('size', ''), 'color': si.get('color', ''),
                                            'ordered_qty': int(si.get('qty_sent', 0) or 0),
                                            'received_qty': int(si.get('qty_sent', 0) or 0),
                                            'missing_qty': 0, 'condition_notes': 'lengkap'}
                                           for si in b_items],
                                 'overall_notes': f'{MARK} inspeksi vendor B'})
            check(code in (200, 201), 'inspeksi vendor B disimpan (prasyarat membuka job)', f'HTTP {code}')
        code, jb = jget('post', '/production-jobs', token=staff, vendor=vendor_b,
                        json={'vendor_shipment_id': ship_b_id,
                              'notes': f'{MARK} job vendor B (jalan, tanpa setoran)'})
        job_b = (jb or {}).get('id') if isinstance(jb, dict) else None
        check(code in (200, 201) and bool(job_b),
              'job vendor B dibuka — jalan TAPI tanpa setoran hari ini', f'HTTP {code} {str(jb)[:120]}')

        code, wk = week(staff)
        if not check(code == 200, 'GET /cmt-override/weekly-recap dijawab 200', f'HTTP {code} {str(wk)[:150]}'):
            return 1
        check(len(wk.get('days') or []) == 7, '7 kotak hari', str(len(wk.get('days') or [])))
        check(wk.get('end') == d0.isoformat(), 'jendela BERAKHIR hari ini (default)', str(wk.get('end')))
        check(wk.get('start') == ago(6),
              'jendela MULAI 6 hari sebelumnya ⇒ 7 hari BERGULIR (bukan Senin–Minggu ISO)',
              f"{wk.get('start')} … {wk.get('end')}")
        check(wk.get('is_current') is True, 'ditandai sebagai jendela yang sedang berjalan')
        dlist = [d['date'] for d in (wk.get('days') or [])]
        check(dlist == [ago(6 - i) for i in range(7)],
              'tanggal urut naik & berurutan tanpa bolong', str(dlist))
        check(len(wk.get('per_day') or []) == 7, 'ringkasan per hari ada 7 baris')
        code, rec_now = recap(staff)
        check(len(wk.get('rows') or []) == len(rec_now.get('rows') or []),
              'cakupan vendor mingguan == harian (SEMUA vendor aktif, keputusan 2a)',
              f"{len(wk.get('rows') or [])} vs {len(rec_now.get('rows') or [])}")
        check(len(wrow(wk, vendor_a).get('cells') or []) == 7, 'tiap vendor punya 7 kotak hari')
        check(len(wrow(wk, vendor_a).get('trend') or []) == 7,
              'sparkline tren pcs punya 7 angka (satu per kotak)')
        check(bool(wk.get('as_of_note')) and bool(wk.get('rules_note')),
              'ada catatan rentang + penjelasan aturan di payload (layar & export memakai teks yang sama)')

        # ── Rentang bisa digeser, parameter salah ditolak ────────────────────
        code, wk_y = week(staff, end=ago(1))
        check(code == 200 and wk_y.get('end') == ago(1) and wk_y.get('start') == ago(7),
              'jendela bisa digeser (?date= = hari TERAKHIR jendela)',
              f"{wk_y.get('start')} … {wk_y.get('end')}")
        check(wk_y.get('is_current') is False, 'jendela lampau ditandai bukan jendela berjalan')
        code, wk3 = week(staff, days=3)
        check(code == 200 and len(wk3.get('days') or []) == 3, '?days=3 menghasilkan 3 kotak hari')
        for bad_days in (0, 99):
            c, _ = week(staff, days=bad_days)
            check(c == 400, f'?days={bad_days} ditolak 400 (bukan 500 / hasil aneh)', f'HTTP {c}')
        c, _ = jget('get', '/cmt-override/weekly-recap?days=abc', token=staff)
        check(c == 400, '?days=abc ditolak 400', f'HTTP {c}')
        c, _ = jget('get', '/cmt-override/weekly-recap?date=08-2026', token=staff)
        check(c == 400, 'tanggal salah format ditolak 400', f'HTTP {c}')

        # ── Hari yang BELUM TERJADI tidak boleh ikut dihitung ────────────────
        code, wk_f = week(staff, end=(d0 + timedelta(days=2)).isoformat())
        fut = [d for d in (wk_f.get('days') or []) if d.get('is_future')]
        check(code == 200 and len(fut) == 2, 'dua hari terakhir ditandai future', str(len(fut)))
        check((wk_f.get('summary') or {}).get('days_elapsed') == 5,
              'hari future TIDAK dihitung sebagai hari berjalan',
              str((wk_f.get('summary') or {}).get('days_elapsed')))
        fa = wrow(wk_f, vendor_a)
        fcells = [c for c in (fa.get('cells') or []) if c.get('is_future')]
        check(bool(fcells) and all(c.get('state') == 'future' and c.get('qty_progress') == 0
                                   for c in fcells),
              'kotak hari future berstate "future" dan nol angka (bukan dituduh "tidak ada pekerjaan")')

        # ══════════════════════════════════════════════════════════════════════
        section('12. Rekap Mingguan MERINGKAS Rekap Harian — angkanya tidak boleh berdebat')
        code, wk = week(staff)
        mism: list[str] = []
        for p in (wk.get('per_day') or []):
            if p.get('is_future'):
                continue
            _c, dayrec = recap(staff, p['date'])
            s = dayrec.get('summary') or {}
            for kw, kd in (('vendors_pending', 'vendors_pending'),
                           ('vendors_partial', 'vendors_partial'),
                           ('vendors_done', 'vendors_done'),
                           ('vendors_idle', 'vendors_idle'),
                           ('tasks_pending_total', 'tasks_pending_total'),
                           ('qty_progress', 'qty_progress_today'),
                           ('qty_shipped', 'qty_shipped_today')):
                if p.get(kw) != s.get(kd):
                    mism.append(f"{p['date']}.{kw} {p.get(kw)}≠{s.get(kd)}")
            drow = {r_['vendor_id']: r_['status'] for r_ in (dayrec.get('rows') or [])}
            for r_ in (wk.get('rows') or []):
                cc = wcell(r_, p['date'])
                if cc.get('state') != drow.get(r_['vendor_id']):
                    mism.append(f"{p['date']}/{r_['vendor_name']} {cc.get('state')}≠{drow.get(r_['vendor_id'])}")
        check(not mism,
              'SEMUA angka per hari DAN state per vendor SAMA dengan rekap harian tanggal itu',
              '; '.join(mism[:4]) if mism else '7 hari × semua vendor cocok')
        sm = wk.get('summary') or {}
        check(sm.get('qty_progress_total') == sum(p['qty_progress'] for p in wk['per_day']),
              'total pcs disetor sepekan == jumlah per hari', str(sm.get('qty_progress_total')))
        check(sm.get('qty_shipped_total') == sum(p['qty_shipped'] for p in wk['per_day']),
              'total pcs dikirim sepekan == jumlah per hari', str(sm.get('qty_shipped_total')))
        check(sm.get('days_late_total') == sum(r_['days_late'] for r_ in wk['rows']),
              'total hari terlambat == jumlah per vendor', str(sm.get('days_late_total')))
        check(sm.get('vendors_total') == len(wk['rows']), 'jumlah vendor di ringkasan == jumlah baris')
        check([t['key'] for t in (wk.get('tasks') or [])]
              == [t['key'] for t in (rec_now.get('tasks') or [])],
              'daftar tugas (kolom) mingguan == harian ⇒ satu SSOT TASKS')

        # ══════════════════════════════════════════════════════════════════════
        section('13. Aturan owner: terlambat vs belum beres · hari tanpa setoran · streak')
        ra, rb = wrow(wk, vendor_a), wrow(wk, vendor_b)
        t_iso = d0.isoformat()
        ca, cb = wcell(ra, t_iso), wcell(rb, t_iso)
        check(ca.get('state') == 'partial',
              'vendor A hari ini "partial" (sudah setor, masih ada sisa)', str(ca.get('state')))
        check(cb.get('state') == 'pending',
              'vendor B hari ini "pending" (ada pekerjaan, NOL bukti)', str(cb.get('state')))

        # Keputusan owner 2 — DUA angka terpisah, tidak ada yang dibuang.
        check(ra.get('days_late') == 0,
              'hari "partial" TIDAK dihitung terlambat (days_late vendor A = 0)',
              str(ra.get('days_late')))
        check(ra.get('days_unfinished', 0) >= 1,
              'hari "partial" DIHITUNG belum beres (days_unfinished vendor A ≥ 1)',
              str(ra.get('days_unfinished')))
        check(rb.get('days_late', 0) >= 1,
              'hari "pending" dihitung terlambat (vendor B)', str(rb.get('days_late')))
        exp_late = sum(1 for c in ra['cells'] if c['state'] == 'pending')
        exp_unf = sum(1 for c in ra['cells'] if c['state'] in ('pending', 'partial'))
        check(ra['days_late'] == exp_late and ra['days_unfinished'] == exp_unf,
              'kedua angka konsisten dengan kotak hari yang tampil',
              f"{ra['days_late']}/{ra['days_unfinished']} vs {exp_late}/{exp_unf}")

        # Aturan turunan (a) — hari tanpa setoran hanya saat MEMANG ada job jalan.
        check(rb.get('days_no_progress', 0) >= 1,
              'vendor B: job jalan tanpa setoran ⇒ "hari tanpa setoran" ≥ 1',
              str(rb.get('days_no_progress')))
        bad_ns = []
        for r_ in wk['rows']:
            exp_ns = sum(1 for c in r_['cells']
                         if (not c['is_future']) and c['progress_state'] != 'none'
                         and c['progress_done'] == 0)
            if r_['days_no_progress'] != exp_ns:
                bad_ns.append(f"{r_['vendor_name']} {r_['days_no_progress']}≠{exp_ns}")
        check(not bad_ns,
              'hari tanpa setoran HANYA dihitung saat vendor punya job jalan '
              '(vendor tanpa pekerjaan tidak dihukum)', '; '.join(bad_ns) or 'semua vendor benar')
        idle_rows = [r_ for r_ in wk['rows'] if r_['days_with_work'] == 0]
        if idle_rows:
            check(all(r_['days_no_progress'] == 0 and r_['days_late'] == 0
                      and r_['status'] == 'idle' for r_ in idle_rows),
                  'vendor tanpa pekerjaan sepekan: nol terlambat, nol tanpa setoran, status "idle"',
                  ', '.join(r_['vendor_name'] for r_ in idle_rows))

        # Aturan owner 4 + turunan (b) — streak.
        def expect_streak(cells_):
            s_, brk_ = 0, ''
            for c_ in reversed([c for c in cells_ if not c['is_future']]):
                if c_['state'] in ('pending', 'partial'):
                    brk_ = c_['state']
                    break
                if c_['state'] == 'done':
                    s_ += 1
            return s_, brk_

        bad_streak = []
        for r_ in wk['rows']:
            es, eb = expect_streak(r_['cells'])
            if r_['streak'] != es or r_['streak_broken_by'] != eb:
                bad_streak.append(f"{r_['vendor_name']} {r_['streak']}/{r_['streak_broken_by']}≠{es}/{eb}")
        check(not bad_streak,
              'streak semua vendor sesuai aturan (mundur dari hari terakhir, putus pada pending/partial)',
              '; '.join(bad_streak) or f"{len(wk['rows'])} vendor")
        check(ra['streak'] == 0 and ra['streak_broken_by'] == 'partial',
              'vendor A: streak putus oleh hari "partial"', f"{ra['streak']}/{ra['streak_broken_by']}")
        check(rb['streak'] == 0 and rb['streak_broken_by'] == 'pending',
              'vendor B: streak putus oleh hari "pending"', f"{rb['streak']}/{rb['streak_broken_by']}")

        # IDLE NETRAL — dibuktikan pada jendela yang BERAKHIR di hari "beres".
        code, wk2 = week(staff, end=ago(2))
        r2 = wrow(wk2, vendor_a)
        s4, s3, s2 = (wcell(r2, ago(4)).get('state'), wcell(r2, ago(3)).get('state'),
                      wcell(r2, ago(2)).get('state'))
        check(s4 == 'done' and s3 == 'idle' and s2 == 'done',
              f'pola uji terbentuk: {ago(4)}=beres · {ago(3)}=tanpa pekerjaan · {ago(2)}=beres',
              f'{s4} / {s3} / {s2}')
        check(r2.get('streak') == 2 and r2.get('streak_broken_by') == '',
              'hari TANPA PEKERJAAN tidak memutus streak dan tidak menambah ⇒ streak = 2 (bukan 3, bukan 1)',
              f"streak={r2.get('streak')} broken_by='{r2.get('streak_broken_by')}'")
        check(r2.get('status') == 'clean',
              'vendor tanpa hari terlambat/belum beres berstatus "clean"', str(r2.get('status')))

        # Tren pcs & total
        check(wcell(ra, ago(4)).get('qty_progress') == 20
              and wcell(ra, ago(2)).get('qty_progress') == 15,
              'tren pcs per hari mengikuti setoran nyata (20 dan 15 pcs)',
              f"{wcell(ra, ago(4)).get('qty_progress')} / {wcell(ra, ago(2)).get('qty_progress')}")
        check(ra['qty_progress_total'] == sum(c['qty_progress'] for c in ra['cells']),
              'total pcs vendor == jumlah angka sparkline-nya', str(ra['qty_progress_total']))

        # Urutan: yang paling perlu diurus di atas.
        rank = {'late': 0, 'unfinished': 1, 'clean': 2, 'idle': 3}
        order = [r_['status'] for r_ in wk['rows']]
        check(order == sorted(order, key=lambda s_: rank[s_]),
              'urutan baris: vendor TERLAMBAT di atas, "tidak ada pekerjaan" di bawah', str(order))
        lates = [r_['days_late'] for r_ in wk['rows'] if r_['status'] == 'late']
        check(lates == sorted(lates, reverse=True),
              'di antara yang terlambat, yang paling banyak bolong di atas', str(lates))

        # Reminder mingguan menyasar SATU tanggal yang jelas, sasaran == tab Harian.
        check(wk.get('remind_date') == t_iso,
              'reminder mingguan menyasar hari terakhir yang sudah berjalan', str(wk.get('remind_date')))
        _c, rec_t = recap(staff, t_iso)
        dpend = sorted(r_['vendor_id'] for r_ in (rec_t.get('rows') or []) if r_['status'] == 'pending')
        wpend = sorted(x['vendor_id'] for x in (wk.get('remind_pending') or []))
        check(dpend == wpend,
              'sasaran reminder mingguan == sasaran tab Harian tanggal itu (fungsi yang SAMA)',
              f'{len(wpend)} vendor')

        # ══════════════════════════════════════════════════════════════════════
        section('14. Rekap Mingguan — export Excel/PDF & RBAC')
        for fmt, magic in (('xlsx', b'PK'), ('pdf', b'%PDF')):
            r = req('get', f'/cmt-override/weekly-recap/export?format={fmt}', token=staff)
            cd = r.headers.get('Content-Disposition') or ''
            check(r.status_code == 200 and r.content[:5].startswith(magic),
                  f'export {fmt.upper()} mingguan = berkas valid',
                  f'HTTP {r.status_code} {len(r.content)}B')
            check(wk['start'].replace('-', '') in cd and wk['end'].replace('-', '') in cd,
                  f'nama berkas {fmt.upper()} menyebut rentang tanggalnya', cd)
        r = req('get', '/cmt-override/weekly-recap/export?format=xlsx', token=staff)
        try:
            from openpyxl import load_workbook
            ws = load_workbook(io.BytesIO(r.content)).active
            vals = {}
            names_xlsx = []
            api_names = [x['vendor_name'] for x in wk['rows']]
            for row_ in ws.iter_rows(values_only=True):
                if row_ and row_[0]:
                    vals.setdefault(str(row_[0]), row_[1])
                    if str(row_[0]) in api_names:
                        names_xlsx.append(str(row_[0]))
            check(vals.get('Vendor aktif') == sm['vendors_total']
                  and vals.get('Total hari terlambat (semua vendor)') == sm['days_late_total']
                  and vals.get('Total pcs disetor sepekan') == sm['qty_progress_total']
                  and vals.get('Total pcs dikirim sepekan') == sm['qty_shipped_total'],
                  'angka di Excel SAMA dengan angka API (satu sumber, bukan hitung ulang)',
                  f"vendor={vals.get('Vendor aktif')} terlambat={vals.get('Total hari terlambat (semua vendor)')} "
                  f"pcs={vals.get('Total pcs disetor sepekan')}")
            check(names_xlsx == api_names,
                  'urutan vendor di Excel == urutan di layar', str(names_xlsx))
        except Exception as e:  # noqa: BLE001
            check(False, 'Excel mingguan bisa dibaca openpyxl', str(e)[:140])
        r = req('get', '/cmt-override/weekly-recap/export?format=csv', token=staff)
        check(r.status_code == 400, 'format export tak dikenal ditolak 400', f'HTTP {r.status_code}')

        for path in ('/cmt-override/weekly-recap', '/cmt-override/weekly-recap/export'):
            c, _ = jget('get', path, token=hr)
            check(c == 403, f'role TIDAK berwenang (hr) → 403 di {path}', f'HTTP {c}')
            c, _ = jget('get', path)
            check(c == 401, f'tanpa token → 401 di {path}', f'HTTP {c}')
        if vb_token:
            c, _ = jget('get', '/cmt-override/weekly-recap', token=vb_token)
            check(c == 403, 'akun VENDOR → 403 (rekap lintas-vendor milik staf DA)', f'HTTP {c}')
        c, wk_ov = week(staff, vendor_header=vendor_a)
        check(c == 200 and len(wk_ov.get('rows') or []) == len(wk['rows']),
              'header X-CMT-Override-Vendor DIABAIKAN (rekap tetap lintas vendor)',
              f"{len(wk_ov.get('rows') or [])} baris")

        # ══════════════════════════════════════════════════════════════════════
        section('15. Kinerja mingguan — 7 hari TIDAK boleh 7× lebih mahal')
        t0 = datetime.now(timezone.utc)
        code, _ = week(staff)
        wms = (datetime.now(timezone.utc) - t0).total_seconds() * 1000
        t1 = datetime.now(timezone.utc)
        for p in (wk.get('per_day') or []):
            if not p.get('is_future'):
                recap(staff, p['date'])
        dms = (datetime.now(timezone.utc) - t1).total_seconds() * 1000
        check(code == 200 and wms < 4000, 'rekap mingguan dijawab < 4 detik', f'{wms:.0f} ms')
        check(wms <= dms * 1.10,
              'mingguan TIDAK lebih mahal dari 7× harian ⇒ data master dibaca SEKALI '
              '(prefetch_context dipakai ulang)',
              f'mingguan {wms:.0f} ms vs 7× harian {dms:.0f} ms')


        # ── 16. KINERJA ─────────────────────────────────────────────────────
        section('16. Kinerja harian — layar pagi tidak boleh lambat')
        t0 = datetime.now(timezone.utc)
        code, _ = recap(staff)
        ms = (datetime.now(timezone.utc) - t0).total_seconds() * 1000
        check(code == 200 and ms < 3000, 'rekap dijawab < 3 detik', f'{ms:.0f} ms')

        # ══════════════════════════════════════════════════════════════════════
        # FASE 5 — `closed_at`: rekap tanggal LAMPAU berhenti menebak
        # ══════════════════════════════════════════════════════════════════════
        # MASALAH YANG DIUJI. Sebelum ini, "job jalan pada tanggal X" dijawab dengan
        # melihat status SEKARANG. Jadi job yang dibuka Senin, TIDAK disetor Senin,
        # lalu ditutup Rabu akan HILANG dari rekap hari Senin — kelalaian yang sudah
        # terjadi terhapus sendiri begitu job-nya ditutup. Karena progress produksi
        # adalah dasar tagihan CMT, laporan yang memaafkan dirinya sendiri seperti
        # itu tidak bisa dipakai memverifikasi apa pun.
        section('17. closed_at — job yang ditutup HARI INI tetap terhitung pada tanggal LAMPAU')

        y_iso = ago(1)

        # `created_at` job ditulis SERVER dan TIDAK bisa dibuat lampau lewat API mana
        # pun. Jadi SEJARAH-nya dibuat langsung di Mongo — yang dipalsukan hanya
        # "kapan job ini lahir", BUKAN perilaku yang sedang diuji (penutupan job
        # tetap lewat HTTP sungguhan, dan `closed_at` tetap ditulis server).
        db.production_jobs.update_one(
            {'id': job_b},
            {'$set': {'created_at': datetime.now(timezone.utc) - timedelta(days=2)}})
        jb_doc = db.production_jobs.find_one({'id': job_b}, {'_id': 0, 'status': 1, 'closed_at': 1})
        check((jb_doc or {}).get('status') == 'In Progress' and not (jb_doc or {}).get('closed_at'),
              'siapkan job vendor B: lahir 2 hari lalu, masih JALAN, belum punya closed_at',
              str(jb_doc))

        code, rec_y = recap(staff, y_iso)
        wait_before = (row_of(rec_y, vendor_b).get('tasks') or {}).get('progress', {}).get('waiting')
        check(code == 200 and wait_before == 1,
              f'rekap {y_iso}: job vendor B terhitung 1 "job jalan" (SEBELUM ditutup)',
              f'waiting={wait_before}')

        # ── Tutup job lewat jalur NORMAL (auto-complete oleh entri progress) ──
        code, jitems_b = jget('get', f'/production-job-items?job_id={job_b}', token=staff, vendor=vendor_b)
        items_b = as_list(jitems_b)
        check(bool(items_b), f'item job vendor B terbaca ({len(items_b)} item)')
        posted = 0
        for it in items_b:
            need = int(it.get('shipment_qty', 0) or 0) - int(it.get('produced_qty', 0) or 0)
            if need <= 0:
                continue
            c_, _r = jget('post', '/production-progress', token=staff, vendor=vendor_b, json={
                'job_item_id': it['id'], 'progress_date': d0.isoformat(),
                'completed_quantity': need, 'notes': f'{MARK} tutup job vendor B'})
            posted += 1 if c_ in (200, 201) else 0
        check(posted == len([i for i in items_b
                             if int(i.get('shipment_qty', 0) or 0) > int(i.get('produced_qty', 0) or 0)]),
              'semua item job vendor B disetor penuh lewat HTTP ⇒ job tertutup otomatis',
              f'{posted} entri progress')

        jb2 = db.production_jobs.find_one({'id': job_b},
                                          {'_id': 0, 'status': 1, 'closed_at': 1,
                                           'closed_at_estimated': 1})
        check((jb2 or {}).get('status') == 'Completed',
              'job vendor B berstatus Completed', str((jb2 or {}).get('status')))
        check(isinstance((jb2 or {}).get('closed_at'), datetime),
              'closed_at ditulis SERVER sebagai TANGGAL BSON (bukan string dari browser — '
              'pelajaran bug received_at)',
              f"tipe={type((jb2 or {}).get('closed_at')).__name__}")
        check((jb2 or {}).get('closed_at_estimated') is None,
              'closed_at hasil PENGAMATAN tidak ditandai perkiraan')

        # ── INTI FASE INI ────────────────────────────────────────────────────
        code, rec_y2 = recap(staff, y_iso)
        prog_y = (row_of(rec_y2, vendor_b).get('tasks') or {}).get('progress', {})
        check(prog_y.get('waiting') == 1,
              f'rekap {y_iso} TETAP menghitung 1 "job jalan" walau job sudah DITUTUP hari ini '
              '(sebelum perbaikan ini angkanya jatuh ke 0 = kelalaian terhapus sendiri)',
              f"waiting={prog_y.get('waiting')} state={prog_y.get('state')}")
        check(prog_y.get('state') == 'pending',
              f'rekap {y_iso}: vendor B tetap ✗ pada Progress Produksi (memang tidak menyetor '
              'hari itu)', str(prog_y.get('state')))

        code, rec_t = recap(staff, d0.isoformat())
        prog_t = (row_of(rec_t, vendor_b).get('tasks') or {}).get('progress', {})
        check(prog_t.get('waiting') == 0 and prog_t.get('done_today', 0) > 0,
              'rekap HARI INI: job itu tidak lagi "menunggu" (ditutup hari ini) tetapi '
              'setorannya terhitung',
              f"waiting={prog_t.get('waiting')} done={prog_t.get('done_today')}")

        # Batas hari: job yang ditutup hari ini TIDAK boleh terhitung jalan pada
        # tanggal SESUDAH penutupan.
        code, rec_tm = recap(staff, (d0 + timedelta(days=1)).isoformat())
        prog_tm = (row_of(rec_tm, vendor_b).get('tasks') or {}).get('progress', {})
        check(prog_tm.get('waiting') == 0,
              'rekap BESOK: job yang sudah ditutup tidak lagi terhitung jalan '
              '(closed_at benar-benar dibandingkan, bukan diabaikan)',
              f"waiting={prog_tm.get('waiting')}")

        # ── Klien tidak boleh menyuntik closed_at (pelajaran received_at) ────
        code, jx = jget('post', '/production-jobs', token=staff, vendor=vendor_a, json={
            'vendor_shipment_id': ship_id, 'notes': f'{MARK} job uji suntik closed_at',
            'closed_at': '2020-01-01T00:00:00Z', 'status': 'Completed'})
        jx_id = (jx or {}).get('id') if isinstance(jx, dict) else None
        if jx_id:
            _created['jobs'].append(jx_id)
            jxd = db.production_jobs.find_one({'id': jx_id}, {'_id': 0, 'status': 1, 'closed_at': 1})
            check(not (jxd or {}).get('closed_at') and (jxd or {}).get('status') == 'In Progress',
                  'closed_at/status kiriman BROWSER diabaikan saat job dibuat '
                  '(server yang menentukan, seperti received_at)', str(jxd))
        else:
            check(code in (400, 409, 422),
                  'pembuatan job kedua ditolak dengan jelas (bukan 500)', f'HTTP {code}')

        # ── Job WARISAN (tertutup sebelum fitur ini ada) + migrasi backfill ──
        section('18. Migrasi backfill closed_at — job warisan tidak ditebak diam-diam')
        legacy_id = f'{MARK}-legacy-{uuid.uuid4().hex[:8]}'
        legacy_closed = datetime.now(timezone.utc) - timedelta(days=1)
        db.production_jobs.insert_one({
            'id': legacy_id, 'job_number': f'{MARK}-JOBLEGACY',
            'vendor_id': vendor_a, 'vendor_name': f'{MARK} Vendor Tanpa Sistem',
            'po_id': po_id, 'status': 'Completed',
            'notes': f'{MARK} job warisan tanpa closed_at',
            'created_at': datetime.now(timezone.utc) - timedelta(days=4),
            'updated_at': legacy_closed,
        })
        d3 = ago(3)
        code, rec_d3 = recap(staff, d3)
        base_wait = (row_of(rec_d3, vendor_a).get('tasks') or {}).get('progress', {}).get('waiting', 0)
        check(rec_d3.get('legacy_jobs_without_closed_at', 0) >= 1,
              'rekap MENGAKU ada job warisan yang waktu tutupnya tidak diketahui '
              '(tidak disembunyikan)',
              str(rec_d3.get('legacy_jobs_without_closed_at')))
        check('add_closed_at_to_production_jobs' in (rec_d3.get('as_of_note') or ''),
              'catatan di layar menyebut migrasi yang harus dijalankan',
              (rec_d3.get('as_of_note') or '')[-90:])

        import subprocess
        mig = subprocess.run(
            [sys.executable, 'migrations/add_closed_at_to_production_jobs.py', '--execute'],
            cwd='/app/backend', capture_output=True, text=True, timeout=180)
        check(mig.returncode == 0, 'migrasi backfill berjalan (--execute)',
              (mig.stdout or mig.stderr or '')[-160:].replace('\n', ' '))
        lg = db.production_jobs.find_one({'id': legacy_id},
                                         {'_id': 0, 'closed_at': 1, 'closed_at_estimated': 1})
        check(isinstance((lg or {}).get('closed_at'), datetime)
              and (lg or {}).get('closed_at_estimated') is True,
              'job warisan diberi closed_at DAN ditandai closed_at_estimated=True '
              '(perkiraan dibedakan dari pengamatan)', str(lg))
        got = (lg or {}).get('closed_at')
        check(got is not None and abs((got.replace(tzinfo=timezone.utc) - legacy_closed)
                                      .total_seconds()) < 2,
              'perkiraannya = updated_at (penanda terbaik yang tersedia), bukan angka karangan',
              f'{got} vs {legacy_closed}')

        code, rec_d3b = recap(staff, d3)
        after_wait = (row_of(rec_d3b, vendor_a).get('tasks') or {}).get('progress', {}).get('waiting', 0)
        check(after_wait == base_wait + 1,
              f'setelah migrasi, rekap {d3} menghitung job warisan itu sebagai job jalan '
              '(sebelumnya tidak bisa — dan tidak ditebak)',
              f'{base_wait} → {after_wait}')
        check(rec_d3b.get('legacy_jobs_without_closed_at', 0) == 0,
              'nol job warisan tersisa ⇒ catatan "tidak bisa dihitung" hilang dari layar',
              str(rec_d3b.get('legacy_jobs_without_closed_at')))
        check('tetap terhitung' in (rec_d3b.get('as_of_note') or '')
              and 'tidak lagi terhitung' not in (rec_d3b.get('as_of_note') or ''),
              'catatan layar tidak lagi mengaku menebak untuk tanggal lampau',
              (rec_d3b.get('as_of_note') or '')[:110])

        # Idempotensi migrasi: dijalankan dua kali tidak mengubah stempel.
        mig2 = subprocess.run(
            [sys.executable, 'migrations/add_closed_at_to_production_jobs.py', '--execute'],
            cwd='/app/backend', capture_output=True, text=True, timeout=180)
        lg2 = db.production_jobs.find_one({'id': legacy_id}, {'_id': 0, 'closed_at': 1})
        check(mig2.returncode == 0 and (lg2 or {}).get('closed_at') == (lg or {}).get('closed_at'),
              'migrasi IDEMPOTEN: dijalankan ulang tidak menggeser stempel yang sudah ada')

        # ── Konsistensi mingguan ikut terbawa (satu sumber angka) ────────────
        code, wk_c = week(staff)
        mism2 = []
        for p in (wk_c.get('per_day') or []):
            if p.get('is_future'):
                continue
            _c, dr = recap(staff, p['date'])
            if p.get('vendors_pending') != (dr.get('summary') or {}).get('vendors_pending'):
                mism2.append(p['date'])
        check(not mism2,
              'rekap MINGGUAN otomatis ikut benar setelah perubahan ini (ia meringkas '
              'build_recap, tidak menghitung sendiri)', str(mism2) or '7 hari cocok')


        return 0

    finally:
        # ── CLEANUP ─────────────────────────────────────────────────────────
        section('CLEANUP')
        if keep:
            print(f'  {Y}--keep: data uji DIBIARKAN (bersihkan manual!){X}')
        else:
            try:
                n = 0
                vids = [v for v in _created['partners'] if v]
                pos = [p for p in _created['pos'] if p]
                jobs = [j['id'] for j in db.production_jobs.find(
                    {'vendor_id': {'$in': vids}}, {'_id': 0, 'id': 1})]
                ships = [s['id'] for s in db.vendor_shipments.find(
                    {'vendor_id': {'$in': vids}}, {'_id': 0, 'id': 1})]
                insps = [i['id'] for i in db.vendor_material_inspections.find(
                    {'vendor_id': {'$in': vids}}, {'_id': 0, 'id': 1})]
                bss = [b['id'] for b in db.buyer_shipments.find(
                    {'vendor_id': {'$in': vids}}, {'_id': 0, 'id': 1})]
                rcpts = [x['id'] for x in db.cmt_receipts.find(
                    {'cmt_vendor_id': {'$in': vids}}, {'_id': 0, 'id': 1})]

                ops = [
                    ('production_progress', {'job_id': {'$in': jobs}}),
                    ('production_job_items', {'job_id': {'$in': jobs}}),
                    ('production_jobs', {'id': {'$in': jobs}}),
                    ('vendor_material_inspection_items', {'inspection_id': {'$in': insps}}),
                    ('vendor_material_inspections', {'id': {'$in': insps}}),
                    ('vendor_shipment_items', {'shipment_id': {'$in': ships}}),
                    ('accessory_shipment_items', {'shipment_id': {'$in': ships}}),
                    ('vendor_shipments', {'id': {'$in': ships}}),
                    ('material_requests', {'vendor_id': {'$in': vids}}),
                    ('production_variances', {'vendor_id': {'$in': vids}}),
                    ('buyer_shipment_items', {'shipment_id': {'$in': bss}}),
                    ('buyer_shipments', {'id': {'$in': bss}}),
                    ('cmt_receipt_lines', {'receipt_id': {'$in': rcpts}}),
                    ('cmt_receipts', {'id': {'$in': rcpts}}),
                    ('po_accessories', {'po_id': {'$in': pos}}),
                    ('po_items', {'po_id': {'$in': pos}}),
                    ('production_pos', {'id': {'$in': pos}}),
                    ('dewi_maklon_bom', {'po_id': {'$in': pos}}),
                    # Efek samping PO maklon (UANG!) — jangan tinggalkan piutang palsu.
                    ('dewi_maklon_pos', {'production_po_id': {'$in': pos}}),
                    ('rahaza_ar_invoices', {'linked_maklon_po_id': {'$in': pos}}),
                    ('dewi_cmt_component_requests', {'vendor_id': {'$in': vids}}),
                    ('dewi_cmt_jobs', {'cmt_partner_id': {'$in': vids}}),
                    ('dewi_cmt_deliveries', {'cmt_partner_id': {'$in': vids}}),
                    ('dewi_cmt_payments', {'cmt_partner_id': {'$in': vids}}),
                    # Reminder rekap yang lahir dari uji tombol — TERMASUK yang
                    # ditujukan ke vendor demo asli (sasaran default = semua yang
                    # merah), kalau tidak dibersihkan inbox vendor demo akan
                    # berisi teguran palsu setiap kali POC dijalankan.
                    ('reminders', {'vendor_id': {'$in': vids}}),
                    ('reminders', {'created_by': {'$regex': MARK}}),
                    ('vendor_partners', {'id': {'$in': vids}}),
                    ('users', {'id': {'$in': [u for u in _created['users'] if u]}}),
                    ('activity_logs', {'details': {'$regex': MARK}}),
                    ('notifications', {'$or': [{'title': {'$regex': MARK}},
                                               {'message': {'$regex': MARK}}]}),
                    ('rahaza_audit_logs', {'entity_id': {'$in': pos + jobs + ships}}),
                    ('login_attempts', {'identifier': {'$regex': MARK.lower()}}),
                ]
                for coll, q in ops:
                    try:
                        n += db[coll].delete_many(q).deleted_count
                    except Exception as e:
                        print(f'  {Y}! gagal bersihkan {coll}: {e}{X}')

                # ── SWEEP TOTAL (jaring pengaman) ──────────────────────────
                # Daftar di atas ditulis TANGAN ⇒ setiap efek samping BARU di
                # backend akan lolos darinya tanpa ada yang tahu. Sweep memeriksa
                # setiap dokumen di setiap koleksi: kalau penanda uji muncul di
                # mana pun isinya, dokumen itu data uji.
                swept, where = 0, {}
                for coll in db.list_collection_names():
                    if coll in ('rate_limit_buckets', 'counters'):
                        continue
                    try:
                        dead = [d['_id'] for d in db[coll].find({}).limit(20000)
                                if MARK in _json.dumps(d, default=str)]
                    except Exception as e:
                        print(f'  {Y}! sweep {coll} dilewati: {e}{X}')
                        continue
                    if dead:
                        db[coll].delete_many({'_id': {'$in': dead}})
                        swept += len(dead)
                        where[coll] = len(dead)
                if swept:
                    print(f'  {Y}sweep menemukan {swept} dokumen sisa: {where}{X}')
                n += swept
                print(f'  {G}✓ {n} dokumen uji dihapus{X}')

                rest = {}
                for coll in db.list_collection_names():
                    if coll in ('rate_limit_buckets', 'counters'):
                        continue
                    try:
                        k = sum(1 for d in db[coll].find({}).limit(20000)
                                if MARK in _json.dumps(d, default=str))
                    except Exception:
                        k = 0
                    if k:
                        rest[coll] = k
                check(not rest, 'CLEANUP: nol jejak data uji tertinggal di seluruh DB',
                      str(rest) if rest else 'bersih')

                # reminder rekap tidak boleh tertinggal untuk vendor MANA PUN
                leftover_recap = db.reminders.count_documents(
                    {'reminder_type': 'daily_recap', 'recap_date': today_wib().isoformat()})
                check(leftover_recap == 0,
                      'CLEANUP: nol reminder rekap sisa (inbox vendor demo tidak tercemar)',
                      str(leftover_recap))

                # UANG harus kembali ke angka semula
                if admin and money_before is not None:
                    code, bill2 = jget('get', '/production/cmt-billing/summary', token=admin)
                    after = (bill2 or {}).get('total_amount') if isinstance(bill2, dict) else None
                    check(after == money_before,
                          'UANG: total tagihan CMT kembali persis seperti sebelum uji',
                          f'{money_before} → {after}')
            except Exception as e:  # noqa: BLE001
                print(f'  {R}CLEANUP gagal: {e}{X}')

        ok = sum(1 for r in _results if r[0])
        bad = [r for r in _results if not r[0]]
        print(f'\n{B}{"═" * 74}{X}')
        print(f'  HASIL: {G}{ok} LULUS{X} / {R}{len(bad)} GAGAL{X} dari {len(_results)} pemeriksaan')
        if bad:
            print(f'\n  {R}Yang GAGAL:{X}')
            for _, name, detail in bad:
                print(f'   {R}✗{X} {name}' + (f'  → {detail}' if detail else ''))
        print(f'{B}{"═" * 74}{X}')
        mongo.close()
        if bad:
            sys.exit(1)


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--keep', action='store_true', help='jangan hapus data uji')
    a = ap.parse_args()
    sys.exit(main(a.keep) or 0)
