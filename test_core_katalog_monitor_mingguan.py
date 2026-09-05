#!/usr/bin/env python3
"""test_core_katalog_monitor_mingguan.py — CORE TEST 4 kemampuan baru (2026-08-12 #5).

SATU berkas uji untuk empat permintaan owner:
  A. **GUDANG PLATFORM belajar dari berkas** — impor pertama sebuah toko menawarkan
     nama gudang dari ekspor; disimpan ke master ⇒ penjaga toko aktif otomatis.
  B. **KATALOG dari MASTER PRODUK** — pilih produk master, tempelkan ke BANYAK toko
     sekaligus; HPP & harga resmi datang dari master (bukan diketik).
  C. **MONITORING PENGIRIMAN** — belum dikirim / lewat batas (bocor) / batal, dengan
     batas hari per toko yang bisa diubah.
  D. **LAPORAN RAPAT MINGGUAN** — per toko + gabungan, vs minggu lalu, vs target
     prorata, pecahan kanal, pemenuhan, ROAS, + catatan kejujuran data; PDF & Excel.

Semua memakai HTTP sungguhan dan diverifikasi ke DB. Berkas contoh:
`samples/TikTok_UntukDikirim_2026-07-19.xlsx` (559 pesanan · Rp 59.783.811).

Pakai:  python3 /app/test_core_katalog_monitor_mingguan.py
"""
from __future__ import annotations

import os
import sys
from datetime import datetime, timedelta, timezone

import requests
from dotenv import load_dotenv

sys.path.insert(0, "/app/backend")
load_dotenv("/app/backend/.env")

BASE = "http://localhost:8001"
SAMPLE = "/app/samples/TikTok_UntukDikirim_2026-07-19.xlsx"
XLSX_MIME = ("application/vnd.openxmlformats-officedocument"
             ".spreadsheetml.sheet")
RESULTS: list[tuple[str, bool, str]] = []


def ok(name, detail=""):
    RESULTS.append((name, True, detail))
    print(f"  \033[92mPASS\033[0m  {name}" + (f" — {detail}" if detail else ""))


def fail(name, detail=""):
    RESULTS.append((name, False, detail))
    print(f"  \033[91mFAIL\033[0m  {name}" + (f" — {detail}" if detail else ""))


def check(name, cond, detail=""):
    (ok if cond else fail)(name, detail)
    return bool(cond)


def mongo():
    from pymongo import MongoClient
    cli = MongoClient(os.environ.get("MONGO_URL", "mongodb://localhost:27017"))
    return cli, cli[os.environ.get("DB_NAME", "test_database")]


def main() -> int:
    tok = requests.post(f"{BASE}/api/auth/login", json={
        "email": "admin@garment.com", "password": "Admin@123"}, timeout=30).json()["token"]
    H = {"Authorization": f"Bearer {tok}"}
    JH = {**H, "Content-Type": "application/json"}
    IMP = f"{BASE}/api/marketing/data-import"
    cli, db = mongo()

    accs = requests.get(f"{BASE}/api/marketing/accounts", headers=H, timeout=30).json()
    accs = accs.get("accounts", accs) if isinstance(accs, dict) else accs
    by_code = {a.get("account_code"): a for a in accs}
    outfit = by_code.get("TIKTOK-OUTFIT")
    moen = by_code.get("TIKTOK-MOEN")
    if not check("prasyarat: 9 toko master ada", bool(outfit and moen), f"{len(accs)} toko"):
        return 1

    sessions: list[str] = []
    created_item_ids: list[str] = []
    try:
        # ══════════════════════════════════════════════════════════════════════
        print("\n[A] GUDANG PLATFORM — belajar dari berkas ekspor")
        # ══════════════════════════════════════════════════════════════════════
        db.marketing_platform_accounts.update_one(
            {"id": outfit["id"]}, {"$set": {"platform_warehouse_name": ""}})
        with open(SAMPLE, "rb") as fh:
            up = requests.post(f"{IMP}/upload", headers=H, files={
                "file": (os.path.basename(SAMPLE), fh, XLSX_MIME)},
                data={"source_type": "marketplace_orders",
                      "account_id": outfit["id"]}, timeout=300)
        if not check("A1 unggah 200 (toko belum punya gudang ⇒ tidak diblokir)",
                     up.status_code == 200, f"{up.status_code} {up.text[:160]}"):
            return 1
        sess = up.json()["session"]
        sessions.append(sess["id"])
        check("A2 sesi membawa nama gudang yang TERBACA di berkas",
              sess.get("shop_guard_warehouse") == "Outfit Boutique",
              repr(sess.get("shop_guard_warehouse")))
        check("A3 sesi memberi peringatan gudang belum bisa dipastikan",
              "Gudang Platform" in (sess.get("shop_guard_hint") or ""),
              (sess.get("shop_guard_hint") or "")[:110])

        r = requests.post(f"{BASE}/api/marketing/accounts/{outfit['id']}/learn-warehouse",
                          headers=JH, json={"warehouse_name": "Outfit Boutique",
                                            "session_id": sess["id"]}, timeout=60)
        check("A4 simpan gudang ke master toko 200", r.status_code == 200,
              f"{r.status_code} {r.text[:140]}")
        acc_db = db.marketing_platform_accounts.find_one({"id": outfit["id"]}, {"_id": 0})
        check("A5 master toko menyimpan gudang + jejak asalnya",
              acc_db.get("platform_warehouse_name") == "Outfit Boutique"
              and acc_db.get("platform_warehouse_source") == "import_file",
              f"{acc_db.get('platform_warehouse_name')} / "
              f"{acc_db.get('platform_warehouse_source')}")

        r2 = requests.post(f"{BASE}/api/marketing/accounts/{moen['id']}/learn-warehouse",
                           headers=JH, json={"warehouse_name": "Outfit Boutique"},
                           timeout=60)
        check("A6 gudang yang sudah dipakai toko lain DITOLAK 409",
              r2.status_code == 409, f"{r2.status_code} "
              f"{(r2.json() or {}).get('detail', '')[:120]}")

        # penjaga toko kini aktif otomatis untuk toko lain
        with open(SAMPLE, "rb") as fh:
            up_bad = requests.post(f"{IMP}/upload", headers=H, files={
                "file": (os.path.basename(SAMPLE), fh, XLSX_MIME)},
                data={"source_type": "marketplace_orders",
                      "account_id": moen["id"]}, timeout=300)
        check("A7 sesudah gudang tersimpan, berkas ke toko lain DITOLAK 400",
              up_bad.status_code == 400, f"{up_bad.status_code}")

        cm = requests.post(f"{IMP}/sessions/{sess['id']}/commit", headers=JH,
                           json={"on_duplicate": "skip"}, timeout=300)
        check("A8 commit 559 pesanan (dasar uji C & D)",
              cm.status_code == 200 and cm.json().get("inserted") == 559,
              str(cm.json().get("inserted") if cm.status_code == 200 else cm.status_code))

        # ══════════════════════════════════════════════════════════════════════
        print("\n[B] KATALOG dari MASTER PRODUK — isi banyak toko sekaligus")
        # ══════════════════════════════════════════════════════════════════════
        mp = requests.get(f"{BASE}/api/marketing/catalogs/master-products",
                          headers=H, params={"account_id": outfit["id"]}, timeout=60)
        if not check("B1 daftar master produk 200", mp.status_code == 200,
                     f"{mp.status_code} {mp.text[:140]}"):
            return 1
        prods = mp.json().get("products") or []
        check("B2 master produk terbaca beserta variannya",
              len(prods) >= 1 and sum(p["variant_count"] for p in prods) >= 2,
              f"{len(prods)} produk · {mp.json().get('total_variants')} varian")
        with_hpp = [p for p in prods if p["hpp"] > 0 and p["retail_price_master"] > 0]
        check("B3 HPP & harga resmi datang dari MASTER (bukan diketik)",
              bool(with_hpp),
              (f"{with_hpp[0]['code']} HPP {with_hpp[0]['hpp']:,.0f} · "
               f"harga {with_hpp[0]['retail_price_master']:,.0f}") if with_hpp else "tidak ada")
        if not with_hpp:
            return 1
        picked = [v["fg_material_id"] for v in with_hpp[0]["variants"][:2]]
        check("B4 stok jual varian dibaca dari SSOT (bukan qty mentah)",
              all("sellable_stock" in v for v in with_hpp[0]["variants"]),
              f"{[v['sellable_stock'] for v in with_hpp[0]['variants'][:2]]}")

        targets = [outfit["id"], moen["id"], by_code["TIKTOK-DALUNA"]["id"]]
        asg = requests.post(f"{BASE}/api/marketing/catalogs/assign-from-master",
                            headers=JH, json={"account_ids": targets,
                                              "fg_material_ids": picked,
                                              "price_mode": "master"}, timeout=180)
        if not check("B5 tempel ke 3 toko sekaligus 200", asg.status_code == 200,
                     f"{asg.status_code} {asg.text[:200]}"):
            return 1
        body = asg.json()
        check("B6 3 toko × 2 varian = 6 item katalog baru",
              body["summary"]["created"] == 6,
              f"created={body['summary']['created']} skipped={body['summary']['skipped']} "
              f"rejected={body['summary']['rejected']}")
        check("B7 katalog toko dibuat otomatis kalau belum ada",
              sum(1 for r in body["results"] if r["catalog_created"]) >= 2,
              f"{[r['catalog_created'] for r in body['results']]}")
        check("B8 nama toko IKUT tercatat di katalog (bukan kosong)",
              all((db.marketing_catalogs.find_one({'id': r['catalog_id']}) or {})
                  .get("account_name") for r in body["results"]),
              str([(db.marketing_catalogs.find_one({'id': r['catalog_id']}) or {})
                   .get("account_name") for r in body["results"]])[:120])
        items = list(db.marketing_catalog_items.find(
            {"fg_material_id": {"$in": picked},
             "account_id": {"$in": targets}}, {"_id": 0}))
        created_item_ids = [i["id"] for i in items]
        check("B9 setiap item bawa HPP + sumber HPP + tautan master FG",
              all(i.get("hpp", 0) > 0 and i.get("hpp_source") not in (None, "none")
                  and i.get("fg_material_id") for i in items),
              f"{len(items)} item · contoh hpp={items[0]['hpp']:,.0f} "
              f"sumber={items[0]['hpp_source']}" if items else "0 item")
        check("B10 harga jual awal = harga resmi master (K-3a)",
              all(i.get("harga_jual", 0) == with_hpp[0]["retail_price_master"]
                  for i in items),
              f"{sorted({i.get('harga_jual') for i in items})}")
        again = requests.post(f"{BASE}/api/marketing/catalogs/assign-from-master",
                              headers=JH, json={"account_ids": targets,
                                                "fg_material_ids": picked,
                                                "price_mode": "master"}, timeout=180)
        check("B11 IDEMPOTEN: ulangi ⇒ 0 baru, 6 'sudah ada'",
              again.json()["summary"]["created"] == 0
              and again.json()["summary"]["skipped"] == 6,
              str(again.json()["summary"]))
        bad = requests.post(f"{BASE}/api/marketing/catalogs/assign-from-master",
                            headers=JH, json={"account_ids": ["tidak-ada"],
                                              "fg_material_ids": picked}, timeout=60)
        check("B12 toko tujuan ngawur ⇒ 404 (bukan diam-diam dilewati)",
              bad.status_code == 404, str(bad.status_code))

        # ══════════════════════════════════════════════════════════════════════
        print("\n[C] MONITORING PENGIRIMAN — belum dikirim / lewat batas / batal")
        # ══════════════════════════════════════════════════════════════════════
        mon = requests.get(f"{BASE}/api/marketing/orders/fulfillment-monitor",
                           headers=H, params={"account_id": outfit["id"],
                                              "bucket": "belum_dikirim"}, timeout=120)
        if not check("C1 monitor 200", mon.status_code == 200,
                     f"{mon.status_code} {mon.text[:160]}"):
            return 1
        m = mon.json()
        check("C2 559 pesanan hasil impor terbaca sebagai BELUM DIKIRIM",
              m["totals"]["belum_dikirim"] == 559,
              f"belum_dikirim={m['totals']['belum_dikirim']} "
              f"dibaca={m['totals']['pesanan_dibaca']}")
        check("C3 nilai yang tertahan dihitung (bukan hanya jumlah pesanan)",
              m["totals"]["nilai_belum_dikirim"] == 59_783_811,
              f"Rp {m['totals']['nilai_belum_dikirim']:,}")
        check("C4 semua sudah LEWAT BATAS (berkas Juli, batas 2/7 hari)",
              m["totals"]["lewat_batas"] == 559,
              f"lewat_batas={m['totals']['lewat_batas']}")
        check("C5 baris diurut dari yang paling tua",
              (m["rows"][0]["age_days"] or 0) >= (m["rows"][-1]["age_days"] or 0),
              f"{m['rows'][0]['age_days']} → {m['rows'][-1]['age_days']} hari")
        row0 = m["rows"][0]
        check("C6 setiap baris menyebut batas, umur, dan lewat berapa hari",
              all(k in row0 for k in ("sla_days", "age_days", "over_by_days", "deadline")),
              f"sla={row0['sla_days']} umur={row0['age_days']} "
              f"lewat={row0['over_by_days']}")
        check("C7 pre-order dikenali (batas kirimnya beda)",
              any(r["is_preorder"] for r in m["rows"]),
              f"{sum(1 for r in m['rows'] if r['is_preorder'])} dari {len(m['rows'])} baris halaman ini")
        check("C8 catatan kejujuran: batal/retur 0 karena Ekspor A tidak memuatnya",
              any("Ekspor C" in n or "batal/retur" in n for n in m["data_notes"]),
              str(m["data_notes"])[:150])

        sla = requests.put(f"{BASE}/api/marketing/accounts/{outfit['id']}/ship-sla",
                           headers=JH, json={"ship_sla_days": 40,
                                             "ship_sla_days_preorder": 60}, timeout=60)
        check("C9 batas kirim bisa diubah per toko", sla.status_code == 200,
              f"{sla.status_code} {sla.text[:120]}")
        mon2 = requests.get(f"{BASE}/api/marketing/orders/fulfillment-monitor",
                            headers=H, params={"account_id": outfit["id"],
                                               "bucket": "lewat_batas"}, timeout=120).json()
        check("C10 batas baru LANGSUNG mengubah daftar 'lewat batas'",
              mon2["totals"]["lewat_batas"] < 559
              and mon2["totals"]["belum_dikirim"] == 559,
              f"lewat_batas {559} → {mon2['totals']['lewat_batas']} "
              f"(batas 40/60 hari)")
        bad_sla = requests.put(f"{BASE}/api/marketing/accounts/{outfit['id']}/ship-sla",
                              headers=JH, json={"ship_sla_days": 10,
                                                "ship_sla_days_preorder": 2}, timeout=60)
        check("C11 batas pre-order lebih pendek dari normal DITOLAK 400",
              bad_sla.status_code == 400, str(bad_sla.status_code))
        requests.put(f"{BASE}/api/marketing/accounts/{outfit['id']}/ship-sla",
                     headers=JH, json={"ship_sla_days": 2,
                                       "ship_sla_days_preorder": 7}, timeout=60)
        allst = requests.get(f"{BASE}/api/marketing/orders/fulfillment-monitor",
                             headers=H, params={"bucket": "lewat_batas"}, timeout=180).json()
        check("C12 tanpa filter toko: rekap per toko ikut keluar",
              len(allst["per_store"]) >= 1
              and allst["per_store"][0]["lewat_batas"] == 559,
              f"{len(allst['per_store'])} toko · teratas "
              f"{allst['per_store'][0]['account_name']}")

        # ══════════════════════════════════════════════════════════════════════
        print("\n[D] LAPORAN RAPAT MINGGUAN")
        # ══════════════════════════════════════════════════════════════════════
        wk = requests.get(f"{BASE}/api/marketing/reports/weekly", headers=H,
                          params={"week_start": "2026-07-15"}, timeout=180)
        if not check("D1 laporan mingguan 200", wk.status_code == 200,
                     f"{wk.status_code} {wk.text[:160]}"):
            return 1
        w = wk.json()
        check("D2 minggu Senin–Minggu terkunci benar (13–19 Juli 2026)",
              w["periode"]["mulai"] == "2026-07-13"
              and w["periode"]["selesai"] == "2026-07-19",
              f"{w['periode']['mulai']} … {w['periode']['selesai']} "
              f"({w['periode']['minggu']})")
        rows = {s["account_code"]: s for s in w["per_toko"]}
        of = rows.get("TIKTOK-OUTFIT", {})
        # kebenaran pembanding: dihitung ulang dari rekap harian di DB
        want = sum(
            (d.get("metrics") or {}).get("revenue", 0)
            for d in db.marketing_sales_data.find(
                {"account_id": outfit["id"], "revenue_type": "total",
                 "date": {"$gte": "2026-07-13", "$lte": "2026-07-19"}}))
        check("D3 omzet toko = jumlah rekap harian minggu itu (bukan angka baru)",
              of.get("omzet") == round(want), f"laporan Rp {of.get('omzet'):,} "
              f"vs DB Rp {round(want):,}")
        check("D4 gabungan = jumlah semua toko",
              w["gabungan"]["omzet"] == sum(s["omzet"] for s in w["per_toko"]),
              f"Rp {w['gabungan']['omzet']:,}")
        check("D5 pembanding minggu lalu ikut dihitung",
              "vs_minggu_lalu" in of and "pembanding" in of["vs_minggu_lalu"]["omzet"],
              f"minggu lalu Rp {of['vs_minggu_lalu']['omzet']['pembanding']:,}")
        check("D6 pecahan kanal live/video/kartu produk terisi",
              of.get("kanal", {}).get("live", 0) > 0,
              f"live Rp {of.get('kanal', {}).get('live', 0):,} · "
              f"video Rp {of.get('kanal', {}).get('video', 0):,}")
        check("D7 ROAS TIDAK dikarang saat iklan belum diimpor (None, bukan 0)",
              of.get("iklan", {}).get("roas") is None
              and of.get("iklan", {}).get("terisi") is False,
              f"roas={of.get('iklan', {}).get('roas')} terisi={of.get('iklan', {}).get('terisi')}")
        check("D8 target prorata melaporkan kalau targetnya belum ada",
              of.get("target", {}).get("lengkap") is False
              and of.get("pencapaian_target_persen") is None,
              f"lengkap={of.get('target', {}).get('lengkap')} "
              f"capai={of.get('pencapaian_target_persen')}")
        # Laporan MINGGUAN hanya boleh menghitung pesanan yang tanggalnya di minggu
        # itu (429 dari 559) — beda dengan Monitoring Pengiriman yang menghitung
        # SEMUA yang masih terbuka. Pembandingnya dihitung ulang dari DB, bukan
        # disalin dari angka laporan.
        want_open = db.marketing_orders.count_documents({
            "account_id": outfit["id"],
            "status": {"$in": ["new", "paid", "packed"]},
            "shipped_at": None,
            "order_date": {
                "$gte": datetime(2026, 7, 13, tzinfo=timezone.utc),
                "$lt": datetime(2026, 7, 20, tzinfo=timezone.utc)}})
        check("D9 pesanan belum dikirim MINGGU ITU dilaporkan (bukan seluruh tunggakan)",
              w["gabungan"]["belum_dikirim"] == want_open and want_open > 0,
              f"laporan {w['gabungan']['belum_dikirim']} = DB {want_open} "
              f"(monitor menghitung semua: 559)")
        notes = " ".join(w["catatan_data"]).lower()
        check("D10 catatan kejujuran menyebut 'sebelum potongan platform'",
              "sebelum potongan platform" in notes)
        check("D11 catatan menyebut toko tanpa data & target belum ada",
              "tidak punya data" in notes and "target" in notes,
              f"{len(w['catatan_data'])} catatan")
        check("D12 catatan menyebut iklan belum diimpor ⇒ ROAS tidak dihitung",
              "roas" in notes and "belum diimpor" in notes)
        one = requests.get(f"{BASE}/api/marketing/reports/weekly", headers=H,
                           params={"week_start": "2026-07-15",
                                   "account_id": outfit["id"]}, timeout=120).json()
        check("D13 bisa dilingkupi satu toko", len(one["per_toko"]) == 1
              and one["per_toko"][0]["account_code"] == "TIKTOK-OUTFIT",
              f"{len(one['per_toko'])} toko")

        pdf = requests.get(f"{BASE}/api/marketing/reports/weekly/export-pdf",
                           headers=H, params={"week_start": "2026-07-15"}, timeout=180)
        check("D14 ekspor PDF 200 & berupa PDF sungguhan",
              pdf.status_code == 200 and pdf.content[:4] == b"%PDF"
              and len(pdf.content) > 5000,
              f"{pdf.status_code} · {len(pdf.content) / 1024:.0f} KB")
        xls = requests.get(f"{BASE}/api/marketing/reports/weekly/export-excel",
                           headers=H, params={"week_start": "2026-07-15"}, timeout=180)
        check("D15 ekspor Excel 200 & berupa xlsx sungguhan",
              xls.status_code == 200 and xls.content[:2] == b"PK"
              and len(xls.content) > 3000,
              f"{xls.status_code} · {len(xls.content) / 1024:.0f} KB")
        if xls.status_code == 200:
            import io
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(xls.content))
            check("D16 Excel punya 3 lembar termasuk 'Catatan Data'",
                  "Catatan Data" in wb.sheetnames and len(wb.sheetnames) == 3,
                  str(wb.sheetnames))
            ws = wb["Ringkas per Toko"]
            found = any(ws.cell(row=r, column=5).value == of.get("omzet")
                        for r in range(6, 6 + len(w["per_toko"])))
            check("D17 angka di Excel SAMA dengan angka layar (tidak dihitung ulang)",
                  found, f"cari Rp {of.get('omzet'):,} di kolom Omzet")
        bad_wk = requests.get(f"{BASE}/api/marketing/reports/weekly", headers=H,
                              params={"week_start": "bukan-tanggal"}, timeout=60)
        check("D18 tanggal ngawur ⇒ 400 (bukan 500)", bad_wk.status_code == 400,
              str(bad_wk.status_code))

    finally:
        print("\n  (bersih-bersih)")
        for sid in reversed(sessions):
            rb = requests.post(f"{IMP}/sessions/{sid}/rollback", headers=H, timeout=300)
            print(f"    rollback {sid[:8]}: {rb.status_code}")
        if created_item_ids:
            n = db.marketing_catalog_items.delete_many(
                {"id": {"$in": created_item_ids}}).deleted_count
            print(f"    item katalog uji dihapus: {n}")
        db.marketing_catalogs.delete_many({"description": {
            "$regex": "Dibuat otomatis saat mengisi katalog dari master produk"}})
        db.marketing_data_import_sessions.delete_many({"status": "rolled_back"})
        db.marketing_platform_accounts.update_many(
            {}, {"$unset": {"ship_sla_days": "", "ship_sla_days_preorder": ""}})
        requests.post(f"{BASE}/api/marketing/accounts/health/recompute-all",
                      headers=H, timeout=180)
        cli.close()

    passed = sum(1 for _, k, _ in RESULTS if k)
    print("\n" + "=" * 90)
    print(f"RINGKAS: {passed}/{len(RESULTS)} PASS")
    print("=" * 90)
    return 0 if passed == len(RESULTS) else 1


if __name__ == "__main__":
    sys.exit(main())
