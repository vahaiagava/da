"""core.marketing_import_engine — **mesin impor marketing yang BEKERJA TANPA AI**.

KENAPA DETERMINISTIK LEBIH DULU, AI BELAKANGAN
----------------------------------------------
Mesin lama (`routes/universal_import.py`) memanggil AI dua kali untuk setiap
impor: sekali untuk menebak jenis data & pemetaan kolom (`_ai_detect_schema`),
sekali lagi untuk menormalkan SETIAP batch baris (`_ai_normalize_rows`). Kalau
panggilan itu gagal — kuota habis, jaringan, atau model berubah perilaku — status
sesi jatuh ke ``queued`` lalu ``failed`` dan **tidak ada jalan lain**: staf tidak
bisa menyelesaikan pekerjaannya hari itu. Impor data adalah pekerjaan rutin harian
staf marketing; membuatnya bergantung pada layanan luar adalah keputusan yang
salah.

Karena itu urutannya dibalik di sini:

1. **Aturan pasti** (`exact`)   — header sama dengan nama/label kolom kanonik.
2. **Kamus sinonim** (`synonym`) — header lazim dari ekspor Shopee/TikTok/
   Tokopedia, ditulis di `marketing_import_schema.py`.
3. **Kemiripan teks** (`fuzzy`) — `difflib` ≥ 0.86 dipakai otomatis; 0.68–0.86
   hanya DIUSULKAN dan harus dikonfirmasi manusia. Ambang ini dipilih supaya
   "Harga Satuan" tidak pernah diam-diam dipetakan ke "Harga Coret".
4. **AI** — hanya kalau staf menekan tombolnya, dan hasilnya pun cuma *usulan*
   yang bisa ditolak. AI tidak pernah menjadi jalur wajib.

DUA CACAT LAIN YANG DIPERBAIKI DI SINI
--------------------------------------
* Mesin lama menulis ``**committed_data`` mentah ke koleksi
  (`universal_import.py:787-797`) — tanpa validasi tipe, tanpa ``account_id``.
  Hasilnya: baris impor tidak pernah muncul di layar yang difilter per toko.
  Di sini setiap nilai **dikonversi & divalidasi**, dan lingkup toko ditempel
  lewat SSOT `core.marketing_account_scope`.
* Mesin lama memetakan dua jenis ke koleksi yang tidak pernah dibaca layar
  (`marketing_discount_campaigns`, `marketing_sample_shipments`). Di sini tujuan
  koleksi hanya boleh berasal dari `marketing_import_schema.SOURCE_TYPES`.

CATATAN ANGKA (ini sumber salah hitung yang paling sering)
----------------------------------------------------------
File dari marketplace Indonesia menulis "Rp 1.250.000" sedangkan ekspor versi
Inggris menulis "1,250,000.50". Keduanya harus jadi 1250000 dan 1250000.5 — bukan
1.25 dan 1250000.50 yang tertukar. :func:`parse_number` memutuskan pemisah desimal
dari **pemisah terakhir** dan dari **panjang grup angka**, lalu menolak (bukan
menebak) kalau tetap ambigu.
"""
from __future__ import annotations

import io
import csv
import re
import difflib
import hashlib
import logging
from datetime import datetime, timezone, date
from typing import Any, Dict, List, Optional, Tuple

from core.marketing_import_schema import Field, SourceType

logger = logging.getLogger(__name__)

MAX_ROWS = 20_000
MAX_PREVIEW_ROWS = 500
FUZZY_AUTO = 0.86        # ≥ ini: dipetakan otomatis
FUZZY_SUGGEST = 0.68     # ≥ ini: hanya diusulkan, wajib konfirmasi


# ═══════════════════════════════════════════════════════════════════════════════
# 1. BACA BERKAS
# ═══════════════════════════════════════════════════════════════════════════════
def _norm_header(s: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").strip().lower()).strip()


def _read_xlsx(raw: bytes, read_only: bool) -> Tuple[List[str], List[Dict[str, Any]]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=read_only, data_only=True)
    ws = wb.active
    if read_only:
        # Berkas ekspor sering menyatakan dimensi yang salah (mis. `A1:A1` untuk
        # 65 kolom). Dalam mode read-only openpyxl mempercayai metadata itu.
        try:
            ws.reset_dimensions()
        except Exception as e:  # noqa: BLE001
            logger.debug("reset_dimensions tidak tersedia: %s", e)
    headers: List[str] = []
    rows: List[Dict[str, Any]] = []
    for values in ws.iter_rows(values_only=True):
        if not headers:
            if values is None or all(v in (None, "") for v in values):
                continue
            headers = [str(v).strip() if v is not None else f"kolom_{i+1}"
                       for i, v in enumerate(values)]
            continue
        if values is None or all(v in (None, "") for v in values):
            continue
        rows.append({headers[i]: values[i] if i < len(values) else None
                     for i in range(len(headers))})
        if len(rows) >= MAX_ROWS:
            break
    try:
        wb.close()
    except Exception as e:  # noqa: BLE001
        logger.debug("gagal menutup workbook: %s", e)
    return headers, rows


def parse_table(raw: bytes, filename: str,
                st: Optional[SourceType] = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    """CSV/XLSX → (headers, rows). Baris kosong dibuang, header kosong diberi nama.

    F7.2 — bila `st` membawa `prenorm`, berkas dinormalkan lebih dulu oleh
    `core.marketing_import_prenorm` (memotong metadata/blok section, memilih baris
    header yang benar, menggabungkan sheet) dan hasilnya sudah berkolom kanonik.
    Tanpa jalur ini, ekspor KPI Seller Center terbaca 0 kolom terpetakan.
    """
    if st is not None and getattr(st, "prenorm", ""):
        from core import marketing_import_prenorm as prenorm
        return prenorm.prenormalize(raw, filename, st.prenorm)

    name = (filename or "").lower()

    if name.endswith((".xlsx", ".xlsm", ".xls")):
        try:
            import openpyxl  # noqa: F401  (dipakai di _read_xlsx)
        except ImportError as e:
            raise ValueError("Server belum punya openpyxl untuk membaca Excel") from e
        headers, rows = _read_xlsx(raw, read_only=True)
        if len(headers) <= 1:
            # CACAT NYATA pada ekspor TikTok Seller Center (dibuktikan 2026-08-11):
            # setiap SEL dibungkus elemen `<row r="1">` sendiri, jadi pembaca aliran
            # (read-only) melihat 65 kolom sebagai 65 "baris" berisi 1 sel.
            # Mode biasa menyusun ulang lewat koordinat sel (`r="B1"`) sehingga
            # tabelnya benar. Tanpa jalur cadangan ini, seluruh berkas terbaca
            # "hanya 1 kolom" dan impor pesanan mustahil.
            logger.warning("xlsx '%s' terbaca 1 kolom di mode aliran — beralih ke "
                           "pembacaan penuh (berkas menulis 1 sel per <row>)", filename)
            headers, rows = _read_xlsx(raw, read_only=False)
        return headers, rows

    # CSV / TSV — deteksi pemisah, karena ekspor Indonesia sering pakai ';'
    text = raw.decode("utf-8-sig", errors="replace")
    sample = text[:4000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delim = dialect.delimiter
    except Exception:
        delim = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    headers = [h.strip() for h in (reader.fieldnames or []) if h is not None]
    rows = []
    for r in reader:
        if all((v is None or str(v).strip() == "") for v in r.values()):
            continue
        rows.append({(k or "").strip(): v for k, v in r.items() if k is not None})
        if len(rows) >= MAX_ROWS:
            break
    return headers, rows


# ═══════════════════════════════════════════════════════════════════════════════
# 2. PEMETAAN KOLOM (tanpa AI)
# ═══════════════════════════════════════════════════════════════════════════════
def auto_map(headers: List[str], st: SourceType) -> List[dict]:
    """Petakan header berkas → field kanonik.

    Hasil per kolom: ``{column, field, field_label, method, score, candidates}``
    ``method``: ``exact`` · ``synonym`` · ``fuzzy`` · ``suggest`` · ``none``
    ``suggest`` berarti **belum** dipakai; layar harus meminta konfirmasi.
    """
    fields = st.input_fields
    # tabel pencarian
    exact: Dict[str, Field] = {}
    for f in fields:
        exact[_norm_header(f.name)] = f
        exact[_norm_header(f.label)] = f
    syn: Dict[str, Field] = {}
    for f in fields:
        for s in f.synonyms:
            syn.setdefault(_norm_header(s), f)

    pool = list(exact.keys()) + list(syn.keys())
    taken: Dict[str, str] = {}           # field.name -> column
    result: List[dict] = []

    for col in headers:
        n = _norm_header(col)
        chosen: Optional[Field] = None
        method = "none"
        score = 0.0
        candidates: List[dict] = []

        if n in exact:
            chosen, method, score = exact[n], "exact", 1.0
        elif n in syn:
            chosen, method, score = syn[n], "synonym", 0.98
        elif n:
            close = difflib.get_close_matches(n, pool, n=3, cutoff=FUZZY_SUGGEST)
            scored = []
            for c in close:
                f = exact.get(c) or syn.get(c)
                if not f:
                    continue
                r = difflib.SequenceMatcher(None, n, c).ratio()
                scored.append((r, f, c))
            scored.sort(key=lambda x: -x[0])
            # jangan usulkan field yang sama dua kali
            seen = set()
            for r, f, c in scored:
                if f.name in seen:
                    continue
                seen.add(f.name)
                candidates.append({"field": f.name, "field_label": f.label,
                                   "matched_on": c, "score": round(r, 3)})
            if candidates:
                top = candidates[0]
                f = st.field(top["field"])
                if top["score"] >= FUZZY_AUTO and f is not None:
                    chosen, method, score = f, "fuzzy", top["score"]
                else:
                    method, score = "suggest", top["score"]

        if chosen is not None and chosen.name in taken:
            # dua kolom menunjuk satu field: yang kedua jadi usulan, tidak dipakai
            candidates.insert(0, {"field": chosen.name, "field_label": chosen.label,
                                  "matched_on": n, "score": round(score, 3)})
            result.append({"column": col, "field": None, "field_label": None,
                           "method": "suggest", "score": round(score, 3),
                           "candidates": candidates[:3],
                           "note": f"kolom '{taken[chosen.name]}' sudah dipetakan "
                                   f"ke {chosen.label}"})
            continue

        if chosen is not None:
            taken[chosen.name] = col
        result.append({
            "column": col,
            "field": chosen.name if chosen else None,
            "field_label": chosen.label if chosen else None,
            "method": method,
            "score": round(score, 3),
            "candidates": _cand_list(chosen, n, score, candidates),
        })
    return result


def _cand_list(chosen: Optional[Field], norm_col: str, score: float,
               candidates: List[dict]) -> List[dict]:
    """Daftar usulan yang SELALU memuat pilihan mesin sebagai usulan #1.

    Tanpa ini, kolom yang dikenali ``exact``/``synonym`` punya ``candidates: []``
    — sehingga begitu staf melepas kolom itu (“— tidak dipakai —”) satu-satunya
    jalan kembali adalah menelusuri 40+ nama field di dropdown, atau mengunggah
    ulang berkasnya. Pilihan ``fuzzy`` sudah ada di dalam ``candidates``, jadi
    di-dedupe agar tidak muncul dua kali.
    """
    out: List[dict] = []
    seen: set = set()
    if chosen is not None:
        out.append({"field": chosen.name, "field_label": chosen.label,
                    "matched_on": norm_col, "score": round(score, 3)})
        seen.add(chosen.name)
    for c in candidates:
        if c.get("field") in seen:
            continue
        seen.add(c.get("field"))
        out.append(c)
    return out[:4]


def mapping_dict(mapping: List[dict]) -> Dict[str, str]:
    """``[{column, field}]`` → ``{column: field}`` (hanya yang benar-benar dipakai)."""
    out = {}
    for m in mapping or []:
        col, fld = m.get("column"), m.get("field")
        if col and fld and m.get("method") != "suggest":
            out[col] = fld
    return out


def mapping_report(mapping: List[dict], st: SourceType) -> dict:
    used = mapping_dict(mapping)
    mapped_fields = set(used.values())
    missing_required = [f.label for f in st.required_fields if f.name not in mapped_fields]
    unmapped_cols = [m["column"] for m in mapping if not m.get("field")
                     or m.get("method") == "suggest"]
    return {
        "mapped": len(used),
        "total_columns": len(mapping),
        "missing_required": missing_required,
        "unmapped_columns": unmapped_cols,
        "ready": len(missing_required) == 0,
        "methods": {k: sum(1 for m in mapping if m.get("method") == k)
                    for k in ("exact", "synonym", "fuzzy", "suggest", "none")},
    }


# ═══════════════════════════════════════════════════════════════════════════════
# 3. KONVERSI NILAI
# ═══════════════════════════════════════════════════════════════════════════════
_CURRENCY = re.compile(r"(?i)\b(rp|idr|rupiah)\b|[^\d,.\-+eE]")


def parse_number(raw: Any) -> Tuple[Optional[float], Optional[str]]:
    """Angka dari teks apa pun yang lazim di ekspor marketplace.

    Menerima: ``1250000`` · ``"Rp 1.250.000"`` · ``"1,250,000.50"`` ·
    ``"1.250.000,50"`` · ``"(1.500)"`` (negatif akuntansi) · ``"12,5%"``.
    Menolak (bukan menebak) kalau pemisahnya tidak konsisten.
    """
    if raw is None or raw == "":
        return None, None
    if isinstance(raw, bool):
        return (1.0 if raw else 0.0), None
    if isinstance(raw, (int, float)):
        return float(raw), None

    s = str(raw).strip()
    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")
    s = _CURRENCY.sub("", s).strip()
    if s in ("", "-", "+"):
        return None, None
    if s.startswith("-"):
        neg = True
        s = s[1:]

    has_dot, has_com = "." in s, "," in s
    if has_dot and has_com:
        dec = "." if s.rfind(".") > s.rfind(",") else ","
        thou = "," if dec == "." else "."
        s = s.replace(thou, "").replace(dec, ".")
    elif has_com:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            s = s.replace(",", ".")                     # desimal
        elif all(len(p) == 3 for p in parts[1:]):
            s = s.replace(",", "")                      # ribuan
        else:
            return None, f"format angka '{raw}' ambigu (pemisah ',' tidak konsisten)"
    elif has_dot:
        parts = s.split(".")
        if len(parts) > 2 and all(len(p) == 3 for p in parts[1:]):
            s = s.replace(".", "")                      # 1.250.000
        elif len(parts) == 2 and len(parts[1]) == 3 and len(parts[0]) <= 3:
            s = s.replace(".", "")                      # 1.250 → 1250 (gaya Indonesia)
        # sisanya dianggap desimal biasa
    try:
        v = float(s)
    except ValueError:
        return None, f"'{raw}' bukan angka"
    return (-v if neg else v), None


_DATE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d/%m/%Y %H:%M:%S",
    "%d/%m/%Y %H:%M", "%d-%m-%Y %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
    "%d %b %Y", "%d %B %Y", "%b %d, %Y",
)
_ID_MONTHS = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4, "mei": 5, "juni": 6,
    "juli": 7, "agustus": 8, "september": 9, "oktober": 10, "november": 11,
    "desember": 12, "jan": 1, "feb": 2, "mar": 3, "apr": 4, "jun": 6, "jul": 7,
    "agu": 8, "ags": 8, "sep": 9, "okt": 10, "nov": 11, "des": 12,
}


def parse_date(raw: Any) -> Tuple[Optional[datetime], Optional[str]]:
    """Tanggal dari teks/serial Excel → datetime (UTC, naive-aware)."""
    if raw is None or raw == "":
        return None, None
    if isinstance(raw, datetime):
        return (raw if raw.tzinfo else raw.replace(tzinfo=timezone.utc)), None
    if isinstance(raw, date):
        return datetime(raw.year, raw.month, raw.day, tzinfo=timezone.utc), None
    if isinstance(raw, (int, float)) and 20000 < float(raw) < 60000:
        # serial Excel (1899-12-30 sebagai hari ke-0)
        from datetime import timedelta
        base = datetime(1899, 12, 30, tzinfo=timezone.utc)
        return base + timedelta(days=float(raw)), None

    s = str(raw).strip()
    if not s:
        return None, None
    s = s.replace("Z", "").strip()
    for fmt in _DATE_FORMATS:
        try:
            d = datetime.strptime(s[:len(datetime.now().strftime(fmt)) + 6], fmt)
            return d.replace(tzinfo=timezone.utc), None
        except Exception:
            pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc), None
        except Exception:
            continue
    # "12 Agustus 2026"
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$", s)
    if m:
        mon = _ID_MONTHS.get(m.group(2).lower())
        if mon:
            return datetime(int(m.group(3)), mon, int(m.group(1)),
                            tzinfo=timezone.utc), None
    try:
        return datetime.fromisoformat(s).replace(tzinfo=timezone.utc), None
    except Exception:
        pass
    return None, f"'{raw}' bukan tanggal yang dikenali (pakai YYYY-MM-DD atau DD/MM/YYYY)"


_TRUE = {"1", "true", "ya", "yes", "y", "aktif", "active", "benar", "ada"}
_FALSE = {"0", "false", "tidak", "no", "n", "nonaktif", "inactive", "salah", "-"}


def _norm_choice(s: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(s or "").strip().lower()).strip("_")


def coerce(f: Field, raw: Any) -> Tuple[Any, Optional[str], Optional[str]]:
    """Konversi satu sel. → ``(nilai, error, warning)``."""
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        if f.required:
            return None, f"{f.label} wajib diisi", None
        return None, None, None

    pre_warn: Optional[str] = None
    # ── F1 — KAMUS NILAI (value_map): "Perlu dikirim" → paid ─────────────────
    # Ditulis, bukan ditebak. Nilai yang tidak ada di kamus:
    #   · field penting (status)      ⇒ baris DITOLAK, pesan memuat nilai aslinya
    #   · field keterangan (kurir/kanal, punya `value_map_fallback`) ⇒ jatuh ke
    #     nilai fallback + PERINGATAN, supaya nama kurir baru tidak menghapus omzet
    if f.value_map:
        vm = {_norm_choice(k): v for k, v in f.value_map}
        canon = {_norm_choice(c) for c in f.choices}
        s_norm = _norm_choice(raw)
        if s_norm in vm:
            raw = vm[s_norm]
        elif s_norm in canon:
            pass                                   # sudah kanonik (mis. hasil re-impor)
        else:
            close = difflib.get_close_matches(s_norm, list(vm.keys()), n=1, cutoff=0.88)
            if close:
                raw = vm[close[0]]
                pre_warn = (f"{f.label}: '{raw}' dipadankan dari nilai mirip "
                            f"'{close[0]}' — periksa sekali")
            elif f.value_map_fallback:
                pre_warn = (f"{f.label}: nilai '{raw}' belum ada di kamus, "
                            f"dicatat sebagai '{f.value_map_fallback}' (nilai asli tetap disimpan)")
                raw = f.value_map_fallback
            else:
                return None, (f"{f.label}: nilai '{raw}' tidak ada di kamus. "
                              f"Nilai yang dikenali: "
                              + ", ".join(sorted({k for k, _ in f.value_map})[:8])
                              + ". Tambahkan ke kamus dulu — jangan dipaksa masuk."), None

    k = f.kind
    if k in ("str", "list"):
        s = str(raw).strip()
        if k == "list":
            return [x.strip() for x in re.split(r"[;,]", s) if x.strip()], None, pre_warn
        return s, None, pre_warn

    if k == "bool":
        s = _norm_choice(raw)
        if s in _TRUE:
            return True, None, pre_warn
        if s in _FALSE:
            return False, None, pre_warn
        return None, f"{f.label}: '{raw}' bukan ya/tidak", None

    if k == "enum":
        s = _norm_choice(raw)
        allowed = {_norm_choice(c): c for c in f.choices}
        if s in allowed:
            return allowed[s], None, pre_warn
        # kecocokan longgar: "Flash Sale!" → flash_sale
        close = difflib.get_close_matches(s, list(allowed.keys()), n=1, cutoff=0.8)
        if close:
            return allowed[close[0]], None, (
                f"{f.label}: '{raw}' ditafsirkan sebagai '{allowed[close[0]]}'")
        return None, (f"{f.label}: '{raw}' tidak dikenali. Pilihan yang sah: "
                      + ", ".join(f.choices)), None

    if k in ("date", "datetime"):
        d, err = parse_date(raw)
        if err:
            return None, f"{f.label}: {err}", None
        if d is None:
            return None, None, pre_warn
        return (d.date().isoformat() if k == "date" else d), None, pre_warn

    v, err = parse_number(raw)
    if err:
        return None, f"{f.label}: {err}", None
    if v is None:
        return None, None, pre_warn
    if k == "int":
        if abs(v - round(v)) > 1e-6:
            return int(round(v)), None, f"{f.label}: {raw} dibulatkan ke {int(round(v))}"
        v = int(round(v))
    if k in ("money", "int", "num") and v < 0:
        return v, None, f"{f.label}: nilai negatif ({v}) — pastikan bukan salah tanda"
    if k == "pct":
        if v < 0:
            return None, f"{f.label}: persen tidak boleh negatif", None
        if v <= 1 and v != 0:
            return v * 100, None, (
                f"{f.label}: {raw} ditafsirkan {v*100:.4g}% (skala 0–1 → 0–100)")
        if v > 100:
            return v, None, f"{f.label}: {v} lebih dari 100% — periksa sekali"
    return v, None, pre_warn


# ═══════════════════════════════════════════════════════════════════════════════
# 4. BANGUN & VALIDASI BARIS
# ═══════════════════════════════════════════════════════════════════════════════
def strip_description_rows(rows: List[dict], mapping: List[dict],
                           st: SourceType) -> Tuple[List[dict], int]:
    """Buang baris DESKRIPSI KOLOM yang dipasang marketplace di baris pertama.

    Ekspor TikTok Seller Center menaruh keterangan kolom di baris ke-2 berkas
    ("Platform unique order ID.", "Current order status.", …). Tanpa ini, baris
    itu ikut dihitung sebagai data: 601 baris menjadi 602, dan satu "pesanan"
    palsu ber-`order_id` = teks keterangan lolos ke laporan.

    Aturan (sengaja ketat supaya baris data asli tidak pernah terbuang):
    baris PERTAMA dianggap deskripsi bila **≥60%** sel pada kolom bertipe
    ``money/int/num/date/datetime`` berisi teks yang **gagal** di-parse sebagai
    angka/tanggal **dan** panjang teksnya > 15 karakter.
    """
    if not rows:
        return rows, 0
    used = mapping_dict(mapping)
    checked = 0
    texty = 0
    first = rows[0]
    for col, fname in used.items():
        f = st.field(fname)
        if f is None or f.kind not in ("money", "int", "num", "date", "datetime"):
            continue
        checked += 1
        raw = first.get(col)
        if raw is None or str(raw).strip() == "":
            continue
        s = str(raw).strip()
        if f.kind in ("date", "datetime"):
            val, err = parse_date(s)
            failed = err is not None or val is None
        else:
            val, err = parse_number(s)
            failed = err is not None or val is None
        if failed and len(s) > 15:
            texty += 1
    if checked >= 3 and texty / checked >= 0.6:
        return rows[1:], 1
    return rows, 0


def format_fingerprint(headers: List[str]) -> str:
    """Sidik jari susunan kolom berkas — dipakai untuk MENGINGAT pemetaan.

    Fingerprint yang sama ⇒ pemetaan langsung dipakai (tanpa AI, tanpa tanya).
    Fingerprint baru ⇒ wizard MEMINTA konfirmasi; tidak pernah menebak diam-diam.
    """
    norm = [_norm_header(h) for h in (headers or []) if str(h or "").strip() != ""]
    return hashlib.sha1("|".join(norm).encode("utf-8")).hexdigest()


def _coerce_row(raw_row: dict, used: Dict[str, str], st: SourceType) -> Tuple[dict, List[str], List[str]]:
    """Satu baris berkas → (data kanonik, errors, warnings). Tanpa database."""
    data: Dict[str, Any] = {}
    errors: List[str] = []
    warnings: List[str] = []
    for col, fname in used.items():
        f = st.field(fname)
        if f is None:
            continue
        raw = raw_row.get(col)
        val, err, warn = coerce(f, raw)
        if err:
            errors.append(err)
        if warn:
            warnings.append(warn)
        if val is not None:
            data[fname] = val
            if f.keep_raw and raw not in (None, ""):
                data[f"{fname}_raw"] = str(raw).strip()
    return data, errors, warnings


def build_rows(rows: List[dict], mapping: List[dict], st: SourceType,
               limit: int = MAX_PREVIEW_ROWS) -> List[dict]:
    """Terapkan mapping + konversi + validasi. Tidak menyentuh database.

    Untuk jenis ber-``group_by`` (F1: pesanan marketplace), BANYAK baris berkas
    menjadi **satu** entri hasil ber-``items[]`` — lihat :func:`build_grouped_rows`.
    """
    used = mapping_dict(mapping)
    if st.is_grouped:
        # Jenis berkelompok: 1 entri = 1 PESANAN, bukan 1 baris berkas. Batas
        # pratinjau per-baris tidak boleh memotong jumlah pesanan, kalau tidak
        # layar melaporkan "500 pesanan" untuk berkas 559 pesanan (dan staf
        # mengira 59 pesanan hilang). Baris sudah dibatasi MAX_ROWS saat dibaca.
        return build_grouped_rows(rows, used, st, limit=MAX_ROWS)
    out: List[dict] = []
    for i, raw_row in enumerate(rows[:limit]):
        data, errors, warnings = _coerce_row(raw_row, used, st)
        # field wajib yang tidak terpetakan sama sekali
        for f in st.required_fields:
            if f.name not in data and f.label not in " ".join(errors):
                errors.append(f"{f.label} wajib diisi")
        out.append({
            "row_id": i,
            "original": {k: (v if not isinstance(v, (datetime, date)) else str(v))
                         for k, v in raw_row.items()},
            "data": data,
            "errors": errors,
            "warnings": warnings,
            "status": "error" if errors else ("warning" if warnings else "valid"),
        })
    return out


def build_grouped_rows(rows: List[dict], used: Dict[str, str], st: SourceType,
                       limit: int = MAX_PREVIEW_ROWS) -> List[dict]:
    """Kelompokkan baris per ``st.group_by`` ⇒ 1 entri = 1 dokumen ber-``items[]``.

    KENAPA (SSOT_KONTRAK_DATA §2): ekspor memberi 1 baris per SKU. Kalau setiap
    baris disimpan sebagai dokumen sendiri, 41 pembaca yang menghitung
    ``$sum: 1`` sebagai "jumlah pesanan" dan ``$sum: '$total_payment'`` sebagai
    omzet akan **salah** (601 pesanan & omzet dobel, padahal 559 pesanan).

    Aturan:
    * field di ``item_fields``    → masuk ``items[]`` (1 elemen per baris)
    * field lain                 → header, nilai dari **baris pertama** pesanan;
                                   nilai berbeda antar baris ⇒ **peringatan**
                                   (bukan galat) dengan menyebut kedua nilainya
    * field di ``per_order_money`` → hanya dari baris pertama, **dilarang** dijumlah
    """
    groups: Dict[tuple, dict] = {}
    order: List[tuple] = []
    orphan: List[dict] = []
    item_required = [f for f in st.required_fields if f.name in st.item_fields]
    header_required = [f for f in st.required_fields if f.name not in st.item_fields]

    for i, raw_row in enumerate(rows):
        data, errors, warnings = _coerce_row(raw_row, used, st)
        key = tuple(str(data.get(g)) if data.get(g) is not None else "" for g in st.group_by)
        if any(k == "" for k in key):
            orphan.append({
                "row_id": i,
                "original": {k: (v if not isinstance(v, (datetime, date)) else str(v))
                             for k, v in raw_row.items()},
                "data": data,
                "errors": errors + [f"{', '.join(st.group_by)} kosong — baris tidak bisa "
                                    f"dikelompokkan ke pesanan mana pun"],
                "warnings": warnings,
                "status": "error",
            })
            continue

        g = groups.get(key)
        if g is None:
            g = groups[key] = {
                "row_id": i, "raw_row_ids": [], "header": {}, "items": [],
                "errors": [], "warnings": [], "first_original": {
                    k: (v if not isinstance(v, (datetime, date)) else str(v))
                    for k, v in raw_row.items()},
            }
            order.append(key)
        g["raw_row_ids"].append(i)
        g["errors"].extend(errors)
        g["warnings"].extend(warnings)

        # ── items[] ──────────────────────────────────────────────────────────
        item = {k: v for k, v in data.items()
                if k in st.item_fields or (k.endswith("_raw") and k[:-4] in st.item_fields)}
        for f in item_required:
            if f.name not in item:
                g["errors"].append(f"baris berkas ke-{i + 2}: {f.label} wajib diisi")
        item["line_no"] = len(g["items"]) + 1
        item["raw_row_id"] = i
        g["items"].append(item)

        # ── header (nilai baris pertama; beda ⇒ peringatan) ──────────────────
        for k, v in data.items():
            if k in st.item_fields or (k.endswith("_raw") and k[:-4] in st.item_fields):
                continue
            if k not in g["header"]:
                g["header"][k] = v
            elif g["header"][k] != v:
                fld = st.field(k)
                lbl = fld.label if fld else k
                note = "dipakai nilai baris pertama"
                if k in st.per_order_money:
                    note = "dipakai nilai baris pertama — uang per pesanan TIDAK dijumlah"
                g["warnings"].append(
                    f"{lbl} berbeda antar baris pesanan ini "
                    f"('{g['header'][k]}' vs '{v}') — {note}")

    out: List[dict] = list(orphan)
    for key in order[:limit]:
        g = groups[key]
        errors = list(dict.fromkeys(g["errors"]))
        warnings = list(dict.fromkeys(g["warnings"]))
        for f in header_required:
            if f.name not in g["header"] and f.label not in " ".join(errors):
                errors.append(f"{f.label} wajib diisi")
        data = dict(g["header"])
        data["items"] = g["items"]
        data["raw_row_ids"] = g["raw_row_ids"]
        out.append({
            "row_id": g["row_id"],
            "group_key": " / ".join(key),
            "rows_in_group": len(g["raw_row_ids"]),
            "items_count": len(g["items"]),
            "original": g["first_original"],
            "data": data,
            "errors": errors,
            "warnings": warnings,
            "status": "error" if errors else ("warning" if warnings else "valid"),
        })
    out.sort(key=lambda r: r["row_id"])
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# 5. TEMPLATE UNDUHAN
# ═══════════════════════════════════════════════════════════════════════════════
def build_template_csv(st: SourceType) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    fields = st.input_fields
    w.writerow([f.label for f in fields])
    w.writerow([f.example or "" for f in fields])
    return buf.getvalue().encode("utf-8-sig")


def build_template_xlsx(st: SourceType) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    fields = st.input_fields

    head_fill = PatternFill("solid", fgColor="1F2937")
    req_fill = PatternFill("solid", fgColor="B91C1C")
    for c, f in enumerate(fields, 1):
        cell = ws.cell(row=1, column=c, value=f.label + (" *" if f.required else ""))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = req_fill if f.required else head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=c, value=f.example or None)
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(34, len(f.label) + 6))
    ws.freeze_panes = "A3"

    doc = wb.create_sheet("Petunjuk")
    doc.column_dimensions["A"].width = 30
    doc.column_dimensions["B"].width = 16
    doc.column_dimensions["C"].width = 10
    doc.column_dimensions["D"].width = 62
    doc["A1"] = f"TEMPLATE IMPOR — {st.label}"
    doc["A1"].font = Font(bold=True, size=13)
    doc["A2"] = st.describe
    doc["A3"] = f"Data masuk ke tabel: {st.collection} · tampil di menu: {st.module_hint}"
    doc["A4"] = ("Toko/akun TIDAK diisi di file ini — dipilih di layar sebelum upload, "
                 "supaya satu file tidak bisa tercampur antar toko.")
    for i, h in enumerate(["Kolom", "Jenis Isi", "Wajib", "Keterangan / Pilihan yang sah"], 1):
        c = doc.cell(row=6, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head_fill
    r = 7
    kind_label = {"str": "teks", "date": "tanggal", "datetime": "tanggal",
                  "int": "bilangan bulat", "num": "angka", "money": "rupiah",
                  "pct": "persen (0-100)", "bool": "ya / tidak", "enum": "pilihan",
                  "list": "daftar (pisah koma)"}
    for f in fields:
        doc.cell(row=r, column=1, value=f.label)
        doc.cell(row=r, column=2, value=kind_label.get(f.kind, f.kind))
        doc.cell(row=r, column=3, value="WAJIB" if f.required else "opsional")
        note = f.note
        if f.choices:
            note = (note + " · " if note else "") + "pilihan: " + ", ".join(f.choices)
        if f.synonyms:
            note = (note + " · " if note else "") + "header lain yang dikenali: " + \
                   ", ".join(f.synonyms[:6])
        doc.cell(row=r, column=4, value=note or "-").alignment = Alignment(wrap_text=True)
        r += 1

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# 6. DETEKSI JENIS DATA & PLATFORM (sesi #34)
# ═══════════════════════════════════════════════════════════════════════════════
# KENAPA: staf memilih jenis data SENDIRI (aturan yang tetap dipertahankan — AI
# tidak boleh memutuskan tabel tujuan). Tetapi pilihan manusia BISA SALAH, dan
# salah pilih jenis berarti berkas "pesanan" masuk sebagai "penjualan harian".
# Modul ini tidak mengambil keputusan; ia MENGUKUR kecocokan header berkas
# terhadap SETIAP jenis dan melaporkannya, supaya layar bisa berkata:
#
#   "Anda memilih Penjualan Harian, tetapi 39 dari 50 kolom berkas ini cocok
#    dengan Pesanan Marketplace (dan hanya 3 kolom cocok dengan pilihan Anda)."
#
# Platform juga tidak ditebak dari nama berkas: ia dibaca dari SIDIK HEADER
# (kolom khas Shopee vs TikTok) dan dibandingkan dengan platform toko yang
# dipilih staf — inilah yang diminta pemilik: "platform apa seharusnya sudah
# terdeteksi oleh system ketika input akun tokonya".

# Sidik header per platform: header yang HANYA muncul di ekspor platform itu.
PLATFORM_FINGERPRINTS = {
    "shopee": (
        "no pesanan", "status pesanan", "sku induk", "nomor referensi sku",
        "no pengembalian", "username pembeli", "voucher ditanggung shopee",
        "diskon dari shopee", "waktu pesanan dibuat", "opsi pengiriman",
        "potongan koin shopee", "cashback koin", "no resi",
        "jumlah produk dikembalikan", "solusi pengembalian barang dana",
    ),
    "tiktok": (
        "order id", "order substatus", "seller sku", "sku id",
        "cancelation return type", "purchase channel", "warehouse name",
        "creator handle", "return order id", "buyer username",
        "nama kampanye", "id campaign", "akun tiktok", "pesanan sku",
        "pendapatan kotor", "jenis materi iklan",
    ),
    "tokopedia": ("tokopedia invoice number", "nomor invoice", "kode pesanan"),
}


def detect_platform(headers: List[str]) -> dict:
    """Platform dari SIDIK HEADER berkas (bukan dari nama berkas)."""
    norm = {_norm_header(h) for h in headers if h}
    hits = {}
    for plat, marks in PLATFORM_FINGERPRINTS.items():
        matched = [m for m in marks if _norm_header(m) in norm]
        if matched:
            hits[plat] = matched
    if not hits:
        return {"platform": "", "confidence": 0.0, "evidence": [], "candidates": {}}
    best = max(hits.items(), key=lambda kv: len(kv[1]))
    total = sum(len(v) for v in hits.values())
    return {
        "platform": best[0],
        "confidence": round(len(best[1]) / total, 3) if total else 0.0,
        "evidence": best[1][:6],
        "candidates": {k: len(v) for k, v in hits.items()},
    }


def score_headers(headers: List[str], st: SourceType) -> dict:
    """Seberapa cocok header berkas dengan satu jenis data. 0..1, dapat diperiksa."""
    mapping = auto_map(headers, st)
    mapped = [m for m in mapping if m.get("field")]
    mapped_names = {m["field"] for m in mapped}
    req = [f for f in st.input_fields if f.required]
    req_hit = [f for f in req if f.name in mapped_names]
    req_cover = (len(req_hit) / len(req)) if req else 1.0
    col_cover = (len(mapped) / len(headers)) if headers else 0.0
    return {
        "source_type": st.key, "label": st.label, "group": st.group,
        "mapped_columns": len(mapped), "total_columns": len(headers),
        "required_hit": len(req_hit), "required_total": len(req),
        "required_missing": [f.label for f in req if f.name not in mapped_names],
        "required_cover": round(req_cover, 3),
        "column_cover": round(col_cover, 3),
        # Kolom WAJIB menimbang paling berat: jenis yang kolom wajibnya tidak
        # lengkap TIDAK PERNAH boleh menang hanya karena banyak kolom nyasar.
        "score": round(req_cover * 0.7 + col_cover * 0.3, 4),
    }


def detect_source_type(raw: bytes, filename: str, catalog: List[SourceType],
                       *, top: int = 5) -> dict:
    """Urutkan SEMUA jenis data menurut kecocokan header berkas.

    Berkas dibaca sekali tanpa prenorm, lalu SEKALI LAGI untuk tiap penormal
    yang dipakai jenis-jenis ber-`prenorm` (ekspor Seller Center yang baris
    pertamanya bukan header). Tanpa itu, jenis ber-prenorm selalu tampak 0 kolom.
    """
    parses: Dict[str, Tuple[List[str], List[dict]]] = {}
    try:
        parses[""] = parse_table(raw, filename, None)
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "error": f"berkas tidak bisa dibaca: {e}", "results": []}
    for st in catalog:
        key = getattr(st, "prenorm", "") or ""
        if key and key not in parses:
            try:
                parses[key] = parse_table(raw, filename, st)
            except Exception:  # noqa: BLE001
                parses[key] = ([], [])

    results = []
    for st in catalog:
        headers, rows = parses.get(getattr(st, "prenorm", "") or "", ([], []))
        if not headers:
            continue
        r = score_headers(headers, st)
        r["prenorm"] = getattr(st, "prenorm", "") or ""
        r["rows_readable"] = len(rows)
        results.append(r)
    results.sort(key=lambda r: (-r["score"], -r["mapped_columns"]))
    base_headers, base_rows = parses.get("", ([], []))
    plat = detect_platform(base_headers)
    if not plat.get("platform"):
        # Jalan mundur yang JUJUR: ekspor yang baris pertamanya judul ("Semua
        # Laporan Iklan CPC - Shopee Indonesia") tidak punya sidik kolom sama
        # sekali. Nama platform di JUDUL/nama berkas dipakai, dan buktinya
        # disebut supaya staf bisa membantah.
        blob = " ".join([str(h) for h in base_headers[:5]] + [filename or ""]).lower()
        for plat_name in ("shopee", "tiktok", "tokopedia", "lazada"):
            if plat_name in blob:
                plat = {"platform": plat_name, "confidence": 0.5,
                        "evidence": [f"kata '{plat_name}' di judul/nama berkas"],
                        "candidates": {plat_name: 1}}
                break
        else:
            if results and results[0]["source_type"].startswith("shopee_"):
                plat = {"platform": "shopee", "confidence": 0.5,
                        "evidence": ["jenis data ini hanya ada di Seller Center Shopee"],
                        "candidates": {"shopee": 1}}
    return {
        "ok": True,
        "headers": base_headers,
        "row_count": len(base_rows),
        "platform": plat,
        "best": results[0] if results else None,
        "results": results[:top],
    }
