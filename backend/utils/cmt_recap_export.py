"""utils.cmt_recap_export — Rekap Harian CMT dalam Excel & PDF.

KENAPA ADA
----------
Rekap "vendor mana yang belum diisi hari ini" bukan cuma untuk dilihat: supervisor
membacakannya di briefing pagi, mengirimkannya ke grup, dan mengarsipkannya sebagai
bukti bahwa vendor sudah/belum ditagih datanya. Karena itu perlu lampiran:

* **Excel** — untuk diolah (filter per vendor, digabung rekap mingguan);
* **PDF**   — untuk dicetak/ditandatangani supervisor.

ISI SAMA PERSIS DENGAN LAYAR. Keduanya menerima objek hasil
``core.cmt_daily_recap.build_recap()`` apa adanya — TIDAK menghitung ulang apa pun.
Kalau file ini ikut menghitung, suatu hari Excel akan bilang 5 vendor merah
sementara layarnya bilang 3, dan tidak ada yang tahu mana yang benar.

Tanpa dependensi baru: ``openpyxl`` & ``reportlab`` sudah ada di ``requirements.txt``.
"""
from __future__ import annotations

import io
from datetime import datetime

from utils.waktu import fmt_wib

# Lambang status — dipakai IDENTIK di Excel & PDF supaya rekap tercetak bisa
# dibandingkan langsung dengan layar tanpa kamus terjemahan di kepala pembaca.
STATE_MARK = {'done': 'OK', 'partial': 'OK*', 'pending': 'BELUM', 'none': '-'}
STATE_LABEL = {
    'pending': 'BELUM DIISI',
    'partial': 'SEBAGIAN',
    'done': 'LENGKAP',
    'idle': 'Tidak ada pekerjaan',
}
SOURCE_LABEL = {'staff': 'staf DA', 'vendor': 'vendor', 'mixed': 'staf DA + vendor', '': ''}

_MONTHS_ID = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli',
              'Agustus', 'September', 'Oktober', 'November', 'Desember']


def date_label(day: str) -> str:
    """'2026-08-08' → 'Sabtu, 8 Agustus 2026' (nama hari ikut — briefing pagi
    menyebut hari, bukan tanggal)."""
    try:
        d = datetime.strptime(day[:10], '%Y-%m-%d')
    except (ValueError, TypeError):
        return str(day)
    hari = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu'][d.weekday()]
    return f'{hari}, {d.day} {_MONTHS_ID[d.month - 1]} {d.year}'


def _filename(fmt: str, day: str) -> str:
    return f"rekap-harian-cmt-{str(day).replace('-', '')}.{fmt}"


def _cell_text(task: dict) -> str:
    """Satu sel tabel: lambang + keterangan (mis. ``BELUM · 1 surat jalan …``)."""
    mark = STATE_MARK.get(task.get('state'), '-')
    detail = task.get('detail') or ''
    src = SOURCE_LABEL.get(task.get('source') or '', '')
    if src:
        detail = f'{detail} ({src})'
    return f'{mark} · {detail}' if detail else mark


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def build_xlsx(*, company: str, recap: dict) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    tasks = recap.get('tasks') or []
    rows = recap.get('rows') or []
    summary = recap.get('summary') or {}
    day = recap.get('date', '')

    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill('solid', fgColor='EDE9FE')
    fill_pending = PatternFill('solid', fgColor='FEE2E2')
    fill_partial = PatternFill('solid', fgColor='FEF3C7')
    fill_done = PatternFill('solid', fgColor='DCFCE7')
    fill_idle = PatternFill('solid', fgColor='F4F4F5')
    thin = Side(style='thin', color='D4D4D8')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {'pending': fill_pending, 'partial': fill_partial,
             'done': fill_done, 'none': fill_idle, 'idle': fill_idle}

    ws = wb.active
    ws.title = 'Rekap Harian'

    r = 1
    for text, size, is_bold in (
        (company, 14, True),
        ('REKAP HARIAN PENGISIAN VENDOR CMT', 12, True),
        (f'Tanggal: {date_label(day)}', 10, False),
        (f"Dicetak: {fmt_wib('%d/%m/%Y %H:%M')} WIB", 9, False),
    ):
        ws.cell(row=r, column=1, value=text).font = Font(bold=is_bold, size=size)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='RINGKASAN').font = bold
    r += 1
    for label, value in (
        ('Vendor aktif', summary.get('vendors_total', 0)),
        ('BELUM diisi (ada tugas merah)', summary.get('vendors_pending', 0)),
        ('Sebagian terisi', summary.get('vendors_partial', 0)),
        ('Lengkap', summary.get('vendors_done', 0)),
        ('Tidak ada pekerjaan', summary.get('vendors_idle', 0)),
        ('Total tugas belum diisi', summary.get('tasks_pending_total', 0)),
        ('Pcs progress masuk hari ini', summary.get('qty_progress_today', 0)),
        ('Pcs dikirim hari ini', summary.get('qty_shipped_today', 0)),
    ):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    headers = ['Vendor', 'Kode', 'Status'] + [t['label'] for t in tasks] + ['Kontak', 'Akun portal']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = bold
        cell.fill = head_fill
        cell.border = box
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r += 1

    for row in rows:
        status = row.get('status', 'idle')
        vals = [row.get('vendor_name', ''), row.get('vendor_code', ''), STATE_LABEL.get(status, status)]
        for t in tasks:
            vals.append(_cell_text((row.get('tasks') or {}).get(t['key'], {})))
        kontak = ' / '.join([x for x in (row.get('contact_name'), row.get('contact_phone')) if x]) or '-'
        vals.append(kontak)
        vals.append('Ada akun aktif' if row.get('has_active_portal_account') else 'Tidak ada akun')
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = box
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if c == 3:
                cell.fill = fills.get(status, fill_idle)
                cell.font = bold
            elif 4 <= c <= 3 + len(tasks):
                st = ((row.get('tasks') or {}).get(tasks[c - 4]['key'], {}) or {}).get('state', 'none')
                cell.fill = fills.get(st, fill_idle)
        r += 1

    if not rows:
        ws.cell(row=r, column=1, value='(tidak ada vendor CMT aktif di master)')
        r += 1

    r += 1
    ws.cell(row=r, column=1,
            value=('Keterangan: OK = ada pengisian pada tanggal ini · OK* = ada pengisian '
                   'tapi masih ada sisa pekerjaan · BELUM = ada pekerjaan menunggu dan '
                   'belum diisi · "-" = memang tidak ada pekerjaan jenis itu.')).font = Font(size=9)

    widths = [30, 12, 16] + [30] * len(tasks) + [24, 16]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), _filename('xlsx', day)


# ─────────────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────────────
def build_pdf(*, company: str, recap: dict) -> tuple[bytes, str]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    tasks = recap.get('tasks') or []
    rows = recap.get('rows') or []
    summary = recap.get('summary') or {}
    day = recap.get('date', '')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm,
                            title=f'Rekap Harian CMT {day}')
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle('h1x', parent=ss['Title'], fontSize=14, leading=17, spaceAfter=2)
    h2 = ParagraphStyle('h2x', parent=ss['Normal'], fontSize=10, leading=13,
                        textColor=colors.HexColor('#52525B'))
    small = ParagraphStyle('smallx', parent=ss['Normal'], fontSize=7.2, leading=9)
    cellst = ParagraphStyle('cellx', parent=ss['Normal'], fontSize=7, leading=8.6)
    cellb = ParagraphStyle('cellb', parent=ss['Normal'], fontSize=7, leading=8.6,
                           fontName='Helvetica-Bold')

    story = [Paragraph(company, h1),
             Paragraph('REKAP HARIAN PENGISIAN VENDOR CMT', h2),
             Paragraph(f'Tanggal: <b>{date_label(day)}</b>', h2),
             Paragraph(f"Dicetak: {fmt_wib('%d/%m/%Y %H:%M')} WIB", small),
             Spacer(1, 5 * mm)]

    sum_data = [[
        Paragraph('<b>Vendor aktif</b>', cellst), Paragraph(str(summary.get('vendors_total', 0)), cellst),
        Paragraph('<b>BELUM diisi</b>', cellst), Paragraph(str(summary.get('vendors_pending', 0)), cellst),
        Paragraph('<b>Sebagian</b>', cellst), Paragraph(str(summary.get('vendors_partial', 0)), cellst),
        Paragraph('<b>Lengkap</b>', cellst), Paragraph(str(summary.get('vendors_done', 0)), cellst),
        Paragraph('<b>Tanpa pekerjaan</b>', cellst), Paragraph(str(summary.get('vendors_idle', 0)), cellst),
        Paragraph('<b>Tugas belum diisi</b>', cellst), Paragraph(str(summary.get('tasks_pending_total', 0)), cellst),
    ]]
    sum_tbl = Table(sum_data, colWidths=[24 * mm, 12 * mm] * 6)
    sum_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D4D4D8')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3), ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story += [sum_tbl, Spacer(1, 5 * mm)]

    head = [Paragraph('<b>Vendor</b>', cellst), Paragraph('<b>Status</b>', cellst)]
    head += [Paragraph(f"<b>{t['label']}</b>", cellst) for t in tasks]
    head += [Paragraph('<b>Kontak</b>', cellst)]
    data = [head]

    fill_map = {'pending': colors.HexColor('#FEE2E2'), 'partial': colors.HexColor('#FEF3C7'),
                'done': colors.HexColor('#DCFCE7'), 'none': colors.HexColor('#F4F4F5'),
                'idle': colors.HexColor('#F4F4F5')}
    style_cmds = [
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D4D4D8')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EDE9FE')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3), ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]

    for i, row in enumerate(rows, start=1):
        status = row.get('status', 'idle')
        name = row.get('vendor_name', '')
        code = row.get('vendor_code', '')
        cells = [Paragraph(f'<b>{name}</b><br/><font size="6">{code}</font>', cellst),
                 Paragraph(STATE_LABEL.get(status, status), cellb)]
        style_cmds.append(('BACKGROUND', (1, i), (1, i), fill_map.get(status, fill_map['idle'])))
        for j, t in enumerate(tasks):
            tk = (row.get('tasks') or {}).get(t['key'], {}) or {}
            cells.append(Paragraph(_cell_text(tk), cellst))
            style_cmds.append(('BACKGROUND', (2 + j, i), (2 + j, i),
                               fill_map.get(tk.get('state'), fill_map['idle'])))
        kontak = ' / '.join([x for x in (row.get('contact_name'), row.get('contact_phone')) if x]) or '-'
        cells.append(Paragraph(kontak, cellst))
        data.append(cells)

    if not rows:
        data.append([Paragraph('(tidak ada vendor CMT aktif di master)', cellst)]
                    + [Paragraph('', cellst)] * (len(tasks) + 1))

    ncol = 3 + len(tasks)
    avail = doc.width
    w_vendor, w_status, w_kontak = 38 * mm, 20 * mm, 28 * mm
    w_task = max(20 * mm, (avail - w_vendor - w_status - w_kontak) / max(1, len(tasks)))
    col_widths = [w_vendor, w_status] + [w_task] * len(tasks) + [w_kontak]
    if len(col_widths) != ncol:  # pragma: no cover — jaga-jaga bila TASKS berubah
        col_widths = [avail / ncol] * ncol

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style_cmds))
    story += [tbl, Spacer(1, 4 * mm),
              Paragraph('Keterangan: <b>OK</b> = ada pengisian pada tanggal ini · '
                        '<b>OK*</b> = ada pengisian tapi masih ada sisa pekerjaan · '
                        '<b>BELUM</b> = ada pekerjaan menunggu dan belum diisi · '
                        '<b>-</b> = memang tidak ada pekerjaan jenis itu.', small)]

    doc.build(story)
    return buf.getvalue(), _filename('pdf', day)



# ═════════════════════════════════════════════════════════════════════════════
# REKAP MINGGUAN (fase 4) — 7 hari bergulir
# ═════════════════════════════════════════════════════════════════════════════
# Sama seperti versi harian: TIDAK menghitung apa pun. Menerima hasil
# ``core.cmt_daily_recap.build_week()`` apa adanya. Supervisor memakai berkas ini
# untuk rapat mingguan ("vendor mana yang belakangan ini sering bolong"), jadi
# angkanya WAJIB identik dengan layar — termasuk urutan barisnya.

# Lambang satu KOTAK HARI. `future` sengaja kosong (bukan "-"): hari yang belum
# terjadi bukan "tidak ada pekerjaan", dan mencetak "-" membuat pembaca menyimpulkan
# vendor tidak punya kerja padahal harinya belum datang.
WEEK_CELL_MARK = {'done': 'OK', 'partial': 'OK*', 'pending': 'BELUM',
                  'idle': '-', 'none': '-', 'future': ''}

WEEK_STATUS_LABEL = {
    'late': 'TERLAMBAT',
    'unfinished': 'BELUM BERES',
    'clean': 'RAPI',
    'idle': 'Tidak ada pekerjaan',
}

# ── F12 — label perbandingan antar-pekan (SATU tempat, dipakai xlsx DAN pdf) ──
# Urutan tuple ini SEKALIGUS urutan tampil. Ditaruh di modul, bukan di dalam
# masing-masing fungsi export, karena dua daftar label yang "kebetulan sama"
# adalah cara paling mudah membuat Excel dan PDF menyebut metrik yang sama
# dengan nama berbeda — dan pembaca rapat menyimpulkan keduanya mengukur hal
# yang berbeda.
WEEK_DELTA_LABEL = (
    ('qty_progress_total', 'Pcs disetor'),
    ('qty_shipped_total', 'Pcs dikirim'),
    ('days_late_total', 'Total hari terlambat'),
    ('days_unfinished_total', 'Total hari belum beres'),
    ('days_no_progress_total', 'Total hari tanpa setoran'),
    ('vendors_late', 'Vendor pernah terlambat'),
    ('vendors_unfinished', 'Vendor belum beres'),
    ('vendors_clean', 'Vendor rapi'),
    ('vendors_idle', 'Vendor tanpa pekerjaan'),
    ('best_streak', 'Streak terbaik'),
)

DELTA_ARROW = {'better': 'MEMBAIK', 'worse': 'MEMBURUK', 'flat': 'SAMA'}


def _delta_verdict(d: dict) -> str:
    """``MEMBAIK / MEMBURUK / SAMA`` untuk satu entri delta.

    Arah baik/buruk TIDAK dihitung ulang di sini — `lower_is_better` dan
    `better` sudah diputuskan `core.cmt_daily_recap.build_week_comparison()`.
    Lampiran hanya menerjemahkannya jadi kata, supaya tidak ada kesempatan
    lampiran menyebut "membaik" untuk angka yang di layar merah.
    """
    if abs(float(d.get('diff') or 0)) < 0.005:
        return DELTA_ARROW['flat']
    return DELTA_ARROW['better' if d.get('better') else 'worse']


def _fmt_diff(value) -> str:
    """Selisih dengan tanda eksplisit: ``+3`` / ``-2`` / ``0``."""
    num = float(value or 0)
    txt = f'{num:.2f}'.rstrip('0').rstrip('.') if num % 1 else f'{int(num)}'
    return f'+{txt}' if num > 0 else txt


def _week_filename(fmt: str, start: str, end: str) -> str:
    a, b = str(start).replace('-', ''), str(end).replace('-', '')
    return f'rekap-mingguan-cmt-{a}-{b}.{fmt}'


def _day_head(d: dict) -> str:
    """Kepala kolom satu hari: ``Sen\\n4/8`` (nama hari + tanggal pendek)."""
    iso = str(d.get('date') or '')
    try:
        dt = datetime.strptime(iso[:10], '%Y-%m-%d')
        tgl = f'{dt.day}/{dt.month}'
    except (ValueError, TypeError):
        tgl = iso
    return f"{d.get('short') or ''}\n{tgl}"


def _week_range_label(start: str, end: str) -> str:
    return f'{date_label(start)} — {date_label(end)}'


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL — mingguan
# ─────────────────────────────────────────────────────────────────────────────
def build_week_xlsx(*, company: str, week: dict) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    days = week.get('days') or []
    rows = week.get('rows') or []
    per_day = week.get('per_day') or []
    summary = week.get('summary') or {}
    start, end = week.get('start', ''), week.get('end', '')

    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill('solid', fgColor='EDE9FE')
    thin = Side(style='thin', color='D4D4D8')
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        'pending': PatternFill('solid', fgColor='FEE2E2'),
        'partial': PatternFill('solid', fgColor='FEF3C7'),
        'done': PatternFill('solid', fgColor='DCFCE7'),
        'idle': PatternFill('solid', fgColor='F4F4F5'),
        'none': PatternFill('solid', fgColor='F4F4F5'),
        'future': PatternFill('solid', fgColor='FFFFFF'),
        'late': PatternFill('solid', fgColor='FEE2E2'),
        'unfinished': PatternFill('solid', fgColor='FEF3C7'),
        'clean': PatternFill('solid', fgColor='DCFCE7'),
    }

    ws = wb.active
    ws.title = 'Rekap Mingguan'

    r = 1
    for text, size, is_bold in (
        (company, 14, True),
        ('REKAP MINGGUAN PENGISIAN VENDOR CMT', 12, True),
        (f'Rentang: {_week_range_label(start, end)}', 10, False),
        (f"{summary.get('days', 0)} hari bergulir · {summary.get('days_elapsed', 0)} hari sudah berjalan", 9, False),
        (f"Dicetak: {fmt_wib('%d/%m/%Y %H:%M')} WIB", 9, False),
    ):
        ws.cell(row=r, column=1, value=text).font = Font(bold=is_bold, size=size)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value='RINGKASAN SEPEKAN').font = bold
    r += 1
    for label, value in (
        ('Vendor aktif', summary.get('vendors_total', 0)),
        ('Vendor pernah TERLAMBAT (ada hari nol bukti)', summary.get('vendors_late', 0)),
        ('Vendor belum beres (masih ada sisa)', summary.get('vendors_unfinished', 0)),
        ('Vendor rapi sepekan', summary.get('vendors_clean', 0)),
        ('Vendor tanpa pekerjaan sepekan', summary.get('vendors_idle', 0)),
        ('Total hari terlambat (semua vendor)', summary.get('days_late_total', 0)),
        ('Total hari belum beres', summary.get('days_unfinished_total', 0)),
        ('Total hari tanpa setoran', summary.get('days_no_progress_total', 0)),
        ('Total pcs disetor sepekan', summary.get('qty_progress_total', 0)),
        ('Total pcs dikirim sepekan', summary.get('qty_shipped_total', 0)),
        ('Streak terbaik (hari beruntun)', summary.get('best_streak', 0)),
    ):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1

    # ── Ringkasan per hari ──────────────────────────────────────────────────
    r += 1
    ws.cell(row=r, column=1, value='PER HARI').font = bold
    r += 1
    ph = ['Tanggal', 'Hari', 'Belum diisi', 'Sebagian', 'Lengkap', 'Tanpa pekerjaan',
          'Tugas belum diisi', 'Pcs disetor', 'Pcs dikirim']
    for c, h in enumerate(ph, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = bold
        cell.fill = head_fill
        cell.border = box
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r += 1
    for d in per_day:
        vals = [d.get('date', ''), d.get('weekday', '')]
        if d.get('is_future'):
            vals += ['(belum terjadi)', '', '', '', '', '', '']
        else:
            vals += [d.get('vendors_pending', 0), d.get('vendors_partial', 0),
                     d.get('vendors_done', 0), d.get('vendors_idle', 0),
                     d.get('tasks_pending_total', 0), d.get('qty_progress', 0),
                     d.get('qty_shipped', 0)]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = box
        r += 1

    # ── Tabel utama ─────────────────────────────────────────────────────────
    r += 2
    headers = (['Vendor', 'Kode', 'Status sepekan']
               + [_day_head(d).replace('\n', ' ') for d in days]
               + ['Hari terlambat', 'Hari belum beres', 'Hari tanpa setoran',
                  'Pcs disetor', 'Pcs dikirim', 'Streak', 'Kontak'])
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = bold
        cell.fill = head_fill
        cell.border = box
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    r += 1

    n_days = len(days)
    for row in rows:
        status = row.get('status', 'idle')
        cells = row.get('cells') or []
        vals = [row.get('vendor_name', ''), row.get('vendor_code', ''),
                WEEK_STATUS_LABEL.get(status, status)]
        for c in cells:
            st = c.get('state', 'idle')
            mark = WEEK_CELL_MARK.get(st, '-')
            qty = int(c.get('qty_progress') or 0)
            vals.append(f'{mark} · {qty} pcs' if (qty and mark) else mark)
        kontak = ' / '.join([x for x in (row.get('contact_name'),
                                         row.get('contact_phone')) if x]) or '-'
        vals += [row.get('days_late', 0), row.get('days_unfinished', 0),
                 row.get('days_no_progress', 0), row.get('qty_progress_total', 0),
                 row.get('qty_shipped_total', 0), row.get('streak', 0), kontak]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = box
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if c == 3:
                cell.fill = fills.get(status, fills['idle'])
                cell.font = bold
            elif 4 <= c <= 3 + n_days:
                st = (cells[c - 4].get('state') if c - 4 < len(cells) else 'idle')
                cell.fill = fills.get(st, fills['idle'])
        r += 1

    if not rows:
        ws.cell(row=r, column=1, value='(tidak ada vendor CMT aktif di master)')
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=f"Keterangan: {week.get('rules_note') or ''}").font = Font(size=9)
    r += 1
    ws.cell(row=r, column=1,
            value=('OK = ada pengisian · OK* = ada pengisian tapi masih ada sisa · '
                   'BELUM = ada pekerjaan menunggu dan belum diisi · "-" = tidak ada '
                   'pekerjaan · kosong = hari belum terjadi.')).font = Font(size=9)

    widths = [30, 10, 18] + [13] * n_days + [13, 14, 15, 12, 12, 9, 24]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # ── F12 — LEMBAR "Perbandingan" (hanya bila ?compare=true) ──────────────
    # Ditaruh di LEMBAR TERSENDIRI, bukan disisipkan ke lembar utama: lembar
    # utama itulah yang dibandingkan gate RK-25 baris-per-baris dengan layar,
    # dan menyelipkan blok baru di tengahnya akan membuat "angka export ==
    # angka layar" jadi soal parsing, bukan soal kebenaran.
    cmp_ = week.get('comparison') or {}
    if cmp_:
        prev = cmp_.get('previous') or {}
        delta = cmp_.get('delta') or {}
        movers = cmp_.get('movers') or {}
        ws2 = wb.create_sheet('Perbandingan')
        rr = 1
        for text, size, is_bold in (
            ('PERBANDINGAN DENGAN PEKAN SEBELUMNYA', 12, True),
            (f"Pekan ini      : {_week_range_label(start, end)}", 10, False),
            (f"Pekan sebelum  : {_week_range_label(prev.get('start', ''), prev.get('end', ''))}", 10, False),
            (cmp_.get('note') or '', 9, False),
        ):
            ws2.cell(row=rr, column=1, value=text).font = Font(bold=is_bold, size=size)
            rr += 1
        if not cmp_.get('comparable', True):
            ws2.cell(row=rr, column=1,
                     value='PERHATIAN: jumlah hari berjalan kedua pekan BERBEDA — '
                           'pakai kolom "per hari" untuk perbandingan yang adil.'
                     ).font = Font(bold=True, size=9, color='B45309')
            rr += 1
        rr += 1

        heads = ['Metrik', 'Pekan ini', 'Pekan lalu', 'Selisih', 'Arah',
                 'Per hari (ini)', 'Per hari (lalu)']
        for c, h in enumerate(heads, start=1):
            cell = ws2.cell(row=rr, column=c, value=h)
            cell.font = bold
            cell.fill = head_fill
            cell.border = box
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        rr += 1
        for key, label in WEEK_DELTA_LABEL:
            d = delta.get(key)
            if not d:
                continue
            verdict = _delta_verdict(d)
            vals = [label, d.get('now', 0), d.get('prev', 0), _fmt_diff(d.get('diff')),
                    verdict, d.get('now_per_day', 0), d.get('prev_per_day', 0)]
            for c, v in enumerate(vals, start=1):
                cell = ws2.cell(row=rr, column=c, value=v)
                cell.border = box
                if c == 5:
                    cell.font = bold
                    cell.fill = (fills['clean'] if verdict == DELTA_ARROW['better']
                                 else fills['late'] if verdict == DELTA_ARROW['worse']
                                 else fills['idle'])
            rr += 1

        # ── Papan peringkat vendor: urutannya DARI BACKEND ──────────────────
        counts = movers.get('counts') or {}
        rr += 1
        ws2.cell(row=rr, column=1, value='VENDOR YANG BERGERAK').font = bold
        rr += 1
        ws2.cell(row=rr, column=1,
                 value=(f"memburuk {counts.get('worsened', 0)} · "
                        f"membaik {counts.get('improved', 0)} · "
                        f"sama {counts.get('flat', 0)} · "
                        f"tidak diperingkat {counts.get('incomparable', 0)} "
                        f"(termasuk {counts.get('new', 0)} vendor baru) "
                        f"dari {counts.get('vendors', 0)} vendor")
                 ).font = Font(size=9)
        rr += 1
        ws2.cell(row=rr, column=1, value=f"Aturan urut: {movers.get('rule') or ''}").font = Font(size=9)
        rr += 2

        mv_heads = ['Arah', 'Vendor', 'Kode', 'Hari terlambat (ini)',
                    'Hari terlambat (lalu)', 'Selisih hari terlambat',
                    'Pcs (ini)', 'Pcs (lalu)', 'Selisih pcs',
                    'Status pekan ini', 'Status pekan lalu']
        for c, h in enumerate(mv_heads, start=1):
            cell = ws2.cell(row=rr, column=c, value=h)
            cell.font = bold
            cell.fill = head_fill
            cell.border = box
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        rr += 1

        def _mv_block(items, arah, fill_key):
            nonlocal rr
            if not items:
                ws2.cell(row=rr, column=1, value=arah).fill = fills[fill_key]
                ws2.cell(row=rr, column=2, value='(tidak ada)')
                rr += 1
                return
            for v in items:
                vals = [arah, v.get('vendor_name', ''), v.get('vendor_code', ''),
                        v.get('days_late_now', 0), v.get('days_late_prev', 0),
                        _fmt_diff(v.get('days_late_diff')),
                        v.get('qty_now', 0), v.get('qty_prev', 0),
                        _fmt_diff(v.get('qty_diff')),
                        WEEK_STATUS_LABEL.get(v.get('status_now'), v.get('status_now') or '-'),
                        WEEK_STATUS_LABEL.get(v.get('status_prev'), v.get('status_prev') or '-')]
                for c, val in enumerate(vals, start=1):
                    cell = ws2.cell(row=rr, column=c, value=val)
                    cell.border = box
                    if c == 1:
                        cell.font = bold
                        cell.fill = fills[fill_key]
                rr += 1

        _mv_block(movers.get('worsened') or [], 'MEMBURUK', 'late')
        _mv_block(movers.get('improved') or [], 'MEMBAIK', 'clean')

        for i, width in enumerate([14, 30, 10, 16, 16, 16, 11, 11, 11, 18, 18], start=1):
            ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), _week_filename('xlsx', start, end)


# ─────────────────────────────────────────────────────────────────────────────
# PDF — mingguan
# ─────────────────────────────────────────────────────────────────────────────
def build_week_pdf(*, company: str, week: dict) -> tuple[bytes, str]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    days = week.get('days') or []
    rows = week.get('rows') or []
    per_day = week.get('per_day') or []
    summary = week.get('summary') or {}
    start, end = week.get('start', ''), week.get('end', '')

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=8 * mm, rightMargin=8 * mm,
                            topMargin=9 * mm, bottomMargin=9 * mm,
                            title=f'Rekap Mingguan CMT {start} sd {end}')
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle('wh1', parent=ss['Title'], fontSize=14, leading=17, spaceAfter=2)
    h2 = ParagraphStyle('wh2', parent=ss['Normal'], fontSize=10, leading=13,
                        textColor=colors.HexColor('#52525B'))
    small = ParagraphStyle('wsm', parent=ss['Normal'], fontSize=7, leading=8.8)
    cellst = ParagraphStyle('wcell', parent=ss['Normal'], fontSize=6.6, leading=8,
                            alignment=1)
    cellleft = ParagraphStyle('wcellL', parent=ss['Normal'], fontSize=6.8, leading=8.2)
    cellb = ParagraphStyle('wcellb', parent=ss['Normal'], fontSize=6.6, leading=8,
                           fontName='Helvetica-Bold', alignment=1)

    story = [Paragraph(company, h1),
             Paragraph('REKAP MINGGUAN PENGISIAN VENDOR CMT', h2),
             Paragraph(f'Rentang: <b>{_week_range_label(start, end)}</b> '
                       f"({summary.get('days', 0)} hari bergulir)", h2),
             Paragraph(f"Dicetak: {fmt_wib('%d/%m/%Y %H:%M')} WIB", small),
             Spacer(1, 4 * mm)]

    sum_pairs = [
        ('Vendor aktif', summary.get('vendors_total', 0)),
        ('Pernah terlambat', summary.get('vendors_late', 0)),
        ('Belum beres', summary.get('vendors_unfinished', 0)),
        ('Rapi', summary.get('vendors_clean', 0)),
        ('Total hari terlambat', summary.get('days_late_total', 0)),
        ('Hari tanpa setoran', summary.get('days_no_progress_total', 0)),
        ('Pcs disetor', summary.get('qty_progress_total', 0)),
        ('Pcs dikirim', summary.get('qty_shipped_total', 0)),
    ]
    sum_data = [[]]
    for label, value in sum_pairs:
        sum_data[0].append(Paragraph(f'<b>{label}</b>', cellst))
        sum_data[0].append(Paragraph(str(value), cellst))
    sum_tbl = Table(sum_data, colWidths=[24 * mm, 10 * mm] * len(sum_pairs))
    sum_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D4D4D8')),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FAFAFA')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2), ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story += [sum_tbl, Spacer(1, 4 * mm)]

    # ── F12 — PERBANDINGAN ANTAR-PEKAN (hanya bila ?compare=true) ───────────
    # Ditaruh SETELAH ringkasan dan SEBELUM tabel vendor: pertanyaan pertama di
    # rapat adalah "arahnya ke mana", lalu langsung "vendor mana". Menaruhnya di
    # halaman terakhir membuat bagian yang paling dipakai justru paling jarang
    # terbaca.
    cmp_ = week.get('comparison') or {}
    if cmp_:
        prev_w = cmp_.get('previous') or {}
        delta = cmp_.get('delta') or {}
        movers = cmp_.get('movers') or {}
        counts = movers.get('counts') or {}
        story += [
            Paragraph('<b>PERBANDINGAN DENGAN PEKAN SEBELUMNYA</b>', h2),
            Paragraph(
                f"Pekan ini <b>{_week_range_label(start, end)}</b> vs pekan sebelumnya "
                f"<b>{_week_range_label(prev_w.get('start', ''), prev_w.get('end', ''))}</b>",
                small),
            Paragraph(cmp_.get('note') or '', small),
        ]
        if not cmp_.get('comparable', True):
            story.append(Paragraph(
                '<b>PERHATIAN:</b> jumlah hari berjalan kedua pekan BERBEDA — '
                'pakai baris "per hari" untuk perbandingan yang adil.', small))
        story.append(Spacer(1, 2 * mm))

        d_head = [Paragraph('<b>Metrik</b>', cellleft), Paragraph('<b>Pekan ini</b>', cellst),
                  Paragraph('<b>Pekan lalu</b>', cellst), Paragraph('<b>Selisih</b>', cellst),
                  Paragraph('<b>Arah</b>', cellst), Paragraph('<b>Per hari<br/>(ini/lalu)</b>', cellst)]
        d_data = [d_head]
        d_style = [
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D4D4D8')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DBEAFE')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]
        di = 0
        for key, label in WEEK_DELTA_LABEL:
            d = delta.get(key)
            if not d:
                continue
            di += 1
            verdict = _delta_verdict(d)
            d_data.append([
                Paragraph(label, cellleft),
                Paragraph(str(d.get('now', 0)), cellb),
                Paragraph(str(d.get('prev', 0)), cellst),
                Paragraph(_fmt_diff(d.get('diff')), cellb),
                Paragraph(verdict, cellb),
                Paragraph(f"{d.get('now_per_day', 0)} / {d.get('prev_per_day', 0)}", cellst),
            ])
            d_style.append(('BACKGROUND', (4, di), (4, di),
                            colors.HexColor('#DCFCE7') if verdict == DELTA_ARROW['better']
                            else colors.HexColor('#FEE2E2') if verdict == DELTA_ARROW['worse']
                            else colors.HexColor('#F4F4F5')))
        d_tbl = Table(d_data, colWidths=[52 * mm, 24 * mm, 24 * mm, 22 * mm, 26 * mm, 30 * mm],
                      repeatRows=1)
        d_tbl.setStyle(TableStyle(d_style))
        story += [d_tbl, Spacer(1, 3 * mm)]

        story += [
            Paragraph('<b>VENDOR YANG BERGERAK</b> '
                      f"<font size=\"7\">(memburuk {counts.get('worsened', 0)} · "
                      f"membaik {counts.get('improved', 0)} · sama {counts.get('flat', 0)} · "
                      f"tidak diperingkat {counts.get('incomparable', 0)} termasuk "
                      f"{counts.get('new', 0)} vendor baru · dari "
                      f"{counts.get('vendors', 0)} vendor)</font>", h2),
            Paragraph(f"Aturan urut: {movers.get('rule') or ''}", small),
            Spacer(1, 1.5 * mm),
        ]
        m_head = [Paragraph('<b>Arah</b>', cellst), Paragraph('<b>Vendor</b>', cellleft),
                  Paragraph('<b>Hari terlambat<br/>ini / lalu</b>', cellst),
                  Paragraph('<b>Selisih<br/>hari</b>', cellst),
                  Paragraph('<b>Pcs<br/>ini / lalu</b>', cellst),
                  Paragraph('<b>Selisih<br/>pcs</b>', cellst),
                  Paragraph('<b>Status<br/>ini / lalu</b>', cellst)]
        m_data = [m_head]
        m_style = [
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D4D4D8')),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DBEAFE')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]
        mi = 0
        for arah, items, bg in (('MEMBURUK', movers.get('worsened') or [], '#FEE2E2'),
                                ('MEMBAIK', movers.get('improved') or [], '#DCFCE7')):
            if not items:
                mi += 1
                m_data.append([Paragraph(arah, cellb), Paragraph('(tidak ada)', cellleft)]
                              + [Paragraph('', cellst)] * 5)
                m_style.append(('BACKGROUND', (0, mi), (0, mi), colors.HexColor(bg)))
                continue
            for v in items:
                mi += 1
                m_data.append([
                    Paragraph(arah, cellb),
                    Paragraph(f"<b>{v.get('vendor_name', '')}</b>"
                              f"<br/><font size=\"5.5\">{v.get('vendor_code') or ''}</font>",
                              cellleft),
                    Paragraph(f"{v.get('days_late_now', 0)} / {v.get('days_late_prev', 0)}", cellst),
                    Paragraph(_fmt_diff(v.get('days_late_diff')), cellb),
                    Paragraph(f"{v.get('qty_now', 0)} / {v.get('qty_prev', 0)}", cellst),
                    Paragraph(_fmt_diff(v.get('qty_diff')), cellb),
                    Paragraph(
                        f"{WEEK_STATUS_LABEL.get(v.get('status_now'), v.get('status_now') or '-')}"
                        f" / {WEEK_STATUS_LABEL.get(v.get('status_prev'), v.get('status_prev') or '-')}",
                        cellst),
                ])
                m_style.append(('BACKGROUND', (0, mi), (0, mi), colors.HexColor(bg)))
        m_tbl = Table(m_data, colWidths=[22 * mm, 46 * mm, 26 * mm, 20 * mm, 26 * mm, 20 * mm,
                                         40 * mm], repeatRows=1)
        m_tbl.setStyle(TableStyle(m_style))
        story += [m_tbl, Spacer(1, 4 * mm)]

    fill_map = {'pending': colors.HexColor('#FEE2E2'), 'partial': colors.HexColor('#FEF3C7'),
                'done': colors.HexColor('#DCFCE7'), 'idle': colors.HexColor('#F4F4F5'),
                'none': colors.HexColor('#F4F4F5'), 'future': colors.white,
                'late': colors.HexColor('#FEE2E2'), 'unfinished': colors.HexColor('#FEF3C7'),
                'clean': colors.HexColor('#DCFCE7')}

    head = [Paragraph('<b>Vendor</b>', cellleft), Paragraph('<b>Status</b>', cellst)]
    head += [Paragraph(f"<b>{_day_head(d).replace(chr(10), '<br/>')}</b>", cellst) for d in days]
    head += [Paragraph('<b>Ter-<br/>lambat</b>', cellst),
             Paragraph('<b>Belum<br/>beres</b>', cellst),
             Paragraph('<b>Tanpa<br/>setoran</b>', cellst),
             Paragraph('<b>Pcs<br/>disetor</b>', cellst),
             Paragraph('<b>Pcs<br/>dikirim</b>', cellst),
             Paragraph('<b>Streak</b>', cellst)]
    data = [head]

    style_cmds = [
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D4D4D8')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EDE9FE')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2), ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
    ]

    n_days = len(days)
    for i, row in enumerate(rows, start=1):
        status = row.get('status', 'idle')
        cells_src = row.get('cells') or []
        line = [Paragraph(f"<b>{row.get('vendor_name', '')}</b>"
                          f"<br/><font size=\"5.5\">{row.get('vendor_code') or ''}</font>", cellleft),
                Paragraph(WEEK_STATUS_LABEL.get(status, status), cellb)]
        style_cmds.append(('BACKGROUND', (1, i), (1, i), fill_map.get(status, fill_map['idle'])))
        for j, c in enumerate(cells_src):
            st = c.get('state', 'idle')
            mark = WEEK_CELL_MARK.get(st, '-')
            qty = int(c.get('qty_progress') or 0)
            txt = f'{mark}<br/><font size="5.5">{qty} pcs</font>' if (qty and mark) else mark
            line.append(Paragraph(txt, cellst))
            style_cmds.append(('BACKGROUND', (2 + j, i), (2 + j, i),
                               fill_map.get(st, fill_map['idle'])))
        line += [Paragraph(str(row.get('days_late', 0)), cellb),
                 Paragraph(str(row.get('days_unfinished', 0)), cellst),
                 Paragraph(str(row.get('days_no_progress', 0)), cellst),
                 Paragraph(str(row.get('qty_progress_total', 0)), cellst),
                 Paragraph(str(row.get('qty_shipped_total', 0)), cellst),
                 Paragraph(str(row.get('streak', 0)), cellst)]
        data.append(line)

    ncol = 2 + n_days + 6
    if not rows:
        data.append([Paragraph('(tidak ada vendor CMT aktif di master)', cellleft)]
                    + [Paragraph('', cellst)] * (ncol - 1))

    avail = doc.width
    w_vendor, w_status = 32 * mm, 17 * mm
    w_tail = 12 * mm
    w_day = max(9 * mm, (avail - w_vendor - w_status - 6 * w_tail) / max(1, n_days))
    col_widths = [w_vendor, w_status] + [w_day] * n_days + [w_tail] * 6
    if len(col_widths) != ncol:  # pragma: no cover — jaga-jaga bila bentuk berubah
        col_widths = [avail / ncol] * ncol

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style_cmds))
    story += [tbl, Spacer(1, 3 * mm),
              Paragraph(f"<b>Keterangan:</b> {week.get('rules_note') or ''}", small),
              Paragraph('<b>OK</b> = ada pengisian · <b>OK*</b> = ada pengisian tapi masih '
                        'ada sisa · <b>BELUM</b> = ada pekerjaan menunggu dan belum diisi · '
                        '<b>-</b> = tidak ada pekerjaan · kosong = hari belum terjadi.', small)]

    # Ringkasan per hari ditaruh di AKHIR: yang dicari lebih dulu di rapat adalah
    # vendor mana yang bolong, bukan totalnya per hari.
    story += [Spacer(1, 4 * mm), Paragraph('<b>RINGKASAN PER HARI</b>', h2)]
    pd_head = [Paragraph('<b>Tanggal</b>', cellst), Paragraph('<b>Hari</b>', cellst),
               Paragraph('<b>Belum diisi</b>', cellst), Paragraph('<b>Sebagian</b>', cellst),
               Paragraph('<b>Lengkap</b>', cellst), Paragraph('<b>Tanpa pekerjaan</b>', cellst),
               Paragraph('<b>Pcs disetor</b>', cellst), Paragraph('<b>Pcs dikirim</b>', cellst)]
    pd_data = [pd_head]
    for d in per_day:
        if d.get('is_future'):
            pd_data.append([Paragraph(d.get('date', ''), cellst),
                            Paragraph(d.get('weekday', ''), cellst),
                            Paragraph('(belum terjadi)', cellst)]
                           + [Paragraph('', cellst)] * 5)
            continue
        pd_data.append([Paragraph(d.get('date', ''), cellst),
                        Paragraph(d.get('weekday', ''), cellst),
                        Paragraph(str(d.get('vendors_pending', 0)), cellb),
                        Paragraph(str(d.get('vendors_partial', 0)), cellst),
                        Paragraph(str(d.get('vendors_done', 0)), cellst),
                        Paragraph(str(d.get('vendors_idle', 0)), cellst),
                        Paragraph(str(d.get('qty_progress', 0)), cellst),
                        Paragraph(str(d.get('qty_shipped', 0)), cellst)])
    pd_tbl = Table(pd_data, colWidths=[26 * mm, 22 * mm] + [22 * mm] * 6, repeatRows=1)
    pd_tbl.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D4D4D8')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EDE9FE')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2), ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    story += [pd_tbl]

    doc.build(story)
    return buf.getvalue(), _week_filename('pdf', start, end)
