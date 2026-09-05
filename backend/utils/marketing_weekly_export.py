"""utils.marketing_weekly_export — Laporan Rapat Mingguan dalam **Excel & PDF**.

KENAPA ADA
----------
Laporan mingguan dibacakan di rapat, dicetak, dan dikirim ke grup. Karena itu
perlu dua lampiran: **Excel** (untuk diolah/diarsipkan) dan **PDF** (untuk
dicetak & ditandatangani).

ISI SAMA PERSIS DENGAN LAYAR. Keduanya menerima hasil
``core.marketing_weekly_report.build_weekly_report()`` **apa adanya** dan tidak
menghitung apa pun. Kalau berkas ini ikut menghitung, suatu hari PDF akan
menyebut omzet yang berbeda dengan layar untuk minggu yang sama.

``catatan_data`` (kejujuran data) IKUT dicetak — laporan yang tidak menyebut
lubangnya sendiri akan dipakai seolah lengkap.

Tanpa dependensi baru: ``openpyxl`` & ``reportlab`` sudah ada di requirements.
"""
from __future__ import annotations

import io
from typing import Any, Dict, List, Tuple

from core.marketing_weekly_report import TRAFFIC_KEYS, TRAFFIC_LABEL


def _rp(v: Any) -> str:
    try:
        return f"Rp {int(round(float(v or 0))):,}".replace(",", ".")
    except (TypeError, ValueError):
        return "Rp 0"


def _num(v: Any) -> str:
    try:
        return f"{int(round(float(v or 0))):,}".replace(",", ".")
    except (TypeError, ValueError):
        return "0"


def _pct(v: Any, suffix: str = "%") -> str:
    if v is None:
        return "—"
    return f"{v}{suffix}"


def _delta_text(d: Dict[str, Any]) -> str:
    if not d:
        return "—"
    pct = d.get("persen")
    arrow = "" if pct is None else ("▲" if pct > 0 else ("▼" if pct < 0 else "="))
    return f"{arrow} {_pct(pct)}" if pct is not None else "— (pembanding 0)"


def filename(fmt: str, report: dict) -> str:
    wk = (report.get("periode") or {}).get("minggu", "minggu")
    return f"laporan-rapat-mingguan-{wk}.{fmt}"


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(*, company: str, report: dict) -> Tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    per = report.get("per_toko") or []
    gab = report.get("gabungan") or {}
    periode = report.get("periode") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Ringkas per Toko"

    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    title_font = Font(bold=True, size=13)

    ws["A1"] = f"{company} — LAPORAN RAPAT MINGGUAN"
    ws["A1"].font = title_font
    ws["A2"] = (f"Minggu {periode.get('minggu')} · {periode.get('label')} "
                f"({periode.get('dasar_minggu')})")
    ws["A3"] = f"Dibuat: {report.get('dibuat_pada', '')[:19].replace('T', ' ')} UTC"

    cols = [
        ("Kode", 18), ("Toko", 26), ("Platform", 12), ("PIC", 16),
        ("Omzet", 16), ("vs Minggu Lalu", 16), ("Pesanan", 10),
        ("Pcs", 8), ("AOV", 13), ("Target prorata", 16), ("Capai %", 10),
        ("Live", 14), ("Video", 14), ("Kartu Produk", 14),
        ("Pemenuhan %", 12), ("Batal", 8), ("Belum kirim", 12),
        ("Iklan", 14), ("ROAS", 8), ("Hari berdata", 12), ("Sumber angka", 22),
        # SESI #9 — dua kolom BARU ditaruh di UJUNG dengan sengaja: menyisipkannya
        # di tengah akan menggeser indeks kolom baris GABUNGAN di bawah (yang
        # ditulis per nomor kolom) dan membuat angka gabungan pindah kolom.
        ("Nilai retur", 14), ("Omzet setelah retur", 18),
    ]
    r0 = 5
    for i, (name, width) in enumerate(cols, start=1):
        c = ws.cell(row=r0, column=i, value=name)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width

    for n, s in enumerate(per, start=1):
        row = r0 + n
        vals = [
            s.get("account_code"), s.get("account_name"), s.get("platform"),
            s.get("pic") or "—",
            s.get("omzet"), _delta_text(s.get("vs_minggu_lalu", {}).get("omzet")),
            s.get("pesanan"), s.get("pcs"), s.get("aov"),
            s.get("target", {}).get("revenue") or 0,
            _pct(s.get("pencapaian_target_persen")),
            s.get("kanal", {}).get("live"), s.get("kanal", {}).get("video"),
            s.get("kanal", {}).get("product_card"),
            _pct((s.get("pemenuhan") or {}).get("fulfillment_rate")),
            s.get("pesanan_mentah", {}).get("batal"),
            s.get("pesanan_mentah", {}).get("belum_dikirim"),
            s.get("iklan", {}).get("spend") if s.get("iklan", {}).get("terisi") else "belum diimpor",
            s.get("iklan", {}).get("roas") if s.get("iklan", {}).get("roas") is not None else "—",
            f"{s.get('hari_berdata')}/7",
            ", ".join(s.get("sumber_angka") or []) or "—",
            s.get("nilai_retur") or 0, s.get("omzet_setelah_retur") or 0,
        ]
        for i, v in enumerate(vals, start=1):
            ws.cell(row=row, column=i, value=v)

    tr = r0 + len(per) + 1
    ws.cell(row=tr, column=1, value="GABUNGAN").font = Font(bold=True)
    ws.cell(row=tr, column=5, value=gab.get("omzet")).font = Font(bold=True)
    ws.cell(row=tr, column=6, value=_delta_text(gab.get("vs_minggu_lalu", {}).get("omzet")))
    ws.cell(row=tr, column=7, value=gab.get("pesanan")).font = Font(bold=True)
    ws.cell(row=tr, column=8, value=gab.get("pcs"))
    ws.cell(row=tr, column=9, value=gab.get("aov"))
    ws.cell(row=tr, column=10, value=gab.get("target_prorata"))
    ws.cell(row=tr, column=11, value=_pct(gab.get("pencapaian_target_persen")))
    ws.cell(row=tr, column=18, value=gab.get("iklan_spend"))
    ws.cell(row=tr, column=19, value=gab.get("roas") if gab.get("roas") is not None else "—")
    ws.cell(row=tr, column=22, value=gab.get("nilai_retur") or 0)
    ws.cell(row=tr, column=23, value=gab.get("omzet_setelah_retur") or 0).font = Font(bold=True)

    # ── Sheet 2: pecahan kanal ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Pecahan Kanal")
    ws2["A1"] = "PECAHAN OMZET PER KANAL"
    ws2["A1"].font = title_font
    hdr = ["Toko"] + [TRAFFIC_LABEL[k] for k in TRAFFIC_KEYS]
    for i, name in enumerate(hdr, start=1):
        c = ws2.cell(row=3, column=i, value=name)
        c.fill = head_fill
        c.font = head_font
        ws2.column_dimensions[get_column_letter(i)].width = 18
    for n, s in enumerate(per, start=1):
        ws2.cell(row=3 + n, column=1, value=s.get("account_name"))
        for i, k in enumerate(TRAFFIC_KEYS, start=2):
            ws2.cell(row=3 + n, column=i, value=(s.get("kanal") or {}).get(k, 0))
    gr = 3 + len(per) + 1
    ws2.cell(row=gr, column=1, value="GABUNGAN").font = Font(bold=True)
    for i, k in enumerate(TRAFFIC_KEYS, start=2):
        ws2.cell(row=gr, column=i, value=(gab.get("kanal") or {}).get(k, 0))

    # ── Sheet 3: catatan kejujuran data ───────────────────────────────────────
    ws3 = wb.create_sheet("Catatan Data")
    ws3["A1"] = "CATATAN KEJUJURAN DATA — WAJIB DIBACA SEBELUM MENYIMPULKAN"
    ws3["A1"].font = title_font
    ws3.column_dimensions["A"].width = 130
    for n, note in enumerate(report.get("catatan_data") or [], start=1):
        c = ws3.cell(row=2 + n, column=1, value=f"{n}. {note}")
        c.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename("xlsx", report)


# ══════════════════════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════════════════════
def build_pdf(*, company: str, report: dict) -> Tuple[bytes, str]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle)

    per = report.get("per_toko") or []
    gab = report.get("gabungan") or {}
    periode = report.get("periode") or {}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm,
                            title=f"Laporan Rapat Mingguan {periode.get('minggu')}")
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=ss["Heading1"], fontSize=14, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=8.5,
                         textColor=colors.HexColor("#444444"))
    h2 = ParagraphStyle("h2", parent=ss["Heading2"], fontSize=10.5, spaceBefore=8,
                        spaceAfter=3)
    small = ParagraphStyle("small", parent=ss["Normal"], fontSize=7.6, leading=9.6)
    cell = ParagraphStyle("cell", parent=ss["Normal"], fontSize=7.2, leading=8.6)

    flow: List[Any] = [
        Paragraph(f"{company} — LAPORAN RAPAT MINGGUAN", h1),
        Paragraph(f"Minggu {periode.get('minggu')} &nbsp;·&nbsp; {periode.get('label')} "
                  f"&nbsp;·&nbsp; dasar minggu: {periode.get('dasar_minggu')} "
                  f"&nbsp;·&nbsp; pembanding: {(periode.get('minggu_sebelumnya') or {}).get('label')}",
                  sub),
        Spacer(1, 5 * mm),
    ]

    # ── Kartu gabungan ────────────────────────────────────────────────────────
    kpi = [[
        Paragraph("<b>OMZET GABUNGAN</b><br/>" + _rp(gab.get("omzet"))
                  + "<br/><font size=6>vs minggu lalu "
                  + _delta_text(gab.get("vs_minggu_lalu", {}).get("omzet")) + "</font>", cell),
        # SESI #9 — omzet setelah retur ikut di kartu ringkas PDF. Tidak
        # ditambahkan ke tabel per-toko PDF supaya lebarnya tetap terbaca; kolom
        # lengkapnya ada di XLSX & CSV layar.
        Paragraph("<b>SETELAH RETUR</b><br/>" + _rp(gab.get("omzet_setelah_retur"))
                  + "<br/><font size=6>retur " + _num(gab.get("retur"))
                  + " · " + _rp(gab.get("nilai_retur")) + "</font>", cell),
        Paragraph("<b>PESANAN</b><br/>" + _num(gab.get("pesanan"))
                  + "<br/><font size=6>" + _num(gab.get("pcs")) + " pcs</font>", cell),
        Paragraph("<b>AOV</b><br/>" + _rp(gab.get("aov")), cell),
        Paragraph("<b>TARGET PRORATA</b><br/>" + _rp(gab.get("target_prorata"))
                  + "<br/><font size=6>capai " + _pct(gab.get("pencapaian_target_persen"))
                  + "</font>", cell),
        Paragraph("<b>IKLAN</b><br/>" + _rp(gab.get("iklan_spend"))
                  + "<br/><font size=6>ROAS "
                  + (str(gab.get("roas")) if gab.get("roas") is not None else "—")
                  + "</font>", cell),
        Paragraph("<b>BELUM DIKIRIM</b><br/>" + _num(gab.get("belum_dikirim"))
                  + "<br/><font size=6>" + _rp(gab.get("nilai_belum_dikirim")) + "</font>", cell),
        Paragraph("<b>TOKO BERDATA</b><br/>" + f"{gab.get('toko_berdata')}/{gab.get('toko')}", cell),
    ]]
    # 8 kartu (sesi #9 menambah "SETELAH RETUR") — lebar dikecilkan supaya total
    # 8×33mm = 264mm tetap masuk halaman A4 landscape (297mm − margin).
    t = Table(kpi, colWidths=[33 * mm] * 8)
    t.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#cccccc")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dddddd")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f7f7fb")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow += [t, Spacer(1, 4 * mm), Paragraph("RINCIAN PER TOKO", h2)]

    head = ["Kode", "Toko", "Omzet", "vs Mg lalu", "Pesanan", "AOV",
            "Target prorata", "Capai", "Live", "Video", "Kartu",
            "Pemenuhan", "Batal", "Blm kirim", "Iklan", "ROAS", "Hari"]
    data = [[Paragraph(f"<b>{h}</b>", cell) for h in head]]
    for s in per:
        ik = s.get("iklan") or {}
        data.append([
            Paragraph(str(s.get("account_code") or ""), cell),
            Paragraph(str(s.get("account_name") or ""), cell),
            Paragraph(_rp(s.get("omzet")), cell),
            Paragraph(_delta_text(s.get("vs_minggu_lalu", {}).get("omzet")), cell),
            Paragraph(_num(s.get("pesanan")), cell),
            Paragraph(_rp(s.get("aov")), cell),
            Paragraph(_rp((s.get("target") or {}).get("revenue")), cell),
            Paragraph(_pct(s.get("pencapaian_target_persen")), cell),
            Paragraph(_rp((s.get("kanal") or {}).get("live")), cell),
            Paragraph(_rp((s.get("kanal") or {}).get("video")), cell),
            Paragraph(_rp((s.get("kanal") or {}).get("product_card")), cell),
            Paragraph(_pct((s.get("pemenuhan") or {}).get("fulfillment_rate")), cell),
            Paragraph(_num((s.get("pesanan_mentah") or {}).get("batal")), cell),
            Paragraph(_num((s.get("pesanan_mentah") or {}).get("belum_dikirim")), cell),
            Paragraph(_rp(ik.get("spend")) if ik.get("terisi") else "belum diimpor", cell),
            Paragraph(str(ik.get("roas")) if ik.get("roas") is not None else "—", cell),
            Paragraph(f"{s.get('hari_berdata')}/7", cell),
        ])
    data.append([
        Paragraph("<b>GABUNGAN</b>", cell), Paragraph("", cell),
        Paragraph(f"<b>{_rp(gab.get('omzet'))}</b>", cell),
        Paragraph(_delta_text(gab.get("vs_minggu_lalu", {}).get("omzet")), cell),
        Paragraph(f"<b>{_num(gab.get('pesanan'))}</b>", cell),
        Paragraph(_rp(gab.get("aov")), cell),
        Paragraph(_rp(gab.get("target_prorata")), cell),
        Paragraph(_pct(gab.get("pencapaian_target_persen")), cell),
        Paragraph(_rp((gab.get("kanal") or {}).get("live")), cell),
        Paragraph(_rp((gab.get("kanal") or {}).get("video")), cell),
        Paragraph(_rp((gab.get("kanal") or {}).get("product_card")), cell),
        Paragraph("", cell),
        Paragraph(_num(gab.get("batal")), cell),
        Paragraph(_num(gab.get("belum_dikirim")), cell),
        Paragraph(_rp(gab.get("iklan_spend")), cell),
        Paragraph(str(gab.get("roas")) if gab.get("roas") is not None else "—", cell),
        Paragraph("", cell),
    ])
    widths = [20, 34, 24, 18, 15, 20, 24, 13, 22, 20, 20, 17, 12, 15, 22, 12, 12]
    tt = Table(data, colWidths=[w * mm for w in widths], repeatRows=1)
    tt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef2ff")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2),
         [colors.white, colors.HexColor("#fafafa")]),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    flow += [tt, Spacer(1, 4 * mm),
             Paragraph("CATATAN KEJUJURAN DATA — wajib dibaca sebelum menyimpulkan", h2)]
    for n, note in enumerate(report.get("catatan_data") or [], start=1):
        flow.append(Paragraph(f"{n}. {note}", small))
    flow += [Spacer(1, 3 * mm),
             Paragraph(f"Dibuat otomatis dari data sistem pada "
                       f"{report.get('dibuat_pada', '')[:19].replace('T', ' ')} UTC. "
                       f"Tidak ada angka pada laporan ini yang diketik untuk rapat.",
                       sub)]

    doc.build(flow)
    return buf.getvalue(), filename("pdf", report)
