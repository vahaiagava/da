"""utils.accessory_valuation_export — FASE 8 lanjutan: rapor valuasi HPP aksesoris (Excel & PDF).

KENAPA ADA
Nilai persediaan aksesoris sudah bisa dilihat di layar (tab "Valuasi HPP"), tapi bagian
keuangan butuh LAMPIRAN yang bisa diarsipkan/dikirim: rekap per item + mutasi bernilai
sebulan + total, dalam format Excel (untuk diolah) dan PDF (untuk ditandatangani).

Isi rapor (dua-duanya sama isinya):
  * Header  : nama perusahaan, judul, periode, metode valuasi, waktu cetak
  * Ringkasan: total nilai persediaan, jumlah item bernilai / belum dinilai, HPP rata-rata
  * Tabel 1 : valuasi per item (kode, nama, kategori, stok, satuan, HPP, nilai, metode)
  * Tabel 2 : mutasi bernilai pada periode (waktu, item, jenis, qty, HPP, nilai, no. jurnal)
  * Catatan : item ber-HPP 0 tidak menghasilkan jurnal — sorot supaya segera dilengkapi

Tanpa dependensi baru: `openpyxl` & `reportlab` sudah ada di requirements.txt.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

MV_LABEL = {
    "receive": "Terima",
    "issue": "Keluar",
    "scrap": "Scrap",
    "opname_adjust": "Opname",
    "adjust": "Penyesuaian",
}
METHOD_LABEL = {"moving_average": "Rata-rata bergerak", "manual": "Manual"}

_MONTHS_ID = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli",
              "Agustus", "September", "Oktober", "November", "Desember"]


def period_label(month: str | None) -> str:
    """'2026-07' → 'Juli 2026'. None → 'Semua periode'."""
    if not month:
        return "Semua periode"
    try:
        y, m = str(month).split("-")[:2]
        return f"{_MONTHS_ID[int(m) - 1]} {y}"
    except (ValueError, IndexError):
        return str(month)


def _rp(v) -> str:
    try:
        return f"Rp {float(v or 0):,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return "Rp 0"


def _num(v) -> str:
    try:
        f = float(v or 0)
        return f"{f:,.2f}".replace(",", "#").replace(".", ",").replace("#", ".")
    except (TypeError, ValueError):
        return "0"


def _dt(v) -> str:
    if not v:
        return "-"
    try:
        if isinstance(v, str):
            v = datetime.fromisoformat(v.replace("Z", "+00:00"))
        return v.strftime("%d/%m/%Y %H:%M")
    except (ValueError, AttributeError):
        return str(v)[:16]


def _filename(fmt: str, month: str | None) -> str:
    suffix = (month or "semua").replace("-", "")
    return f"valuasi-aksesoris-{suffix}.{fmt}"


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL
# ─────────────────────────────────────────────────────────────────────────────
def build_xlsx(*, company: str, summary: dict, movements: list, month: str | None) -> tuple[bytes, str]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="EDE9FE")
    warn_font = Font(color="B45309")
    thin = Side(style="thin", color="D4D4D8")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.active
    ws.title = "Valuasi"
    rows_meta = [
        (company, 14, True),
        ("RAPOR VALUASI PERSEDIAAN AKSESORIS", 12, True),
        (f"Periode: {period_label(month)}", 10, False),
        (f"Metode valuasi: {METHOD_LABEL.get(summary.get('cost_method'), summary.get('cost_method', '-'))}", 10, False),
        (f"Dicetak: {_dt(datetime.now(timezone.utc))} UTC", 9, False),
    ]
    r = 1
    for text, size, is_bold in rows_meta:
        ws.cell(row=r, column=1, value=text).font = Font(bold=is_bold, size=size)
        r += 1

    t = summary.get("totals", {})
    r += 1
    ws.cell(row=r, column=1, value="RINGKASAN").font = bold
    r += 1
    for label, value in (
        ("Total nilai persediaan", _rp(t.get("total_value"))),
        ("Total kuantitas", _num(t.get("total_qty"))),
        ("Item bernilai", t.get("valued_items", 0)),
        ("Item BELUM dinilai", t.get("unvalued_items", 0)),
        ("Kuantitas belum dinilai", _num(t.get("unvalued_qty"))),
        ("HPP rata-rata", _rp(t.get("avg_unit_cost"))),
    ):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=value)
        r += 1

    r += 1
    headers = ["Kode", "Nama", "Kategori", "Stok", "Satuan", "HPP (Rp)", "Nilai Stok (Rp)", "Metode"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = bold
        cell.fill = head_fill
        cell.border = box
        cell.alignment = Alignment(horizontal="center")
    r += 1
    for it in summary.get("items", []):
        vals = [it.get("code", ""), it.get("name", ""), it.get("category", ""),
                float(it.get("stock_qty") or 0), it.get("unit", ""),
                float(it.get("unit_cost") or 0), float(it.get("stock_value") or 0),
                METHOD_LABEL.get(it.get("cost_method"), it.get("cost_method") or "-")]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = box
            if c in (6, 7):
                cell.number_format = "#,##0"
            if not it.get("valued"):
                cell.font = warn_font
        r += 1
    ws.cell(row=r, column=5, value="TOTAL").font = bold
    total_cell = ws.cell(row=r, column=7, value=float(t.get("total_value") or 0))
    total_cell.font = bold
    total_cell.number_format = "#,##0"
    r += 2
    if t.get("unvalued_items"):
        ws.cell(row=r, column=1,
                value=(f"CATATAN: {t.get('unvalued_items')} item belum punya harga satuan (baris berwarna). "
                       "Selama HPP masih 0, mutasinya TIDAK menghasilkan jurnal keuangan.")).font = warn_font

    for col, width in zip("ABCDEFGH", (16, 34, 14, 12, 10, 16, 18, 20)):
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("Mutasi Bernilai")
    ws2.cell(row=1, column=1, value=f"MUTASI BERNILAI — {period_label(month)}").font = bold
    mv_headers = ["Waktu", "Kode", "Aksesoris", "Jenis", "Qty", "Satuan", "HPP (Rp)", "Nilai (Rp)", "No. Jurnal"]
    for c, h in enumerate(mv_headers, start=1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = bold
        cell.fill = head_fill
        cell.border = box
    rr = 4
    for m in movements:
        vals = [_dt(m.get("created_at")), m.get("material_code", ""), m.get("material_name", ""),
                MV_LABEL.get(m.get("movement_type"), m.get("movement_type") or ""),
                float(m.get("qty_signed") or 0), m.get("unit", ""),
                float(m.get("unit_cost") or 0), float(m.get("value") or 0),
                m.get("je_number") or "(tanpa jurnal)"]
        for c, v in enumerate(vals, start=1):
            cell = ws2.cell(row=rr, column=c, value=v)
            cell.border = box
            if c in (7, 8):
                cell.number_format = "#,##0"
        rr += 1
    if not movements:
        ws2.cell(row=4, column=1, value="(tidak ada mutasi pada periode ini)")
    for col, width in zip("ABCDEFGHI", (18, 16, 30, 14, 10, 10, 14, 16, 20)):
        ws2.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), _filename("xlsx", month)


# ─────────────────────────────────────────────────────────────────────────────
# PDF
# ─────────────────────────────────────────────────────────────────────────────
def build_pdf(*, company: str, summary: dict, movements: list, month: str | None) -> tuple[bytes, str]:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title="Rapor Valuasi Persediaan Aksesoris")
    ss = getSampleStyleSheet()
    h_center = ParagraphStyle("hc", parent=ss["Title"], fontSize=13, alignment=TA_CENTER, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=9, alignment=TA_CENTER, textColor=colors.grey)
    small = ParagraphStyle("sm", parent=ss["Normal"], fontSize=8)
    warn = ParagraphStyle("warn", parent=ss["Normal"], fontSize=8, textColor=colors.HexColor("#B45309"))

    t = summary.get("totals", {})
    story = [
        Paragraph(company, h_center),
        Paragraph("RAPOR VALUASI PERSEDIAAN AKSESORIS", h_center),
        Paragraph(f"Periode: {period_label(month)} · Metode: "
                  f"{METHOD_LABEL.get(summary.get('cost_method'), '-')} · "
                  f"Dicetak {_dt(datetime.now(timezone.utc))} UTC", sub),
        Spacer(1, 6 * mm),
    ]

    sum_data = [
        ["Total nilai persediaan", _rp(t.get("total_value")),
         "Item bernilai", str(t.get("valued_items", 0))],
        ["HPP rata-rata", _rp(t.get("avg_unit_cost")),
         "Item BELUM dinilai", str(t.get("unvalued_items", 0))],
        ["Total kuantitas", _num(t.get("total_qty")),
         "Qty belum dinilai", _num(t.get("unvalued_qty"))],
    ]
    st = Table(sum_data, colWidths=[52 * mm, 42 * mm, 42 * mm, 30 * mm])
    st.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.grey),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.grey),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story += [st, Spacer(1, 5 * mm), Paragraph("<b>Valuasi per item</b>", small), Spacer(1, 2 * mm)]

    head = ["Kode", "Nama", "Kategori", "Stok", "Sat.", "HPP", "Nilai Stok", "Metode"]
    data = [head]
    warn_rows = []
    for i, it in enumerate(summary.get("items", []), start=1):
        data.append([it.get("code", ""), (it.get("name") or "")[:38], it.get("category", ""),
                     _num(it.get("stock_qty")), it.get("unit", ""),
                     _rp(it.get("unit_cost")) if it.get("valued") else "belum diisi",
                     _rp(it.get("stock_value")),
                     METHOD_LABEL.get(it.get("cost_method"), "-")])
        if not it.get("valued"):
            warn_rows.append(i)
    data.append(["", "", "", "", "TOTAL", "", _rp(t.get("total_value")), ""])
    tbl = Table(data, colWidths=[26 * mm, 66 * mm, 24 * mm, 20 * mm, 14 * mm, 26 * mm, 30 * mm, 32 * mm],
                repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EDE9FE")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -2), 0.25, colors.HexColor("#D4D4D8")),
        ("ALIGN", (3, 1), (6, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for i in warn_rows:
        style.append(("TEXTCOLOR", (0, i), (-1, i), colors.HexColor("#B45309")))
    tbl.setStyle(TableStyle(style))
    story.append(tbl)

    if t.get("unvalued_items"):
        story += [Spacer(1, 3 * mm),
                  Paragraph(f"Catatan: {t.get('unvalued_items')} item belum punya harga satuan "
                            "(baris berwarna). Selama HPP masih 0, mutasinya TIDAK menghasilkan "
                            "jurnal keuangan sehingga nilai persediaan tidak sinkron dengan buku besar.", warn)]

    story += [Spacer(1, 7 * mm),
              Paragraph(f"<b>Mutasi bernilai — {period_label(month)}</b>", small), Spacer(1, 2 * mm)]
    mv_head = ["Waktu", "Kode", "Aksesoris", "Jenis", "Qty", "HPP", "Nilai", "No. Jurnal"]
    mv_data = [mv_head]
    for m in movements[:200]:
        mv_data.append([_dt(m.get("created_at")), m.get("material_code", ""),
                        (m.get("material_name") or "")[:30],
                        MV_LABEL.get(m.get("movement_type"), m.get("movement_type") or ""),
                        _num(m.get("qty_signed")), _rp(m.get("unit_cost")), _rp(m.get("value")),
                        m.get("je_number") or "(tanpa jurnal)"])
    if not movements:
        mv_data.append(["(tidak ada mutasi pada periode ini)", "", "", "", "", "", "", ""])
    mv_tbl = Table(mv_data, colWidths=[28 * mm, 24 * mm, 52 * mm, 20 * mm, 20 * mm, 26 * mm, 28 * mm, 36 * mm],
                   repeatRows=1)
    mv_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EDE9FE")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D4D4D8")),
        ("ALIGN", (4, 1), (6, -1), "RIGHT"),
    ]))
    story.append(mv_tbl)

    doc.build(story)
    return buf.getvalue(), _filename("pdf", month)
