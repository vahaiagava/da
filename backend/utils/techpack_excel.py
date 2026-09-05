"""
Parser for DA "Data Techpack Ringkasan Produk V5" Excel.

Layout: single sheet, multi-row per product. A product starts on a row where
col B (Nama Produk) is non-empty; subsequent rows with empty B are continuation
rows that add construction points (H), fabric consumption (K) and measurement
categories (L).

Columns (1-indexed):
  A SKU | B Nama Produk | C Kategori | D Buyer/Brand | E Season/Tanggal
  F Deskripsi | G Variasi Warna (';') | H Tipe Jahitan (per poin/baris)
  I Size | J Material ('&' = komposit) | K Penggunaan Bahan (per size, + Kombinasi)
  L Ukuran Detail (per kategori: "KATEGORI: titik: nilai; ...")

Returns normalized product dicts ready to build styles + variants + tech packs.
Defensive: unparseable cells are kept as `raw` and never crash the import.
"""
import re
import io
import unicodedata

import openpyxl

# Header row is row 1; data starts row 2.
COL = {"sku": 1, "name": 2, "category": 3, "buyer": 4, "season": 5, "desc": 6,
       "colors": 7, "construction": 8, "size": 9, "material": 10,
       "consumption": 11, "measurement": 12}

_SIZE_TOKENS = ["XXXL", "XXL", "XL", "S", "M", "L", "ALLSIZE"]


def _s(v):
    return "" if v is None else str(v).strip()


def _slug_code(name: str) -> str:
    """Derive a stable style code from product name (UPPER, alnum + dashes)."""
    n = unicodedata.normalize("NFKD", name or "").encode("ascii", "ignore").decode()
    n = re.sub(r"[^A-Za-z0-9]+", "-", n).strip("-").upper()
    return n[:40] or "STYLE"


def _norm_size(tok: str) -> str:
    t = re.sub(r"[^A-Z0-9]", "", (tok or "").upper())
    if not t:
        return ""
    if "ALLSIZE" in t or t in ("ALL", "ALLSZ", "OS", "ONESIZE"):
        return "ALLSIZE"
    for s in _SIZE_TOKENS:
        if t == s:
            return s
    return t  # keep custom size code as-is (uppercased)


def _parse_colors(cell: str):
    return [c.strip() for c in re.split(r"[;\n]", _s(cell)) if c.strip()]


def _parse_material(cell: str):
    """'A & B' -> [{name:A, role:main}, {name:B, role:combination}]."""
    parts = [p.strip() for p in re.split(r"\s*&\s*|\s*\+\s*", _s(cell)) if p.strip()]
    out = []
    for i, p in enumerate(parts):
        out.append({"name": p, "role": "main" if i == 0 else "combination"})
    return out


_NUM = r"(\d+(?:[.,]\d+)?)"


def _num(x):
    if x is None:
        return None
    try:
        return float(str(x).replace(",", "."))
    except ValueError:
        return None


def _parse_consumption_cell(cell: str):
    """
    'M: 463 cm - 150 (5 pcs)'                          -> 1 entry (main)
    'M: 168 cm (4 pcs), Kombinasi: 27 cm (3 pcs)'      -> 2 entries (main + combination)
    'All size: 338 cm (3 pcs)'                         -> ALLSIZE main
    """
    raw = _s(cell)
    if not raw:
        return []
    # size prefix = text before the FIRST colon of the whole cell
    size = ""
    m = re.match(r"\s*([A-Za-z0-9 ]+?)\s*:", raw)
    if m:
        size = _norm_size(m.group(1))
    entries = []
    # split into segments: on comma OR right before a "Kombinasi"/"Combination" keyword
    # (source data uses both "..., Kombinasi:" and "... Kombinasi:" with just a space).
    segments = [seg.strip() for seg in
                re.split(r"\s*,\s*|\s+(?=(?:kombinasi|combination)\b)", raw, flags=re.I)
                if seg.strip()]
    for idx, seg in enumerate(segments):
        low = seg.lower()
        role = "combination" if ("kombinasi" in low or "combination" in low) else "main"
        length = re.search(_NUM + r"\s*cm", seg, re.I)
        # width: number right after a '-' (e.g. '463 cm - 150')
        width = re.search(r"-\s*" + _NUM, seg)
        yield_m = re.search(r"\(\s*" + _NUM + r"\s*pcs", seg, re.I)
        entries.append({
            "size": size or "ALLSIZE",
            "fabric_role": role,
            "length_cm": _num(length.group(1)) if length else None,
            "width_cm": _num(width.group(1)) if width else None,
            "yield_pcs": int(_num(yield_m.group(1))) if yield_m else None,
            "raw": seg,
        })
    return entries or [{"size": size or "ALLSIZE", "fabric_role": "main",
                        "length_cm": None, "width_cm": None, "yield_pcs": None, "raw": raw}]


def _parse_measurement_cell(cell: str):
    """
    'STANDAR: LD: 110 CM; Panjang Dress: 60 CM; Bahu: 12 CM'
    -> (category='STANDAR', [(point, value), ...])
    """
    raw = _s(cell)
    if not raw:
        return None, []
    # category = text before first colon
    m = re.match(r"\s*([^:]+?)\s*:\s*(.*)$", raw, re.S)
    if not m:
        return None, []
    category = m.group(1).strip().upper()
    rest = m.group(2)
    points = []
    for chunk in rest.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        pm = re.match(r"\s*(.+?)\s*:\s*(.+)$", chunk)
        if pm:
            points.append((pm.group(1).strip(), pm.group(2).strip()))
    return category, points


def parse_techpack_v5(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    max_row, max_col = ws.max_row, ws.max_column

    def cell(r, key):
        c = COL[key]
        if c > max_col:
            return ""
        return _s(ws.cell(row=r, column=c).value)

    # product start rows = rows with non-empty Nama Produk (skip header row 1)
    starts = [r for r in range(2, max_row + 1) if cell(r, "name")]
    starts.append(max_row + 1)

    products, errors = [], []
    for pi in range(len(starts) - 1):
        r1, r2 = starts[pi], starts[pi + 1] - 1
        try:
            name = cell(r1, "name")
            sku = cell(r1, "sku")
            code = sku.upper() if sku else _slug_code(name)

            colors = _parse_colors(cell(r1, "colors"))

            construction_points = []
            for r in range(r1, r2 + 1):
                txt = cell(r, "construction")
                if txt:
                    construction_points.append({
                        "seq": len(construction_points) + 1,
                        "title": "",
                        "description": txt,
                    })

            fabrics = _parse_material(cell(r1, "material"))

            consumption = []
            for r in range(r1, r2 + 1):
                consumption.extend(_parse_consumption_cell(cell(r, "consumption")))

            # measurement categories + merged points
            categories = []
            points_map = {}   # point -> {cat: value}
            point_order = []
            for r in range(r1, r2 + 1):
                cat, pts = _parse_measurement_cell(cell(r, "measurement"))
                if not cat:
                    continue
                if cat not in categories:
                    categories.append(cat)
                for p, val in pts:
                    if p not in points_map:
                        points_map[p] = {}
                        point_order.append(p)
                    points_map[p][cat] = val
            measurements = [{"point": p, "values": points_map[p]} for p in point_order]

            # sizes: prefer consumption size prefixes; fallback to col I parse
            sizes = []
            for e in consumption:
                sz = e.get("size")
                if sz and sz not in sizes:
                    sizes.append(sz)
            size_raw = cell(r1, "size")
            if not sizes:
                if re.search(r"all\s*size", size_raw, re.I):
                    sizes = ["ALLSIZE"]
                else:
                    found = re.findall(r"\b(XXXL|XXL|XL|S|M|L)\b", size_raw.upper())
                    sizes = list(dict.fromkeys(found)) or ["ALLSIZE"]

            # #2b: fit categories (info saja, tidak mengubah SKU/varian).
            # Ambil dari label kategori pengukuran (STANDAR/JUMBO) atau dari kolom Size.
            fit_categories = []
            for label in list(categories) + re.findall(
                    r"(standar|standard|jumbo|reguler|regular|plus|big\s*size)", size_raw, re.I):
                lab = str(label).strip().upper()
                if lab in ("STANDAR", "STANDARD", "JUMBO", "REGULER", "REGULAR", "PLUS", "BIG SIZE") \
                        and lab not in fit_categories:
                    fit_categories.append("STANDARD" if lab == "STANDARD" else lab)

            products.append({
                "row_start": r1, "row_end": r2,
                "sku": sku, "style_code": code, "style_name": name,
                "category": cell(r1, "category"),
                "buyer": cell(r1, "buyer") or "DA",
                "season": cell(r1, "season"),
                "description": cell(r1, "desc"),
                "colors": colors,
                "sizes": sizes,
                "size_raw": size_raw,
                "fit_categories": fit_categories,
                "fabrics": fabrics,
                "construction_points": construction_points,
                "fabric_consumption": consumption,
                "measurement_categories": categories,
                "measurements": measurements,
            })
        except Exception as e:  # never let one bad product break the batch
            errors.append({"row": r1, "name": cell(r1, "name"), "error": str(e)})

    return {
        "sheet": ws.title,
        "total": len(products),
        "products": products,
        "errors": errors,
    }
