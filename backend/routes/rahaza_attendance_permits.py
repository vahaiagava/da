"""rahaza_attendance_permits.py — PERSETUJUAN IZIN KELUAR (FASE 16).

═══════════════════════════════════════════════════════════════════════════════
KENAPA MODUL INI ADA
═══════════════════════════════════════════════════════════════════════════════
Keputusan user 2026-07-26: *"Izin keluar butuh persetujuan atasan/HR dulu, bukan
langsung tercatat."* Sebelumnya `POST /attendance/sessions/start {kind:'izin'}`
langsung membuka sesi ⇒ siapa pun bisa keluar kapan pun dan jam kerjanya
otomatis terpotong tanpa ada yang menyetujui.

Alur sekarang:

    karyawan  ── ajukan ──▶  sesi izin `approval_status='pending'` (out_at=None)
    HR/atasan ── setujui ─▶  `approved`, out_at = waktu persetujuan → izin BERJALAN
              └─ tolak ───▶  `rejected` (menit TIDAK pernah dihitung)
    karyawan  ── batalkan ▶  `cancelled` (hanya selagi masih pending)

Data tetap tinggal di `rahaza_attendance_events.sessions[]` — TIDAK ada koleksi
baru (menghindari SSOT tandingan; lihat AGENT_DEVELOPMENT_RULES.md).

Identitas & otorisasi memakai SSOT `utils/employee_identity.py`
(`resolve_employee_for` / `is_hr`) — jangan menulis query identitas sendiri.
"""
from __future__ import annotations

import logging
from datetime import date, datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, HTTPException, Request

from auth import log_activity, require_auth, serialize_doc
from database import get_db
from routes.rahaza_attendance_sessions import (APPROVAL_APPROVED,
                                               APPROVAL_CANCELLED,
                                               APPROVAL_PENDING,
                                               APPROVAL_REJECTED,
                                               recompute_session_totals,
                                               session_approval)
from utils.employee_identity import is_hr, resolve_my_employee

router = APIRouter(prefix="/api/rahaza", tags=["rahaza-hr"])
log = logging.getLogger(__name__)

# Zona waktu operasional (WIB) — jam pada laporan ditampilkan waktu lokal.
WIB = timezone(timedelta(hours=7))


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _today_iso() -> str:
    return date.today().isoformat()


def _require_approver(user: dict) -> None:
    """Hanya HR/atasan yang boleh memutuskan izin orang lain."""
    if not is_hr(user):
        raise HTTPException(
            403, "Hanya HR/atasan yang boleh menyetujui atau menolak izin keluar.")


async def _find_permit(db, session_id: str) -> tuple[dict, dict]:
    """Cari (record absensi, sesi) berdasarkan id sesi. 404 bila tidak ada."""
    rec = await db.rahaza_attendance_events.find_one(
        {"sessions.id": session_id}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "Pengajuan izin tidak ditemukan.")
    for s in rec.get("sessions") or []:
        if s.get("id") == session_id:
            return rec, s
    raise HTTPException(404, "Pengajuan izin tidak ditemukan.")


async def _employee_brief(db, emp_id: str) -> dict:
    e = await db.rahaza_employees.find_one(
        {"id": emp_id},
        {"_id": 0, "id": 1, "name": 1, "employee_code": 1, "job_title": 1,
         "department": 1, "user_id": 1},
    ) or {}
    return e


async def _notify_employee(db, emp: dict, sess: dict, decided_by: str,
                           approved: bool) -> None:
    """Kabari karyawan hasil keputusan (gagal notif tidak boleh membatalkan aksi)."""
    try:
        from routes.rahaza_notifications import publish_notification
        target_ids = []
        uid = emp.get("user_id")
        if uid:
            target_ids.append(uid)
        else:
            u = await db.users.find_one({"employee_id": emp.get("id")}, {"_id": 0, "id": 1})
            if u:
                target_ids.append(u["id"])
        await publish_notification(
            db,
            type_="attendance_permit_decision",
            severity="success" if approved else "warning",
            title="Izin keluar disetujui" if approved else "Izin keluar ditolak",
            message=(f"Izin \"{sess.get('reason', '')}\" "
                     f"{'DISETUJUI' if approved else 'DITOLAK'} oleh {decided_by}."
                     + ("" if approved else
                        f" Catatan: {sess.get('decision_notes') or '-'}")),
            link_module="portal-absen",
            link_id=sess.get("id"),
            target_user_ids=target_ids or None,
            target_roles=None if target_ids else ["hr"],
            dedup_key=f"permit-dec-{sess.get('id')}",
        )
    except Exception:  # noqa: BLE001
        log.debug("notif keputusan izin gagal", exc_info=True)


# ═════════════════════════════════════════════════════════════════════════════
# DAFTAR
# ═════════════════════════════════════════════════════════════════════════════
@router.get("/attendance/permits")
async def list_permits(request: Request,
                       status: str = "pending",
                       from_date: Optional[str] = None,
                       to_date: Optional[str] = None,
                       employee_id: Optional[str] = None,
                       limit: int = 500):
    """Daftar pengajuan izin.

    · HR/atasan → semua karyawan (bisa disaring `employee_id`).
    · Karyawan biasa → otomatis hanya miliknya sendiri.
    `status`: pending | approved | rejected | cancelled | all
    """
    user = await require_auth(request)
    db = get_db()

    q: dict = {"sessions.kind": "izin"}
    scope_emp_id = None
    if not is_hr(user):
        mine = await resolve_my_employee(db, user)
        if not mine:
            return {"ok": True, "total": 0, "items": [], "scope": "self-unlinked"}
        scope_emp_id = mine["id"]
    elif employee_id:
        scope_emp_id = employee_id
    if scope_emp_id:
        q["employee_id"] = scope_emp_id

    if from_date or to_date:
        rng = {}
        if from_date:
            rng["$gte"] = from_date
        if to_date:
            rng["$lte"] = to_date
        q["date"] = rng

    limit = max(1, min(int(limit or 500), 2000))
    rows = await db.rahaza_attendance_events.find(q, {"_id": 0}).sort("date", -1).to_list(limit)
    emp_ids = list({r.get("employee_id") for r in rows if r.get("employee_id")})
    emps = await db.rahaza_employees.find(
        {"id": {"$in": emp_ids}},
        {"_id": 0, "id": 1, "name": 1, "employee_code": 1, "job_title": 1, "department": 1},
    ).to_list(2000) if emp_ids else []
    e_map = {e["id"]: e for e in emps}

    items = []
    for r in rows:
        e = e_map.get(r.get("employee_id")) or {}
        for s in r.get("sessions") or []:
            if s.get("kind") != "izin":
                continue
            appr = session_approval(s)
            if status != "all" and appr != status:
                continue
            items.append({
                **s,
                "approval_status": appr,
                "date": r.get("date"),
                "employee_id": r.get("employee_id"),
                "employee_name": e.get("name", "?"),
                "employee_code": e.get("employee_code", "-"),
                "job_title": e.get("job_title", ""),
                "department": e.get("department", ""),
            })
    items.sort(key=lambda x: str(x.get("requested_at") or x.get("out_at") or ""), reverse=True)
    return serialize_doc({"ok": True, "total": len(items), "items": items,
                          "scope": "self" if scope_emp_id and not is_hr(user) else "all"})


@router.get("/attendance/permits/pending-count")
async def pending_count(request: Request):
    """Angka lencana untuk menu HR (murah, tanpa memuat daftar)."""
    user = await require_auth(request)
    db = get_db()
    if not is_hr(user):
        return {"ok": True, "count": 0}
    rows = await db.rahaza_attendance_events.find(
        {"sessions.kind": "izin"}, {"_id": 0, "sessions": 1}).to_list(2000)
    n = sum(1 for r in rows for s in (r.get("sessions") or [])
            if s.get("kind") == "izin" and session_approval(s) == APPROVAL_PENDING)
    return {"ok": True, "count": n}


# ═════════════════════════════════════════════════════════════════════════════
# KEPUTUSAN
# ═════════════════════════════════════════════════════════════════════════════
async def _decide(db, user: dict, session_id: str, approve: bool,
                  notes: str = "") -> dict:
    rec, sess = await _find_permit(db, session_id)
    if sess.get("kind") != "izin":
        raise HTTPException(400, "Sesi ini bukan pengajuan izin.")
    appr = session_approval(sess)
    if appr != APPROVAL_PENDING:
        raise HTTPException(
            400, f"Pengajuan sudah berstatus '{appr}' — tidak bisa diputuskan lagi.")

    now = _now()
    sessions = list(rec.get("sessions") or [])
    updated = None
    for s in sessions:
        if s.get("id") != session_id:
            continue
        s["approval_status"] = APPROVAL_APPROVED if approve else APPROVAL_REJECTED
        s["approved_by"] = user["id"]
        s["approved_by_name"] = user.get("name", "")
        s["approved_at"] = now
        s["decision_notes"] = (notes or "").strip() or None
        if approve:
            # Izin BERJALAN sejak disetujui — bukan sejak diajukan.
            s["out_at"] = now
        else:
            s["out_at"] = None
            s["in_at"] = None
            s["minutes"] = 0
        updated = s
        break

    rec["sessions"] = sessions
    totals = recompute_session_totals(rec)
    await db.rahaza_attendance_events.update_one(
        {"id": rec["id"]},
        {"$set": {"sessions": sessions, "updated_at": now, **totals}},
    )
    emp = await _employee_brief(db, rec.get("employee_id"))
    await log_activity(user["id"], user.get("name", ""),
                       "izin-disetujui" if approve else "izin-ditolak",
                       "attendance", rec.get("employee_id"))
    await _notify_employee(db, emp, updated, user.get("name", "HR"), approve)
    return serialize_doc({
        "ok": True,
        "session": updated,
        "employee": emp,
        "message": ("Izin disetujui — durasi dihitung mulai sekarang."
                    if approve else "Izin ditolak."),
        **totals,
    })


@router.post("/attendance/permits/{session_id}/approve")
async def approve_permit(session_id: str, request: Request):
    user = await require_auth(request)
    _require_approver(user)
    body = await _safe_body(request)
    return await _decide(get_db(), user, session_id, True, body.get("notes", ""))


@router.post("/attendance/permits/{session_id}/reject")
async def reject_permit(session_id: str, request: Request):
    user = await require_auth(request)
    _require_approver(user)
    body = await _safe_body(request)
    notes = (body.get("notes") or body.get("reason") or "").strip()
    if not notes:
        raise HTTPException(400, "Alasan penolakan wajib diisi.")
    return await _decide(get_db(), user, session_id, False, notes)


@router.post("/attendance/permits/{session_id}/cancel")
async def cancel_permit(session_id: str, request: Request):
    """Karyawan membatalkan pengajuannya sendiri (selagi masih `pending`)."""
    user = await require_auth(request)
    db = get_db()
    rec, sess = await _find_permit(db, session_id)
    mine = await resolve_my_employee(db, user)
    owner = mine and mine.get("id") == rec.get("employee_id")
    if not owner and not is_hr(user):
        raise HTTPException(403, "Anda hanya bisa membatalkan pengajuan izin sendiri.")
    if session_approval(sess) != APPROVAL_PENDING:
        raise HTTPException(400, "Hanya pengajuan yang masih menunggu yang bisa dibatalkan.")

    now = _now()
    sessions = list(rec.get("sessions") or [])
    for s in sessions:
        if s.get("id") == session_id:
            s["approval_status"] = APPROVAL_CANCELLED
            s["out_at"] = None
            s["in_at"] = None
            s["minutes"] = 0
            s["decision_notes"] = "Dibatalkan oleh pemohon"
            break
    rec["sessions"] = sessions
    totals = recompute_session_totals(rec)
    await db.rahaza_attendance_events.update_one(
        {"id": rec["id"]},
        {"$set": {"sessions": sessions, "updated_at": now, **totals}},
    )
    await log_activity(user["id"], user.get("name", ""), "izin-dibatalkan",
                       "attendance", rec.get("employee_id"))
    return serialize_doc({"ok": True, "message": "Pengajuan izin dibatalkan.", **totals})


async def _safe_body(request: Request) -> dict:
    """Body JSON opsional — approve tanpa body tidak boleh 500."""
    try:
        b = await request.json()
        return b if isinstance(b, dict) else {}
    except Exception:  # noqa: BLE001
        return {}


# ═════════════════════════════════════════════════════════════════════════════
# EXPORT EXCEL — rekap istirahat & izin (permintaan user 2026-07-26)
# ═════════════════════════════════════════════════════════════════════════════
@router.get("/attendance/sessions/export.xlsx")
async def export_sessions_excel(request: Request,
                                from_date: Optional[str] = None,
                                to_date: Optional[str] = None,
                                employee_id: Optional[str] = None,
                                kind: Optional[str] = None,
                                status: Optional[str] = None):
    """Unduh rekap istirahat & izin sebagai XLSX (filter sama dengan layar)."""
    import io

    from fastapi.responses import StreamingResponse

    from routes.rahaza_attendance_sessions import list_sessions

    data = await list_sessions(request, from_date=from_date, to_date=to_date,
                               employee_id=employee_id, kind=kind, status=status,
                               limit=5000)
    items = data.get("items", [])
    summary = data.get("summary", {})

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:  # pragma: no cover
        raise HTTPException(500, "openpyxl belum terpasang di server.")

    def _jam(v) -> str:
        if not v:
            return "-"
        try:
            dt = v if isinstance(v, datetime) else datetime.fromisoformat(
                str(v).replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(WIB).strftime("%H:%M")
        except Exception:  # noqa: BLE001
            return "-"

    wb = Workbook()
    ws = wb.active
    ws.title = "Istirahat & Izin"
    ws.append(["CV. DEWI ADITYA — Rekap Istirahat & Izin Keluar"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([f"Periode: {from_date or 'awal'} s/d {to_date or 'hari ini'}"])
    ws.append([f"Istirahat: {summary.get('istirahat_count', 0)} sesi "
               f"({summary.get('istirahat_minutes', 0)} menit) · "
               f"Izin: {summary.get('izin_count', 0)} sesi "
               f"({summary.get('izin_minutes', 0)} menit disetujui)"])
    ws.append([])

    headers = ["Tanggal", "NIK", "Nama", "Jenis", "Keluar", "Kembali",
               "Durasi (menit)", "Alasan", "Status", "Diputuskan oleh",
               "Catatan", "Dihitung?"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[5]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    for it in items:
        ws.append([
            it.get("date", "-"),
            it.get("employee_code", "-"),
            it.get("employee_name", "-"),
            (it.get("kind") or "").capitalize(),
            _jam(it.get("out_at")),
            _jam(it.get("in_at")),
            int(it.get("minutes") or 0),
            it.get("reason") or "-",
            it.get("approval_status", "-"),
            it.get("approved_by_name") or "-",
            it.get("decision_notes") or "-",
            "Ya" if it.get("counted") else "Tidak",
        ])

    for col, width in zip("ABCDEFGHIJKL",
                          [12, 12, 24, 12, 10, 10, 14, 34, 14, 20, 28, 11]):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A6"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"rekap_istirahat_izin_{from_date or 'awal'}_{to_date or _today_iso()}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )
