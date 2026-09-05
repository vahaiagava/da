# ruff: noqa: F401
"""
rahaza_payroll_runs.py — Payroll Run Operations
Extracted from rahaza_payroll.py (1539 LOC monolith)

Refactored: Session #11.19 Phase 3.2 Batch #4
Endpoints: GET /payroll-runs, POST /payroll-runs, GET /payroll-runs/{id}, POST finalize/post-to-gl/retry-post/pay/void-payment, DELETE, POST pay-bpjs/pay-pph21, GET export/pdf
"""
from fastapi import APIRouter, Request, HTTPException, Depends
from routes.shared import require_portal_dep
from fastapi.responses import StreamingResponse
from database import get_db
from auth import require_auth, serialize_doc, log_activity
from routes.rahaza_payroll_shared import (
    _uid, _now, VALID_SCHEMES, VALID_PERIOD_TYPES, VALID_RUN_STATUS,
    _get_applicable_allowances,
    _require_hr, _to_date, _date_range_filter,
    _generate_run_number, _compute_payslip_for_employee,
)
from routes.rahaza_posting import post_payroll_run, payroll_deduction_totals
from utils.saga import SagaExecutor
import uuid
import io
import csv
import logging
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers as xl_numbers
)
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, date, timedelta
from typing import Optional

log = logging.getLogger(__name__)

router = APIRouter(prefix="/api/rahaza", tags=["rahaza-payroll-runs"],
                   dependencies=[Depends(require_portal_dep("hr", "finance"))])  # RBAC: HR/finance (BUG-RBAC-1)

@router.get("/payroll-runs")
async def list_runs(request: Request, status: Optional[str] = None, limit: int = 50, skip: int = 0):
    await require_auth(request)
    db = get_db()
    q = {}
    if status:
        q["status"] = status
    rows = await db.rahaza_payroll_runs.find(q, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(500)
    return serialize_doc(rows)


@router.post("/payroll-runs")
async def create_run(request: Request):
    user = await _require_hr(request)
    db = get_db()
    body = await request.json()
    period_from = (body.get("period_from") or "").strip()
    period_to = (body.get("period_to") or "").strip()
    if not (period_from and period_to):
        raise HTTPException(400, "period_from & period_to wajib (YYYY-MM-DD).")
    try:
        _to_date(period_from)
        _to_date(period_to)
    except Exception:
        raise HTTPException(400, "Format tanggal harus YYYY-MM-DD.")
    if period_from > period_to:
        raise HTTPException(400, "period_from tidak boleh > period_to.")

    # Ambil profile aktif
    employee_ids = body.get("employee_ids") or []
    q = {"active": True}
    if employee_ids:
        q["employee_id"] = {"$in": employee_ids}
    profiles = await db.rahaza_payroll_profiles.find(q, {"_id": 0}).to_list(500)
    if not profiles:
        raise HTTPException(400, "Tidak ada payroll profile aktif untuk diproses. Buat profile dulu di menu Payroll Profiles.")

    emp_ids = [p["employee_id"] for p in profiles]
    emps = await db.rahaza_employees.find({"id": {"$in": emp_ids}}, {"_id": 0}).to_list(500)
    e_map = {e["id"]: e for e in emps}

    # Create run header
    run_number = await _generate_run_number(db)
    run_id = _uid()
    now = _now()

    # Generate payslips
    payslips = []
    for p in profiles:
        emp = e_map.get(p["employee_id"])
        if not emp:
            continue
        slip = await _compute_payslip_for_employee(db, p, period_from, period_to, emp)
        slip.update({
            "run_id": run_id,
            "run_number": run_number,
            "created_at": now,
            "updated_at": now,
        })
        payslips.append(slip)

    # ── Saga pattern: atomic payslip insert + run header insert ─────────────────
    # Since MongoDB standalone doesn't support multi-document transactions,
    # we use a compensation saga: if run header insert fails, payslips are deleted.
    payslips_inserted = False

    total_gross = sum(s["gross_pay"] for s in payslips)
    total_ded = sum(s["deductions_total"] for s in payslips)
    total_net = sum(s["net_pay"] for s in payslips)

    run_doc = {
        "id": run_id,
        "run_number": run_number,
        "period_from": period_from,
        "period_to": period_to,
        "status": "draft",
        "total_employees": len(payslips),
        "total_gross": total_gross,
        "total_deductions": total_ded,
        "total_net": total_net,
        "notes": body.get("notes") or "",
        "created_at": now,
        "created_by": user["id"],
        "created_by_name": user.get("name", ""),
        "updated_at": now,
    }

    async def _insert_payslips():
        nonlocal payslips_inserted
        if payslips:
            await db.rahaza_payslips.insert_many(payslips)
        payslips_inserted = True

    async def _compensate_payslips():
        await db.rahaza_payslips.delete_many({"run_id": run_id})

    async def _insert_run_header():
        await db.rahaza_payroll_runs.insert_one(run_doc)

    saga = SagaExecutor(name="create_payroll_run")
    saga.add_step(
        name="insert_payslips",
        action=_insert_payslips,
        compensate=_compensate_payslips,
    )
    saga.add_step(
        name="insert_run_header",
        action=_insert_run_header,
        compensate=lambda: db.rahaza_payroll_runs.delete_one({"id": run_id}),
    )
    saga_result = await saga.execute()
    if not saga_result.success:
        log.error(f"Saga failed creating payroll run {run_number}: {saga_result.error_detail}")
        raise HTTPException(500, f"Gagal membuat payroll run: {saga_result.error_detail}")

    await log_activity(user["id"], user.get("name", ""), "create", "rahaza.payroll_run", run_number)
    out = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    return serialize_doc(out)


@router.get("/payroll-runs/{run_id}")
async def get_run(run_id: str, request: Request):
    await require_auth(request)
    db = get_db()
    run = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Payroll run tidak ditemukan.")
    payslips = await db.rahaza_payslips.find({"run_id": run_id}, {"_id": 0}).sort("employee_code", 1).to_list(500)
    return serialize_doc({"run": run, "payslips": payslips})


@router.post("/payroll-runs/{run_id}/finalize")
async def finalize_run(run_id: str, request: Request):
    user = await _require_hr(request)
    db = get_db()
    run = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Payroll run tidak ditemukan.")
    if run.get("status") != "draft":
        raise HTTPException(400, f"Run sudah ber-status '{run.get('status')}', tidak bisa finalize.")
    # Recalc totals dari payslips (in case deductions diubah)
    payslips = await db.rahaza_payslips.find({"run_id": run_id}, {"_id": 0}).to_list(500)
    total_gross = sum(s.get("gross_pay", 0) for s in payslips)
    total_ded = sum(s.get("deductions_total", 0) for s in payslips)
    total_net = sum(s.get("net_pay", 0) for s in payslips)
    comp = payroll_deduction_totals(payslips)  # H-02: komponen potongan disimpan di run
    await db.rahaza_payroll_runs.update_one({"id": run_id}, {"$set": {
        "status": "finalized",
        "total_gross": total_gross,
        "total_deductions": total_ded,
        "total_net": total_net,
        "total_pph21": comp["pph21"],
        "total_bpjs_employee": comp["bpjs"],
        "total_kasbon": comp["kasbon"],
        "total_other_deductions": comp["other"],
        "deductions_by_type": comp["by_type"],
        "finalized_at": _now(),
        "finalized_by": user["id"],
        "finalized_by_name": user.get("name", ""),
        "updated_at": _now(),
    }})
    await log_activity(user["id"], user.get("name", ""), "finalize", "rahaza.payroll_run", run.get("run_number"))
    out = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})

    # ── F3 Auto-post Payroll JE
    posting_result = None
    try:
        posting_result = await post_payroll_run(db, out, user)
    except Exception as e:
        log.exception("Payroll auto-post failed")
        posting_result = {"ok": False, "error": str(e)}
    out = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})

    # ── Notifikasi payslip siap ke semua karyawan dalam run ini ──────────────
    try:
        from routes.rahaza_notifications import publish_notification
        # Kumpulkan user_id dari employees dalam run
        payslip_emps = await db.rahaza_payslips.find(
            {"run_id": run_id}, {"_id": 0, "employee_id": 1, "net_pay": 1}
        ).to_list(500)
        emp_ids_in_run = [s["employee_id"] for s in payslip_emps]
        if emp_ids_in_run:
            linked_emps = await db.rahaza_employees.find(
                {"id": {"$in": emp_ids_in_run}, "user_id": {"$exists": True, "$ne": None}},
                {"_id": 0, "user_id": 1, "name": 1}
            ).to_list(500)
            for le in linked_emps:
                net = next((s["net_pay"] for s in payslip_emps if s["employee_id"] == le.get("id")), 0)
                await publish_notification(
                    db,
                    type_="payslip_ready",
                    severity="info",
                    title="Slip Gaji Tersedia",
                    message=(
                        f"Slip gaji periode {out.get('period_from','')[:7]} sudah tersedia. "
                        f"Take-home: Rp {net:,.0f}."
                    ),
                    link_module="self-dashboard",
                    target_user_ids=[le["user_id"]],
                    dedup_key=f"payslip_ready_{run_id}_{le['user_id']}",
                )
    except Exception as ne:
        log.warning(f"[payroll] payslip notif failed: {ne}")

    out["_posting_result"] = posting_result
    return serialize_doc(out)


@router.post("/payroll-runs/{run_id}/post-to-gl")
async def retry_post_payroll(run_id: str, request: Request):
    """F3: manual retry post payroll run to GL (idempotent)."""
    user = await _require_hr(request)
    db = get_db()
    run = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Payroll run tidak ditemukan.")
    if run.get("status") != "finalized":
        raise HTTPException(400, "Hanya run yang sudah finalized yang bisa di-post.")
    result = await post_payroll_run(db, run, user)
    out = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    out["_posting_result"] = result
    return serialize_doc(out)


@router.post("/payroll-runs/{run_id}/retry-post")
async def retry_post_alias(run_id: str, request: Request):
    """Alias untuk /post-to-gl (backward compat frontend)."""
    return await retry_post_payroll(run_id, request)


@router.post("/payroll-runs/{run_id}/pay")
async def pay_payroll_run(run_id: str, request: Request):
    """
    Tandai gaji sebagai sudah dibayar dan buat Payment JE.
    Dr 2-1200 Hutang Gaji / Cr [bank_account_code].
    Body: { payment_date, bank_account_code, payment_method, notes }
    """
    user = await _require_hr(request)
    db   = get_db()
    body = await request.json()

    run = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Payroll run tidak ditemukan.")
    if run.get("status") != "finalized":
        raise HTTPException(400, "Hanya run FINALIZED yang bisa dibayar.")
    if run.get("payment_status") == "paid":
        raise HTTPException(400,
            f"Penggajian {run.get('run_number')} sudah dibayar "
            f"({run.get('payment_gl_je_number')}). "
            "Gunakan void-payment untuk membatalkan.")

    payment_date   = (body.get("payment_date") or str(date.today()))[:10]
    bank_code      = (body.get("bank_account_code") or "1-1201").strip()
    payment_method = body.get("payment_method") or "bank_transfer"
    notes          = (body.get("notes") or "").strip()

    # Validate bank CoA exists
    bank_acc = await db.rahaza_coa_accounts.find_one(
        {"code": bank_code, "active": True}, {"_id": 0, "name": 1}
    )
    if not bank_acc:
        raise HTTPException(400, f"Akun GL '{bank_code}' tidak ditemukan atau tidak aktif.")

    from routes.rahaza_posting import post_payroll_payment
    result = await post_payroll_payment(db, run, payment_date, bank_code, user)

    update = {
        "payment_status":       "paid" if result.get("ok") else "payment_error",
        "payment_method":       payment_method,
        "payment_date":         payment_date,
        "payment_bank_code":    bank_code,
        "payment_bank_name":    bank_acc.get("name", ""),
        "payment_notes":        notes,
        "payment_gl_je_id":     result.get("je_id"),
        "payment_gl_je_number": result.get("je_number"),
        "payment_error":        result.get("error"),
        "paid_at":              _now(),
        "paid_by":              user["id"],
        "paid_by_name":         user.get("name", ""),
        "updated_at":           _now(),
    }
    await db.rahaza_payroll_runs.update_one({"id": run_id}, {"$set": update})
    await log_activity(user["id"], user.get("name", ""), "pay_payroll", "rahaza.payroll_run",
                       f"{run.get('run_number')} → {bank_code} {payment_date}")
    out = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    out["_payment_result"] = result
    return serialize_doc(out)


@router.post("/payroll-runs/{run_id}/void-payment")
async def void_payroll_payment_endpoint(run_id: str, request: Request):
    """
    Batalkan jurnal pembayaran gaji (void payment JE).
    Hanya bisa dilakukan jika payment JE masih aktif.
    Body: { reason }
    """
    user = await _require_hr(request)
    db   = get_db()
    body = await request.json()
    reason = body.get("reason") or "Pembatalan pembayaran gaji"

    run = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Payroll run tidak ditemukan.")
    if run.get("payment_status") != "paid":
        raise HTTPException(400, "Tidak ada pembayaran aktif yang bisa dibatalkan.")

    from routes.rahaza_posting import void_payroll_payment
    await void_payroll_payment(db, run_id, user, reason)
    await db.rahaza_payroll_runs.update_one(
        {"id": run_id},
        {"$set": {"payment_status": "void", "updated_at": _now()}}
    )
    await log_activity(user["id"], user.get("name", ""), "void_payment", "rahaza.payroll_run",
                       f"{run.get('run_number')} — {reason}")
    out = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    return serialize_doc(out)


@router.delete("/payroll-runs/{run_id}")
async def delete_run(run_id: str, request: Request):
    await _require_hr(request)
    db = get_db()
    run = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Payroll run tidak ditemukan.")
    if run.get("status") == "finalized":
        raise HTTPException(400, "Run yang sudah finalized tidak bisa dihapus. Gunakan cancel atau buat run baru.")
    await db.rahaza_payslips.delete_many({"run_id": run_id})
    await db.rahaza_payroll_runs.delete_one({"id": run_id})
    return {"status": "deleted"}


@router.post("/payroll-runs/{run_id}/pay-bpjs")
async def pay_bpjs(run_id: str, request: Request):
    """
    Bayar BPJS dari payroll run ini.
    Dr 2-1500 Hutang BPJS / Cr [bank_code].
    Body: { payment_date, bank_account_code, notes }
    """
    user = await _require_hr(request)
    db   = get_db()
    body = await request.json()
    run  = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Run tidak ditemukan.")
    if run.get("status") != "finalized":
        raise HTTPException(400, "Hanya run FINALIZED.")
    if run.get("bpjs_payment_status") == "paid":
        raise HTTPException(400, "BPJS run ini sudah dibayar.")

    payment_date = (body.get("payment_date") or str(date.today()))[:10]
    from routes.rahaza_posting_profiles import get_mapping
    fin_map = await get_mapping(db, "payroll_finalize")
    pay_map = await get_mapping(db, "payroll_payment")
    liab_code = fin_map.get("credit_bpjs_payable")
    bank_code = (body.get("bank_account_code") or pay_map.get("credit_bank_default") or "").strip()
    if not liab_code or not bank_code:
        raise HTTPException(400, "Mapping payroll_finalize.credit_bpjs_payable / payroll_payment.credit_bank_default belum lengkap.")
    body.get("notes") or ""

    # H-02: komponen BPJS dari SATU agregator (type bpjs_*) — sama dgn JE finalize
    slips  = await db.rahaza_payslips.find({"run_id": run_id}, {"_id": 0, "deductions": 1}).to_list(500)
    bpjs_total = payroll_deduction_totals(slips)["bpjs"]

    if bpjs_total <= 0:
        raise HTTPException(400, "Tidak ada potongan BPJS di run ini.")

    # Build JE: Dr Hutang BPJS / Cr Bank
    bank_acc = await db.rahaza_coa_accounts.find_one({"code": bank_code, "active": True}, {"_id": 0, "name": 1})
    if not bank_acc:
        raise HTTPException(400, f"Akun GL '{bank_code}' tidak ditemukan.")

    from routes.rahaza_posting import _create_posted_je
    run_id_ref = f"bpjspay:{run_id}"
    try:
        je_date = date.fromisoformat(payment_date)
    except Exception:
        je_date = date.today()
    memo  = f"Bayar BPJS {run.get('run_number')} · {run.get('period_from')}–{run.get('period_to')}"
    lines = [
        {"account_code": liab_code, "debit": bpjs_total, "credit": 0, "description": memo},
        {"account_code": bank_code, "debit": 0, "credit": bpjs_total, "description": memo},
    ]
    result = await _create_posted_je(db, je_date, memo, "bpjs_payment", run_id_ref, lines, user)
    update = {
        "bpjs_payment_status":  "paid" if result.get("ok") else "error",
        "bpjs_payment_date":    payment_date,
        "bpjs_payment_amount":  bpjs_total,
        "bpjs_payment_je":      result.get("je_number"),
        "bpjs_payment_error":   result.get("error"),
        "updated_at":           _now(),
    }
    await db.rahaza_payroll_runs.update_one({"id": run_id}, {"$set": update})
    await log_activity(user["id"], user.get("name",""), "pay_bpjs", "rahaza.payroll_run",
                       f"{run.get('run_number')} BPJS Rp {bpjs_total:,.0f}")
    out = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    out["_payment_result"] = result
    return serialize_doc(out)


@router.post("/payroll-runs/{run_id}/pay-pph21")
async def pay_pph21(run_id: str, request: Request):
    """
    Bayar PPh21 dari payroll run ini ke DJP.
    Dr 2-1301 Hutang PPh21 / Cr [bank_code].
    Body: { payment_date, bank_account_code, notes }
    """
    user = await _require_hr(request)
    db   = get_db()
    body = await request.json()
    run  = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Run tidak ditemukan.")
    if run.get("status") != "finalized":
        raise HTTPException(400, "Hanya run FINALIZED.")
    if run.get("pph21_payment_status") == "paid":
        raise HTTPException(400, "PPh21 run ini sudah dibayar.")

    payment_date = (body.get("payment_date") or str(date.today()))[:10]
    from routes.rahaza_posting_profiles import get_mapping
    fin_map = await get_mapping(db, "payroll_finalize")
    pay_map = await get_mapping(db, "payroll_payment")
    liab_code = fin_map.get("credit_tax_pph21")
    bank_code = (body.get("bank_account_code") or pay_map.get("credit_bank_default") or "").strip()
    if not liab_code or not bank_code:
        raise HTTPException(400, "Mapping payroll_finalize.credit_tax_pph21 / payroll_payment.credit_bank_default belum lengkap.")

    slips = await db.rahaza_payslips.find({"run_id": run_id}, {"_id": 0, "deductions": 1}).to_list(500)
    pph_total = payroll_deduction_totals(slips)["pph21"]

    if pph_total <= 0:
        raise HTTPException(400, "Tidak ada potongan PPh21 di run ini.")

    bank_acc = await db.rahaza_coa_accounts.find_one({"code": bank_code, "active": True}, {"_id": 0, "name": 1})
    if not bank_acc:
        raise HTTPException(400, f"Akun GL '{bank_code}' tidak ditemukan.")

    from routes.rahaza_posting import _create_posted_je
    run_id_ref = f"pph21pay:{run_id}"
    try:
        je_date = date.fromisoformat(payment_date)
    except Exception:
        je_date = date.today()
    memo  = f"Bayar PPh21 {run.get('run_number')} · {run.get('period_from')}–{run.get('period_to')}"
    lines = [
        {"account_code": liab_code, "debit": pph_total, "credit": 0, "description": memo},
        {"account_code": bank_code, "debit": 0, "credit": pph_total, "description": memo},
    ]
    result = await _create_posted_je(db, je_date, memo, "pph21_payment", run_id_ref, lines, user)
    update = {
        "pph21_payment_status": "paid" if result.get("ok") else "error",
        "pph21_payment_date":   payment_date,
        "pph21_payment_amount": pph_total,
        "pph21_payment_je":     result.get("je_number"),
        "pph21_payment_error":  result.get("error"),
        "updated_at":           _now(),
    }
    await db.rahaza_payroll_runs.update_one({"id": run_id}, {"$set": update})
    await log_activity(user["id"], user.get("name",""), "pay_pph21", "rahaza.payroll_run",
                       f"{run.get('run_number')} PPh21 Rp {pph_total:,.0f}")
    out = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    out["_payment_result"] = result
    return serialize_doc(out)


@router.get("/payroll-runs/{run_id}/export-excel")
async def export_run_excel(run_id: str, request: Request):
    """Export payroll run ke Excel — 1 sheet rekapitulasi + 1 sheet per karyawan."""
    await require_auth(request)
    db = get_db()
    run = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Payroll run tidak ditemukan.")
    payslips = await db.rahaza_payslips.find(
        {"$or": [{"run_id": run_id}, {"payroll_run_id": run_id}]}, {"_id": 0}
    ).sort("employee_code", 1).to_list(500)

    # ── Style helpers ────────────────────────────────────────────────────────
    BG_HEADER   = PatternFill("solid", fgColor="1A3A5C")   # Navy gelap
    BG_SUBHDR   = PatternFill("solid", fgColor="2D6A9F")   # Biru sedang
    BG_ALTROW   = PatternFill("solid", fgColor="EBF3FB")   # Biru muda alt row
    BG_TOTAL    = PatternFill("solid", fgColor="FFF3CD")   # Kuning total
    BG_GREEN    = PatternFill("solid", fgColor="D4EDDA")   # Hijau net
    FG_WHITE    = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    FG_DARK     = Font(name="Calibri", color="1A1A1A", size=9)
    FG_DARK_B   = Font(name="Calibri", color="1A1A1A", bold=True, size=9)
    FG_TITLE    = Font(name="Calibri", color="1A3A5C", bold=True, size=14)
    FG_SUBTITLE = Font(name="Calibri", color="2D6A9F", bold=False, size=10)
    BORDER_THIN = Border(
        left=Side(style="thin", color="BBBBBB"),
        right=Side(style="thin", color="BBBBBB"),
        top=Side(style="thin", color="BBBBBB"),
        bottom=Side(style="thin", color="BBBBBB"),
    )
    BORDER_MED  = Border(
        left=Side(style="medium", color="1A3A5C"),
        right=Side(style="medium", color="1A3A5C"),
        top=Side(style="medium", color="1A3A5C"),
        bottom=Side(style="medium", color="1A3A5C"),
    )
    IDR_FMT = '#,##0'
    CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT   = Alignment(horizontal="right",  vertical="center")

    def _cell(ws, row, col, value, font=None, fill=None, align=None, border=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if border:
            c.border = border
        if num_fmt:
            c.number_format = num_fmt
        return c

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    run_num    = run.get("run_number", "PAYROLL")
    period_from = run.get("period_from", "")
    period_to   = run.get("period_to", "")
    period_lbl  = f"{period_from} s/d {period_to}"

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1: Rekapitulasi
    # ─────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Rekapitulasi"
    ws1.freeze_panes = "A8"
    ws1.sheet_view.showGridLines = False

    # Company header
    ws1.merge_cells("A1:P1")
    _cell(ws1, 1, 1, "CV. DEWI ADITYA", font=FG_TITLE, align=CENTER)
    ws1.row_dimensions[1].height = 26

    ws1.merge_cells("A2:P2")
    _cell(ws1, 2, 1, "Jl. Garmen Indah No. 1, Solo, Jawa Tengah | Telp: (0271) 123456", font=FG_SUBTITLE, align=CENTER)

    ws1.merge_cells("A3:P3")
    _cell(ws1, 3, 1, f"SLIP GAJI REKAPITULASI — {run_num} | Periode: {period_lbl}", font=Font(name="Calibri", bold=True, size=11, color="1A3A5C"), align=CENTER)
    ws1.row_dimensions[3].height = 20

    ws1.merge_cells("A4:P4")
    ws1.row_dimensions[4].height = 8

    # Run summary box
    summary_items = [
        ("Run Number", run_num), ("Status", run.get("status","").upper()),
        ("Total Karyawan", run.get("total_employees", len(payslips))),
        ("Total Bruto", run.get("total_gross_pay", 0)),
        ("Total Potongan", run.get("total_deductions", 0)),
        ("Total Neto", run.get("total_net_pay", 0)),
    ]
    for i, (lbl, val) in enumerate(summary_items):
        col = (i % 3) * 2 + 1
        row = 5 + (i // 3)
        _cell(ws1, row, col, lbl, font=FG_DARK_B, fill=BG_SUBHDR, align=LEFT,
              border=BORDER_THIN)
        ws1.cell(row=row, column=col).font = Font(name="Calibri", color="FFFFFF", bold=True, size=9)
        c = _cell(ws1, row, col+1, val, font=FG_DARK_B, align=RIGHT, border=BORDER_THIN)
        if isinstance(val, (int, float)) and val > 1000:
            c.number_format = IDR_FMT
    ws1.row_dimensions[5].height = 18
    ws1.row_dimensions[6].height = 18

    # Empty row
    ws1.row_dimensions[7].height = 6

    # Table headers — row 8
    RECAP_COLS = [
        ("No", 4), ("NIK", 10), ("Nama Karyawan", 24), ("Departemen", 18),
        ("Skema", 10), ("Gaji Pokok", 14), ("Tunjangan", 13), ("Lembur", 11),
        ("Total Bruto", 14), ("Terlambat", 11), ("Kasbon", 11), ("BPJS Kes", 11),
        ("BPJS TK JHT", 12), ("BPJS TK JP", 11), ("PPh 21", 11),
        ("Total Potongan", 15), ("Gaji Bersih", 14),
    ]
    for ci, (hdr, w) in enumerate(RECAP_COLS, 1):
        _cell(ws1, 8, ci, hdr, font=FG_WHITE, fill=BG_HEADER, align=CENTER, border=BORDER_THIN)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[8].height = 36

    # FASE 15 — ambil angka dari STRUKTUR YANG MEMANG DITULIS mesin hitung
    # (`deductions[]` / `allowances[]`), bukan field datar yang tidak pernah ada.
    def _ded(slip: dict, *types: str) -> float:
        total = 0.0
        for d in (slip.get("deductions") or []):
            if d.get("type") in types:
                total += float(d.get("amount") or 0)
        return total

    # Data rows
    NUM_COLS = {6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17}  # IDR columns
    total_row = {i: 0 for i in NUM_COLS}

    for idx, slip in enumerate(payslips):
        r = 9 + idx
        fill = BG_ALTROW if idx % 2 == 1 else None
        bsalary    = slip.get("earnings_total", 0) or 0
        allowances = slip.get("allowance_total", 0) or 0
        ot_amt     = slip.get("overtime_amount", 0) or 0
        gross      = slip.get("gross_pay", 0) or 0
        late_ded   = _ded(slip, "late")
        kasbon     = _ded(slip, "kasbon")
        bpjs_kes   = _ded(slip, "bpjs_kesehatan")
        bpjs_jht   = _ded(slip, "bpjs_jht")
        bpjs_jp    = _ded(slip, "bpjs_jp")
        pph21      = _ded(slip, "pph21")
        ded_total  = slip.get("deductions_total", 0) or 0
        net        = slip.get("net_pay", 0) or 0

        vals = [
            idx+1,
            slip.get("employee_code",""),
            slip.get("employee_name",""),
            slip.get("department",""),
            slip.get("pay_scheme","monthly"),
            bsalary, allowances, ot_amt, gross,
            late_ded, kasbon, bpjs_kes, bpjs_jht, bpjs_jp, pph21, ded_total, net,
        ]
        for ci, val in enumerate(vals, 1):
            c = _cell(ws1, r, ci, val, font=FG_DARK, fill=fill, border=BORDER_THIN,
                      align=CENTER if ci <= 2 else (LEFT if ci <= 5 else RIGHT))
            if ci in NUM_COLS:
                c.number_format = IDR_FMT
                total_row[ci] = total_row.get(ci, 0) + (val or 0)

        # Net gaji hijau
        net_cell = ws1.cell(row=r, column=17)
        net_cell.fill = BG_GREEN
        net_cell.font = Font(name="Calibri", bold=True, size=9, color="155724")
        ws1.row_dimensions[r].height = 16

    # Total row
    tr = 9 + len(payslips)
    ws1.merge_cells(f"A{tr}:E{tr}")
    _cell(ws1, tr, 1, "TOTAL", font=FG_DARK_B, fill=BG_TOTAL, align=CENTER, border=BORDER_MED)
    for ci in NUM_COLS:
        c = _cell(ws1, tr, ci, total_row.get(ci, 0), font=FG_DARK_B, fill=BG_TOTAL,
                  align=RIGHT, border=BORDER_MED, num_fmt=IDR_FMT)
    ws1.row_dimensions[tr].height = 20

    # Signature block
    sig_row = tr + 3
    ws1.merge_cells(f"A{sig_row}:D{sig_row}")
    _cell(ws1, sig_row, 1, "Disetujui Oleh:", font=FG_DARK_B, align=LEFT)
    ws1.merge_cells(f"N{sig_row}:Q{sig_row}")
    _cell(ws1, sig_row, 14, "Mengetahui:", font=FG_DARK_B, align=LEFT)
    ws1.merge_cells(f"A{sig_row+4}:D{sig_row+4}")
    _cell(ws1, sig_row+4, 1, "( Manager Finance )", font=FG_DARK, align=CENTER)
    ws1.merge_cells(f"N{sig_row+4}:Q{sig_row+4}")
    _cell(ws1, sig_row+4, 14, "( Direktur )", font=FG_DARK, align=CENTER)
    ws1.merge_cells(f"A{sig_row+5}:D{sig_row+5}")
    _cell(ws1, sig_row+5, 1, "CV. Dewi Aditya", font=FG_DARK, align=CENTER)
    ws1.merge_cells(f"N{sig_row+5}:Q{sig_row+5}")
    _cell(ws1, sig_row+5, 14, "CV. Dewi Aditya", font=FG_DARK, align=CENTER)

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 2: Slip Gaji Individual (satu baris per karyawan, format slip)
    # ─────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Slip Individual")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 3
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 20
    ws2.column_dimensions["D"].width = 3
    ws2.column_dimensions["E"].width = 26
    ws2.column_dimensions["F"].width = 20

    current_row = 1
    for slip in payslips:
        # FASE 15 — BACA STRUKTUR YANG BENAR.
        # Dulu blok ini membaca field DATAR (`transport_allowance`, `meal_allowance`,
        # `bpjs_kes_employee`, `pph21_amount`, `kasbon_deduction`, …) yang TIDAK
        # PERNAH ditulis mesin hitung — mesin hitung menyimpannya sebagai DAFTAR
        # (`earnings[]`, `allowances[]`, `deductions[]`). Terbukti 2026-07-26:
        # tunjangan transport Rp 600.000 ADA di payslip tapi sel Excel-nya 0,
        # sehingga kolom Penghasilan tidak menjumlah ke Total Bruto (slip yang
        # dibagikan ke karyawan terlihat salah). Slip PDF sudah benar; sekarang
        # kedua kanal cetak membaca SUMBER YANG SAMA.
        ot_hrs    = slip.get("overtime_hours", 0) or 0
        ot_amt    = slip.get("overtime_amount", slip.get("overtime_pay", 0)) or 0
        gross     = slip.get("gross_pay", slip.get("gross_salary", 0))
        ded_total = slip.get("deductions_total", 0)
        net       = slip.get("net_pay", slip.get("net_salary", 0))

        # Sisi PENGHASILAN: setiap baris earnings + setiap tunjangan + lembur.
        left_items = [
            (e.get("label") or e.get("process_code") or "Penghasilan", e.get("amount") or 0)
            for e in (slip.get("earnings") or [])
        ]
        for a in (slip.get("allowances") or []):
            nm = a.get("name") or a.get("label") or "Tunjangan"
            if a.get("calc_type") == "per_day_attendance":
                nm = f"{nm} ({a.get('qty', 0)} hr × {int(a.get('rate') or 0):,})".replace(",", ".")
            left_items.append((nm, a.get("amount") or 0))
        if ot_amt:
            left_items.append((f"Lembur ({float(ot_hrs):.1f} jam)", ot_amt))
        if not left_items:
            left_items = [("Gaji Pokok", slip.get("earnings_total", 0) or 0)]

        # Sisi POTONGAN: apa adanya dari mesin hitung (BPJS, PPh21, kasbon,
        # LWOP, keterlambatan) — tidak lagi menebak nama field.
        right_items = [
            (d.get("label") or d.get("type") or "Potongan", d.get("amount") or 0)
            for d in (slip.get("deductions") or [])
        ]

        r = current_row
        # Header slip
        ws2.merge_cells(f"B{r}:F{r}")
        _cell(ws2, r, 2, "CV. DEWI ADITYA — SLIP GAJI", font=Font(name="Calibri", bold=True, size=11, color="1A3A5C"), align=CENTER)
        ws2.row_dimensions[r].height = 20
        r += 1

        ws2.merge_cells(f"B{r}:F{r}")
        _cell(ws2, r, 2, f"Periode: {period_lbl}  |  {run_num}", font=FG_SUBTITLE, align=CENTER)
        ws2.row_dimensions[r].height = 14
        r += 1

        # Identitas karyawan
        info = [
            ("NIK", slip.get("employee_code","")),
            ("Nama", slip.get("employee_name","")),
            ("Departemen", slip.get("department","")),
            ("Skema", slip.get("pay_scheme","monthly")),
            ("Hari Hadir", f"{slip.get('days_hadir', slip.get('days_present', 0))} hari"),
        ]
        for lbl, val in info:
            _cell(ws2, r, 2, lbl, font=FG_DARK_B, align=LEFT)
            _cell(ws2, r, 3, val, font=FG_DARK, align=LEFT)
            ws2.row_dimensions[r].height = 14
            r += 1
        r += 1  # spacer

        # Penghasilan
        _cell(ws2, r, 2, "PENGHASILAN", font=FG_WHITE, fill=BG_HEADER, align=LEFT, border=BORDER_THIN)
        _cell(ws2, r, 3, "", fill=BG_HEADER, border=BORDER_THIN)
        _cell(ws2, r, 5, "POTONGAN", font=FG_WHITE, fill=BG_HEADER, align=LEFT, border=BORDER_THIN)
        _cell(ws2, r, 6, "", fill=BG_HEADER, border=BORDER_THIN)
        ws2.row_dimensions[r].height = 16
        r += 1

        max_rows = max(len(left_items), len(right_items))
        for i in range(max_rows):
            lbl_l, val_l = left_items[i] if i < len(left_items) else ("", "")
            lbl_r, val_r = right_items[i] if i < len(right_items) else ("", "")
            _cell(ws2, r, 2, lbl_l, font=FG_DARK, align=LEFT, border=BORDER_THIN)
            c = _cell(ws2, r, 3, val_l or "", font=FG_DARK, align=RIGHT, border=BORDER_THIN)
            if isinstance(val_l, (int, float)) and val_l:
                c.number_format = IDR_FMT
            _cell(ws2, r, 5, lbl_r, font=FG_DARK, align=LEFT, border=BORDER_THIN)
            c2 = _cell(ws2, r, 6, val_r or "", font=FG_DARK, align=RIGHT, border=BORDER_THIN)
            if isinstance(val_r, (int, float)) and val_r:
                c2.number_format = IDR_FMT
            ws2.row_dimensions[r].height = 14
            r += 1

        # Total baris
        _cell(ws2, r, 2, "Total Bruto", font=FG_DARK_B, fill=BG_TOTAL, align=LEFT, border=BORDER_THIN)
        c = _cell(ws2, r, 3, gross, font=FG_DARK_B, fill=BG_TOTAL, align=RIGHT, border=BORDER_THIN, num_fmt=IDR_FMT)
        _cell(ws2, r, 5, "Total Potongan", font=FG_DARK_B, fill=BG_TOTAL, align=LEFT, border=BORDER_THIN)
        c2 = _cell(ws2, r, 6, ded_total, font=FG_DARK_B, fill=BG_TOTAL, align=RIGHT, border=BORDER_THIN, num_fmt=IDR_FMT)
        ws2.row_dimensions[r].height = 16
        r += 1

        # GAJI BERSIH
        ws2.merge_cells(f"B{r}:E{r}")
        _cell(ws2, r, 2, "GAJI BERSIH DITERIMA", font=Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
              fill=PatternFill("solid", fgColor="1D6F42"), align=LEFT, border=BORDER_MED)
        c = _cell(ws2, r, 6, net, font=Font(name="Calibri", bold=True, size=11, color="155724"),
                  fill=BG_GREEN, align=RIGHT, border=BORDER_MED, num_fmt=IDR_FMT)
        ws2.row_dimensions[r].height = 22
        r += 3  # gap antar slip

        current_row = r

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 3: Data Bank (untuk proses transfer)
    # ─────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Data Transfer Bank")
    ws3.sheet_view.showGridLines = False
    BANK_COLS = [("NIK", 10), ("Nama", 28), ("Bank", 10), ("No. Rekening", 20),
                 ("Gaji Bersih", 16), ("Keterangan", 22)]
    ws3.merge_cells("A1:F1")
    _cell(ws3, 1, 1, f"DATA TRANSFER GAJI — {run_num} | Periode: {period_lbl}",
          font=Font(name="Calibri", bold=True, size=11, color="1A3A5C"),
          fill=PatternFill("solid", fgColor="D6E4F0"), align=CENTER)
    ws3.row_dimensions[1].height = 22

    for ci, (hdr, w) in enumerate(BANK_COLS, 1):
        _cell(ws3, 2, ci, hdr, font=FG_WHITE, fill=BG_HEADER, align=CENTER, border=BORDER_THIN)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[2].height = 18

    emp_cache: dict = {}
    for idx, slip in enumerate(payslips):
        nik = slip.get("employee_code", "")
        if nik not in emp_cache:
            emp_doc = await db.rahaza_employees.find_one({"employee_code": nik}, {"_id": 0})
            emp_cache[nik] = emp_doc or {}
        emp = emp_cache[nik]
        net = slip.get("net_pay", slip.get("net_salary", 0))
        r = 3 + idx
        fill = BG_ALTROW if idx % 2 == 1 else None
        vals = [
            nik, slip.get("employee_name",""),
            emp.get("bank_name",""), emp.get("bank_account",""),
            net, f"Gaji {period_lbl}",
        ]
        for ci, val in enumerate(vals, 1):
            c = _cell(ws3, r, ci, val, font=FG_DARK, fill=fill, border=BORDER_THIN,
                      align=CENTER if ci == 1 else RIGHT if ci == 5 else LEFT)
            if ci == 5:
                c.number_format = IDR_FMT
        ws3.row_dimensions[r].height = 15

    # Total
    tr3 = 3 + len(payslips)
    ws3.merge_cells(f"A{tr3}:D{tr3}")
    _cell(ws3, tr3, 1, "TOTAL TRANSFER", font=FG_DARK_B, fill=BG_TOTAL, align=CENTER, border=BORDER_MED)
    total_net = sum(s.get("net_pay", s.get("net_salary", 0)) for s in payslips)
    _cell(ws3, tr3, 5, total_net, font=FG_DARK_B, fill=BG_TOTAL, align=RIGHT, border=BORDER_MED, num_fmt=IDR_FMT)
    ws3.row_dimensions[tr3].height = 20

    # ── Stream ke response ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_num = run_num.replace("/", "-")
    filename = f"payroll_{safe_num}_{period_from}_{period_to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── EXPORT CSV ──────────────────────────────────────────────────────────────
# FASE 20 — BUG NYATA YANG LOLOS SEMUA GATE:
# Badan fungsi di bawah ini dulu MENYATU di dalam `export_run_excel()`, tepat
# SETELAH `return StreamingResponse(...)`-nya. Artinya 31 baris ini adalah KODE
# MATI dan dekorator `@router.get(".../export")`-nya HILANG, sehingga
# `GET /api/rahaza/payroll-runs/{id}/export` TIDAK PERNAH terdaftar di FastAPI.
# Akibat: tombol "Download CSV" di `RahazaPayrollRunModule` selalu 404 senyap.
#
# Kenapa tak terdeteksi:
#   - CHECK D (orphan handler) butuh sebuah `def` tanpa dekorator → di sini
#     tidak ada `def` baru sama sekali, jadi tak ada yang bisa dilihat.
#   - CHECK B hanya bisa bilang "FE memanggil path yang tak ada", tanpa tahu
#     implementasinya SUDAH ADA namun tak terjangkau.
# Sentinel baru: scripts/probe_unreachable_after_return.py (terdaftar di gate.sh)
@router.get("/payroll-runs/{run_id}/export")
async def export_run_csv(run_id: str, request: Request):
    """Export payroll run ke CSV (1 baris per karyawan)."""
    await require_auth(request)
    db = get_db()
    run = await db.rahaza_payroll_runs.find_one({"id": run_id}, {"_id": 0})
    if not run:
        raise HTTPException(404, "Payroll run tidak ditemukan.")
    payslips = await db.rahaza_payslips.find({"run_id": run_id}, {"_id": 0}).sort("employee_code", 1).to_list(500)

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "run_number", "period_from", "period_to",
        "employee_code", "employee_name", "pay_scheme",
        "earnings_total", "overtime_hours", "overtime_amount",
        "gross_pay", "deductions_total", "manual_deduction", "net_pay",
        "days_hadir", "total_hours_worked", "adjustment_notes",
    ])
    for s in payslips:
        w.writerow([
            run.get("run_number"), run.get("period_from"), run.get("period_to"),
            s.get("employee_code"), s.get("employee_name"), s.get("pay_scheme"),
            s.get("earnings_total", 0), s.get("overtime_hours", 0), s.get("overtime_amount", 0),
            s.get("gross_pay", 0), s.get("deductions_total", 0),
            s.get("manual_deduction", 0), s.get("net_pay", 0),
            s.get("days_hadir", 0), s.get("total_hours_worked", 0),
            s.get("adjustment_notes", ""),
        ])
    buf.seek(0)
    # Nomor run bisa memuat "/" (mis. PR/2026/07/001) → harus disanitasi, kalau
    # tidak header Content-Disposition rusak & browser menyimpan nama aneh.
    safe_num = str(run.get("run_number") or run_id).replace("/", "-")
    filename = f"payroll_{safe_num}.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


