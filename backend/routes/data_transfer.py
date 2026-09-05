"""
Universal Export / Import — generic, registry-driven data transfer.

- Export any whitelisted collection to CSV / XLSX.
- Download a blank import TEMPLATE (headers only).
- Import (upsert-by-key) with a DRY-RUN preview step (valid/invalid rows + errors)
  then a COMMIT step.

Design:
- REGISTRY is the single source of truth for what is exportable/importable, the
  column schema (for validation/coercion), the natural key (for upsert), and RBAC.
- Master tables => import + export.  Transaction tables => export only.
- Nested/complex columns are carried as JSON strings in a single cell.

Endpoints (prefix /api/data-transfer):
  GET  /registry                      -> tables the current user may use
  GET  /export/{key}?format=csv|xlsx  -> file download
  GET  /template/{key}?format=csv|xlsx-> blank template (headers [+ 1 example])
  POST /import/{key}?mode=dry_run|commit (multipart file) -> preview / apply
"""
# ruff: noqa: E402
import io
import csv
import json
import logging
import uuid
from datetime import datetime, timezone, date

from fastapi import APIRouter, Request, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse

from database import get_db
from auth import require_auth, log_activity, hash_password

import openpyxl
from openpyxl.utils import get_column_letter
from utils.waktu import now_wib

router = APIRouter(prefix="/api/data-transfer", tags=["data-transfer"])

ADMIN_ROLES = {"superadmin", "admin", "owner"}


def _uid():
    return str(uuid.uuid4())


def _now():
    return datetime.now(timezone.utc)


# ── column helper ────────────────────────────────────────────────────────────
def col(name, type="str", required=False, label=None, example=""):
    return {"name": name, "type": type, "required": required, "label": label or name, "example": example}


# ── REGISTRY ─────────────────────────────────────────────────────────────────
# type: str | num | bool | date | json
REGISTRY = {
    # ===== GROUP A — MASTER (export + import) =====
    "materials": {
        "collection": "rahaza_materials", "label": "Bahan / Material", "group": "Master Produksi",
        "key": ["code"], "importable": True, "export_roles": ["spv", "gudang", "manager", "produksi"],
        "columns": [
            col("code", "str", True, example="MAT-001"), col("name", "str", True, example="Kain Katun"),
            col("type", "str", example="yarn"), col("unit", "str", example="kg"),
            col("composition", "str"), col("color", "str"), col("min_stock", "num", example="0"),
            col("unit_cost", "num", example="0"), col("active", "bool", example="true"),
            # ── Kemasan / multi-satuan (INV-UOM-4) ──
            col("base_uom", "str", label="satuan dasar", example="pcs"),
            col("pack_unit", "str", label="satuan kemasan", example="pak"),
            col("pack_size", "num", label="isi per kemasan", example="12"),
            col("display_in_packs", "bool", label="tampilkan per kemasan", example="true"),
        ],
    },
    "models": {
        "collection": "rahaza_models", "label": "Model / Artikel", "group": "Master Produksi",
        "key": ["code"], "importable": True, "export_roles": ["spv", "produksi", "manager"],
        "columns": [col("code", "str", True, example="MDL-001"), col("name", "str", True), col("description", "str"), col("active", "bool", example="true")],
    },
    "sizes": {
        "collection": "rahaza_sizes", "label": "Ukuran", "group": "Master Produksi",
        "key": ["code"], "importable": True, "export_roles": ["spv", "produksi"],
        "columns": [col("code", "str", True, example="S"), col("name", "str", True), col("order_seq", "num", example="1"), col("active", "bool", example="true")],
    },
    "processes": {
        "collection": "rahaza_processes", "label": "Proses Produksi", "group": "Master Produksi",
        "key": ["code"], "importable": True, "export_roles": ["spv", "produksi"],
        "columns": [col("code", "str", True), col("name", "str", True), col("order_seq", "num", example="1"), col("is_rework", "bool", example="false"), col("description", "str"), col("active", "bool", example="true")],
    },
    "lines": {
        "collection": "rahaza_lines", "label": "Lini Produksi", "group": "Master Produksi",
        "key": ["code"], "importable": True, "export_roles": ["spv", "produksi"],
        "columns": [col("code", "str", True), col("name", "str", True), col("location_id", "str"), col("process_id", "str"), col("capacity_per_hour", "num", example="0"), col("active", "bool", example="true")],
    },
    "locations": {
        "collection": "rahaza_locations", "label": "Lokasi / Gudang", "group": "Master Gudang",
        "key": ["code"], "importable": True, "export_roles": ["gudang", "spv"],
        "columns": [col("code", "str", True, example="GDG-A1"), col("name", "str", True), col("type", "str", example="rack"), col("parent_id", "str"), col("active", "bool", example="true")],
    },
    "shifts": {
        "collection": "rahaza_shifts", "label": "Shift", "group": "Master Produksi",
        "key": ["code"], "importable": True, "export_roles": ["spv", "hr"],
        "columns": [col("code", "str", True), col("name", "str", True), col("start_time", "str", example="08:00"), col("end_time", "str", example="16:00"), col("active", "bool", example="true")],
    },
    "coa_accounts": {
        "collection": "rahaza_coa_accounts", "label": "Chart of Accounts (COA)", "group": "Finance",
        "key": ["code"], "importable": True, "export_roles": ["finance"],
        "columns": [
            col("code", "str", True, example="1-101"), col("name", "str", True, example="Kas"),
            col("type", "str", True, example="ASSET"), col("parent_code", "str", example="1-100"),
            col("is_group", "bool", example="false"), col("normal_balance", "str", example="DEBIT"),
            col("active", "bool", example="true"), col("flags", "json", example="{}"),
        ],
    },
    "employees": {
        "collection": "rahaza_employees", "label": "Karyawan", "group": "HR / SDM",
        "key": ["employee_code"], "importable": True, "export_roles": ["hr", "spv"],
        "columns": [col("employee_code", "str", True, example="DA-001"), col("name", "str", True), col("role_hint", "str"), col("phone", "str"), col("join_date", "date", example="2025-01-01"), col("active", "bool", example="true")],
    },
    "salary_grades": {
        "collection": "rahaza_salary_grades", "label": "Grade Gaji", "group": "HR / SDM",
        "key": ["grade_code"], "importable": True, "export_roles": ["hr"],
        "columns": [col("grade_code", "str", True), col("grade_name", "str", True), col("level", "num", example="1"), col("min_salary", "num"), col("mid_salary", "num"), col("max_salary", "num"), col("currency", "str", example="IDR"), col("is_active", "bool", example="true")],
    },
    "leave_types": {
        "collection": "rahaza_leave_types", "label": "Jenis Cuti", "group": "HR / SDM",
        "key": ["code"], "importable": True, "export_roles": ["hr"],
        "columns": [col("code", "str", True), col("name", "str", True), col("quota_default", "num", example="12"), col("unpaid", "bool", example="false"), col("request_type", "str", example="leave"), col("requires_document", "bool", example="false"), col("active", "bool", example="true")],
    },
    "cmt_partners": {
        "collection": "dewi_cmt_partners", "label": "Vendor CMT", "group": "Partner",
        "key": ["code"], "importable": True, "export_roles": ["maklon", "spv", "finance"],
        "columns": [
            col("code", "str", True, example="CMT-001"), col("name", "str", True), col("owner_name", "str"),
            col("phone", "str"), col("address", "str"), col("city", "str"), col("specialization", "str"),
            col("rate_per_pcs", "num"), col("capacity_per_week", "num"), col("bank_name", "str"),
            col("bank_account", "str"), col("bank_holder", "str"), col("rating", "num"),
            col("status", "str", example="active"), col("penalty_per_day", "num"), col("notes", "str"),
        ],
    },
    "vendor_partners": {
        "collection": "vendor_partners", "label": "Vendor CMT (Partner)", "group": "Partner",
        "key": ["code"], "importable": True, "export_roles": ["maklon", "spv", "produksi", "manager", "finance"],
        "columns": [
            col("code", "str", True, example="CMT-001"), col("name", "str", True, example="CMT Berkah Jaya"),
            col("contact_name", "str"), col("contact_phone", "str"), col("address", "str"),
            col("notes", "str"), col("is_active", "bool", example="true"),
        ],
    },
    "maklon_clients": {
        "collection": "dewi_maklon_clients", "label": "Klien Maklon", "group": "Partner",
        "key": ["code"], "importable": True, "export_roles": ["maklon", "marketing"],
        "columns": [col("code", "str", True, example="CL-001"), col("name", "str", True), col("contact_name", "str"), col("contact_phone", "str"), col("address", "str"), col("notes", "str"), col("active", "bool", example="true")],
    },
    "platform_accounts": {
        "collection": "marketing_platform_accounts", "label": "Akun Toko / Channel", "group": "Partner",
        "key": ["account_code"], "importable": True, "export_roles": ["marketing"],
        "columns": [col("account_code", "str", True, example="SHP-01"), col("account_name", "str", True), col("platform", "str", example="shopee"), col("username", "str"), col("status", "str", example="active"), col("group", "str")],
    },
    "users": {
        "collection": "users", "label": "Pengguna Sistem", "group": "Admin",
        "key": ["email"], "importable": True, "export_roles": [],  # admin only
        "import_roles": list(ADMIN_ROLES),
        "exclude_export": ["password"],
        "import_defaults": {"password": "Dewi@123", "status": "active"},
        "columns": [col("name", "str", True), col("email", "str", True, example="user@dewiaditya.id"), col("role", "str", example="staff"), col("status", "str", example="active")],
    },

    # ===== GROUP B — MASTER advanced (nested → JSON) =====
    "posting_profiles": {
        "collection": "rahaza_posting_profiles", "label": "Profil Posting GL", "group": "Finance",
        "key": ["event_type"], "importable": True, "export_roles": ["finance"],
        "columns": [col("event_type", "str", True, example="cmt_ap_invoice"), col("description", "str"), col("mapping", "json", True, example='{"debit":"...","credit":"..."}'), col("active", "bool", example="true")],
    },
    "payroll_profiles": {
        "collection": "rahaza_payroll_profiles", "label": "Profil Payroll", "group": "HR / SDM",
        "key": ["employee_id"], "importable": True, "export_roles": ["hr"],
        "columns": [col("employee_id", "str", True), col("pay_scheme", "str", example="monthly"), col("period_type", "str"), col("base_rate", "num"), col("overtime_rate", "num"), col("pcs_process_rates", "json", example="{}"), col("notes", "str"), col("active", "bool", example="true")],
    },
    "boms": {
        "collection": "rahaza_boms", "label": "BOM (Bill of Materials)", "group": "Master Produksi",
        "key": ["model_id", "size_id"], "importable": True, "export_roles": ["spv", "produksi"],
        "columns": [col("model_id", "str", True), col("size_id", "str", True), col("color", "str"), col("materials", "json", example="[]"), col("notes", "str"), col("active", "bool", example="true")],
    },

    # ===== GROUP C — TRANSACTIONS (export only; columns=None => all fields) =====
    "maklon_pos": {"collection": "dewi_maklon_pos", "label": "PO Maklon", "group": "Transaksi", "importable": False, "export_roles": ["maklon", "spv"], "columns": None},
    "production_pos": {"collection": "production_pos", "label": "PO Produksi", "group": "Transaksi", "importable": False, "export_roles": ["spv", "produksi"], "columns": None},
    "cutting_requests": {"collection": "dewi_cutting_requests", "label": "Permintaan Cutting", "group": "Transaksi", "importable": False, "export_roles": ["spv", "produksi"], "columns": None},
    "cmt_jobs": {"collection": "dewi_cmt_jobs", "label": "Job CMT", "group": "Transaksi", "importable": False, "export_roles": ["maklon", "spv"], "columns": None},
    "ar_invoices": {"collection": "rahaza_ar_invoices", "label": "Invoice AR", "group": "Transaksi", "importable": False, "export_roles": ["finance"], "columns": None},
    "maklon_invoices": {"collection": "dewi_maklon_invoices", "label": "Invoice Maklon", "group": "Transaksi", "importable": False, "export_roles": ["finance", "maklon"], "columns": None},
    "material_issues": {"collection": "rahaza_material_issues", "label": "Pengeluaran Material", "group": "Transaksi", "importable": False, "export_roles": ["gudang", "spv"], "columns": None},
    "material_stock": {"collection": "rahaza_material_stock", "label": "Stok Material", "group": "Transaksi", "importable": False, "export_roles": ["gudang", "spv"], "columns": None},
    "attendance_events": {"collection": "rahaza_attendance_events", "label": "Absensi", "group": "Transaksi", "importable": False, "export_roles": ["hr", "spv"], "columns": None},
    "activity_logs": {"collection": "activity_logs", "label": "Log Aktivitas", "group": "Transaksi", "importable": False, "export_roles": [], "columns": None},
    "buyer_shipments": {"collection": "buyer_shipments", "label": "Pengiriman Buyer", "group": "Transaksi", "importable": False, "export_roles": ["spv", "produksi"], "columns": None},
}


# ── RBAC helpers ─────────────────────────────────────────────────────────────
def _role(user):
    return (user.get("role") or "").lower()


def _can_export(user, entry):
    r = _role(user)
    if r in ADMIN_ROLES:
        return True
    roles = entry.get("export_roles") or []
    return "*" in roles or r in roles


def _can_import(user, entry):
    if not entry.get("importable"):
        return False
    r = _role(user)
    allowed = set(entry.get("import_roles") or ADMIN_ROLES)
    return r in allowed


def _get_entry(key):
    entry = REGISTRY.get(key)
    if not entry:
        raise HTTPException(404, f"Tabel '{key}' tidak terdaftar untuk export/import")
    return entry


# ── value (de)serialization ──────────────────────────────────────────────────
def _to_cell(value, ctype=None):
    """Convert a stored value into a flat cell string for export."""
    if value is None:
        return ""
    if ctype == "json" or isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


TRUE_SET = {"true", "1", "yes", "ya", "y", "t", "aktif", "active"}
FALSE_SET = {"false", "0", "no", "tidak", "n", "f", "nonaktif", "inactive", ""}


def _coerce(raw, ctype, colname):
    """Coerce a raw cell (string) into the target type. Returns (value, error)."""
    if raw is None:
        return None, None
    s = str(raw).strip()
    if s == "":
        return None, None
    try:
        if ctype == "num":
            s2 = s.replace(",", "") if s.count(",") and "." in s else s.replace(",", ".")
            f = float(s2)
            return (int(f) if f.is_integer() else f), None
        if ctype == "bool":
            low = s.lower()
            if low in TRUE_SET:
                return True, None
            if low in FALSE_SET:
                return False, None
            return None, f"'{colname}': nilai boolean tidak valid ('{s}')"
        if ctype == "json":
            return json.loads(s), None
        if ctype == "date":
            # keep as ISO string; validate it parses
            datetime.fromisoformat(s.replace("Z", "+00:00"))
            return s, None
        return s, None  # str
    except Exception as e:
        return None, f"'{colname}': gagal parse ({ctype}) — {e}"


# ── column resolution (dynamic for columns=None) ─────────────────────────────
async def _resolve_columns(db, entry):
    if entry["columns"]:
        return entry["columns"], [c["name"] for c in entry["columns"]]
    # dynamic: union of keys from a sample of documents
    cursor = db[entry["collection"]].find({}, limit=200)
    keys = []
    seen = set()
    async for d in cursor:
        for k in d.keys():
            if k in ("_id",) or k in seen:
                continue
            seen.add(k)
            keys.append(k)
    # stable, friendly ordering
    priority = ["id", "code", "po_number", "invoice_number", "job_code", "request_code", "name"]
    keys.sort(key=lambda k: (priority.index(k) if k in priority else 999, k))
    cols = [{"name": k, "type": "auto", "required": False, "label": k, "example": ""} for k in keys]
    return cols, keys


# ── ENDPOINTS ────────────────────────────────────────────────────────────────
@router.get("/registry")
async def get_registry(request: Request):
    user = await require_auth(request)
    out = []
    for key, entry in REGISTRY.items():
        if not _can_export(user, entry):
            continue
        out.append({
            "key": key,
            "label": entry["label"],
            "group": entry["group"],
            "importable": entry.get("importable", False) and _can_import(user, entry),
            "exportable": True,
            "key_fields": entry.get("key", []),
        })
    out.sort(key=lambda x: (x["group"], x["label"]))
    return {"tables": out, "count": len(out)}


@router.get("/export/{key}")
async def export_table(key: str, request: Request, format: str = Query("xlsx")):
    user = await require_auth(request)
    entry = _get_entry(key)
    if not _can_export(user, entry):
        raise HTTPException(403, "Anda tidak punya akses export tabel ini")
    db = get_db()
    cols, colnames = await _resolve_columns(db, entry)
    exclude = set(entry.get("exclude_export") or [])
    colnames = [c for c in colnames if c not in exclude]
    cols = [c for c in cols if c["name"] not in exclude]

    rows = []
    async for doc in db[entry["collection"]].find({}):
        rows.append(doc)

    ctype_map = {c["name"]: c["type"] for c in cols}
    fname_base = f"{key}_{now_wib().strftime('%Y%m%d_%H%M')}"

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(colnames)
        for doc in rows:
            w.writerow([_to_cell(doc.get(cn), ctype_map.get(cn)) for cn in colnames])
        data = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(io.BytesIO(data), media_type="text/csv",
                                 headers={"Content-Disposition": f'attachment; filename="{fname_base}.csv"'})
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = key[:31]
        ws.append(colnames)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for doc in rows:
            ws.append([_to_cell(doc.get(cn), ctype_map.get(cn)) for cn in colnames])
        for i, cn in enumerate(colnames, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(cn) + 2, 12), 40)
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f'attachment; filename="{fname_base}.xlsx"'})


@router.get("/template/{key}")
async def template_table(key: str, request: Request, format: str = Query("xlsx")):
    user = await require_auth(request)
    entry = _get_entry(key)
    if not entry.get("importable"):
        raise HTTPException(400, "Tabel ini tidak mendukung import (export-only)")
    if not _can_import(user, entry):
        raise HTTPException(403, "Anda tidak punya akses import tabel ini")
    cols = entry["columns"]
    headers = [c["name"] for c in cols]
    examples = [c.get("example", "") for c in cols]

    if format == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        w.writerow(examples)
        data = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(io.BytesIO(data), media_type="text/csv",
                                 headers={"Content-Disposition": f'attachment; filename="template_{key}.csv"'})
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = key[:31]
    ws.append(headers)
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=j)
        cell.font = openpyxl.styles.Font(bold=True)
        req = " *" if c.get("required") else ""
        ws.column_dimensions[get_column_letter(j)].width = min(max(len(c["name"]) + len(req) + 2, 12), 40)
    ws.append(examples)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="template_{key}.xlsx"'})


async def _material_uom_fields(db, doc: dict, existing: dict | None):
    """Susun field UOM (`uoms` + cermin lama) untuk satu baris impor material.

    Kembalian: (fields_to_set, errors). Kolom kemasan yang KOSONG = tidak
    menyentuh konfigurasi satuan yang sudah ada (perilaku impor lama).
    Mengganti SATUAN DASAR item yang masih berstok DITOLAK — harus lewat
    endpoint rebase resmi supaya angka stok ikut dikonversi (INV-UOM-3).
    """
    from core import uom as _uom
    from core import stock_service

    given = {k: doc[k] for k in ("unit", "base_uom", "pack_unit", "pack_size", "display_in_packs")
             if k in doc and doc[k] not in (None, "")}
    if not given:
        return {}, []

    cur = existing or {}
    cur_base = _uom.base_uom_of(cur) if existing else ""
    base = _uom.normalize_code(given.get("base_uom") or given.get("unit") or cur_base or _uom.DEFAULT_BASE)

    if existing and cur_base and base != cur_base:
        onhand = await stock_service.get_onhand(existing.get("id"), db=db)
        if abs(float(onhand or 0)) > 1e-9:
            return {}, [f"satuan dasar '{cur_base}' → '{base}' ditolak: stok masih {onhand:g} {cur_base}. "
                        f"Pakai tombol 'Ubah Satuan Dasar' di form material (rebase)."]

    pack_unit = _uom.normalize_code(given.get("pack_unit"))
    pack_size = float(given.get("pack_size") or 0)
    if pack_unit == base:
        pack_unit, pack_size = "", 0.0
    if pack_unit and pack_size <= 1:
        return {}, [f"'pack_size' harus lebih besar dari 1 bila 'pack_unit' diisi ('{pack_unit}')"]
    if pack_size > 1 and not pack_unit:
        return {}, ["'pack_unit' wajib diisi bila 'pack_size' diisi"]

    rows = [{"code": base, "factor": 1, "is_base": True}]
    if existing and base == cur_base:
        # pertahankan tingkat kemasan lain (mis. karton) yang tidak disebut di file
        for r in _uom.resolve_uoms(cur):
            code = _uom.normalize_code(r.get("code"))
            if code in (base, pack_unit) or float(r.get("factor") or 0) <= 1:
                continue
            rows.append(dict(r))
    if pack_unit:
        rows.append({"code": pack_unit, "factor": pack_size, "parent": base,
                     "is_purchase_default": True,
                     "is_display_default": bool(given.get("display_in_packs"))})

    body = {"base_uom": base, "uoms": rows}
    if pack_unit:
        # dibeli per kemasan, dipakai per satuan dasar — inti kasus owner
        body["purchase_uom"] = pack_unit
        body["issue_uom"] = base
        body["display_uom"] = pack_unit if given.get("display_in_packs") else base
    if "display_in_packs" in given:
        body["display_in_packs"] = bool(given["display_in_packs"])
    try:
        return _uom.apply_payload(body, cur), []
    except _uom.UomError as e:
        return {}, [str(e)]


def _parse_upload(filename, content):
    """Return list[dict] rows from CSV or XLSX bytes."""
    name = (filename or "").lower()
    if name.endswith(".csv") or (b"," in content[:200] and not content[:4] == b"PK\x03\x04"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(r) for r in reader]
    # xlsx
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return out


@router.post("/import/{key}")
async def import_table(key: str, request: Request, file: UploadFile = File(...), mode: str = Query("dry_run")):
    user = await require_auth(request)
    entry = _get_entry(key)
    if not entry.get("importable"):
        raise HTTPException(400, "Tabel ini tidak mendukung import (export-only)")
    if not _can_import(user, entry):
        raise HTTPException(403, "Import hanya untuk admin")
    db = get_db()
    content = await file.read()
    try:
        raw_rows = _parse_upload(file.filename, content)
    except Exception as e:
        raise HTTPException(400, f"Gagal membaca file: {e}")

    cols = entry["columns"]
    keyf = entry["key"]
    defaults = entry.get("import_defaults") or {}
    colnames = {c["name"]: c for c in cols}

    valid, invalid = [], []
    for idx, raw in enumerate(raw_rows):
        errors = []
        doc = {}
        for c in cols:
            cn = c["name"]
            v, err = _coerce(raw.get(cn, raw.get(c.get("label"))), c["type"], cn)
            if err:
                errors.append(err)
            if v is not None:
                doc[cn] = v
        # per-collection normalization (before key/validation so upsert is consistent)
        if entry["collection"] == "users" and doc.get("email"):
            doc["email"] = str(doc["email"]).strip().lower()
        # required checks
        for c in cols:
            if c.get("required") and (doc.get(c["name"]) in (None, "")):
                errors.append(f"'{c['name']}' wajib diisi")
        # key present
        for kf in keyf:
            if doc.get(kf) in (None, ""):
                errors.append(f"key '{kf}' kosong")
        row_info = {"row": idx + 2, "data": {k: _to_cell(v, colnames.get(k, {}).get("type")) for k, v in doc.items()}}
        if errors:
            row_info["errors"] = errors
            invalid.append(row_info)
        else:
            valid.append((doc, row_info))

    # detect insert vs update for preview
    would_insert = would_update = 0
    still_valid = []
    for doc, info in valid:
        q = {kf: doc[kf] for kf in keyf}
        existing = await db[entry["collection"]].find_one(q)
        if entry["collection"] == "rahaza_materials":
            uom_fields, uom_errs = await _material_uom_fields(db, doc, existing)
            if uom_errs:
                info["errors"] = uom_errs
                invalid.append(info)
                continue
            if uom_fields:
                doc.update(uom_fields)
                info["data"].update({
                    "base_uom": uom_fields.get("base_uom"),
                    "pack_unit": uom_fields.get("pack_unit"),
                    "pack_size": uom_fields.get("pack_size"),
                    "satuan": " · ".join(f"{r['code']}×{r['factor']:g}" for r in uom_fields.get("uoms", [])),
                })
        if existing:
            info["action"] = "update"
            would_update += 1
        else:
            info["action"] = "insert"
            would_insert += 1
        still_valid.append((doc, info))
    valid = still_valid

    summary = {
        "key": key, "label": entry["label"], "mode": mode,
        "total_rows": len(raw_rows), "valid": len(valid), "invalid": len(invalid),
        "would_insert": would_insert, "would_update": would_update,
        "preview_valid": [i for _, i in valid][:100],
        "preview_invalid": invalid[:100],
    }

    if mode != "commit":
        return summary

    if invalid:
        raise HTTPException(400, f"Ada {len(invalid)} baris invalid. Perbaiki dulu sebelum commit (atau hapus baris invalid).")

    inserted = updated = 0
    for doc, info in valid:
        q = {kf: doc[kf] for kf in keyf}
        existing = await db[entry["collection"]].find_one(q)
        if existing:
            doc["updated_at"] = _now()
            await db[entry["collection"]].update_one({"_id": existing["_id"]}, {"$set": doc})
            updated += 1
        else:
            for dk, dv in defaults.items():
                doc.setdefault(dk, dv)
            # per-collection insert hooks (e.g. hash user passwords so imported logins work)
            if entry["collection"] == "users":
                doc["password"] = hash_password(str(doc.get("password") or "Dewi@123"))
                doc.setdefault("permissions", [])
            doc.setdefault("id", _uid())
            doc.setdefault("active", True)
            doc["created_at"] = _now()
            doc["updated_at"] = _now()
            doc["created_by"] = user.get("id")
            doc["created_by_name"] = user.get("name")
            await db[entry["collection"]].insert_one(doc)
            inserted += 1

        # Phase 5/6: auto-create COA subledger on import (idempotent, non-fatal)
        _COA_AUTO_COLLECTIONS = {
            "dewi_cmt_partners": "cmt_vendor",
            "rahaza_vendors": "supplier",
            "rahaza_customers": "customer",
            "marketing_platform_accounts": "channel",
            "rahaza_cash_accounts": "bank",
        }
        _et = _COA_AUTO_COLLECTIONS.get(entry["collection"])
        if _et:
            try:
                from routes.coa_auto import ensure_subledger_for_entity
                await ensure_subledger_for_entity(db, _et, doc, user)
            except Exception as _e:  # noqa: BLE001
                # 2026-08-07 — DULU `pass` (variabel `_e` bahkan tak dipakai).
                # Impor massal yang gagal membuat subledger COA menghasilkan
                # entitas uang tanpa akun Buku Besar — dan karena impornya
                # ratusan baris, tanpa log tidak mungkin tahu mana yang bolong.
                logging.getLogger(__name__).error(
                    "[coa] impor: subledger COA (%s) GAGAL untuk %s — entitas ini "
                    "tidak punya akun Buku Besar: %s",
                    _et, doc.get("name") or doc.get("id"), _e)

    await log_activity(user.get("id"), user.get("name"), "import",
                       f"data-transfer:{key}", f"import {inserted} baru, {updated} update")
    summary.update({"committed": True, "inserted": inserted, "updated": updated})
    return summary
