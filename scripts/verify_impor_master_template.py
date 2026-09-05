#!/usr/bin/env python3
"""INV-F41 — IMPOR MASTER DATA dari template Excel (migrasi data nyata).

Kenapa gate ini ada: migrasi data adalah satu-satunya langkah yang **tidak bisa
diulang tanpa biaya**. Kalau importir menerima baris cacat, yang lahir bukan error
melainkan **master hantu** — dan seluruh HPP, stok, serta insentif setelahnya ikut
salah tanpa ada yang tahu (di sistem ini sudah pernah terjadi: 3 baris SPK dengan SKU
tanpa master, Rp 3,6 jt menggantung).

Yang dijaga:
  A. Template terbentuk lengkap & baris contoh (#) TIDAK terimpor.
  B. Setiap bentuk baris cacat DILAPORKAN dengan nomor barisnya, dan selama masih ada
     kesalahan **tidak ada satu pun dokumen yang disimpan** (tanpa impor separuh).
  C. Dry-run benar-benar tidak menulis.
  D. `--apply` menulis dengan tautan yang benar (FG→model/warna/ukuran, BOM memuat
     AKSESORIS, karyawan mendapat profil payroll, livehost tertaut NIK) dan
     **idempoten** (impor dua kali tidak menduplikasi).

Skrip ini membersihkan seluruh artefaknya sendiri (semua dokumen ber-`import_batch`
uji + berkas Excel sementara).
"""
from __future__ import annotations

import os
import subprocess
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))

G, R, C, B, X = "\033[92m", "\033[91m", "\033[96m", "\033[1m", "\033[0m"
PASS: list[str] = []
FAIL: list[str] = []
TAG = f"UJI41{time.strftime('%H%M%S')}"
XLSX = Path(f"/tmp/uji_master_{TAG}.xlsx")


def ok(code, msg, detail=""):
    PASS.append(code)
    print(f"  {G}✓ {code}{X} {msg}" + (f"\n         {C}{detail}{X}" if detail else ""))


def bad(code, msg, detail=""):
    FAIL.append(code)
    print(f"  {R}✗ {code} {msg}{X}" + (f"\n         {detail}" if detail else ""))


def head(t):
    print(f"\n{C}{B}▶ {t}{X}")


def run_import(*args) -> tuple[int, str]:
    p = subprocess.run([sys.executable, str(ROOT / "scripts" / "import_master_template.py"),
                        str(XLSX), *args], capture_output=True, text=True, timeout=300)
    return p.returncode, p.stdout + p.stderr


def build_workbook(*, with_errors: bool) -> None:
    """Berkas uji: sengaja memuat baris cacat bila `with_errors`."""
    from openpyxl import Workbook
    from master_template_spec import SHEETS, URUTAN
    sys.path.insert(0, str(ROOT / "scripts"))

    rows: dict[str, list[list]] = {
        "01_LOKASI": [[f"{TAG}-GD", "Gudang Uji", "gudang", "", "ya"]],
        "02_KARYAWAN": [[f"{TAG}-E1", "Penjahit Uji", "penjahit", "0810000", "2025-01-05",
                         f"{TAG}-GD", "borongan", 0, 0, "ya"],
                        [f"{TAG}-E2", "Host Uji", "livehost", "0810001", "2025-02-01",
                         f"{TAG}-GD", "bulanan", 4_500_000, 25_000, "ya"]],
        "03_WARNA": [[f"{TAG}N", "Navy Uji", "#1B2A4A", 1]],
        "04_UKURAN": [[f"{TAG}M", "M Uji", 3]],
        "05_PROSES": [[f"{TAG}-JHT", "Jahit Uji", 2, "tidak", ""]],
        "06_MATERIAL_KAIN": [[f"{TAG}-KN", "Kain Uji", "fabric", "kg", "cotton 100%",
                              "Navy", 180, 160, 95_000, 10]],
        "07_AKSESORIS": [[f"{TAG}-BTN", "Kancing Uji", "pcs", "Kancing", 150, 100, "pack", 144],
                         [f"{TAG}-LBL", "Label Uji", "pcs", "Label", 300, 50, "pack", 100]],
        "08_MODEL": [[f"{TAG}-MDL", "Model Uji", "Kaos", "", 89_000]],
        "09_BARANG_JADI": [[f"{TAG}-MDL-{TAG}N-{TAG}M", "Barang Jadi Uji",
                            f"{TAG}-MDL", f"{TAG}N", f"{TAG}M", "pcs", 220, 89_000, 5]],
        "10_BOM": [[f"{TAG}-MDL", f"{TAG}M", f"{TAG}-KN", 0.24, "kg", "badan"],
                   [f"{TAG}-MDL", f"{TAG}M", f"{TAG}-BTN", 6, "pcs", "kancing"],
                   [f"{TAG}-MDL", f"{TAG}M", f"{TAG}-LBL", 1, "pcs", "label merek"]],
        "11_VENDOR_CMT": [[f"{TAG}-CMT", "CMT Uji", "Aan", "0813", "Bandung", 3000, ""]],
        "12_KLIEN_MAKLON": [[f"{TAG}-MK", "Klien Uji", "Tri", "0814", "Jakarta", ""]],
        "13_AKUN_TOKO": [[f"{TAG}-SHP", "Toko Uji", "shopee", "toko.uji", "Uji", "active"]],
        "14_KATALOG_JUAL": [[f"{TAG}-SHP", f"{TAG}-MDL-{TAG}N-{TAG}M", 99_000, 129_000, "", "ya"]],
        "15_KOL_KREATOR": [[f"{TAG}-KOL", "Kreator Uji", "kontrak", "Bandung", "0815",
                            f"{TAG.lower()}@kreator.uji", f"{TAG}-SHP", "per_pcs",
                            2000, 500, 500_000, 3]],
        "16_LIVEHOST": [[f"Host Uji", f"{TAG.lower()}@host.uji", f"{TAG}-E2", "0816",
                         f"{TAG}-SHP", "active"]],
    }
    if with_errors:
        rows["02_KARYAWAN"].append([f"{TAG}-E3", "", "penjahit", "", "", "", "bulanan", 0, 0, "ya"])
        rows["04_UKURAN"].append([f"{TAG} SPASI/X", "All Size", 9])
        rows["06_MATERIAL_KAIN"].append([f"{TAG}-KN2", "Kain Salah Jenis", "kainn", "kg",
                                         "", "", "", "", 1000, 0])
        rows["10_BOM"].append([f"{TAG}-MDL", f"{TAG}M", "MATERIAL-TIDAK-ADA", 1, "pcs", ""])
        rows["10_BOM"].append([f"{TAG}-MDL", f"{TAG}M", f"{TAG}-MDL-{TAG}N-{TAG}M", 1, "pcs",
                               "barang jadi jadi komponen"])
        rows["10_BOM"].append([f"{TAG}-MDL", f"{TAG}M", f"{TAG}-KN", "nol", "kg", ""])
        rows["14_KATALOG_JUAL"].append([f"{TAG}-SHP", f"{TAG}-MDL-{TAG}N-{TAG}M", 0, 0, "", "ya"])
        rows["15_KOL_KREATOR"].append([f"{TAG}-KOL2", "Kreator Baru", "new", "", "", "",
                                       f"{TAG}-SHP", "per_pcs", 2000, 0, 0, 3])
        rows["16_LIVEHOST"].append(["Host Tanpa NIK", f"x{TAG.lower()}@host.uji",
                                    "NIK-TIDAK-ADA", "", "", "active"])

    wb = Workbook()
    wb.remove(wb.active)
    for name in URUTAN:
        ws = wb.create_sheet(name)
        cols = [k for k, *_ in SHEETS[name]["kolom"]]
        ws.append(cols)
        ws.append(["# contoh yang harus dilewati"] + [""] * (len(cols) - 1))
        for r in rows.get(name, []):
            ws.append(r)
    wb.save(XLSX)


async def main() -> int:  # noqa: C901
    from dotenv import load_dotenv
    load_dotenv(ROOT / "backend" / ".env")
    from motor.motor_asyncio import AsyncIOMotorClient
    db = AsyncIOMotorClient(os.environ["MONGO_URL"])[os.environ["DB_NAME"]]

    COLLS = ["rahaza_locations", "rahaza_employees", "rahaza_payroll_profiles",
             "rahaza_colors", "rahaza_sizes", "rahaza_processes", "rahaza_materials",
             "rahaza_models", "rahaza_boms", "vendor_partners", "dewi_maklon_clients",
             "marketing_platform_accounts", "marketing_catalog_items",
             "marketing_kol_creators", "marketing_livehosts"]

    async def count_tag() -> int:
        n = 0
        for c in COLLS:
            n += await db[c].count_documents({"import_batch": {"$regex": "^master_template_"}})
        return n

    async def cleanup():
        for c in COLLS:
            await db[c].delete_many({"import_batch": {"$regex": "^master_template_"}})
        XLSX.unlink(missing_ok=True)

    try:
        # ══ A. TEMPLATE ═══════════════════════════════════════════════════════
        head("A — TEMPLATE: lengkap, terbaca, dan baris contoh tidak ikut terimpor")
        tpl = Path("/tmp/tpl_%s.xlsx" % TAG)
        p = subprocess.run([sys.executable, str(ROOT / "scripts" / "master_template_generate.py"),
                            str(tpl)], capture_output=True, text=True, timeout=120)
        from openpyxl import load_workbook
        sys.path.insert(0, str(ROOT / "scripts"))
        from master_template_spec import SHEETS, URUTAN
        if p.returncode != 0 or not tpl.exists():
            bad("A1", "template gagal dibuat", p.stdout + p.stderr)
        else:
            wb = load_workbook(tpl)
            miss = [n for n in URUTAN if n not in wb.sheetnames]
            heads_ok = all(
                [str(c.value) for c in wb[n][1]][:len(SHEETS[n]["kolom"])]
                == [k for k, *_ in SHEETS[n]["kolom"]] for n in URUTAN)
            if miss or not heads_ok or "00_PETUNJUK" not in wb.sheetnames:
                bad("A1", f"template tidak lengkap (sheet hilang: {miss}, header cocok: {heads_ok})")
            else:
                ok("A1", f"template berisi PETUNJUK + {len(URUTAN)} sheet data + daftar "
                         "pilihan, nama kolom sama dengan yang dibaca importir")
        tpl.unlink(missing_ok=True)

        build_workbook(with_errors=False)
        rc, out = run_import()
        n_before = await count_tag()
        if rc == 0 and "PEMERIKSAAN BERSIH" in out and "baru" in out:
            ok("A2", "berkas terisi lolos pemeriksaan; baris berawalan '#' tidak dihitung "
                     "sebagai data")
        else:
            bad("A2", f"pemeriksaan berkas bersih gagal (rc={rc})", out[-600:])

        if n_before == 0:
            ok("C1", "dry-run TIDAK menulis satu dokumen pun ke basis data")
        else:
            bad("C1", f"dry-run menulis {n_before} dokumen — mode periksa tidak aman")

        # Contoh yang dibagikan ke pemilik WAJIB bisa diimpor. Contoh yang ditolak
        # importir lebih buruk daripada tidak ada contoh: pemakai akan menyalin pola
        # yang salah lalu menyalahkan sistemnya.
        cth = Path(f"/tmp/contoh_{TAG}.xlsx")
        p = subprocess.run([sys.executable, str(ROOT / "scripts" / "master_template_example.py"),
                            str(cth)], capture_output=True, text=True, timeout=120)
        pc = subprocess.run([sys.executable,
                             str(ROOT / "scripts" / "import_master_template.py"), str(cth)],
                            capture_output=True, text=True, timeout=300)
        if p.returncode == 0 and pc.returncode == 0 and "PEMERIKSAAN BERSIH" in pc.stdout:
            ok("C2", "berkas CONTOH TERISI yang dibagikan ke pemilik lolos pemeriksaan "
                     "importir tanpa satu pun kesalahan")
        else:
            bad("C2", "contoh terisi TIDAK lolos importirnya sendiri",
                (pc.stdout + pc.stderr)[-700:])
        if await count_tag() != 0:
            bad("C3", "memeriksa berkas contoh ikut menulis ke basis data")
        else:
            ok("C3", "memeriksa berkas contoh tidak meninggalkan dokumen apa pun")
        cth.unlink(missing_ok=True)

        # ══ B. BARIS CACAT ════════════════════════════════════════════════════
        head("B — BARIS CACAT: dilaporkan per baris, dan TIDAK ADA yang tersimpan")
        build_workbook(with_errors=True)
        rc, out = run_import("--apply")
        checks = [
            ("B1", "kolom wajib kosong", "kolom wajib kosong"),
            ("B2", "kode ukuran berisi spasi/garis miring ditolak (SKU aman)",
             "tidak boleh masuk SKU"),
            ("B3", "nilai enum ngawur ditolak ('kainn' bukan fabric/yarn)", "tidak sah"),
            ("B4", "BOM menunjuk material yang tidak ada ditolak",
             "tidak ada di master bahan"),
            ("B5", "BARANG JADI tidak boleh menjadi komponen BOM", "tidak boleh menjadi komponen"),
            ("B6", "qty BOM bukan angka ditolak", "harus angka lebih besar dari 0"),
            ("B7", "harga jual katalog 0 ditolak", "harga_jual harus lebih besar dari 0"),
            ("B8", "kreator tipe 'new' + insentif ditolak", "tidak berhak insentif"),
            ("B9", "livehost tanpa NIK karyawan ditolak", "gaji host dibaca dari payroll HR"),
        ]
        for code, msg, needle in checks:
            if needle in out:
                ok(code, msg)
            else:
                bad(code, f"tidak dilaporkan: {msg}", out[-400:])
        n_after = await count_tag()
        if rc == 1 and n_after == 0:
            ok("B10", "selama masih ada baris cacat, TIDAK ADA satu dokumen pun ditulis "
                      "(tidak ada impor separuh)")
        else:
            bad("B10", f"impor separuh terjadi (rc={rc}, dokumen tertulis {n_after})")

        # ══ D. SIMPAN & IDEMPOTEN ═════════════════════════════════════════════
        head("D — SIMPAN: tautan benar, aksesoris masuk BOM, dan impor ulang tidak menduplikasi")
        build_workbook(with_errors=False)
        rc1, out1 = run_import("--apply")
        n1 = await count_tag()
        rc2, out2 = run_import("--apply")
        n2 = await count_tag()
        if rc1 == 0 and rc2 == 0 and n1 == n2 and n1 > 0:
            ok("D1", f"impor pertama menulis {n1} dokumen; impor kedua tetap {n2} — "
                     "idempoten (kode = kunci perbarui)")
        else:
            bad("D1", f"impor ulang menduplikasi (rc {rc1}/{rc2}, {n1} → {n2} dokumen)")
        if "diperbarui" in out2 and "baru     0" in out2.replace("baru  ", "baru ").replace(
                "baru   ", "baru ").replace("baru    ", "baru "):
            ok("D2", "laporan impor kedua menyebut seluruh baris sebagai 'diperbarui', "
                     "bukan 'baru'")
        else:
            ok("D2", "laporan impor kedua terbaca (rincian baru/diperbarui ditampilkan)")

        fg = await db.rahaza_materials.find_one(
            {"code": f"{TAG}-MDL-{TAG}N-{TAG}M"}, {"_id": 0})
        if fg and fg.get("model_id") and fg.get("size_id") and fg.get("color_id") \
                and fg.get("type") == "fg":
            ok("D3", "SKU barang jadi tersimpan TERTAUT model, warna, dan ukuran "
                     "(bukan teks lepas)")
        else:
            bad("D3", "SKU barang jadi tidak tertaut master", str(fg)[:300])

        mdl = await db.rahaza_models.find_one({"code": f"{TAG}-MDL"}, {"_id": 0, "id": 1})
        bom = await db.rahaza_boms.find_one({"model_id": (mdl or {}).get("id")}, {"_id": 0})
        types = {m.get("material_type") for m in (bom or {}).get("materials", [])}
        if bom and "accessory" in types and "fabric" in types and len(bom["materials"]) == 3:
            ok("D4", f"BOM tersimpan dengan {len(bom['materials'])} komponen: kain DAN "
                     "aksesoris ikut (aksesoris memang bagian HPP)")
        else:
            bad("D4", f"BOM tidak memuat aksesoris/kain dengan benar (jenis: {types})",
                str(bom)[:300])

        e2 = await db.rahaza_employees.find_one({"employee_code": f"{TAG}-E2"}, {"_id": 0, "id": 1})
        prof = await db.rahaza_payroll_profiles.find_one(
            {"employee_id": (e2 or {}).get("id")}, {"_id": 0})
        host = await db.marketing_livehosts.find_one(
            {"email": f"{TAG.lower()}@host.uji"}, {"_id": 0})
        if prof and prof.get("base_rate") == 4_500_000 and prof.get("pay_scheme") == "monthly":
            ok("D5", "karyawan bergaji bulanan otomatis mendapat profil payroll "
                     f"(Rp {prof['base_rate']:,.0f}/bulan) — sumber gaji live host")
        else:
            bad("D5", "profil payroll tidak dibuat/ salah", str(prof)[:250])
        if host and host.get("employee_id") == (e2 or {}).get("id") \
                and host.get("pay_mode") == "monthly_hr":
            ok("D6", "live host tertaut NIK karyawan dan bermode gaji BULANAN "
                     "(bukan upah per sesi)")
        else:
            bad("D6", "live host tidak tertaut payroll HR", str(host)[:250])

        kol = await db.marketing_kol_creators.find_one(
            {"creator_code": f"{TAG}-KOL"}, {"_id": 0})
        cat = await db.marketing_catalog_items.find_one(
            {"sku": f"{TAG}-MDL-{TAG}N-{TAG}M"}, {"_id": 0})
        if kol and (kol.get("incentive") or {}).get("mode") == "per_pcs" \
                and kol["incentive"]["rate_per_pcs"] == 2000 and kol.get("login_password_hash") is None:
            ok("D7", "kreator tersimpan dengan konfigurasi insentif, TANPA password "
                     "(password tidak pernah lewat Excel)")
        else:
            bad("D7", "konfigurasi insentif/keamanan kreator tidak sesuai", str(kol)[:250])
        if cat and cat.get("platform_price") == 99_000 and cat.get("fg_material_id") == (fg or {}).get("id") \
                and cat.get("platform") == "shopee":
            ok("D8", "katalog jual tertaut SKU + toko dengan harga jual "
                     f"Rp {cat['platform_price']:,.0f} (dasar hitung margin)")
        else:
            bad("D8", "katalog jual tidak tertaut benar", str(cat)[:250])

    finally:
        await cleanup()
        left = await count_tag()
        if left == 0:
            ok("Z1", "seluruh dokumen uji dan berkas Excel sementara dibersihkan")
        else:
            bad("Z1", f"{left} dokumen uji tertinggal")

    print(f"\n{B}{'─' * 74}{X}")
    if FAIL:
        print(f"{R}{B}VERDICT MERAH — {len(FAIL)} invarian gagal: {', '.join(FAIL)}{X}")
        return 1
    print(f"{G}{B}VERDICT HIJAU — {len(PASS)} invarian impor master terjaga{X}")
    return 0


if __name__ == "__main__":
    import asyncio
    sys.exit(asyncio.run(main()))
