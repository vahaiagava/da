#!/usr/bin/env python3
"""INV-REKAP — invarian **Rekap Harian + Rekap Mingguan CMT**.

Harian menjawab "vendor mana yang belum diisi HARI INI"; mingguan (fase 4)
menjawab "vendor mana yang BELAKANGAN INI sering bolong" untuk 7 hari bergulir.

KELAS MASALAH YANG DIJAGA (semuanya nyata, bukan hipotesis)
-----------------------------------------------------------
Layar ini dipakai staf setiap pagi untuk memutuskan vendor mana yang dikejar.
Kalau ia salah, akibatnya bukan "tampilan jelek":

  1. **UANG.** Vendor yang belum setor progress dianggap beres ⇒ progress hari itu
     tidak masuk ⇒ tagihan CMT tidak bisa ditagih/diverifikasi (progress = dasar
     perhitungan tagihan, lihat audit sesi Portal CMT Override).
  2. **KEPERCAYAAN ALAT.** Vendor yang sudah setor tapi ditandai merah akan
     dikejar/ditegur; sekali itu terjadi, staf berhenti memakai layarnya dan
     kembali ke WhatsApp — fiturnya mati walau kodenya hidup.
  3. **BATAS HARI.** Jam sistem container UTC. Kalau batas hari bukan WIB, maka
     selama **07 jam setiap hari** (00:00–07:00 WIB — persis jam produksi mulai)
     rekap "hari ini" menampilkan hari sebelumnya dan SEMUA vendor tampak belum
     mengisi.
  4. **SATU ANGKA.** Layar, berkas Excel/PDF, dan sasaran tombol reminder wajib
     menyebut angka yang sama; kalau tidak, staf berdebat dengan lampirannya.
  5. **TIPE DATA `received_at`.** Bug yang ditutup sesi ini: field itu dulu hanya
     ditulis BROWSER sebagai STRING, sementara field waktu lain bertipe Date ⇒
     query rentang tanggal tidak pernah cocok ⇒ kolom "Terima" abadi ✗. Kalau
     suatu hari ada yang "merapikan" server agar percaya kiriman browser lagi,
     gate ini harus MERAH.
  6. **ANTI-SPAM.** Tombol reminder harus idempoten per vendor per tanggal, dan
     reminder yang dilahirkannya sendiri TIDAK BOLEH membuat vendor abadi-merah
     (kalau ikut dihitung, kolom "Balas Reminder" langsung ✗ pada hari yang sama).

TAMBAHAN FASE 4 — REKAP MINGGUAN (RK-20 … RK-27)
------------------------------------------------
  7. **DUA SUMBER ANGKA.** `build_week()` WAJIB hanya meringkas `build_recap()`
     7×. Kalau suatu hari ia "dioptimalkan" dengan agregasi sendiri, tab Mingguan
     dan tab Harian akan mulai berbeda pada kasus pinggir — dan itu angka tagihan.
  8. **DEFINISI DILEBUR.** `days_late` (hari NOL bukti) dan `days_unfinished`
     (termasuk masih ada sisa) adalah dua keputusan owner yang berbeda; meleburnya
     menghilangkan dasar keputusan "vendor mana yang ditegur".
  9. **HUKUMAN PALSU.** "Hari tanpa setoran" pada vendor yang memang tidak diberi
     job = tuduhan tanpa dasar. Begitu juga menghitung hari yang BELUM TERJADI.
 10. **KINERJA MUNDUR.** `prefetch_context()` ada supaya 7 hari tidak membaca
     master 7×; tanpa gate, refactor berikutnya akan menghapusnya tanpa ada yang
     sadar (data demo terlalu kecil untuk terasa).

TAMBAHAN FASE 5 — `closed_at` (RK-28, RK-28b, RK-29, RK-30)
-----------------------------------------------------------
 11. **LAPORAN YANG MEMAAFKAN DIRINYA SENDIRI.** Sebelum `closed_at` ada, "job
     jalan pada tanggal X" dijawab dari status SEKARANG ⇒ job yang dibuka Senin
     dan ditutup Rabu HILANG dari rekap Senin, sehingga kelalaian yang sudah
     terjadi terhapus sendiri begitu job-nya ditutup.
 12. **JALUR TUTUP BARU YANG LUPA MENULIS STEMPEL.** Ada DUA jalur penutup job
     (auto-complete `production_execution.py` + Quick Complete `production_pos.py`),
     keduanya WAJIB lewat `core.production_job_lifecycle.close_job()`. RK-29
     memeriksa seluruh DB: nol job tertutup tanpa `closed_at`.
 13. **STEMPEL DARI BROWSER.** Sama seperti `received_at`: `closed_at` kiriman
     klien harus DIABAIKAN.
 14. **KETIDAKTAHUAN YANG DISEMBUNYIKAN.** Sisa satu-satunya sesudah fase 5 adalah
     job WARISAN (tertutup sebelum fitur ini ada, tanpa `closed_at`): waktu
     tutupnya tidak diketahui dan SENGAJA tidak ditebak. Justru karena itu
     jumlahnya WAJIB sampai ke layar — harian DAN mingguan — lengkap dengan
     perintah migrasi yang menyembuhkannya. Angka yang tidak pernah ditampilkan
     sama saja dengan tidak ada: staf akan membaca rekap tanggal lampau sebagai
     kebenaran penuh, padahal masih ada job yang tidak terhitung. RK-30 sekaligus
     mengunci `as_of_note` (dibaca berkas export) = `as_of_note_base` +
     `legacy_note`, supaya layar dan lampirannya tidak pernah mengatakan hal
     berbeda tentang data yang sama.

TAMBAHAN F12 — PERBANDINGAN ANTAR-PEKAN (RK-31 … RK-36)
-------------------------------------------------------
 15. **DUA ANGKA DI SATU LAYAR.** Panel perbandingan adalah cara termudah membuat
     "pcs pekan ini" di kartu delta berbeda dari "pcs disetor sepekan" di kartu
     ringkasan — padahal jendelanya sama. RK-31 mengunci: menyalakan `compare`
     TIDAK menggeser satu pun angka/urutan jendela berjalan.
 16. **JENDELA PEMBANDING YANG CURANG.** Kalau jendela "pekan lalu" tumpang tindih
     atau panjangnya berbeda, panah naik/turun membandingkan rentang yang tidak
     setara (RK-32).
 17. **WARNA YANG BERBOHONG.** Arah baik/buruk sudah diputuskan backend
     (`lower_is_better`). Kalau layar/lampiran menghitungnya ulang, suatu hari
     angka yang memburuk akan diberi warna hijau (RK-33).
 18. **SATU BARIS, DUA KEBENARAN.** Kolom "vs pekan lalu" duduk PERSIS di sebelah
     angka mingguannya; kalau sumbernya berbeda, satu baris tabel menampilkan dua
     kebenaran sekaligus (RK-34).
 19. **TEGURAN SALAH SASARAN.** Vendor yang pekan lalu TIDAK DIBERI pekerjaan akan
     selalu tampak "paling membaik", dan yang pekan ini tidak diberi pekerjaan
     tampak "paling memburuk" — keduanya menghukum/memuji vendor atas keputusan
     order KITA. RK-35 mengunci: hanya vendor yang punya pekerjaan di KEDUA pekan
     yang diperingkat, dan yang dikeluarkan WAJIB membawa alasan tertulis.
 20. **JANJI YANG JADI BOHONG.** Legenda layar berbunyi "Excel/PDF isinya sama
     dengan layar ini". Begitu layar punya panel perbandingan sementara lampiran
     tidak, yang dibawa ke rapat justru kehilangan bagian yang dipakai mengambil
     keputusan (RK-36) — sekaligus menjaga biayanya tetap ~2 jendela, bukan dua
     kali seluruh konteks master.

SIFAT SKRIP
-----------
Self-contained: membuat sendiri vendor uji + staf uji + PO + surat jalan lewat API
sungguhan, lalu MENGHAPUS seluruh jejaknya di ``finally`` langsung ke Mongo —
termasuk turunan UANG (AR invoice maklon + mirror PO maklon) dan **sweep seluruh
koleksi**, karena riwayat repo ini menunjukkan daftar-hapus manual selalu
ketinggalan satu efek samping (dan pernah meninggalkan piutang palsu).

Pakai:
    cd /app && python3 scripts/verify_rekap_harian.py
    cd /app && python3 scripts/verify_rekap_harian.py --keep   # sisakan data uji
"""
from __future__ import annotations

import json
import os
import sys
import uuid
from datetime import datetime, timedelta, timezone

import requests

BASE = os.environ.get("POC_BASE", "http://localhost:8001")
API = f"{BASE}/api"
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME = os.environ.get("DB_NAME", "test_database")
KEEP = "--keep" in sys.argv

MARK = "__REKAPTEST__"
OVH = "X-CMT-Override-Vendor"
WIB = timezone(timedelta(hours=7))

# FASE G (2026-08-16): nomor PO uji WAJIB mengikuti pola resmi jenis dokumennya.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "lib"))
from gr_common import test_doc_number  # noqa: E402

G, R, Y, B, X = "\033[92m", "\033[91m", "\033[93m", "\033[94m", "\033[0m"
PASSES: list[str] = []
FAILS: list[str] = []


def ok(code: str, msg: str, ev=None):
    PASSES.append(code)
    ex = f" · {json.dumps(ev, default=str)[:150]}" if ev else ""
    print(f"  {G}[OK]{X} {code} — {msg}{ex}")


def bad(code: str, msg: str, ev=None):
    FAILS.append(code)
    ex = f" · {json.dumps(ev, default=str)[:240]}" if ev else ""
    print(f"  {R}[FAIL]{X} {code} — {msg}{ex}")


def expect(cond, code, msg_ok, msg_bad, ev=None):
    (ok if cond else bad)(code, msg_ok if cond else msg_bad, ev)
    return bool(cond)


def H(tok, vendor=None):
    h = {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}
    if vendor:
        h[OVH] = vendor
    return h


def call(method, path, tok=None, vendor=None, body=None, timeout=90, raw=False):
    fn = getattr(requests, method.lower())
    kw = {"headers": H(tok, vendor) if tok else {"Content-Type": "application/json"},
          "timeout": timeout}
    if body is not None:
        kw["json"] = body
    r = fn(f"{API}{path}", **kw)
    if raw:
        return r
    try:
        return r.status_code, r.json()
    except Exception:
        return r.status_code, r.text


def login(email, password):
    c, b = call("post", "/auth/login", body={"email": email, "password": password})
    return b.get("token") if c == 200 and isinstance(b, dict) else None


def get_db():
    from pymongo import MongoClient
    return MongoClient(MONGO_URL)[DB_NAME]


def today_wib_str() -> str:
    return datetime.now(WIB).date().isoformat()


def row_of(rec, vid):
    for r in (rec or {}).get("rows", []):
        if r.get("vendor_id") == vid:
            return r
    return {}


# ═══════════════════════════════════════════════════════════════════════════
def main() -> int:
    print(f"{B}{'=' * 78}{X}")
    print(f"  {B}INV-REKAP{X} — Rekap Harian + Mingguan CMT: batas WIB · definisi terisi · SSOT · anti-spam · UANG")
    print(f"{B}{'=' * 78}{X}")

    db = get_db()
    created_users: list[str] = []
    vid = None
    po_id = ship_id = None

    try:
        admin = login("admin@garment.com", "Admin@123")
        if not admin:
            bad("RK-0", "login superadmin gagal — gate tidak bisa berjalan")
            return 1
        hr = login("hr@dewiaditya.id", "Dewi@123")

        # ── setup ────────────────────────────────────────────────────────────
        c, vb = call("post", "/vendor-portal/partners", admin, body={
            "name": f"{MARK} Vendor Rekap", "code": f"RK{uuid.uuid4().hex[:5].upper()}",
            "notes": f"{MARK} data uji gate", "capacity_pcs": 100})
        vid = vb.get("id") if isinstance(vb, dict) else None
        if not expect(c in (200, 201) and vid, "RK-0", "vendor uji dibuat",
                      "gagal membuat vendor uji", {"http": c}):
            return 1

        s_email = f"rekaptest.staff.{uuid.uuid4().hex[:6]}@example.test"
        c, su = call("post", "/users", admin, body={
            "name": f"{MARK} Staf", "email": s_email, "password": "GateRk@123",
            "role": "admin_produksi", "department": "Produksi"})
        if isinstance(su, dict) and su.get("id"):
            created_users.append(su["id"])
        staff = login(s_email, "GateRk@123")
        if not expect(bool(staff), "RK-0", "staf admin_produksi uji siap",
                      "login staf uji gagal"):
            return 1

        po_number = test_doc_number("production_pos.po_number_maklon", admin)
        dl = (datetime.now(timezone.utc) + timedelta(days=7)).date().isoformat()
        c, po = call("post", "/production-pos", admin, body={
            "po_number": po_number, "business_type": "maklon", "vendor_id": vid,
            "customer_name": f"{MARK} Buyer", "status": "Confirmed",
            "deadline": dl, "delivery_deadline": dl, "notes": f"{MARK}",
            "items": [{"product_name": f"{MARK} Kaos", "sku": f"{MARK}-M", "size": "M",
                       "color": "Navy", "qty": 40, "serial_number": f"{MARK}-SN1",
                       "cmt_price_snapshot": 8000}]})
        po_id = po.get("id") if isinstance(po, dict) else None
        if not expect(c in (200, 201) and po_id, "RK-0", "PO maklon uji dibuat",
                      "gagal membuat PO uji", {"http": c}):
            return 1
        po_items = list(db.po_items.find({"po_id": po_id}, {"_id": 0, "id": 1, "qty": 1}))
        c, sh = call("post", "/vendor-shipments", admin, body={
            "shipment_number": f"{MARK}-SJ-{uuid.uuid4().hex[:6].upper()}",
            "vendor_id": vid, "po_id": po_id, "shipment_type": "NORMAL", "notes": f"{MARK}",
            "items": [{"po_id": po_id, "po_item_id": it["id"], "qty_sent": it["qty"]}
                      for it in po_items]})
        ship_id = sh.get("id") if isinstance(sh, dict) else None
        if not expect(c in (200, 201) and ship_id, "RK-0", "surat jalan uji dibuat (Sent)",
                      "gagal membuat surat jalan uji", {"http": c}):
            return 1

        money_before = (call("get", "/production/cmt-billing/summary", admin)[1] or {}) \
            .get("total_amount")

        # ── RK-1 — KEWENANGAN ────────────────────────────────────────────────
        c1, _ = call("get", "/cmt-override/daily-recap", hr)
        c2, _ = call("post", "/cmt-override/daily-recap/remind", hr, body={})
        c3, _ = call("get", "/cmt-override/daily-recap/export?format=xlsx", hr)
        c4, _ = call("get", "/cmt-override/daily-recap")
        expect(c1 == 403 and c2 == 403 and c3 == 403 and c4 in (401, 403), "RK-1",
               "role tak berwenang & tanpa token DITOLAK di ketiga pintu rekap",
               "rekap bisa dibuka pihak tak berwenang — daftar pekerjaan seluruh vendor bocor",
               {"recap": c1, "remind": c2, "export": c3, "anon": c4})

        # ── RK-2 — BATAS HARI WIB ────────────────────────────────────────────
        c, rec = call("get", "/cmt-override/daily-recap", staff)
        expect(c == 200 and (rec or {}).get("date") == today_wib_str(), "RK-2",
               "tanggal default = hari ini menurut WIB (bukan jam UTC container)",
               "batas hari bukan WIB — selama 00:00–07:00 WIB rekap menampilkan hari kemarin "
               "dan semua vendor tampak belum mengisi",
               {"api": (rec or {}).get("date"), "wib": today_wib_str(),
                "utc": datetime.now(timezone.utc).date().isoformat()})

        # ── RK-3 — CAKUPAN: tidak ada vendor yang hilang tanpa penjelasan ────
        active = db.vendor_partners.count_documents(
            {"$and": [{"is_active": {"$ne": False}}, {"active": {"$ne": False}}]})
        got = len((rec or {}).get("rows", []))
        expect(got == active, "RK-3",
               f"semua vendor aktif tampil ({got} baris = {active} master) — keputusan owner 2a",
               "ada vendor aktif yang tidak muncul di rekap ⇒ staf tidak akan pernah tahu "
               "vendor itu terlewat", {"rows": got, "master_aktif": active})

        # ── RK-4 — RINGKASAN konsisten dengan barisnya ───────────────────────
        s = (rec or {}).get("summary", {})
        rws = (rec or {}).get("rows", [])
        expect(s.get("vendors_pending") == sum(1 for r in rws if r.get("status") == "pending")
               and s.get("vendors_total") == len(rws)
               and s.get("tasks_pending_total") == sum(r.get("pending_count", 0) for r in rws),
               "RK-4", "kartu ringkasan == hitungan baris (layar tidak berdebat dengan dirinya)",
               "ringkasan tidak cocok dengan tabelnya", {"summary": s})

        # ── RK-5 — DEFINISI "belum diisi" untuk pekerjaan yang menunggu ──────
        r0 = row_of(rec, vid)
        expect((r0.get("tasks", {}).get("terima", {}).get("state")) == "pending"
               and r0.get("status") == "pending", "RK-5",
               "surat jalan menunggu ⇒ kolom Terima ✗ dan vendor bertanda belum diisi",
               "pekerjaan yang menunggu TIDAK muncul merah — vendor terlewat tanpa ada yang tahu",
               {"tasks": {k: v.get("state") for k, v in (r0.get("tasks") or {}).items()}})

        # ── RK-6 — received_at ditulis SERVER sebagai TANGGAL ────────────────
        c, _ = call("put", f"/vendor-shipments/{ship_id}", staff, vendor=vid,
                    body={"status": "Received", "received_at": "JAM-PALSU-BROWSER"})
        sdoc = db.vendor_shipments.find_one({"id": ship_id}, {"_id": 0}) or {}
        ra = sdoc.get("received_at")
        expect(c == 200 and isinstance(ra, datetime), "RK-6",
               "received_at diisi SERVER bertipe tanggal (kiriman browser diabaikan)",
               "received_at bukan tanggal ⇒ query rentang hari tidak akan pernah cocok dan "
               "kolom Terima abadi ✗ walau barang diterima",
               {"http": c, "type": type(ra).__name__, "value": str(ra)[:40]})

        c, rec2 = call("get", "/cmt-override/daily-recap", staff)
        t2 = row_of(rec2, vid).get("tasks", {})
        expect(t2.get("terima", {}).get("state") == "done"
               and t2.get("terima", {}).get("source") == "staff"
               and t2.get("inspeksi", {}).get("state") == "pending", "RK-7",
               "✗ → ✓ setelah dikerjakan, sumbernya ditandai, dan tugas berikutnya otomatis ✗",
               "kolom tidak berubah setelah pekerjaannya dikerjakan",
               {"terima": t2.get("terima"), "inspeksi": t2.get("inspeksi", {}).get("state")})

        # ── RK-8 — PROGRESS diresolusi lewat job, bukan vendor_id ────────────
        # `production_progress` TIDAK menyimpan vendor_id. Kalau suatu hari
        # filternya "dirapikan" memakai vendor_id, kolom Progress akan NOL
        # selamanya — dan justru progress inilah dasar tagihan.
        ship_items = list(db.vendor_shipment_items.find({"shipment_id": ship_id}, {"_id": 0}))
        call("post", "/vendor-material-inspections", staff, vendor=vid, body={
            "shipment_id": ship_id, "overall_notes": f"{MARK} inspeksi",
            "items": [{"shipment_item_id": si["id"], "sku": si.get("sku", ""),
                       "product_name": si.get("product_name", ""), "size": si.get("size", ""),
                       "color": si.get("color", ""), "ordered_qty": int(si.get("qty_sent", 0) or 0),
                       "received_qty": int(si.get("qty_sent", 0) or 0), "missing_qty": 0}
                      for si in ship_items]})
        c, job = call("post", "/production-jobs", staff, vendor=vid,
                      body={"vendor_shipment_id": ship_id, "notes": f"{MARK} job"})
        job_id = job.get("id") if isinstance(job, dict) else None
        c, jit = call("get", f"/production-job-items?job_id={job_id}", staff, vendor=vid)
        jitems = jit if isinstance(jit, list) else (jit or {}).get("items", [])
        if jitems:
            call("post", "/production-progress", staff, vendor=vid, body={
                "job_item_id": jitems[0]["id"], "progress_date": today_wib_str(),
                "completed_quantity": 12, "notes": f"{MARK} setoran"})
        c, rec3 = call("get", "/cmt-override/daily-recap", staff)
        p3 = row_of(rec3, vid).get("tasks", {}).get("progress", {})
        expect(p3.get("state") in ("done", "partial") and p3.get("qty_today") == 12, "RK-8",
               "progress terhitung lewat production_jobs (12 pcs) — bukan lewat vendor_id yang "
               "tidak pernah ada di production_progress",
               "setoran progress tidak terbaca rekap ⇒ vendor yang sudah setor tetap merah",
               {"progress": p3})
        expect(row_of(rec3, vid).get("tasks", {}).get("kirim", {}).get("waiting") == 12, "RK-9",
               "barang selesai yang belum dikirim terhitung 12 pcs (satuan pcs, bukan dokumen)",
               "kolom Kirim tidak melihat barang selesai yang menganggur",
               {"kirim": row_of(rec3, vid).get("tasks", {}).get("kirim")})

        # ── RK-10 — SSOT: angka export == angka layar ────────────────────────
        r = call("get", f"/cmt-override/daily-recap/export?format=xlsx&date={today_wib_str()}",
                 staff, raw=True)
        xlsx_ok = r.status_code == 200 and r.content[:2] == b"PK"
        same = False
        if xlsx_ok:
            try:
                import io

                import openpyxl
                ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
                for row in ws.iter_rows(values_only=True):
                    if row and row[0] == "BELUM diisi (ada tugas merah)":
                        same = int(row[1]) == int((rec3 or {}).get("summary", {})
                                                  .get("vendors_pending", -1))
                        break
            except Exception as e:  # noqa: BLE001
                print(f"  {Y}! gagal membaca xlsx: {e}{X}")
        rp = call("get", "/cmt-override/daily-recap/export?format=pdf", staff, raw=True)
        expect(xlsx_ok and same and rp.status_code == 200 and rp.content[:5] == b"%PDF-", "RK-10",
               "Excel & PDF valid DAN angkanya sama dengan API (satu sumber kebenaran)",
               "berkas export tidak valid atau angkanya beda dengan layar ⇒ staf berdebat "
               "dengan lampirannya sendiri",
               {"xlsx": r.status_code, "angka_sama": same, "pdf": rp.status_code})

        # ── RK-11 — TANGGAL LAIN benar-benar disaring ────────────────────────
        yday = (datetime.now(WIB).date() - timedelta(days=1)).isoformat()
        c, recy = call("get", f"/cmt-override/daily-recap?date={yday}", staff)
        ty = row_of(recy, vid).get("tasks", {})
        c400, _ = call("get", "/cmt-override/daily-recap?date=bukan-tanggal", staff)
        expect(c == 200 and (recy or {}).get("date") == yday
               and all((ty.get(k) or {}).get("state") == "none" for k in ty)
               and c400 == 400, "RK-11",
               "rekap tanggal lain menyaring dengan benar & format salah ditolak 400",
               "navigasi tanggal tidak mengubah isi (atau format salah diterima diam-diam) ⇒ "
               "supervisor mengecek hari yang salah",
               {"kemarin": {k: (v or {}).get('state') for k, v in ty.items()}, "http400": c400})

        # ── RK-12 — rekap MENGABAIKAN header override ────────────────────────
        c, rech = call("get", "/cmt-override/daily-recap", staff, vendor=vid)
        expect(c == 200 and len((rech or {}).get("rows", [])) == got, "RK-12",
               "rekap tetap lintas-vendor walau header override terpasang",
               "rekap ikut ter-scope satu vendor ⇒ staf yakin hanya satu vendor yang perlu diisi",
               {"rows": len((rech or {}).get("rows", []))})

        # ── RK-13 — reminder: idempoten & tidak membuat abadi-merah ──────────
        c, res1 = call("post", "/cmt-override/daily-recap/remind", staff,
                       body={"vendor_ids": [vid]})
        c, res2 = call("post", "/cmt-override/daily-recap/remind", staff,
                       body={"vendor_ids": [vid]})
        n_rem = db.reminders.count_documents(
            {"vendor_id": vid, "reminder_type": "daily_recap", "recap_date": today_wib_str()})
        expect((res1 or {}).get("sent_count") == 1 and (res2 or {}).get("sent_count") == 0
               and (res2 or {}).get("skipped_count") == 1 and n_rem == 1, "RK-13",
               "reminder rekap idempoten: satu vendor satu reminder per tanggal",
               "reminder tergandakan ⇒ inbox vendor dibanjiri dan teguran kehilangan arti",
               {"pertama": (res1 or {}).get("sent_count"),
                "kedua": (res2 or {}).get("sent_count"), "dokumen": n_rem})

        c, rec4 = call("get", "/cmt-override/daily-recap", staff)
        rw = row_of(rec4, vid).get("tasks", {}).get("reminder", {})
        c, rec5 = call("get",
                       f"/cmt-override/daily-recap?date="
                       f"{(datetime.now(WIB).date() + timedelta(days=1)).isoformat()}", staff)
        rw2 = row_of(rec5, vid).get("tasks", {}).get("reminder", {})
        expect(rw.get("waiting") == 0 and rw2.get("waiting", 0) >= 1, "RK-14",
               "reminder rekap tidak dihitung sebagai pekerjaan pada tanggalnya sendiri, "
               "tetapi DIHITUNG pada tanggal berikutnya",
               "reminder rekap membuat vendor mustahil hijau (tombolnya jadi jebakan) atau "
               "hilang dari hitungan selamanya",
               {"hari_ini": rw.get("waiting"), "besok": rw2.get("waiting")})

        c404, _ = call("post", "/cmt-override/daily-recap/remind", staff,
                       body={"vendor_ids": ["vendor-tidak-ada"]})
        expect(c404 == 404, "RK-15",
               "vendor tak dikenal pada reminder → 404 jelas",
               "vendor tak dikenal dilewati diam-diam ⇒ staf yakin sudah menegur padahal tidak",
               {"http": c404})

        # ══════════════════════════════════════════════════════════════════════
        # REKAP MINGGUAN (fase 4) — RK-20 … RK-27
        # ══════════════════════════════════════════════════════════════════════
        # Kelas masalah yang dijaga di sini, semuanya sudah pernah terjadi di repo
        # lain bentuknya:
        #   * DUA SUMBER ANGKA. Kalau suatu hari `build_week()` "dioptimalkan"
        #     dengan menghitung agregasinya sendiri (bukan memanggil
        #     `build_recap()` 7×), tab Mingguan dan tab Harian akan mulai berbeda
        #     pada kasus pinggir — dan angka itu dasar tagihan CMT.
        #   * DEFINISI DILEBUR. `days_late` (hari NOL bukti) dan `days_unfinished`
        #     (termasuk yang masih ada sisa) adalah keputusan owner yang eksplisit
        #     berbeda. Meleburnya jadi satu angka menghilangkan informasi yang
        #     dipakai memutuskan vendor mana yang ditegur.
        #   * HUKUMAN PALSU. "Hari tanpa setoran" pada vendor yang memang tidak
        #     diberi job = menuduh vendor tanpa dasar.
        #   * KINERJA MUNDUR. `prefetch_context()` ada supaya 7 hari tidak membaca
        #     master 7×; tanpa gate, refactor berikutnya akan menghapusnya.
        def wk_row(wkd, vendor_id):
            for r_ in (wkd or {}).get("rows", []):
                if r_.get("vendor_id") == vendor_id:
                    return r_
            return {}

        today_iso = today_wib_str()
        d_end = datetime.now(WIB).date()

        # ── RK-20 — jendela 7 hari BERGULIR + validasi parameter ─────────────
        cw, wk = call("get", "/cmt-override/weekly-recap", staff)
        exp_start = (d_end - timedelta(days=6)).isoformat()
        c_bad1, _ = call("get", "/cmt-override/weekly-recap?days=0", staff)
        c_bad2, _ = call("get", "/cmt-override/weekly-recap?days=99", staff)
        c_bad3, _ = call("get", "/cmt-override/weekly-recap?date=08-2026", staff)
        expect(cw == 200 and (wk or {}).get("end") == today_iso
               and (wk or {}).get("start") == exp_start
               and len((wk or {}).get("days") or []) == 7
               and c_bad1 == 400 and c_bad2 == 400 and c_bad3 == 400, "RK-20",
               "jendela = 7 hari BERGULIR yang berakhir di ?date (default hari ini WIB); "
               "parameter salah ditolak 400",
               "rentang mingguan salah atau parameter buruk tidak ditolak ⇒ laporan "
               "menyebut pekan yang bukan diminta",
               {"start": (wk or {}).get("start"), "end": (wk or {}).get("end"),
                "days": len((wk or {}).get("days") or []),
                "http_days0/99/date": [c_bad1, c_bad2, c_bad3]})

        # ── RK-21 — INVARIAN UTAMA: mingguan hanya MERINGKAS harian ──────────
        beda: list[str] = []
        for p in (wk or {}).get("per_day", []):
            if p.get("is_future"):
                continue
            _c, dayrec = call("get", f"/cmt-override/daily-recap?date={p['date']}", staff)
            s = (dayrec or {}).get("summary") or {}
            for kw, kd in (("vendors_pending", "vendors_pending"),
                           ("vendors_partial", "vendors_partial"),
                           ("vendors_done", "vendors_done"),
                           ("vendors_idle", "vendors_idle"),
                           ("tasks_pending_total", "tasks_pending_total"),
                           ("qty_progress", "qty_progress_today"),
                           ("qty_shipped", "qty_shipped_today")):
                if p.get(kw) != s.get(kd):
                    beda.append(f"{p['date']}.{kw} {p.get(kw)}!={s.get(kd)}")
            dstat = {r_["vendor_id"]: r_["status"] for r_ in (dayrec or {}).get("rows", [])}
            for r_ in (wk or {}).get("rows", []):
                cell = next((c_ for c_ in r_.get("cells", []) if c_.get("date") == p["date"]), {})
                if cell.get("state") != dstat.get(r_["vendor_id"]):
                    beda.append(f"{p['date']}/{r_['vendor_name'][:12]} "
                                f"{cell.get('state')}!={dstat.get(r_['vendor_id'])}")
        sm = (wk or {}).get("summary") or {}
        totals_ok = (sm.get("qty_progress_total")
                     == sum(p["qty_progress"] for p in (wk or {}).get("per_day", []))
                     and sm.get("days_late_total")
                     == sum(r_["days_late"] for r_ in (wk or {}).get("rows", [])))
        expect(not beda and totals_ok, "RK-21",
               "SETIAP angka per hari & state per vendor pada rekap mingguan == rekap "
               "harian tanggal itu (mingguan hanya MERINGKAS, tidak menghitung ulang)",
               "tab Mingguan dan tab Harian menyebut angka BERBEDA untuk tanggal yang "
               "sama ⇒ tidak ada yang tahu mana yang benar, dan angka ini dasar tagihan",
               {"selisih": beda[:6], "total_konsisten": totals_ok})

        # ── RK-22 — dua angka "terlambat" tetap TERPISAH ─────────────────────
        salah22 = []
        for r_ in (wk or {}).get("rows", []):
            cells = r_.get("cells") or []
            e_late = sum(1 for c_ in cells if c_.get("state") == "pending")
            e_unf = sum(1 for c_ in cells if c_.get("state") in ("pending", "partial"))
            if r_.get("days_late") != e_late or r_.get("days_unfinished") != e_unf:
                salah22.append(f"{r_['vendor_name'][:14]} {r_.get('days_late')}/"
                               f"{r_.get('days_unfinished')} != {e_late}/{e_unf}")
        expect(not salah22, "RK-22",
               "days_late HANYA menghitung hari 'pending' (nol bukti) dan days_unfinished "
               "menghitung 'pending'+'partial' — dua keputusan owner yang berbeda tetap "
               "dilaporkan terpisah",
               "kedua definisi dilebur ⇒ vendor yang sudah menyetor tapi menyisakan "
               "pekerjaan dituduh sama dengan vendor yang tidak menyetor sama sekali",
               {"selisih": salah22[:5]})

        # ── RK-23 — "hari tanpa setoran" tidak menghukum vendor tanpa job ────
        salah23 = []
        for r_ in (wk or {}).get("rows", []):
            e_ns = sum(1 for c_ in (r_.get("cells") or [])
                       if not c_.get("is_future")
                       and c_.get("progress_state") != "none"
                       and int(c_.get("progress_done") or 0) == 0)
            if r_.get("days_no_progress") != e_ns:
                salah23.append(f"{r_['vendor_name'][:14]} {r_.get('days_no_progress')} != {e_ns}")
        idle_bad = [r_["vendor_name"] for r_ in (wk or {}).get("rows", [])
                    if r_.get("days_with_work") == 0
                    and (r_.get("days_no_progress") or r_.get("days_late"))]
        expect(not salah23 and not idle_bad, "RK-23",
               "'hari tanpa setoran' hanya dihitung pada hari vendor MEMANG punya job "
               "jalan; vendor tanpa pekerjaan tidak dihukum",
               "vendor yang tidak diberi pekerjaan ikut dihitung 'tanpa setoran' ⇒ "
               "peringkat vendor terburuk jadi karangan",
               {"selisih": salah23[:5], "idle_dihukum": idle_bad[:5]})

        # ── RK-24 — aturan STREAK (putus pada pending/partial, idle NETRAL) ──
        salah24 = []
        for r_ in (wk or {}).get("rows", []):
            s_, brk_ = 0, ""
            for c_ in reversed([c for c in (r_.get("cells") or []) if not c.get("is_future")]):
                if c_.get("state") in ("pending", "partial"):
                    brk_ = c_["state"]
                    break
                if c_.get("state") == "done":
                    s_ += 1
            if r_.get("streak") != s_ or r_.get("streak_broken_by") != brk_:
                salah24.append(f"{r_['vendor_name'][:14]} {r_.get('streak')}/"
                               f"{r_.get('streak_broken_by')} != {s_}/{brk_}")
        expect(not salah24, "RK-24",
               "streak = rentetan beruntun TERAKHIR, putus pada hari 'pending' ATAU "
               "'partial', hari tanpa pekerjaan NETRAL (tidak memutus, tidak menambah)",
               "streak dihitung dengan aturan lain ⇒ vendor dipuji/dihukum tanpa dasar "
               "yang bisa dijelaskan ke vendornya",
               {"selisih": salah24[:5]})

        # ── RK-25 — SSOT export mingguan (angka & URUTAN sama dengan layar) ──
        rx = call("get", f"/cmt-override/weekly-recap/export?format=xlsx&date={today_iso}",
                  staff, raw=True)
        rp2 = call("get", f"/cmt-override/weekly-recap/export?format=pdf&date={today_iso}",
                   staff, raw=True)
        xl_ok = rx.status_code == 200 and rx.content[:2] == b"PK"
        same_w, order_w = False, False
        cd_ok = (today_iso.replace("-", "") in (rx.headers.get("Content-Disposition") or ""))
        if xl_ok:
            try:
                import io

                import openpyxl
                ws2 = openpyxl.load_workbook(io.BytesIO(rx.content)).active
                api_names = [r_["vendor_name"] for r_ in (wk or {}).get("rows", [])]
                vals2, seen_names = {}, []
                for row_ in ws2.iter_rows(values_only=True):
                    if row_ and row_[0]:
                        vals2.setdefault(str(row_[0]), row_[1])
                        if str(row_[0]) in api_names:
                            seen_names.append(str(row_[0]))
                same_w = (vals2.get("Vendor aktif") == sm.get("vendors_total")
                          and vals2.get("Total hari terlambat (semua vendor)")
                          == sm.get("days_late_total")
                          and vals2.get("Total pcs disetor sepekan")
                          == sm.get("qty_progress_total"))
                order_w = seen_names == api_names
            except Exception as e:  # noqa: BLE001
                print(f"  {Y}! Excel mingguan tak terbaca: {e}{X}")
        expect(xl_ok and same_w and order_w and cd_ok
               and rp2.status_code == 200 and rp2.content[:5] == b"%PDF-", "RK-25",
               "Excel & PDF mingguan valid, angkanya == API, URUTAN barisnya == layar, "
               "nama berkas menyebut rentangnya",
               "lampiran mingguan menghitung sendiri / mengurutkan sendiri ⇒ rapat "
               "memakai angka yang berbeda dari layar",
               {"xlsx": rx.status_code, "angka_sama": same_w, "urutan_sama": order_w,
                "pdf": rp2.status_code})

        # ── RK-26 — RBAC, header override diabaikan, hari depan tidak dihitung ─
        c_hr1, _ = call("get", "/cmt-override/weekly-recap", hr) if hr else (403, None)
        c_hr2, _ = call("get", "/cmt-override/weekly-recap/export", hr) if hr else (403, None)
        c_no = requests.get(f"{API}/cmt-override/weekly-recap", timeout=30).status_code
        c_ovh, wk_ovh = call("get", "/cmt-override/weekly-recap", staff, vendor=vid)
        fut_end = (d_end + timedelta(days=2)).isoformat()
        _c, wk_fut = call("get", f"/cmt-override/weekly-recap?date={fut_end}", staff)
        fut_cells = [c_ for r_ in (wk_fut or {}).get("rows", [])
                     for c_ in (r_.get("cells") or []) if c_.get("is_future")]
        fut_ok = ((wk_fut or {}).get("summary", {}).get("days_elapsed") == 5
                  and bool(fut_cells)
                  and all(c_.get("state") == "future" and not c_.get("qty_progress")
                          for c_ in fut_cells))
        expect(c_hr1 == 403 and c_hr2 == 403 and c_no in (401, 403)
               and c_ovh == 200
               and len((wk_ovh or {}).get("rows") or []) == len((wk or {}).get("rows") or [])
               and fut_ok, "RK-26",
               "rekap mingguan: role tak berwenang 403 · tanpa token 401 · header "
               "X-CMT-Override-Vendor DIABAIKAN (tetap lintas vendor) · hari yang belum "
               "terjadi ditandai 'future' dan tidak ikut dihitung",
               "kewenangan bocor, atau hari yang belum terjadi dihitung sebagai hari "
               "bolong ⇒ vendor dituduh terlambat untuk hari yang belum datang",
               {"hr": [c_hr1, c_hr2], "tanpa_token": c_no, "override_diabaikan": c_ovh,
                "future_benar": fut_ok})

        # ── RK-27 — KINERJA: 7 hari tidak boleh 7× lebih mahal ──────────────
        t_w0 = datetime.now(timezone.utc)
        call("get", "/cmt-override/weekly-recap", staff)
        ms_w = (datetime.now(timezone.utc) - t_w0).total_seconds() * 1000
        t_d0 = datetime.now(timezone.utc)
        for p in (wk or {}).get("per_day", []):
            if not p.get("is_future"):
                call("get", f"/cmt-override/daily-recap?date={p['date']}", staff)
        ms_d = (datetime.now(timezone.utc) - t_d0).total_seconds() * 1000
        expect(ms_w < 5000 and ms_w <= ms_d * 1.25, "RK-27",
               "rekap mingguan < 5 detik DAN tidak lebih mahal daripada 7× rekap harian "
               "⇒ data master benar-benar dibaca sekali (prefetch_context dipakai ulang)",
               "mingguan membaca master berulang kali ⇒ layar pagi melambat seiring "
               "jumlah vendor bertambah (dan itu tidak akan terlihat di data demo kecil)",
               {"mingguan_ms": round(ms_w), "7x_harian_ms": round(ms_d)})

        # ══════════════════════════════════════════════════════════════════════
        # F12 — PERBANDINGAN ANTAR-PEKAN (RK-31 … RK-36)
        # ══════════════════════════════════════════════════════════════════════
        # Kelas masalah yang dijaga di sini: sebuah panel perbandingan adalah cara
        # paling mudah membuat DUA angka yang berdebat muncul di satu layar.
        # Begitu "pcs pekan ini" di kartu delta berbeda dari "pcs disetor sepekan"
        # di kartu ringkasan — padahal keduanya jendela yang sama — tidak ada lagi
        # yang bisa dipakai mengambil keputusan, dan angka itu dasar tagihan CMT.
        #
        # POLA DUA PEKAN DIBUAT SENDIRI OLEH GATE INI — dan itu bukan kemewahan.
        # Kalau RK-35 hanya memeriksa apa pun yang kebetulan ada di DB, papan
        # peringkat bisa KOSONG dan gate lulus tanpa sekali pun menguji aturannya
        # (persis cara sebuah gate berubah jadi hiasan). Jadi vendor uji dibuat
        # punya pekerjaan di KEDUA jendela, dengan pekan LALU terisi rapi tiap hari
        # dan pekan INI hampir kosong ⇒ arahnya WAJIB 'worse' dengan tambahan hari
        # terlambat yang bisa dihitung tangan.
        db.production_jobs.update_one(
            {"id": job_id},
            {"$set": {"created_at": datetime.now(timezone.utc) - timedelta(days=16)}})
        _cji, jit_f12 = call("get", f"/production-job-items?job_id={job_id}", staff, vendor=vid)
        items_f12 = jit_f12 if isinstance(jit_f12, list) else (jit_f12 or {}).get("items", [])
        f12_item = items_f12[0] if items_f12 else None
        seeded_prev = 0
        if f12_item:
            for off in range(13, 6, -1):          # 13..7 = SELURUH jendela sebelumnya
                day_ = (d_end - timedelta(days=off)).isoformat()
                # Setoran DAN kiriman bertanggal sama — kalau kirimannya tidak ada,
                # hari-hari sesudahnya jadi merah ("pcs selesai belum dikirim") dan
                # pola yang mau diuji rusak.
                cA, _ = call("post", "/production-progress", staff, vendor=vid, body={
                    "job_item_id": f12_item["id"], "progress_date": day_,
                    "completed_quantity": 2, "notes": f"{MARK} setoran pekan lalu"})
                cB, _ = call("post", "/buyer-shipments", staff, vendor=vid, body={
                    "shipment_number": f"{MARK}-SJB-{off}-{uuid.uuid4().hex[:5].upper()}",
                    "job_id": job_id, "po_id": po_id, "shipment_date": day_,
                    "notes": f"{MARK} kirim pekan lalu",
                    "items": [{"po_item_id": f12_item.get("po_item_id"),
                               "sku": f12_item.get("sku", ""),
                               "product_name": f12_item.get("product_name", ""),
                               "size": f12_item.get("size", ""),
                               "color": f12_item.get("color", ""), "qty_shipped": 2}]})
                if cA in (200, 201) and cB in (200, 201):
                    seeded_prev += 1

        # Kedua jendela di-ambil ULANG setelah pola dibuat — kalau RK-31
        # membandingkan `wk` lama dengan `wcmp` baru, ia akan merah karena ulah
        # gate-nya sendiri, bukan karena bug produk.
        t_p0 = datetime.now(timezone.utc)
        c_plain2, wk2 = call("get", "/cmt-override/weekly-recap", staff)
        ms_w2 = (datetime.now(timezone.utc) - t_p0).total_seconds() * 1000
        t_c0 = datetime.now(timezone.utc)
        c_cmp, wcmp = call("get", "/cmt-override/weekly-recap?compare=true", staff)
        ms_cmp = (datetime.now(timezone.utc) - t_c0).total_seconds() * 1000
        cmp_ = (wcmp or {}).get("comparison") or {}

        # ── RK-31 — `compare=true` TIDAK BOLEH menggeser angka jendela ini ────
        keys_now = ["qty_progress_total", "qty_shipped_total", "days_late_total",
                    "days_unfinished_total", "days_no_progress_total", "vendors_late",
                    "vendors_unfinished", "vendors_clean", "vendors_idle",
                    "vendors_total", "days", "days_elapsed", "best_streak"]
        s_plain = (wk2 or {}).get("summary") or {}
        s_cmp = (wcmp or {}).get("summary") or {}
        same_summary = all(s_plain.get(k) == s_cmp.get(k) for k in keys_now)
        same_rows = ([r.get("vendor_id") for r in (wk2 or {}).get("rows", [])]
                     == [r.get("vendor_id") for r in (wcmp or {}).get("rows", [])])
        expect(c_plain2 == 200 and c_cmp == 200 and bool(cmp_)
               and same_summary and same_rows, "RK-31",
               "?compare=true menambahkan blok perbandingan TANPA mengubah satu pun "
               "angka/urutan jendela berjalan (panel tidak menghitung ulang rekapnya)",
               "menyalakan perbandingan menggeser angka jendela berjalan ⇒ kartu delta "
               "dan kartu ringkasan di layar yang SAMA saling berdebat",
               {"http": [c_plain2, c_cmp], "ringkasan_sama": same_summary,
                "urutan_baris_sama": same_rows, "hari_pekan_lalu_diisi": seeded_prev})

        # ── RK-32 — jendela pembanding bersebelahan & sama panjang ───────────
        prev_blk = cmp_.get("previous") or {}
        adj = fair_len = False
        try:
            cur_start = datetime.strptime((wcmp or {}).get("start"), "%Y-%m-%d").date()
            p_start = datetime.strptime(prev_blk.get("start"), "%Y-%m-%d").date()
            p_end = datetime.strptime(prev_blk.get("end"), "%Y-%m-%d").date()
            adj = (p_end == cur_start - timedelta(days=1))
            fair_len = ((p_end - p_start).days == (len((wcmp or {}).get("days") or []) - 1))
        except Exception:  # noqa: BLE001
            pass
        expect(adj and fair_len, "RK-32",
               "jendela pembanding tepat BERSEBELAHAN (berakhir sehari sebelum jendela ini) "
               "dan sama panjang ⇒ tidak ada hari yang dihitung dua kali atau terlewat",
               "jendela pembanding tumpang tindih / panjangnya berbeda ⇒ 'naik/turun' "
               "membandingkan rentang yang tidak setara",
               {"ini_mulai": (wcmp or {}).get("start"), "lalu": f"{prev_blk.get('start')}…{prev_blk.get('end')}"})

        # ── RK-33 — delta = sekarang − sebelumnya, dan arah baik/buruk benar ──
        delta = cmp_.get("delta") or {}
        lower_better = {"days_late_total", "days_unfinished_total", "days_no_progress_total",
                        "vendors_late", "vendors_unfinished", "vendors_idle"}
        salah33 = []
        for k, d in delta.items():
            if round(float(d.get("now", 0)) - float(d.get("prev", 0)), 2) != round(float(d.get("diff", 0)), 2):
                salah33.append(f"{k}: diff bukan now-prev")
            if float(d.get("now", 0)) != float((s_cmp or {}).get(k, 0)):
                salah33.append(f"{k}: 'now' berbeda dari ringkasan jendela ini")
            if float(d.get("prev", 0)) != float((prev_blk.get("summary") or {}).get(k, 0)):
                salah33.append(f"{k}: 'prev' berbeda dari ringkasan pekan lalu")
            if bool(d.get("lower_is_better")) != (k in lower_better):
                salah33.append(f"{k}: lower_is_better salah")
            diff = float(d.get("diff", 0))
            want_better = (diff < 0) if (k in lower_better) else (diff > 0)
            if bool(d.get("better")) != want_better:
                salah33.append(f"{k}: penilaian better/worse salah arah")
        expect(bool(delta) and not salah33, "RK-33",
               "setiap delta = (jendela ini − pekan lalu), angkanya diambil dari kedua "
               "ringkasan, dan 'membaik/memburuk' searah dengan aturan "
               "lower_is_better (hari terlambat NAIK = buruk)",
               "delta dihitung/diberi warna sendiri ⇒ layar bisa menghijaukan angka "
               "yang sebenarnya memburuk",
               {"pelanggaran": salah33[:6], "n_metrik": len(delta)})

        # ── RK-34 — per-vendor: cakupan penuh & selisihnya benar ────────────
        pv = cmp_.get("per_vendor") or []
        cur_ids = [r.get("vendor_id") for r in (wcmp or {}).get("rows", [])]
        by_id = {v.get("vendor_id"): v for v in pv}
        salah34 = []
        if [v.get("vendor_id") for v in pv] != cur_ids:
            salah34.append("cakupan/urutan per_vendor != baris tabel")
        for r in (wcmp or {}).get("rows", []):
            v = by_id.get(r.get("vendor_id"))
            if not v:
                continue
            if v.get("qty_now") != r.get("qty_progress_total"):
                salah34.append(f"{r.get('vendor_code')}: qty_now != tabel")
            if v.get("days_late_now") != r.get("days_late"):
                salah34.append(f"{r.get('vendor_code')}: days_late_now != tabel")
            if v.get("qty_diff") != (v.get("qty_now", 0) - v.get("qty_prev", 0)):
                salah34.append(f"{r.get('vendor_code')}: qty_diff bukan now-prev")
            if v.get("days_late_diff") != (v.get("days_late_now", 0) - v.get("days_late_prev", 0)):
                salah34.append(f"{r.get('vendor_code')}: days_late_diff bukan now-prev")
        expect(bool(pv) and not salah34, "RK-34",
               "baris perbandingan per vendor menutup SEMUA vendor pada urutan yang sama "
               "dengan tabel, dan angka 'sekarang'-nya identik dengan baris tabelnya",
               "kolom 'vs pekan lalu' memakai angka lain daripada baris di sebelahnya ⇒ "
               "satu baris tabel menampilkan dua kebenaran",
               {"pelanggaran": salah34[:6], "n_vendor": len(pv)})

        # ── RK-35 — papan peringkat JUJUR: yang tanpa pekerjaan tidak dinilai ─
        movers = cmp_.get("movers") or {}
        counts = movers.get("counts") or {}
        salah35 = []
        worsened_ids = [v.get("vendor_id") for v in (movers.get("worsened") or [])]
        ranked_ids = set(worsened_ids) | {v.get("vendor_id") for v in (movers.get("improved") or [])}
        for vid_ in ranked_ids:
            v = by_id.get(vid_) or {}
            if v.get("direction") == "incomparable":
                salah35.append(f"{v.get('vendor_code')}: masuk peringkat padahal tak sebanding")
            if int(v.get("days_with_work_prev") or 0) <= 0 or int(v.get("days_with_work_now") or 0) <= 0:
                salah35.append(f"{v.get('vendor_code')}: diperingkat tanpa pekerjaan di salah satu pekan")
        for v in pv:
            if v.get("is_new") and v.get("direction") != "incomparable":
                salah35.append(f"{v.get('vendor_code')}: vendor baru ikut dinilai naik/turun")
            if v.get("direction") == "incomparable" and not v.get("incomparable_reason"):
                salah35.append(f"{v.get('vendor_code')}: dikeluarkan tanpa alasan tertulis")
        # Urutan "paling memburuk" harus benar-benar menurun.
        wl = [int(v.get("days_late_diff") or 0) for v in (movers.get("worsened") or [])]
        if wl != sorted(wl, reverse=True):
            salah35.append("urutan 'memburuk' tidak menurun menurut tambahan hari terlambat")
        if counts and (counts.get("ranked", 0) + counts.get("incomparable", 0)
                       != counts.get("vendors", -1)):
            salah35.append("counts tidak menjumlah: ranked + incomparable != vendors")
        # POLA YANG DIBUAT GATE INI harus benar-benar terbaca — inilah yang membuat
        # RK-35 menguji aturannya, bukan cuma membaca DB apa adanya.
        own = by_id.get(vid) or {}
        if seeded_prev >= 6:
            if int(own.get("days_with_work_prev") or 0) <= 0:
                salah35.append("vendor uji tidak tercatat punya pekerjaan di pekan lalu "
                               "⇒ pola gate tidak terbentuk")
            elif int(own.get("days_late_diff") or 0) <= 0:
                salah35.append(f"vendor uji seharusnya menambah hari terlambat, "
                               f"dapat {own.get('days_late_diff')}")
            elif own.get("direction") != "worse":
                salah35.append(f"vendor uji arahnya {own.get('direction')}, seharusnya 'worse'")
            elif vid not in worsened_ids:
                salah35.append("vendor uji memburuk tapi TIDAK masuk daftar 'paling memburuk'")
        expect(bool(movers) and not salah35, "RK-35",
               "papan 'vendor yang bergerak' hanya memperingkat vendor yang punya "
               "pekerjaan di KEDUA pekan, urut menurun menurut tambahan hari terlambat, "
               "yang dikeluarkan selalu membawa alasan, dan pola dua-pekan buatan gate "
               "ini benar-benar muncul sebagai 'memburuk'",
               "papan peringkat memuji/menuduh vendor karena keputusan order kita sendiri "
               "(pekan lalu tidak diberi pekerjaan ⇒ tampak 'paling membaik') ⇒ teguran "
               "salah sasaran dan daftarnya berhenti dipercaya",
               {"pelanggaran": salah35[:6], "counts": counts,
                "vendor_uji": {k: own.get(k) for k in
                               ("direction", "days_late_prev", "days_late_now",
                                "days_late_diff", "days_with_work_prev", "days_with_work_now")}})

        # ── RK-36 — biaya perbandingan & lampiran ikut membawa perbandingan ──
        rx = call("get", "/cmt-override/weekly-recap/export?format=xlsx&compare=true",
                  staff, raw=True)
        rp = call("get", "/cmt-override/weekly-recap/export?format=pdf&compare=true",
                  staff, raw=True)
        xlsx_has_sheet = False
        try:
            import io as _io
            import zipfile as _zip
            with _zip.ZipFile(_io.BytesIO(rx.content)) as zf:
                wbxml = zf.read("xl/workbook.xml").decode("utf-8", "ignore")
            xlsx_has_sheet = "Perbandingan" in wbxml
        except Exception:  # noqa: BLE001
            pass
        cost_ok = ms_cmp <= max(2500.0, ms_w2 * 2.6)
        expect(rx.status_code == 200 and xlsx_has_sheet
               and rp.status_code == 200 and rp.content[:5] == b"%PDF-" and cost_ok, "RK-36",
               "lampiran Excel/PDF ikut membawa bagian perbandingan (janji 'isinya sama "
               "dengan layar' tetap benar) DAN membangun dua jendela tidak lebih mahal "
               "dari ~2,6× satu jendela (ctx tetap dipakai ulang)",
               "layar punya panel naik/turun tapi lampiran yang dibawa ke rapat tidak — "
               "atau perbandingan membangun ulang seluruh konteks master dua kali",
               {"xlsx": rx.status_code, "lembar_perbandingan": xlsx_has_sheet,
                "pdf": rp.status_code, "compare_ms": round(ms_cmp),
                "satu_jendela_ms": round(ms_w2)})

        # ══════════════════════════════════════════════════════════════════════
        # FASE 5 — `closed_at`: rekap tanggal LAMPAU tidak boleh menebak (RK-28, RK-29)
        # ══════════════════════════════════════════════════════════════════════
        # Kelas masalah yang dijaga:
        #   * **LAPORAN YANG MEMAAFKAN DIRINYA SENDIRI.** Sebelum `closed_at` ada,
        #     "job jalan pada tanggal X" dijawab dari status SEKARANG, sehingga job
        #     yang dibuka Senin dan ditutup Rabu HILANG dari rekap Senin — kelalaian
        #     yang sudah terjadi terhapus sendiri. Progress = dasar tagihan CMT, jadi
        #     laporan seperti itu tidak bisa dipakai memverifikasi apa pun.
        #   * **JALUR TUTUP BARU YANG LUPA MENULIS STEMPEL.** Ada DUA jalur penutup
        #     job (auto-complete + Quick Complete). Kalau nanti ada jalur ketiga yang
        #     tidak memakai `close_job()`, RK-29 akan MERAH.
        #   * **STEMPEL DARI BROWSER.** Pelajaran `received_at`: stempel waktu yang
        #     dipakai laporan wajib ditulis SERVER, bukan diterima dari body.
        # Daftar status penutup diimpor dari SSOT-nya (jangan disalin ke sini —
        # salinan itulah yang suatu hari akan berbeda dari yang dipakai backend).
        # Gate dijalankan dari `/app`, jadi `backend/` perlu ditambahkan ke path.
        if "/app/backend" not in sys.path:
            sys.path.insert(0, "/app/backend")
        from core.production_job_lifecycle import JOB_CLOSED_STATUSES as _CLOSED

        # `created_at` job ditulis server dan tidak bisa dibuat lampau lewat API —
        # jadi SEJARAH-nya dibuat langsung di Mongo. Yang dipalsukan hanya "kapan job
        # lahir", bukan perilaku yang diuji (penutupan tetap lewat HTTP).
        db.production_jobs.update_one(
            {"id": job_id},
            {"$set": {"created_at": datetime.now(timezone.utc) - timedelta(days=2)}})

        jdoc = db.production_jobs.find_one(
            {"id": job_id}, {"_id": 0, "status": 1, "closed_at": 1, "closed_at_estimated": 1})
        if not (jdoc or {}).get("closed_at"):
            # Belum tertutup (item belum penuh) → tutup lewat jalur NORMAL: setor sisa.
            c_, jit2 = call("get", f"/production-job-items?job_id={job_id}", staff, vendor=vid)
            for it in (jit2 if isinstance(jit2, list) else (jit2 or {}).get("items", [])):
                need = int(it.get("shipment_qty", 0) or 0) - int(it.get("produced_qty", 0) or 0)
                if need > 0:
                    call("post", "/production-progress", staff, vendor=vid, body={
                        "job_item_id": it["id"], "progress_date": today_iso,
                        "completed_quantity": need, "notes": f"{MARK} tutup job"})
            jdoc = db.production_jobs.find_one(
                {"id": job_id}, {"_id": 0, "status": 1, "closed_at": 1, "closed_at_estimated": 1})

        closed_ok = isinstance((jdoc or {}).get("closed_at"), datetime)
        _c, rec_y = call("get", f"/cmt-override/daily-recap?date={(d_end - timedelta(days=1)).isoformat()}",
                         staff)
        wait_y = (row_of(rec_y, vid).get("tasks") or {}).get("progress", {}).get("waiting", 0)
        _c, rec_tm = call("get", f"/cmt-override/daily-recap?date={(d_end + timedelta(days=1)).isoformat()}",
                          staff)
        wait_tm = (row_of(rec_tm, vid).get("tasks") or {}).get("progress", {}).get("waiting", 0)
        expect(closed_ok and (jdoc or {}).get("status") in _CLOSED
               and wait_y >= 1 and wait_tm == 0, "RK-28",
               "job yang DITUTUP hari ini tetap terhitung 'job jalan' pada tanggal SEBELUM "
               "penutupan, dan tidak terhitung lagi pada tanggal SESUDAHNYA "
               "(`closed_at` bertipe tanggal & ditulis server)",
               "rekap tanggal lampau memaafkan kelalaian yang sudah terjadi begitu job "
               "ditutup ⇒ progress yang tidak diisi jadi tak terlihat, padahal itu dasar "
               "tagihan CMT",
               {"status": (jdoc or {}).get("status"),
                "closed_at_tipe": type((jdoc or {}).get("closed_at")).__name__,
                "menunggu_kemarin": wait_y, "menunggu_besok": wait_tm})

        # Klien tidak boleh menyuntik `closed_at` saat membuat job.
        c_inj, jinj = call("post", "/production-jobs", staff, vendor=vid, body={
            "vendor_shipment_id": ship_id, "notes": f"{MARK} suntik closed_at",
            "closed_at": "2020-01-01T00:00:00Z", "status": "Completed"})
        inj_id = (jinj or {}).get("id") if isinstance(jinj, dict) else None
        if inj_id:
            ijd = db.production_jobs.find_one({"id": inj_id}, {"_id": 0, "status": 1, "closed_at": 1})
            inj_ok = not (ijd or {}).get("closed_at") and (ijd or {}).get("status") not in _CLOSED
        else:
            inj_ok = c_inj in (400, 409, 422)   # ditolak dengan jelas, bukan 500
        expect(inj_ok, "RK-28b",
               "`closed_at`/`status` kiriman BROWSER diabaikan saat job dibuat — server yang "
               "menentukan (pelajaran bug `received_at`)",
               "klien bisa menentukan kapan job 'ditutup' ⇒ rekap tanggal lampau bisa "
               "dikarang dari luar",
               {"http": c_inj, "job": inj_id})

        # RK-29 — integritas seluruh DB: tidak boleh ada job tertutup tanpa stempel.
        orphan_closed = db.production_jobs.count_documents(
            {"status": {"$in": list(_CLOSED)},
             "$or": [{"closed_at": {"$exists": False}}, {"closed_at": None}]})
        expect(orphan_closed == 0, "RK-29",
               "nol job berstatus tertutup yang tidak punya `closed_at` (semua jalur penutup "
               "memakai `close_job()`, dan job warisan sudah di-backfill)",
               "ada job tertutup tanpa stempel waktu tutup ⇒ rekap tanggal lampau tidak bisa "
               "menghitungnya. Kalau ini job LAMA: jalankan "
               "`python3 migrations/add_closed_at_to_production_jobs.py --execute`. Kalau ini "
               "job BARU: ada jalur penutup yang tidak memakai core.production_job_lifecycle.close_job()",
               {"job_tertutup_tanpa_closed_at": orphan_closed})

        # ── RK-30 — layar tidak boleh DIAM soal batasnya sendiri ─────────────
        # Sesudah fase 5 tinggal SATU hal yang benar-benar tidak diketahui: job
        # WARISAN (tertutup sebelum fitur ini ada, tanpa `closed_at`). Waktu tutupnya
        # sengaja TIDAK ditebak — dan justru karena itu jumlahnya wajib sampai ke
        # layar. Angka yang hanya ada di respons API tapi tak pernah ditampilkan sama
        # saja dengan tidak ada: staf akan membaca rekap tanggal lampau sebagai
        # kebenaran penuh padahal masih ada job yang tak terhitung.
        # Diuji dengan benar-benar MELEPAS stempel job uji (bukan hanya memeriksa
        # nol), karena cabang "ada warisan" itulah yang dipakai pengguna sungguhan.
        # `as_of_note` (dibaca berkas export) dikunci = base + catatan, supaya layar
        # dan lampirannya tidak pernah bercerita beda tentang data yang sama.
        d_cnt = w_cnt = -1
        d_note = w_note = ""
        compose_ok = False
        _cur = db.production_jobs.find_one({"id": job_id}, {"_id": 0, "closed_at": 1}) or {}
        _restore = _cur.get("closed_at")
        try:
            db.production_jobs.update_one({"id": job_id}, {"$unset": {"closed_at": ""}})
            _c, rl = call("get", f"/cmt-override/daily-recap?"
                                 f"date={(d_end - timedelta(days=1)).isoformat()}", staff)
            _c, wl = call("get", "/cmt-override/weekly-recap", staff)
            d_cnt = int((rl or {}).get("legacy_jobs_without_closed_at") or 0)
            w_cnt = int((wl or {}).get("legacy_jobs_without_closed_at") or 0)
            d_note = str((rl or {}).get("legacy_note") or "")
            w_note = str((wl or {}).get("legacy_note") or "")
            base = str((rl or {}).get("as_of_note_base") or "")
            full = str((rl or {}).get("as_of_note") or "")
            compose_ok = full == (f"{base} Catatan: {d_note}" if d_note else base)
        finally:
            # Stempel dipulihkan APA PUN yang terjadi. Tanpa ini, RK-29 pada putaran
            # berikutnya MERAH karena ulah gate-nya sendiri — persis kelas bug
            # "alatnya jadi sumber bug" yang membuat 52 skrip lama dihapus.
            if _restore is not None:
                db.production_jobs.update_one({"id": job_id},
                                              {"$set": {"closed_at": _restore}})
        expect(d_cnt >= 1 and w_cnt == d_cnt and w_note == d_note
               and "add_closed_at_to_production_jobs" in d_note and compose_ok,
               "RK-30",
               "job WARISAN (tertutup tanpa `closed_at`) dilaporkan APA ADANYA ke layar "
               "harian DAN mingguan — jumlah + perintah migrasinya — dan `as_of_note` "
               "(dipakai export) = `as_of_note_base` + `legacy_note`",
               "rekap tanggal lampau menyembunyikan ketidaktahuannya sendiri ⇒ dibaca "
               "sebagai kebenaran penuh padahal ada job yang tidak terhitung; atau layar "
               "dan berkas export mengatakan hal yang berbeda tentang data yang sama",
               {"harian": d_cnt, "mingguan": w_cnt, "catatan_sama": w_note == d_note,
                "as_of_note_gabungan_benar": compose_ok,
                "menyebut_migrasi": "add_closed_at_to_production_jobs" in d_note})



        # ── RK-16 — UANG tidak bergeser ──────────────────────────────────────
        money_after = (call("get", "/production/cmt-billing/summary", admin)[1] or {}) \
            .get("total_amount")
        # PO uji memang menambah tagihan selama uji; yang dijaga adalah tidak ada
        # angka yang bergeser SETELAH cleanup (diperiksa di blok finally).
        print(f"  {Y}·{X} tagihan CMT selama uji: {money_before} → {money_after} "
              f"(dicek ulang setelah cleanup)")

        return 0

    finally:
        # ── CLEANUP ─────────────────────────────────────────────────────────
        print(f"\n{B}-- cleanup --{X}")
        if KEEP:
            print(f"  {Y}--keep: data uji DIBIARKAN{X}")
        else:
            try:
                vids = [v for v in [vid] if v]
                pos = [p for p in [po_id] if p]
                jobs = [j["id"] for j in db.production_jobs.find(
                    {"vendor_id": {"$in": vids}}, {"_id": 0, "id": 1})]
                ships = [s["id"] for s in db.vendor_shipments.find(
                    {"vendor_id": {"$in": vids}}, {"_id": 0, "id": 1})]
                insps = [i["id"] for i in db.vendor_material_inspections.find(
                    {"vendor_id": {"$in": vids}}, {"_id": 0, "id": 1})]
                bss = [b["id"] for b in db.buyer_shipments.find(
                    {"vendor_id": {"$in": vids}}, {"_id": 0, "id": 1})]
                rcpts = [x["id"] for x in db.cmt_receipts.find(
                    {"cmt_vendor_id": {"$in": vids}}, {"_id": 0, "id": 1})]
                ops = [
                    ("production_progress", {"job_id": {"$in": jobs}}),
                    ("production_job_items", {"job_id": {"$in": jobs}}),
                    ("production_jobs", {"id": {"$in": jobs}}),
                    ("vendor_material_inspection_items", {"inspection_id": {"$in": insps}}),
                    ("vendor_material_inspections", {"id": {"$in": insps}}),
                    ("vendor_shipment_items", {"shipment_id": {"$in": ships}}),
                    ("accessory_shipment_items", {"shipment_id": {"$in": ships}}),
                    ("vendor_shipments", {"id": {"$in": ships}}),
                    ("material_requests", {"vendor_id": {"$in": vids}}),
                    ("production_variances", {"vendor_id": {"$in": vids}}),
                    ("buyer_shipment_items", {"shipment_id": {"$in": bss}}),
                    ("buyer_shipments", {"id": {"$in": bss}}),
                    ("cmt_receipt_lines", {"receipt_id": {"$in": rcpts}}),
                    ("cmt_receipts", {"id": {"$in": rcpts}}),
                    ("po_accessories", {"po_id": {"$in": pos}}),
                    ("po_items", {"po_id": {"$in": pos}}),
                    ("production_pos", {"id": {"$in": pos}}),
                    ("dewi_maklon_bom", {"po_id": {"$in": pos}}),
                    # turunan UANG — jangan tinggalkan piutang palsu
                    ("dewi_maklon_pos", {"production_po_id": {"$in": pos}}),
                    ("rahaza_ar_invoices", {"linked_maklon_po_id": {"$in": pos}}),
                    ("dewi_cmt_component_requests", {"vendor_id": {"$in": vids}}),
                    ("dewi_cmt_jobs", {"cmt_partner_id": {"$in": vids}}),
                    ("dewi_cmt_deliveries", {"cmt_partner_id": {"$in": vids}}),
                    ("dewi_cmt_payments", {"cmt_partner_id": {"$in": vids}}),
                    ("reminders", {"vendor_id": {"$in": vids}}),
                    ("reminders", {"created_by": {"$regex": MARK}}),
                    ("vendor_partners", {"id": {"$in": vids}}),
                    ("users", {"id": {"$in": created_users}}),
                    ("activity_logs", {"details": {"$regex": MARK}}),
                    ("notifications", {"$or": [{"title": {"$regex": MARK}},
                                               {"message": {"$regex": MARK}}]}),
                    ("rahaza_audit_logs", {"entity_id": {"$in": pos + jobs + ships}}),
                    ("login_attempts", {"identifier": {"$regex": "rekaptest"}}),
                ]
                n = 0
                for coll, q in ops:
                    try:
                        n += db[coll].delete_many(q).deleted_count
                    except Exception as e:  # noqa: BLE001
                        print(f"  {Y}! gagal bersihkan {coll}: {e}{X}")

                # sweep seluruh koleksi — jaring pengaman terakhir
                swept, where = 0, {}
                for coll in db.list_collection_names():
                    if coll in ("rate_limit_buckets", "counters"):
                        continue
                    try:
                        dead = [d["_id"] for d in db[coll].find({}).limit(20000)
                                if MARK in json.dumps(d, default=str)]
                    except Exception:
                        continue
                    if dead:
                        db[coll].delete_many({"_id": {"$in": dead}})
                        swept += len(dead)
                        where[coll] = len(dead)
                if swept:
                    print(f"  {Y}sweep: {swept} dokumen sisa {where}{X}")
                print(f"  {G}{n + swept} dokumen uji dihapus{X}")

                rest = {}
                for coll in db.list_collection_names():
                    if coll in ("rate_limit_buckets", "counters"):
                        continue
                    try:
                        k = sum(1 for d in db[coll].find({}).limit(20000)
                                if MARK in json.dumps(d, default=str))
                    except Exception:
                        k = 0
                    if k:
                        rest[coll] = k
                expect(not rest, "RK-17", "nol jejak data uji tertinggal di seluruh DB",
                       "gate ini sendiri meninggalkan sampah di database", rest)

                # RK-18 — reminder rekap yang DIBUAT GATE INI tidak boleh tersisa.
                #
                # DULU query ini menghitung SEMUA `daily_recap` bertanggal hari ini
                # tanpa memandang vendornya, dan itu salah dua kali:
                #   1. gate jadi merah karena teguran SAH milik orang lain — staf
                #      yang benar-benar menekan tombol "Kirim reminder" di layar
                #      siang itu sudah cukup untuk menggagalkannya;
                #   2. sejak F12 ada PENJADWAL 16:00 WIB yang mengirim reminder
                #      rekap otomatis setiap hari, gate akan MERAH setiap kali
                #      dijalankan sesudah jam 16:00 — kegagalan yang tidak ada
                #      hubungannya dengan mutu produk. Gate yang merah karena
                #      sebab palsu adalah gate yang mulai diabaikan, dan setelah
                #      itu ia tidak menjaga apa pun.
                # Yang benar-benar dijaga: gate ini sendiri tidak mencemari inbox
                # vendor sungguhan. Jadi lingkupnya vendor uji + jejak MARK.
                leftover = db.reminders.count_documents({
                    "reminder_type": "daily_recap",
                    "$or": [
                        {"vendor_id": {"$in": [v for v in [vid] if v]}},
                        {"message": {"$regex": MARK}},
                        {"subject": {"$regex": MARK}},
                        {"created_by": {"$regex": MARK}},
                    ],
                })
                expect(leftover == 0, "RK-18",
                       "nol reminder rekap sisa DARI GATE INI (inbox vendor sungguhan "
                       "tidak tercemar, dan teguran sah milik orang lain / penjadwal "
                       "16:00 WIB tidak ikut dihitung)",
                       "gate meninggalkan teguran palsu di inbox vendor",
                       {"sisa": leftover})

                adm = login("admin@garment.com", "Admin@123")
                if adm:
                    after = (call("get", "/production/cmt-billing/summary", adm)[1] or {}) \
                        .get("total_amount")
                    orphan_ar = db.rahaza_ar_invoices.count_documents(
                        {"source": "maklon", "linked_maklon_po_id": {"$in": [None, ""]}})
                    expect(orphan_ar == 0, "RK-19",
                           "nol AR invoice maklon yatim (pelajaran gate lama yang membocorkan "
                           "uang palsu)",
                           "ada piutang maklon tanpa induk — uang palsu di laporan",
                           {"orphan_ar": orphan_ar, "tagihan_cmt": after})
            except Exception as e:  # noqa: BLE001
                print(f"  {R}cleanup gagal: {e}{X}")

        print(f"\n{B}{'=' * 78}{X}")
        print(f"  INV-REKAP: {G}{len(PASSES)} OK{X} / {R}{len(FAILS)} FAIL{X}")
        if FAILS:
            print(f"  {R}gagal: {', '.join(FAILS)}{X}")
        print(f"{B}{'=' * 78}{X}")


if __name__ == "__main__":
    rc = main()
    sys.exit(1 if (rc or FAILS) else 0)
