#!/usr/bin/env python3
"""ANALISIS SAJA (read-only) — bagian 2: dekomposisi uang + stabilitas kunci produk."""
import sys
from collections import Counter, defaultdict

import openpyxl

PATH = sys.argv[1] if len(sys.argv) > 1 else "/app/samples/TikTok_UntukDikirim_2026-07-19.xlsx"

wb = openpyxl.load_workbook(PATH, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = [list(r) for r in ws.iter_rows(values_only=True)]
header = [str(h) if h is not None else "" for h in rows[0]]
data = [r for r in rows[2:] if any(c not in (None, "") for c in r)]
I = {h: i for i, h in enumerate(header)}


def num(v):
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("Rp", "").replace(" ", "").strip()
    s = s.replace(".", "") if not ("," in s and "." in s) else s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def g(r, name):
    i = I.get(name)
    return r[i] if i is not None else None


print("=" * 104)
print("A. DEKOMPOSISI 'Order Amount' — apakah itu omzet penjual atau yang dibayar pembeli?")
print("=" * 104)
per_order = defaultdict(lambda: {"amount": None, "sku_after": 0.0, "ship_after": 0.0,
                                 "buyer_fee": None, "handling": None, "pay_disc": None,
                                 "item_ins": None, "ship_ins": None, "rows": 0})
for r in data:
    oid = str(g(r, "Order ID")).strip()
    d = per_order[oid]
    d["rows"] += 1
    d["amount"] = num(g(r, "Order Amount"))
    d["sku_after"] += num(g(r, "SKU Subtotal After Discount"))
    if d["ship_after"] == 0.0:
        d["ship_after"] = num(g(r, "Shipping Fee After Discount"))
    for key, colname in [("buyer_fee", "Buyer Service Fee"), ("handling", "Handling Fee"),
                         ("pay_disc", "Payment platform discount"),
                         ("item_ins", "Item Insurance"), ("ship_ins", "Shipping Insurance")]:
        if d[key] is None:
            d[key] = num(g(r, colname))

tot = {k: 0.0 for k in ["amount", "sku_after", "ship_after", "buyer_fee", "handling", "pay_disc", "item_ins", "ship_ins"]}
for d in per_order.values():
    for k in tot:
        tot[k] += (d[k] or 0.0)

print(f"  Order Amount            (1x per order)  = Rp {tot['amount']:>15,.0f}   <- yang DIBAYAR PEMBELI")
print(f"  SKU Subtotal AfterDisc  (jumlah baris)  = Rp {tot['sku_after']:>15,.0f}   <- omzet PRODUK penjual")
print(f"  Shipping Fee AfterDisc  (1x per order)  = Rp {tot['ship_after']:>15,.0f}   <- ongkir dibayar pembeli")
print(f"  Buyer Service Fee       (1x per order)  = Rp {tot['buyer_fee']:>15,.0f}")
print(f"  Handling Fee            (1x per order)  = Rp {tot['handling']:>15,.0f}")
print(f"  Payment platform disc   (1x per order)  = Rp {tot['pay_disc']:>15,.0f}")
print(f"  Item + Shipping Ins     (1x per order)  = Rp {tot['item_ins'] + tot['ship_ins']:>15,.0f}")
rekon = tot["sku_after"] + tot["ship_after"] + tot["buyer_fee"] + tot["handling"] - tot["pay_disc"]
print(f"  ---")
print(f"  sku_after + ship + buyer_fee + handling - pay_disc = Rp {rekon:,.0f}")
print(f"  selisih vs Order Amount                            = Rp {tot['amount'] - rekon:,.0f}")

exact = sum(1 for d in per_order.values()
            if abs(d["amount"] - (d["sku_after"] + d["ship_after"] + d["buyer_fee"] + d["handling"] - d["pay_disc"])) < 1)
print(f"  pesanan yang rumusnya PAS (selisih <Rp1): {exact}/{len(per_order)}")
print()
print(f"  ==> Kalau 'omzet marketing' pakai Order Amount, angkanya {tot['amount'] - tot['sku_after']:,.0f} "
      f"({(tot['amount'] / tot['sku_after'] - 1) * 100:.1f}%) LEBIH TINGGI dari omzet produk,")
print(f"      karena ikut menghitung ongkir+biaya layanan yang BUKAN pendapatan penjual.")
print()

print("=" * 104)
print("B. STABILITAS KUNCI PRODUK — SKU ID vs Product Name vs Variation")
print("=" * 104)
sku_names = defaultdict(set)
sku_vars = defaultdict(set)
sku_price = defaultdict(set)
namevar_sku = defaultdict(set)
for r in data:
    s = str(g(r, "SKU ID")).strip()
    sku_names[s].add(str(g(r, "Product Name")).strip())
    sku_vars[s].add(str(g(r, "Variation")).strip())
    sku_price[s].add(num(g(r, "SKU Unit Original Price")))
    namevar_sku[(str(g(r, "Product Name")).strip(), str(g(r, "Variation")).strip())].add(s)

print(f"  SKU ID unik                                  : {len(sku_names)}")
print(f"  SKU ID dgn >1 Product Name (nama BERUBAH)    : {sum(1 for v in sku_names.values() if len(v) > 1)}")
print(f"  SKU ID dgn >1 Variation                      : {sum(1 for v in sku_vars.values() if len(v) > 1)}")
print(f"  SKU ID dgn >1 harga satuan                   : {sum(1 for v in sku_price.values() if len(v) > 1)}")
print(f"  pasangan (Nama,Variasi) unik                 : {len(namevar_sku)}")
print(f"  (Nama,Variasi) yang menunjuk >1 SKU ID       : {sum(1 for v in namevar_sku.values() if len(v) > 1)}  <== bahaya kalau cocokkan pakai nama")
print()
print("  Contoh nama BERUBAH pada SKU ID yang sama (bukti nama tidak boleh jadi kunci):")
n = 0
for s, names in sku_names.items():
    if len(names) > 1:
        ns = sorted(names, key=len)
        print(f"    SKU {s}")
        for x in ns:
            print(f"        - {x[-70:]}")
        n += 1
        if n >= 2:
            break
print()
print("  Contoh (Nama,Variasi) SAMA tapi SKU ID BEDA:")
n = 0
for (nm, vr), skus in namevar_sku.items():
    if len(skus) > 1:
        print(f"    {nm[-58:]} / {vr[:34]}  ==> {sorted(skus)}")
        n += 1
        if n >= 3:
            break
print()

print("=" * 104)
print("C. PRODUK INDUK (untuk pemetaan) — 8 nama produk, 83 SKU")
print("=" * 104)
prod = defaultdict(lambda: {"sku": set(), "qty": 0, "rev": 0.0})
for r in data:
    nm = str(g(r, "Product Name")).strip()
    base = nm.split("/")[0].strip()
    p = prod[base]
    p["sku"].add(str(g(r, "SKU ID")).strip())
    p["qty"] += int(num(g(r, "Quantity")))
    p["rev"] += num(g(r, "SKU Subtotal After Discount"))
for k, v in sorted(prod.items(), key=lambda x: -x[1]["rev"]):
    print(f"  {v['qty']:>4} pcs  Rp {v['rev']:>12,.0f}  {len(v['sku']):>3} SKU  {k[:64]}")
print()

print("=" * 104)
print("D. KOLOM RETUR/BATAL — ada di skema yang SAMA (semua kosong di ekspor 'Perlu dikirim')")
print("=" * 104)
for c in ["Cancelation/Return Type", "Cancelled Time", "Cancel By", "Cancel Reason",
          "Order Refund Amount", "Sku Quantity of return", "Shipped Time", "Delivered Time"]:
    if c in I:
        vals = [g(r, c) for r in data]
        filled = sum(1 for v in vals if v not in (None, ""))
        nonzero = sum(1 for v in vals if num(v) != 0) if c in ("Order Refund Amount", "Sku Quantity of return") else "-"
        print(f"  ADA  {c:<28} terisi {filled}/{len(vals)}  bukan-nol={nonzero}")
    else:
        print(f"  TIDAK ADA  {c}")
print()
print("  ==> Skema 65 kolom ini SUDAH memuat kolom kirim, batal, dan retur.")
print("      Artinya ekspor 'Dikirim/Selesai' & 'Batal/Retur' dari menu yang sama = kolom IDENTIK,")
print("      hanya beda BARIS TERISI. Satu peta kolom cukup untuk ketiganya.")
print()

print("=" * 104)
print("E. GEOGRAFI & ATRIBUSI — yang TIDAK disamarkan (nilai tambah gratis)")
print("=" * 104)
for c in ["Province", "Regency and City", "Order Channel", "Creator Handle", "Payment Method", "Product Category"]:
    vals = [str(g(r, c)) for r in data if g(r, c) not in (None, "")]
    print(f"  {c:<20} terisi {len(vals):>4} | unik {len(set(vals)):>4} | top: "
          + ", ".join(f"{k}({v})" for k, v in Counter(vals).most_common(3)))
print()
rev_ch = defaultdict(float)
for r in data:
    rev_ch[str(g(r, "Order Channel"))] += num(g(r, "SKU Subtotal After Discount"))
print("  Omzet produk per Order Channel (bisa otomatis, tanpa input manual):")
for k, v in sorted(rev_ch.items(), key=lambda x: -x[1]):
    print(f"      {k:<16} Rp {v:>13,.0f}  ({v / sum(rev_ch.values()) * 100:.1f}%)")
rev_cr = defaultdict(float)
for r in data:
    ch = g(r, "Creator Handle")
    if ch not in (None, ""):
        rev_cr[str(ch)] += num(g(r, "SKU Subtotal After Discount"))
print(f"\n  Kreator terisi: {len(rev_cr)} handle. Top 5 kontribusi omzet:")
for k, v in sorted(rev_cr.items(), key=lambda x: -x[1])[:5]:
    print(f"      {k:<24} Rp {v:>12,.0f}")
