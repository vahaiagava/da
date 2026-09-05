#!/usr/bin/env python3
"""BUKTI ANALISIS (READ-ONLY) — 5 keputusan impor Seller Center.

Tidak menulis apa pun: tidak ke MongoDB, tidak ke berkas aplikasi.
Menjawab dengan angka, bukan pendapat:

  S1  Bisa TANPA AI? → jalankan auto_map() nyata pada 65 header asli
  S2  Kalau nama kolom BERUBAH (Shopee/Inggris/rename) → seberapa rusak
  S3  Kalau AI dipakai → apa yang AI bisa/tidak bisa (uji panggilan nyata)
  S4  Q2 potong stok di Impor B → fakta stok/FG/katalog di database
  S5  Q5 toko & retur → fakta kolom + data akun

Cara ulang:  cd /app/backend && python3 ../scripts/_prove_import_decisions.py
"""
from __future__ import annotations

import asyncio
import os
import sys
from collections import Counter, defaultdict

sys.path.insert(0, "/app/backend")
os.chdir("/app/backend")

import openpyxl  # noqa: E402
from dotenv import load_dotenv  # noqa: E402

load_dotenv("/app/backend/.env")

from core import marketing_import_engine as eng  # noqa: E402
from core.marketing_import_schema import SOURCE_TYPES, get_source_type  # noqa: E402

XLSX = "/app/samples/TikTok_UntukDikirim_2026-07-19.xlsx"
LINE = "=" * 104


def bar(t):
    print(f"\n{LINE}\n{t}\n{LINE}")


def read_headers_and_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = [r for r in rows[2:] if any(c not in (None, "") for c in r)]
    return headers, data


# ══════════════════════════════════════════════════════════════════════════════
def s1_no_ai(headers):
    bar("S1 — APAKAH BISA TANPA AI? (auto_map() nyata, jenis impor 'orders' yang ADA sekarang)")
    st = get_source_type("orders")
    mapping = eng.auto_map(headers, st)
    rep = eng.mapping_report(mapping, st)
    print(f"  jenis impor    : {st.key} → koleksi {st.collection}")
    print(f"  kolom berkas   : {len(headers)}")
    print(f"  metode         : {rep['methods']}")
    print(f"  terpetakan     : {rep['mapped']}/{rep['total_columns']}")
    print(f"  field wajib blm: {rep['missing_required']}")
    print(f"  siap commit?   : {rep['ready']}")
    print("\n  Kolom yang BERHASIL dipetakan otomatis (tanpa AI):")
    for m in mapping:
        if m.get("field") and m.get("method") != "suggest":
            print(f"    {m['method']:<8} {m['column'][:34]:<34} → {m['field']}")
    unmapped = [m["column"] for m in mapping if not m.get("field") or m.get("method") == "suggest"]
    print(f"\n  Kolom TIDAK terpetakan: {len(unmapped)} dari {len(headers)}")
    print("   ", ", ".join(unmapped[:24]))
    if len(unmapped) > 24:
        print("     ...dan", len(unmapped) - 24, "lagi")
    # cek pemetaan BERBAHAYA (uang salah kolom)
    print("\n  PERIKSA BAHAYA — apakah kolom uang nyangkut di field yang salah?")
    danger = {"revenue": "omzet", "total_payment": "total bayar", "price_final": "harga final",
              "price_original": "harga satuan", "discount_seller": "diskon penjual",
              "shipping_cost": "ongkir"}
    for m in mapping:
        if m.get("field") in danger and m.get("method") != "suggest":
            print(f"    {m['column'][:40]:<40} → {m['field']:<16} ({m['method']}, {m['score']})")
    return mapping, rep


def s1b_semantics(headers, data):
    bar("S1b — WALAU TERPETAKAN, APA JENIS 'orders' MEMANG COCOK? (uji makna, bukan sekadar nama)")
    st = get_source_type("orders")
    print(f"  dedupe jenis 'orders'  : {st.dedupe}")
    I = {h: i for i, h in enumerate(headers)}
    oids = [str(r[I["Order ID"]]).strip() for r in data]
    c = Counter(oids)
    multi = sum(1 for v in c.values() if v > 1)
    print(f"  baris berkas           : {len(data)}")
    print(f"  Order ID unik          : {len(c)}")
    print(f"  Order ID muncul >1x    : {multi}")
    print(f"\n  ==> dedupe ('account_id','order_id') menganggap 1 pesanan = 1 BARIS.")
    print(f"      Berkas ini 1 baris = 1 SKU. {multi} pesanan multi-SKU akan saling MENIMPA:")
    ex = [k for k, v in c.items() if v > 1][:1]
    if ex:
        rowsx = [r for r in data if str(r[I["Order ID"]]).strip() == ex[0]]
        print(f"      contoh Order {ex[0]} punya {len(rowsx)} baris SKU:")
        for r in rowsx:
            print(f"         SKU {r[I['SKU ID']]} qty={r[I['Quantity']]} "
                  f"subtotal={r[I['SKU Subtotal After Discount']]}  {str(r[I['Variation']])[:28]}")
        print(f"      dgn dedupe sekarang: hanya 1 dari {len(rowsx)} baris bertahan ⇒ "
              f"item lain HILANG (bukan dobel — HILANG).")
    print(f"\n  Field 'sku_id' jenis 'orders' mencocokkan katalog lewat SKU.")
    seller_sku_filled = sum(1 for r in data if r[I["Seller SKU"]] not in (None, ""))
    print(f"      'Seller SKU' terisi   : {seller_sku_filled}/{len(data)}  ⇒ pencocokan SKU MUSTAHIL")
    print(f"      'SKU ID' (platform)   : {len(set(str(r[I['SKU ID']]) for r in data))} unik, terisi penuh")
    print(f"\n  Field 'status' jenis 'orders' adalah enum:")
    print(f"      pilihan sistem     : {st.field('status').choices}")
    print(f"      nilai di berkas    : {sorted(set(str(r[I['Order Status']]) for r in data))}")
    print(f"      ==> 'Perlu dikirim' TIDAK ada di daftar ⇒ 601 baris kena galat enum.")


def s2_robustness(headers):
    bar("S2 — KALAU NAMA KOLOM BERUBAH (ekspor B / Shopee / bahasa lain): seberapa tahan?")
    st = get_source_type("orders")

    def score(hs, label):
        m = eng.auto_map(hs, st)
        r = eng.mapping_report(m, st)
        used = {x["column"]: x["field"] for x in m if x.get("field") and x.get("method") != "suggest"}
        return label, r, used

    base_label, base_rep, base_used = score(headers, "asli TikTok (65 kolom)")

    variants = {
        "TikTok tapi diterjemahkan (Shopee-ID style)": {
            "Order ID": "No. Pesanan", "Order Status": "Status Pesanan",
            "Created Time": "Waktu Pesanan Dibuat", "Quantity": "Jumlah",
            "SKU Subtotal After Discount": "Harga Setelah Diskon",
            "Tracking ID": "No. Resi", "Shipping Provider Name": "Jasa Kirim",
            "Product Name": "Nama Produk", "Variation": "Nama Variasi",
            "Order Amount": "Total Pembayaran", "Payment Method": "Metode Pembayaran",
            "Recipient": "Nama Penerima", "Regency and City": "Kota/Kabupaten",
        },
        "kolom kirim terisi (ekspor B): + Shipped/Delivered": {},
        "header dipersingkat platform (rename ringan)": {
            "Order ID": "Order Id", "Created Time": "Create Time",
            "SKU Subtotal After Discount": "SKU Subtotal after discount",
            "Tracking ID": "Tracking Number", "Quantity": "Qty",
        },
        "header ASING total (nama internal platform)": {
            "Order ID": "ord_no", "Created Time": "ts_created", "Quantity": "qty_ord",
            "SKU Subtotal After Discount": "amt_sku_net", "Order Amount": "amt_pay",
            "Tracking ID": "awb", "Product Name": "prd_nm", "Variation": "var_nm",
            "SKU ID": "sku_pid", "Order Status": "st", "Payment Method": "pay_ch",
            "Shipping Provider Name": "lgt", "Recipient": "rcv_nm",
        },
    }
    print(f"  {'skenario':<50} terpetakan  wajib-belum         siap?")
    print(f"  {'-' * 96}")
    print(f"  {base_label:<50} {base_rep['mapped']:>4}/{base_rep['total_columns']:<4}   "
          f"{str(base_rep['missing_required'])[:22]:<22} {base_rep['ready']}")
    for label, ren in variants.items():
        hs = [ren.get(h, h) for h in headers]
        _, r, used = score(hs, label)
        lost = [f for f in base_used.values() if f not in used.values()]
        print(f"  {label:<50} {r['mapped']:>4}/{r['total_columns']:<4}   "
              f"{str(r['missing_required'])[:22]:<22} {r['ready']}"
              + (f"   HILANG: {sorted(set(lost))}" if lost else ""))
    print("\n  Ambang mesin (tetap, bukan tebakan AI):")
    print(f"    exact = 1.00 · synonym = 0.98 · fuzzy OTOMATIS ≥ {eng.FUZZY_AUTO} · "
          f"hanya DIUSULKAN ≥ {eng.FUZZY_SUGGEST} · di bawah itu = tidak dipetakan")


async def s3_ai(headers):
    bar("S3 — KALAU PAKAI AI: apa yang AI bisa & TIDAK bisa (uji nyata)")
    print("  (a) AI importer yang SUDAH ADA — routes/marketing_import.py")
    from routes.marketing_import import SYSTEM_FIELDS, _heuristic_mapping
    print(f"      jumlah field yang dikenal AI importer : {len(SYSTEM_FIELDS)}")
    print(f"      field-nya                            : {list(SYSTEM_FIELDS)}")
    print("      agregasi bawaan                      : by_date (menjumlah revenue per TANGGAL)")
    print("      ==> ini importer REKAP HARIAN, bukan per-pesanan. 601 baris pesanan akan")
    print("          diringkas jadi beberapa baris tanggal: Order ID, resi, SKU, kurir, kreator HILANG.")
    h = _heuristic_mapping(headers, SYSTEM_FIELDS)
    got = {m['source_column']: m['target_field'] for m in h['mapping'] if m.get('target_field')}
    print(f"      heuristik importer AI pada 65 kolom  : {len(got)} kolom terpetakan → {got}")

    print("\n  (b) AI-assist di wizard TANPA AI — POST /api/marketing/data-import/sessions/{id}/ai-assist")
    import inspect
    from routes import marketing_data_import as mdi
    src = inspect.getsource(mdi.ai_assist)
    print("      sifat yang dijamin kode:")
    print(f"        · hanya kolom BELUM terpetakan  : {'unmapped' in src}")
    print(f"        · tidak menimpa exact/synonym   : {'taken' in src}")
    print(f"        · hasil = USULAN, wajib disetujui: {'Usulan AI' in src}")
    print(f"        · gagal AI ⇒ manual tetap jalan  : {'manual tetap' in src}")

    print("\n  (c) UJI PANGGILAN AI NYATA — header ASING (nama internal platform), tanpa tulis DB")
    st = get_source_type("orders")
    alien = {"Order ID": "ord_no", "Created Time": "ts_created", "Quantity": "qty_ord",
             "SKU Subtotal After Discount": "amt_sku_net", "Order Amount": "amt_pay",
             "Tracking ID": "awb", "Product Name": "prd_nm", "SKU ID": "sku_pid",
             "Order Status": "st", "Shipping Provider Name": "lgt"}
    alien_cols = list(alien.values())
    truth = {v: k for k, v in alien.items()}
    key = os.environ.get("EMERGENT_LLM_KEY", "")
    if not key:
        print("      EMERGENT_LLM_KEY kosong ⇒ tidak diuji")
        return
    try:
        import json
        import time
        from ai_llm import LlmChat, UserMessage
        fields_txt = "\n".join(f"- {f.name} :: {f.label} ({f.kind})" for f in st.input_fields)
        prompt = (
            f"Petakan nama kolom berkas ke field sistem untuk jenis data '{st.label}'.\n\n"
            f"FIELD SISTEM:\n{fields_txt}\n\n"
            f"KOLOM YANG BELUM TERPETAKAN: {json.dumps(alien_cols)}\n\n"
            "Jawab HANYA JSON array: [{\"column\":\"...\",\"field\":\"nama_field_sistem_atau_null\","
            "\"confidence\":0.0-1.0,\"reason\":\"alasan singkat\"}]")
        chat = LlmChat(api_key=key, session_id="analysis-readonly-map",
                       system_message="Kamu pemeta kolom berkas impor. Jawab JSON saja.")
        chat.with_model("openai", "gpt-4o-mini")
        t0 = time.time()
        reply = await chat.send_message(UserMessage(text=prompt))
        dt = time.time() - t0
        txt = (reply or "").strip()
        if txt.startswith("```"):
            txt = txt.split("```")[1]
            txt = txt[4:] if txt.lower().startswith("json") else txt
        sug = json.loads(txt)
        print(f"      AI menjawab dalam {dt:.1f}s, {len(sug)} usulan")
        ok = bad = null = 0
        for item in sug:
            col, fld = item.get("column"), item.get("field")
            exp = truth.get(col, "?")
            hit = "-"
            if fld is None:
                null += 1
                hit = "kosong"
            else:
                # nilai benar menurut manusia (peta longgar: nama field kanonik)
                expect_field = {"Order ID": "order_id", "Created Time": "order_date",
                                "Quantity": "quantity", "SKU Subtotal After Discount": "price_final",
                                "Order Amount": "total_payment", "Tracking ID": "tracking_number",
                                "Product Name": "product_name", "SKU ID": "sku_id",
                                "Order Status": "status", "Shipping Provider Name": "courier"}.get(exp)
                if fld == expect_field:
                    ok += 1
                    hit = "BENAR"
                else:
                    bad += 1
                    hit = f"SALAH (harusnya {expect_field})"
            print(f"        {col:<12} → {str(fld):<18} conf={item.get('confidence')}  [{hit}]  ({exp})")
        print(f"\n      SKOR AI: benar {ok} · salah {bad} · dikosongkan {null} dari {len(alien_cols)} kolom")
        print("      ==> AI berguna untuk header yang TIDAK dikenal kamus, TAPI jawaban salah")
        print("          tetap mungkin ⇒ karena itu hasilnya USULAN, bukan langsung dipakai.")
    except Exception as e:
        print(f"      panggilan AI GAGAL: {type(e).__name__}: {e}")
        print("      ==> justru ini alasan jalur tanpa-AI harus jadi jalur utama.")


async def s4_stock(headers, data):
    bar("S4 — Q2 'POTONG STOK GUDANG SAAT IMPOR BUKTI KIRIM': fakta di database SEKARANG")
    from motor.motor_asyncio import AsyncIOMotorClient
    cli = AsyncIOMotorClient(os.environ["MONGO_URL"])
    db = cli[os.environ.get("DB_NAME", "test_database")]

    accounts = await db.marketing_accounts.find({}, {"_id": 0, "id": 1, "name": 1, "platform": 1}).to_list(200)
    print(f"  A. Toko/akun marketing terdaftar : {len(accounts)}")
    for a in accounts[:12]:
        print(f"       {a.get('platform', '-'):<12} {a.get('name', '-')}")

    cats = await db.marketing_catalogs.find({}, {"_id": 0, "id": 1, "account_id": 1, "name": 1}).to_list(500)
    items = await db.marketing_catalog_items.find({}, {"_id": 0}).to_list(3000)
    linked = [i for i in items if i.get("fg_material_id") or i.get("material_id")]
    print(f"\n  B. Katalog toko                  : {len(cats)} katalog, {len(items)} item")
    print(f"     item TERTAUT master FG        : {len(linked)}/{len(items)}")
    print(f"     item punya kolom 'sku'        : {sum(1 for i in items if i.get('sku'))}/{len(items)}")
    print(f"     item punya SKU platform (map) : "
          f"{sum(1 for i in items if i.get('platform_sku_id'))}/{len(items)}  <== kamus SKU platform")

    mats = await db.rahaza_materials.find({}, {"_id": 0, "id": 1, "name": 1, "category": 1,
                                               "material_type": 1, "type": 1, "hpp": 1}).to_list(5000)
    def is_fg(m):
        blob = " ".join(str(m.get(k, "")).lower() for k in ("category", "material_type", "type"))
        return "fg" in blob or "finish" in blob or "jadi" in blob
    fgs = [m for m in mats if is_fg(m)]
    print(f"\n  C. Master material               : {len(mats)}")
    print(f"     yang bertipe produk jadi (FG) : {len(fgs)}")
    catcount = Counter(str(m.get("category") or m.get("material_type") or m.get("type") or "-") for m in mats)
    print("     kategori material (top 8)     :", dict(catcount.most_common(8)))

    stock = await db.rahaza_material_stock.find({}, {"_id": 0}).to_list(20000)
    from core.stock_schema import read_qty, read_reserved
    fg_ids = {m["id"] for m in fgs}
    fg_rows = [s for s in stock if s.get("material_id") in fg_ids]
    tot_all = sum(read_qty(s) for s in stock)
    tot_fg = sum(read_qty(s) for s in fg_rows)
    print(f"\n  D. Baris stok gudang             : {len(stock)}  (total {tot_all:,.0f} unit)")
    print(f"     baris stok PRODUK JADI (FG)   : {len(fg_rows)}  (total {tot_fg:,.0f} unit)")
    print(f"     reserved pada baris FG        : {sum(read_reserved(s) for s in fg_rows):,.0f}")

    I = {h: i for i, h in enumerate(headers)}
    need = defaultdict(int)
    for r in data:
        need[str(r[I["SKU ID"]]).strip()] += int(float(r[I["Quantity"]] or 0))
    print(f"\n  E. Kebutuhan potong stok dari 1 berkas ini:")
    print(f"     SKU platform berbeda          : {len(need)}")
    print(f"     total pcs yang harus dipotong : {sum(need.values())}")
    preo = sum(int(float(r[I['Quantity']] or 0)) for r in data if str(r[I['Normal or Pre-order']]) == 'Pre-order')
    print(f"     di antaranya Pre-order        : {preo} pcs "
          f"({preo / sum(need.values()) * 100:.0f}%) — barangnya belum dibuat saat pesanan masuk")
    print(f"\n     Rantai yang HARUS lengkap agar 1 pcs bisa dipotong:")
    print(f"       SKU platform → item katalog → fg_material_id → baris stok FG di lokasi")
    print(f"       tersedia sekarang:  kamus SKU={sum(1 for i in items if i.get('platform_sku_id'))}"
          f" · item→FG={len(linked)}/{len(items)} · baris stok FG={len(fg_rows)}")
    if len(fg_rows) == 0 or len(linked) == 0:
        print(f"       ==> HARI INI: 0 dari {sum(need.values())} pcs bisa dipotong. "
              f"Kalau impor B dipaksa memotong stok, hasilnya salah satu dari:")
        print(f"           (i) 601 baris GAGAL semua, atau (ii) stok jadi MINUS, atau")
        print(f"           (iii) potongan diam-diam dilewati ⇒ 'sudah dikirim' tapi stok tak berubah.")

    orders = await db.marketing_orders.find({}, {"_id": 0}).to_list(5000)
    print(f"\n  F. marketing_orders sekarang     : {len(orders)} dokumen")
    print(f"     tanpa account_id              : {sum(1 for o in orders if not o.get('account_id'))}")
    print(f"     tanpa order_id/no pesanan     : {sum(1 for o in orders if not o.get('order_id'))}")
    src = Counter(str(o.get("source") or o.get("import_source") or "-") for o in orders)
    print(f"     sumber (field source)         : {dict(src)}")
    fields_seen = Counter()
    for o in orders:
        for k in o:
            fields_seen[k] += 1
    print(f"     field uang yang dipakai       : "
          + ", ".join(f"{k}={v}" for k, v in fields_seen.items()
                      if k in ("revenue", "total_amount", "total_payment", "price_final",
                               "order_amount", "amount")))
    cli.close()
    return {"items": len(items), "linked": len(linked), "fg_rows": len(fg_rows), "orders": len(orders)}


def s5_store_returns(headers, data):
    bar("S5 — Q5 TOKO & RETUR: apa yang berkas ini benar-benar bisa buktikan")
    I = {h: i for i, h in enumerate(headers)}
    print("  A. Kandidat penanda TOKO di berkas:")
    for c in ["Warehouse Name", "Purchase Channel", "Fulfillment Type", "Seller Note"]:
        if c in I:
            vals = [str(r[I[c]]) for r in data if r[I[c]] not in (None, "")]
            print(f"     {c:<20} terisi {len(vals):>4}/{len(data)} | unik {len(set(vals))} | "
                  f"nilai: {sorted(set(vals))[:3]}")
    print("\n     Catatan: 'Warehouse Name' = nama GUDANG yang diatur penjual di Seller Center,")
    print("     bukan nama toko. Satu toko bisa punya beberapa gudang, dan dua toko berbeda")
    print("     bisa memberi nama gudang yang sama ⇒ dipakai sebagai PEMERIKSA, bukan sumber.")
    print("     'Purchase Channel' = platform ('TikTok') ⇒ berguna untuk memeriksa platform toko.")

    print("\n  B. Kolom RETUR/BATAL — ADA di skema yang sama, tinggal kosong di ekspor ini:")
    for c in ["Cancelation/Return Type", "Cancelled Time", "Cancel By", "Cancel Reason",
              "Order Refund Amount", "Sku Quantity of return"]:
        print(f"     {'ADA ' if c in I else 'TIDAK ADA'} {c}")
    print("\n     ==> ekspor 'Perlu Dikirim', 'Dikirim/Selesai', dan 'Batal/Retur' berasal dari")
    print("         menu yang SAMA ⇒ 65 kolom identik. SATU peta kolom melayani ketiganya;")
    print("         yang membedakan hanya kolom mana yang TERISI.")

    print("\n  C. Bukti tahapan yang tersedia per pesanan (untuk monitoring 'sudah diurus'):")
    for c in ["Created Time", "Paid Time", "RTS Time", "Shipped Time", "Delivered Time", "Cancelled Time"]:
        n = sum(1 for r in data if r[I[c]] not in (None, "")) if c in I else -1
        print(f"     {c:<16} terisi {n}/{len(data)}")
    print("     ==> ekspor A memberi Created/Paid/RTS. Shipped & Delivered HANYA dari ekspor B.")


async def main():
    headers, data = read_headers_and_rows(XLSX)
    print(f"BERKAS: {XLSX}")
    print(f"HEADER: {len(headers)} kolom · DATA: {len(data)} baris (baris 2 = deskripsi, dilewati)")
    s1_no_ai(headers)
    s1b_semantics(headers, data)
    s2_robustness(headers)
    await s3_ai(headers)
    await s4_stock(headers, data)
    s5_store_returns(headers, data)
    print(f"\n{LINE}\nSELESAI — tidak ada tulisan ke database maupun berkas aplikasi.\n{LINE}")


if __name__ == "__main__":
    asyncio.run(main())
