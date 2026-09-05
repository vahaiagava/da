"""
Importer: DATA_PRODUKSI_MAKLOON_SPLIT_3.xlsx  →  ERP CV. Dewi Aditya.

Ruang lingkup (disetujui user):
  - Sheet "Produksi Internal" (INVOICE DA) -> production_pos business_type=internal
  - Sheet "Produksi Maklon"  (INVOICE AE) -> production_pos business_type=maklon
  - No PO      -> production_pos.po_number
  - No Invoice -> po_items.serial_number (SN); 1 PO punya banyak SN
  - Snapshot (tanpa rincian tiap setor): 1 vendor_shipment (potongan dikirim) +
    1 cmt_receipt Approved (rekap setoran) per PO.
  - Auto-create master yang belum ada (CMT + produk). CMT kosong -> placeholder.
  - TIDAK memicu efek samping finance/BOM/FG-stock (insert langsung ke koleksi kanonik
    dengan bentuk yang dibaca services.cmt_kejar / cmt_intake / ProductionPOModule).

Idempoten: semua dokumen yang dibuat ditandai import_source=IMPORT_TAG dan dihapus
lebih dulu saat re-run. Master yang dibuat import juga ditandai & hanya yang bertanda
itu yang dibersihkan (master seed asli tidak disentuh).
"""
import asyncio
import os
import sys
import uuid
from datetime import datetime, timezone

import openpyxl
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient

load_dotenv("/app/backend/.env")

XLSX = "/app/data_import/DATA_PRODUKSI_MAKLOON_SPLIT_3.xlsx"
IMPORT_TAG = "excel_produksi_maklon_v1"
PLACEHOLDER_CMT_NAME = "(Belum Ditentukan)"

NOW = datetime.now(timezone.utc)
NOW_ISO = NOW.isoformat()


def nid():
    return str(uuid.uuid4())


def num(v):
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def txt(v):
    return "" if v is None else str(v).strip()


def as_dt(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def parse_sheet(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[2]
    idx = {h: i for i, h in enumerate(hdr) if h}
    out = []
    for r in rows[3:]:
        if all(c is None for c in r):
            continue
        d = {k: r[i] for k, i in idx.items()}
        if not txt(d.get("No Invoice")):
            continue
        out.append(d)
    return out


def parse_daftar_cmt(wb):
    ws = wb["daftar CMT"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[2]
    idx = {h: i for i, h in enumerate(hdr) if h}
    out = []
    for r in rows[3:]:
        if all(c is None for c in r):
            continue
        d = {k: r[i] for k, i in idx.items()}
        if not txt(d.get("Nama CMT")):
            continue
        out.append(d)
    return out


async def clean_previous(db):
    """Hapus artefak import sebelumnya (idempoten)."""
    tx = {"import_source": IMPORT_TAG}
    cols = [
        "production_pos", "po_items", "vendor_shipments", "vendor_shipment_items",
        "vendor_material_inspections", "vendor_material_inspection_items",
        "cmt_receipts", "cmt_receipt_lines",
        # master yang dibuat oleh import (hanya yang bertanda import ini):
        "vendor_partners", "rahaza_models", "dewi_maklon_clients",
    ]
    for c in cols:
        res = await db[c].delete_many(tx)
        if res.deleted_count:
            print(f"  cleaned {c}: {res.deleted_count}")


async def build_vendor_map(db, wb, prod_rows):
    """Kembalikan dict {UPPER(name): vendor_doc}. Buat yang belum ada + placeholder."""
    existing = {}
    async for v in db.vendor_partners.find({}, {"_id": 0}):
        existing[txt(v.get("name")).upper()] = v

    created = 0
    updated = 0

    # 1) Master resmi dari sheet "daftar CMT"
    for d in parse_daftar_cmt(wb):
        name = txt(d.get("Nama CMT"))
        key = name.upper()
        status = txt(d.get("Status")).lower()
        is_active = not status.startswith("non")
        cap = num(d.get("Kapasitas (pcs)"))
        loc = txt(d.get("Lokasi"))
        contact = txt(d.get("Kontak / HP"))
        code = txt(d.get("Kode CMT")) or name
        if key not in existing:
            doc = {
                "id": nid(), "name": name, "code": code,
                "contact_name": "", "contact_phone": contact, "address": loc,
                "notes": "Auto dari import produksi (daftar CMT)",
                "is_active": is_active, "capacity_pcs": cap, "capacity_note": "",
                "import_source": IMPORT_TAG, "created_at": NOW, "created_by": "import",
            }
            await db.vendor_partners.insert_one(doc)
            existing[key] = doc
            created += 1
        else:
            # sinkronkan status aktif + isi kapasitas/lokasi bila kosong (jangan timpa data baik)
            v = existing[key]
            patch = {}
            if v.get("is_active") != is_active:
                patch["is_active"] = is_active
            if not num(v.get("capacity_pcs")) and cap:
                patch["capacity_pcs"] = cap
            if not txt(v.get("address")) and loc:
                patch["address"] = loc
            if patch:
                await db.vendor_partners.update_one({"id": v["id"]}, {"$set": patch})
                updated += 1

    # 2) CMT yang muncul di data produksi tapi belum ada master (mis. P Aan, P Suratno)
    for d in prod_rows:
        name = txt(d.get("Nama CMT"))
        if not name:
            continue
        key = name.upper()
        if key not in existing:
            doc = {
                "id": nid(), "name": name, "code": name,
                "contact_name": "", "contact_phone": "", "address": "",
                "notes": "Auto dari import produksi (penjahit belum ada di master — mohon lengkapi)",
                "is_active": True, "capacity_pcs": 0, "capacity_note": "",
                "import_source": IMPORT_TAG, "created_at": NOW, "created_by": "import",
            }
            await db.vendor_partners.insert_one(doc)
            existing[key] = doc
            created += 1

    # 3) Placeholder untuk baris tanpa CMT (Belum Kirim)
    if PLACEHOLDER_CMT_NAME.upper() not in existing:
        doc = {
            "id": nid(), "name": PLACEHOLDER_CMT_NAME, "code": "CMT-PLACEHOLDER",
            "contact_name": "", "contact_phone": "", "address": "",
            "notes": "Placeholder — CMT belum ditentukan. Silakan edit & pindahkan order.",
            "is_active": True, "capacity_pcs": 0, "capacity_note": "",
            "import_source": IMPORT_TAG, "created_at": NOW, "created_by": "import",
        }
        await db.vendor_partners.insert_one(doc)
        existing[PLACEHOLDER_CMT_NAME.upper()] = doc

    print(f"  vendor_partners: created={created} updated={updated} total_map={len(existing)}")
    return existing


async def build_product_map(db, prod_rows):
    """Peta {sku_upper: model_id}. Buat rahaza_models untuk SKU yang belum ada di master."""
    model_by_code = {}
    async for m in db.rahaza_models.find({}, {"_id": 0, "id": 1, "code": 1}):
        c = txt(m.get("code")).upper()
        if c:
            model_by_code[c] = m["id"]
    mat_codes = set()
    async for m in db.rahaza_materials.find({}, {"_id": 0, "code": 1, "sku": 1}):
        if m.get("code"):
            mat_codes.add(txt(m["code"]).upper())
        if m.get("sku"):
            mat_codes.add(txt(m["sku"]).upper())

    # SKU -> nama produk (ambil nama pertama yang muncul)
    sku_name = {}
    for d in prod_rows:
        sku = txt(d.get("SKU"))
        if sku and sku.upper() not in sku_name:
            sku_name[sku.upper()] = (sku, txt(d.get("Nama Produk")))

    created = 0
    for up, (sku, pname) in sku_name.items():
        if up in model_by_code:
            continue
        if up in mat_codes:
            # sudah ada di material master (FG/kain) — tidak perlu buat model, biarkan sku teks
            continue
        doc = {
            "id": nid(), "code": sku, "name": pname or sku, "category": "IMPORT PRODUKSI",
            "material_kg_per_pcs": 0, "bundle_size": 0,
            "description": "Auto dari import produksi (produk belum ada di master)",
            "cmt_cost_per_pcs": 0, "sop_steps": [], "reference_videos": [],
            "reference_images": [], "source": "excel_produksi_import", "active": True,
            "spec": {}, "import_source": IMPORT_TAG,
            "created_at": NOW, "updated_at": NOW,
        }
        await db.rahaza_models.insert_one(doc)
        model_by_code[up] = doc["id"]
        created += 1
    print(f"  rahaza_models: created={created} (missing products)")
    return model_by_code


async def ensure_maklon_client(db, name):
    c = await db.dewi_maklon_clients.find_one({"name": name}, {"_id": 0, "id": 1})
    if c:
        return c["id"]
    doc = {
        "id": nid(), "name": name, "code": "KOHTRI", "contact_name": "", "phone": "",
        "address": "", "notes": "Auto dari import produksi maklon",
        "is_active": True, "import_source": IMPORT_TAG,
        "created_at": NOW, "updated_at": NOW,
    }
    await db.dewi_maklon_clients.insert_one(doc)
    return doc["id"]


def derive_po_status(items):
    """internal: semua Lunas->Completed; ada kirim & belum semua lunas->In Production;
    semua Belum Kirim->Draft."""
    statuses = [txt(i.get("Status")).lower() for i in items]
    any_dispatch = any(i.get("Tgl Kirim ke CMT") for i in items)
    all_lunas = all(s == "lunas" for s in statuses) and statuses
    if all_lunas:
        return "Completed"
    if not any_dispatch and all(s in ("belum kirim", "") for s in statuses):
        return "Draft"
    return "In Production"


async def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    internal = parse_sheet(wb, "Produksi Internal")
    maklon = parse_sheet(wb, "Produksi Maklon")
    all_rows = internal + maklon

    db = AsyncIOMotorClient(os.environ["MONGO_URL"])[os.environ["DB_NAME"]]

    print("== CLEAN previous import ==")
    await clean_previous(db)

    print("== MASTER data ==")
    vmap = await build_vendor_map(db, wb, all_rows)
    pmap = await build_product_map(db, all_rows)
    kohtri_id = await ensure_maklon_client(db, "Koh Tri (SnBM)")

    def vendor_for(name):
        v = vmap.get(txt(name).upper()) if txt(name) else vmap[PLACEHOLDER_CMT_NAME.upper()]
        return v or vmap[PLACEHOLDER_CMT_NAME.upper()]

    stats = dict(pos=0, items=0, shipments=0, ship_items=0, receipts=0, receipt_lines=0,
                 inspections=0)

    for sheet_rows, btype in ((internal, "internal"), (maklon, "maklon")):
        # group by No PO
        groups = {}
        for d in sheet_rows:
            groups.setdefault(txt(d.get("No PO")), []).append(d)

        for po_number, items in groups.items():
            if not po_number:
                continue
            vname = next((txt(i.get("Nama CMT")) for i in items if txt(i.get("Nama CMT"))), "")
            vendor = vendor_for(vname)
            deadlines = [as_dt(i.get("Deadline Mitra")) for i in items if as_dt(i.get("Deadline Mitra"))]
            ship_dates = [as_dt(i.get("Tgl Kirim ke CMT")) for i in items if as_dt(i.get("Tgl Kirim ke CMT"))]
            delivery_deadline = min(deadlines) if deadlines else None
            po_id = nid()
            customer_name = "Koh Tri (SnBM)" if btype == "maklon" else "DA Group (Internal)"
            po = {
                "id": po_id, "po_number": po_number, "customer_name": customer_name,
                "buyer_id": kohtri_id if btype == "maklon" else None,
                "vendor_id": vendor["id"], "vendor_name": vendor.get("name", ""),
                "po_date": min(ship_dates) if ship_dates else NOW,
                "deadline": delivery_deadline, "delivery_deadline": delivery_deadline,
                "status": derive_po_status(items), "notes": "Impor dari Excel produksi",
                "business_type": btype,
                "import_source": IMPORT_TAG,
                "created_by": "import", "created_at": NOW, "updated_at": NOW,
            }
            await db.production_pos.insert_one(po)
            stats["pos"] += 1

            ship_item_docs = []
            receipt_line_docs = []
            for d in items:
                sku = txt(d.get("SKU"))
                qty = num(d.get("Jml Order"))
                serial = txt(d.get("No Invoice"))
                pi_id = nid()
                total_disetor = num(d.get("Total Disetor"))
                reject = num(d.get("Reject Potongan"))
                retur = num(d.get("Retur ke Penjahit"))
                diterima = num(d.get("Diterima Bersih"))
                po_item = {
                    "id": pi_id, "po_id": po_id, "po_number": po_number,
                    "product_id": None, "catalog_item_id": None,
                    "product_name": txt(d.get("Nama Produk")),
                    "model_id": pmap.get(sku.upper()), "size_id": None,
                    "sku": sku, "serial_number": serial,
                    "size": txt(d.get("Ukuran")), "color": txt(d.get("Warna")),
                    "qty": qty,
                    "selling_price_snapshot": 0.0, "cmt_price_snapshot": 0.0,
                    # nilai monitoring asli dari Excel (fidelitas penuh):
                    "excel_total_disetor": total_disetor,
                    "excel_reject": reject,
                    "excel_retur_penjahit": retur,
                    "excel_diterima_bersih": diterima,
                    "excel_sisa_potongan": num(d.get("Sisa Potongan")),
                    "excel_kali_setor": num(d.get("Kali Setor")),
                    "excel_status": txt(d.get("Status")),
                    "excel_alert": txt(d.get("ALERT")),
                    "excel_deadline_mitra": as_dt(d.get("Deadline Mitra")),
                    "excel_tgl_kirim": as_dt(d.get("Tgl Kirim ke CMT")),
                    "catatan": txt(d.get("Catatan")),
                    "import_source": IMPORT_TAG, "created_at": NOW,
                }
                await db.po_items.insert_one(po_item)
                stats["items"] += 1

                dispatched = bool(as_dt(d.get("Tgl Kirim ke CMT"))) or txt(d.get("Status")).lower() != "belum kirim"
                if dispatched and qty > 0:
                    ship_item_docs.append({
                        "po_item_id": pi_id, "serial_number": serial, "sku": sku,
                        "product_name": po_item["product_name"], "size": po_item["size"],
                        "color": po_item["color"], "qty_sent": qty,
                    })
                if total_disetor > 0 or diterima > 0:
                    receipt_line_docs.append({
                        "po_item_id": pi_id, "sku_code": sku,
                        "product_name": po_item["product_name"], "size": po_item["size"],
                        "color": po_item["color"],
                        "qty_shipped_by_cmt": total_disetor or diterima,
                        "qty_actual": diterima,
                        "reject_qty": reject, "retur_qty": retur,
                    })

            # vendor_shipment (potongan dikirim ke CMT)
            if ship_item_docs:
                sid = nid()
                shipment = {
                    "id": sid, "shipment_number": f"SJ-IMP-{po_number}",
                    "delivery_note_number": "",
                    "vendor_id": vendor["id"], "vendor_name": vendor.get("name", ""),
                    "po_id": po_id, "po_number": po_number,
                    "shipment_date": min(ship_dates) if ship_dates else NOW,
                    "shipment_type": "NORMAL", "parent_shipment_id": None,
                    "business_type": btype, "status": "Received",
                    "inspection_status": "Inspected",
                    "notes": "Impor snapshot", "import_source": IMPORT_TAG,
                    "created_by": "import", "created_at": NOW, "updated_at": NOW,
                }
                await db.vendor_shipments.insert_one(shipment)
                stats["shipments"] += 1
                insp_id = nid()
                insp_items = []
                for si in ship_item_docs:
                    vsi_id = nid()
                    await db.vendor_shipment_items.insert_one({
                        "id": vsi_id, "shipment_id": sid, "shipment_number": shipment["shipment_number"],
                        "po_id": po_id, "po_number": po_number, "po_item_id": si["po_item_id"],
                        "source_po_item_id": si["po_item_id"],
                        "product_name": si["product_name"], "serial_number": si["serial_number"],
                        "size": si["size"], "color": si["color"], "sku": si["sku"],
                        "qty_sent": si["qty_sent"], "ordered_qty": si["qty_sent"],
                        "shipment_type": "NORMAL", "parent_shipment_id": None,
                        "import_source": IMPORT_TAG, "created_at": NOW,
                    })
                    stats["ship_items"] += 1
                    insp_items.append({
                        "id": nid(), "inspection_id": insp_id, "shipment_item_id": vsi_id,
                        "item_type": "material", "sku": si["sku"], "size": si["size"],
                        "color": si["color"], "ordered_qty": si["qty_sent"],
                        "received_qty": si["qty_sent"], "missing_qty": 0,
                        "condition_notes": "", "import_source": IMPORT_TAG, "created_at": NOW,
                    })
                # inspeksi (potongan diterima CMT penuh)
                await db.vendor_material_inspections.insert_one({
                    "id": insp_id, "shipment_id": sid, "po_id": po_id,
                    "total_received": sum(x["received_qty"] for x in insp_items),
                    "total_missing": 0, "total_acc_received": 0, "total_acc_missing": 0,
                    "status": "Submitted", "import_source": IMPORT_TAG,
                    "created_by": "import", "created_at": NOW,
                })
                stats["inspections"] += 1
                if insp_items:
                    await db.vendor_material_inspection_items.insert_many(insp_items)

            # cmt_receipt (rekap setoran / diterima DA) — snapshot, status Approved
            if receipt_line_docs:
                rid = nid()
                total_shipped = sum(l["qty_shipped_by_cmt"] for l in receipt_line_docs)
                total_actual = sum(l["qty_actual"] for l in receipt_line_docs)
                total_reject = sum(l["reject_qty"] for l in receipt_line_docs)
                await db.cmt_receipts.insert_one({
                    "id": rid, "receipt_code": f"CMT-RCV-IMP-{po_number}",
                    "cmt_name": vendor.get("name", ""), "cmt_vendor_id": vendor["id"],
                    "wo_number": "", "wo_id": "", "po_id": po_id, "po_number": po_number,
                    "business_type": btype, "receipt_date": NOW_ISO[:10],
                    "delivery_note": "", "notes": "Impor snapshot setoran",
                    "status": "Approved",
                    "submitted_at": NOW_ISO, "submitted_by": "import",
                    "approved_at": NOW_ISO, "approved_by": "import", "reject_reason": "",
                    "related_shipment_id": "",
                    "total_shipped_by_cmt": total_shipped, "total_actual": total_actual,
                    "total_rejected": total_reject, "variance_reason": "",
                    "defect_photos": [], "import_source": IMPORT_TAG,
                    "created_by": "import", "created_at": NOW_ISO, "updated_at": NOW_ISO,
                })
                stats["receipts"] += 1
                for l in receipt_line_docs:
                    await db.cmt_receipt_lines.insert_one({
                        "id": nid(), "receipt_id": rid,
                        "sku_code": l["sku_code"], "product_name": l["product_name"],
                        "color": l["color"], "size": l["size"],
                        "qty_expected": l["qty_shipped_by_cmt"],
                        "qty_shipped_by_cmt": l["qty_shipped_by_cmt"],
                        "qty_actual": l["qty_actual"], "reject_qty": l["reject_qty"],
                        "reject_reason": "", "photos": [], "po_item_id": l["po_item_id"],
                        "notes": "", "import_source": IMPORT_TAG, "created_at": NOW_ISO,
                    })
                    stats["receipt_lines"] += 1

    print("== DONE ==")
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    asyncio.run(main())
