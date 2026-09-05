#!/usr/bin/env python3
"""
SEED MASTER DATA NYATA — CV. Dewi Aditya
========================================
Mengganti seluruh seed demo lama dengan MASTER DATA ASLI dari 7 berkas Excel milik owner.

Keputusan owner (dikonfirmasi 2026-07-26):
  · DB lama DIHAPUS TOTAL, lalu diisi MASTER DATA SAJA.
  · Transaksi (PO, setoran CMT, penjualan harian, absensi, dsb.) DIKOSONGKAN
    supaya tim mulai dari data bersih.
  · Sheet `Sheet1` pada berkas Aksesoris (1.645 baris hijab/kaos kaki) DIABAIKAN —
    dinyatakan tidak relevan oleh owner.

CATATAN PENTING soal "saldo awal stok"
--------------------------------------
Stok awal kain / aksesoris / barang jadi TETAP diisi, karena tanpanya modul Gudang,
Cutting, dan Produksi tidak bisa dijalankan sama sekali (semua nol). Saldo awal ini
dicatat sebagai `saldo_awal` di ledger sehingga jelas terlihat bukan transaksi
operasional. Bila owner ingin nol semua, jalankan dengan flag `--no-stock`.

Sumber → tujuan (hasil pemetaan database, bukan asumsi)
--------------------------------------------------------
| Berkas / sheet                          | Koleksi tujuan                              |
|-----------------------------------------|---------------------------------------------|
| HR · Data Karyawan DA Grosir            | rahaza_employees, users, dewi_org_units,     |
|                                         | dewi_org_positions, rahaza_locations         |
| HR · Data THP                           | rahaza_payroll_profiles, da_payroll_allowances |
| STOCK BAHAN KAIN (11 sheet)             | rahaza_materials(type=fabric) + stok gudang  |
| Aksesoris · Stok (+harga dari ALL Juli) | rahaza_materials(type=accessory) + stok      |
| Sistem CMT · MASTER CMT                 | vendor_partners                              |
| Sistem CMT · GALERI PRODUK              | rahaza_models (+ ongkos jahit)               |
| Sistem CMT · SPEK PRODUK                | rahaza_models.spec (bahan + aksesoris)       |
| Techpack V5                             | rahaza_styles (+ varian warna/size)          |
| Gudang · Master Data Stok Produk        | rahaza_materials(type=fg) + stok rak         |
| Dashboard Marketing (8 sheet platform)  | marketing_platform_accounts, marketing_account_targets |

Pemakaian:
    python scripts/seed_da_master_from_excel.py --wipe        # hapus DB + seed ulang penuh
    python scripts/seed_da_master_from_excel.py               # seed saja (idempoten)
    python scripts/seed_da_master_from_excel.py --no-stock    # tanpa saldo awal stok
"""
from __future__ import annotations

import argparse
import asyncio
import os
import re
import subprocess
import sys
import time
import unicodedata
import uuid
from datetime import datetime, timezone

sys.path.insert(0, "/app/backend")

import openpyxl  # noqa: E402
from dotenv import load_dotenv  # noqa: E402
from motor.motor_asyncio import AsyncIOMotorClient  # noqa: E402

load_dotenv("/app/backend/.env")

DATA = "/app/data_import"
FILES = {
    "hr": f"{DATA}/hr_erp.xlsx",
    "kain": f"{DATA}/stock_kain.xlsx",
    "aksesoris": f"{DATA}/aksesoris_pemakaian.xlsx",
    "cmt": f"{DATA}/sistem_cmt.xlsx",
    "techpack": f"{DATA}/techpack_v5.xlsx",
    "fg": f"{DATA}/gudang_stok_produk.xlsx",
    "marketing": f"{DATA}/dashboard_marketing.xlsx",
}
DEFAULT_PASSWORD = "Dewi@123"
SUMMARY: dict[str, int] = {}
BACKLOG: list[str] = []


# ═════════════════════════════════════════════════ util
def now():
    return datetime.now(timezone.utc)


def uid():
    return str(uuid.uuid4())


def s(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def num(v, default=0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        return float(v)
    txt = re.sub(r"[^\d,.\-]", "", str(v)).replace(".", "").replace(",", ".")
    try:
        return float(txt) if txt not in ("", "-", ".") else default
    except ValueError:
        return default


def slug(v: str, maxlen: int = 18) -> str:
    v = unicodedata.normalize("NFKD", str(v or "")).encode("ascii", "ignore").decode()
    v = re.sub(r"[^A-Za-z0-9]+", "-", v).strip("-").upper()
    return v[:maxlen] or "NA"


def iso_date(v) -> str | None:
    if v is None or s(v) in ("", "-"):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    txt = s(v)
    for fmt in ("%d %B %Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d %b %Y"):
        try:
            return datetime.strptime(txt, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def rows_of(path: str, sheet: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    out = list(ws.iter_rows(values_only=True))
    wb.close()
    return out


def cell(row, i):
    return row[i] if row and i < len(row) else None


def bump(key, n=1):
    SUMMARY[key] = SUMMARY.get(key, 0) + n


# ═════════════════════════════════════════════════ stok
async def set_stock(db, material_id: str, location_id: str, qty: float, meta: dict):
    """Saldo awal via SSOT stock_service supaya ledger + alias konsisten."""
    if qty <= 0:
        return
    from core import stock_service
    await stock_service.add(
        material_id, location_id, qty,
        meta=meta,
        ref={"source": "seed_master", "ref_type": "saldo_awal", "ref_id": "excel-da-2026"},
        actor={"id": "seed", "email": "seed@dewiaditya.id"},
        db=db,
    )


# ═════════════════════════════════════════════════ 1. WIPE
def wipe_and_restart():
    from pymongo import MongoClient
    url = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
    name = os.environ.get("DB_NAME", "test_database")
    cli = MongoClient(url)
    cols = cli[name].list_collection_names()
    for c in cols:
        cli[name].drop_collection(c)
    print(f"  · {len(cols)} koleksi dihapus dari '{name}'")
    cli.close()
    print("  · restart backend supaya seed baseline (COA, lokasi, shift, role, admin) jalan ulang…")
    subprocess.run(["sudo", "supervisorctl", "restart", "backend"], capture_output=True)
    import urllib.request
    for _ in range(60):
        time.sleep(2)
        try:
            with urllib.request.urlopen("http://localhost:8001/api/health", timeout=5) as r:
                if r.status == 200:
                    print("  · backend siap.")
                    time.sleep(3)
                    return
        except Exception:
            continue
    raise SystemExit("Backend tidak kunjung siap setelah wipe.")


# ═════════════════════════════════════════════════ 2. LOKASI & ORG
ROLE_BY_DIVISION = {
    "AKUNTANSI": "accounting",
    "HRD": "hr",
    "RND": "rnd_staff",
    "MARKETING": "pic_toko",
    "GUDANG": "admin_gudang",
    "PRODUKSI": "admin_produksi",
}
ROLE_BY_JABATAN = {
    "SPV AKUNTANSI & KEUANGAN": "accounting",
    "ASSISTEN ACCOUNTING": "staff_keuangan",
    "HR GENERALIST": "hr",
    "SUPERVISOR MARKETING": "pic_toko",
    "CONTENT CREATOR": "marketing_kol",
    "HOST LIVE": "marketing_kol",
    "PIC AKUN": "pic_toko",
    "CUSTOMER SERVICE": "pic_toko",
    "SUPERVISOR GUDANG": "admin_gudang",
    "ADMIN GUDANG": "admin_gudang",
    "PACKING": "tim_packing",
    "ADMIN AKSESORIS": "admin_aksesoris",
    "SUPERVISOR CUTING": "spv_cuting",
    "KENEK CUTING": "operator_cuting",
    "ADMIN PRODUKSI": "admin_produksi",
    "QUALITY CONTROL": "operator",
    "RESEARCH & DEVELOPMENT": "rnd_staff",
    "DRIVER & PETUGAS UMUM": "operator",
}


async def seed_locations_and_org(db, emp_rows):
    """Lokasi kerja nyata + unit organisasi (divisi) + posisi (jabatan)."""
    loc_map = {}
    for name in sorted({s(cell(r, 6)) for r in emp_rows if s(cell(r, 6))}):
        code = slug(name, 20)
        ex = await db.rahaza_locations.find_one({"code": code}, {"_id": 0})
        if ex:
            loc_map[name] = ex["id"]
            continue
        doc = {"id": uid(), "code": code, "name": name, "type": "office",
               "active": True, "created_at": now(), "updated_at": now()}
        await db.rahaza_locations.insert_one(doc)
        loc_map[name] = doc["id"]
        bump("lokasi_kerja")

    root = await db.dewi_org_units.find_one({"code": "DA"}, {"_id": 0})
    if not root:
        root = {"unit_id": uid(), "name": "CV. Dewi Aditya Official", "code": "DA",
                "type": "company", "parent_id": None, "level": 0,
                "head_employee_id": None, "head_employee_name": "Direktur Utama",
                "headcount_actual": 0, "headcount_target": 0, "color": "",
                "description": "Perusahaan induk", "is_active": True,
                "created_by": "seed", "created_at": now(), "updated_at": now()}
        await db.dewi_org_units.insert_one(dict(root))
        bump("org_unit")

    unit_map = {}
    for div in sorted({s(cell(r, 3)).upper() for r in emp_rows if s(cell(r, 3))}):
        ex = await db.dewi_org_units.find_one({"code": slug(div)}, {"_id": 0})
        if ex:
            unit_map[div] = ex
            continue
        doc = {"unit_id": uid(), "name": div.title(), "code": slug(div),
               "type": "division", "parent_id": root["unit_id"], "level": 1,
               "head_employee_id": None, "head_employee_name": "",
               "headcount_actual": sum(1 for r in emp_rows if s(cell(r, 3)).upper() == div),
               "headcount_target": 0, "color": "", "description": "",
               "is_active": True, "created_by": "seed",
               "created_at": now(), "updated_at": now()}
        await db.dewi_org_units.insert_one(dict(doc))
        unit_map[div] = doc
        bump("org_unit")

    for r in emp_rows:
        title, div = s(cell(r, 4)), s(cell(r, 3)).upper()
        if not title:
            continue
        if await db.dewi_org_positions.find_one({"title": title}, {"_id": 0}):
            continue
        unit = unit_map.get(div) or {}
        await db.dewi_org_positions.insert_one({
            "position_id": uid(), "title": title,
            "unit_id": unit.get("unit_id", ""), "unit_name": unit.get("name", ""),
            "grade": 1, "reports_to_position_id": None,
            "reports_to_title": s(cell(r, 5)), "headcount_target": 1,
            "headcount_actual": sum(1 for x in emp_rows if s(cell(x, 4)) == title),
            "salary_grade": "", "is_active": True, "created_by": "seed",
            "created_at": now(), "updated_at": now(),
        })
        bump("org_posisi")
    return loc_map


# ═════════════════════════════════════════════════ 3. KARYAWAN + USER + PAYROLL
def gender_of(v):
    t = s(v).lower()
    return "L" if t.startswith("laki") else ("P" if t.startswith("perem") else "")


def marital_of(v):
    t = s(v).lower()
    if "belum" in t:
        return "single"
    if "menikah" in t or "kawin" in t:
        return "married"
    return ""


def email_of(name: str, used: set) -> str:
    base = re.sub(r"[^a-z0-9]+", ".", s(name).lower()).strip(".")
    base = ".".join([p for p in base.split(".") if p][:2]) or "karyawan"
    mail = f"{base}@dewiaditya.id"
    i = 2
    while mail in used:
        mail = f"{base}{i}@dewiaditya.id"
        i += 1
    used.add(mail)
    return mail


async def seed_employees(db, emp_rows, thp_rows, loc_map):
    from auth import hash_password
    used_mail = {u["email"] async for u in db.users.find({}, {"_id": 0, "email": 1})}
    creds = []

    thp = {}
    for r in thp_rows:
        nm = s(cell(r, 1))
        if nm and nm.upper() not in ("NAMA", ""):
            thp[nm.lower()] = r

    pending_manager = []
    for r in emp_rows:
        code = s(cell(r, 0)).upper()
        name = s(cell(r, 1))
        if not code or not name:
            continue
        if await db.rahaza_employees.find_one({"employee_code": code}, {"_id": 0}):
            continue
        div = s(cell(r, 3))
        jab = s(cell(r, 4))
        mail = email_of(name, used_mail)
        role = ROLE_BY_JABATAN.get(jab.upper()) or ROLE_BY_DIVISION.get(div.upper()) or "operator"
        emp = {
            "id": uid(), "employee_code": code, "name": name,
            "department": div, "job_title": jab or "Operator",
            "location_id": loc_map.get(s(cell(r, 6))),
            "location_name": s(cell(r, 6)),
            "phone": s(cell(r, 7)), "email": mail,
            "contract_type": (s(cell(r, 8)) if s(cell(r, 8)) not in ("", "-") else None),
            "contract_start_date": iso_date(cell(r, 9)),
            "contract_end_date": iso_date(cell(r, 10)),
            "wage_scheme": "monthly" if "bulan" in s(cell(r, 11)).lower() else "borongan_pcs",
            "base_rate": num(cell(r, 12)),
            "joined_at": iso_date(cell(r, 9)) or now().isoformat(),
            "manager_id": None, "manager_name": s(cell(r, 5)),
            "gender": gender_of(cell(r, 13)),
            "birth_date": iso_date(cell(r, 16)), "birth_place": s(cell(r, 15)),
            "marital_status": marital_of(cell(r, 14)),
            "religion": s(cell(r, 17)), "nationality": s(cell(r, 18)) or "Indonesia",
            "ktp_address": s(cell(r, 19)), "current_address": s(cell(r, 20)),
            "education_level": s(cell(r, 21)), "education_institution": s(cell(r, 22)),
            "education_major": s(cell(r, 23)), "photo_url": "",
            "ktp_number": s(cell(r, 24)).replace("'", ""), "npwp_number": s(cell(r, 25)),
            "tax_ptkp": s(cell(r, 26)) or "TK/0",
            "bpjs_kesehatan_number": s(cell(r, 27)),
            "bpjs_ketenagakerjaan_number": s(cell(r, 28)),
            "bank_name": s(cell(r, 29)), "bank_account_number": s(cell(r, 30)),
            "bank_account_holder": s(cell(r, 31)) or name,
            "emergency_contact_name": s(cell(r, 32)),
            "emergency_phone": s(cell(r, 33)), "emergency_relation": s(cell(r, 34)),
            "active": True, "created_at": now(), "updated_at": now(),
        }
        await db.rahaza_employees.insert_one(dict(emp))
        bump("karyawan")
        pending_manager.append((emp["id"], s(cell(r, 5))))

        if not await db.users.find_one({"email": mail}, {"_id": 0}):
            await db.users.insert_one({
                "id": uid(), "name": name, "email": mail,
                "password": hash_password(DEFAULT_PASSWORD), "role": role,
                "status": "active", "employee_id": emp["id"],
                "employee_code": code, "department": div,
                "created_at": now(), "updated_at": now(),
            })
            bump("user_login")
            creds.append((code, name, jab, mail, role))

        # ── payroll profile + komponen THP
        base = num(cell(r, 12))
        await db.rahaza_payroll_profiles.insert_one({
            "id": uid(), "employee_id": emp["id"],
            "pay_scheme": "monthly", "period_type": "monthly",
            "cutoff_config": {"start_day": 1},
            "base_rate": base, "overtime_rate": 0.0, "pcs_process_rates": [],
            "notes": "Seed dari HR - ERP System.xlsx (kolom Rate/Base)",
            "active": True, "created_at": now(), "updated_at": now(),
            "created_by": "seed", "created_by_name": "Seed Master",
        })
        bump("payroll_profile")

        t = thp.get(name.lower())
        if t:
            gaji_pokok = num(cell(t, 4))
            await db.rahaza_payroll_profiles.update_one(
                {"employee_id": emp["id"]},
                {"$set": {"base_rate": gaji_pokok or base,
                          "notes": "Seed dari sheet 'Data THP' (Gaji Pokok)"}},
            )
            for idx, (nm, calc, fixed_wage) in [
                (5, ("Tunjangan Jabatan", "fixed", True)),
                (6, ("Insentif Makan", "per_day_attendance", False)),
                (7, ("Tunjangan Kesehatan", "fixed", True)),
                (8, ("Bonus Kehadiran", "fixed", False)),
                (9, ("Tunjangan Transport", "fixed", False)),
            ]:
                amount = num(cell(t, idx))
                if amount <= 0:
                    continue
                if nm == "Insentif Makan":
                    amount = 10000.0  # header sheet: Rp10.000 / hari hadir
                await db.da_payroll_allowances.insert_one({
                    "allowance_id": uid(), "name": nm, "amount": amount,
                    "calc_type": calc, "is_fixed_wage": fixed_wage,
                    "applicable_to": "employee", "department": div,
                    "employee_ids": [emp["id"]],
                    "description": f"Seed dari sheet 'Data THP' — {name}",
                    "is_active": True, "created_by": "seed",
                    "created_at": now(), "updated_at": now(),
                })
                bump("tunjangan")

    # resolusi atasan (setelah semua karyawan ada)
    for emp_id, mgr_title in pending_manager:
        if not mgr_title or mgr_title.lower() == "direktur utama":
            continue
        mgr = await db.rahaza_employees.find_one({"job_title": mgr_title, "active": True}, {"_id": 0})
        if mgr:
            await db.rahaza_employees.update_one(
                {"id": emp_id},
                {"$set": {"manager_id": mgr["id"], "manager_name": mgr["name"]}})
            bump("relasi_atasan")
    return creds


# ═════════════════════════════════════════════════ 4. KAIN
UNIT_MAP = {"kg": "kg", "yard": "yard", "yd": "yard", "meter": "m", "m": "m",
            "gram": "gram", "rol": "rol", "roll": "rol", "pcs": "pcs"}


async def seed_fabrics(db, loc_id, with_stock=True):
    wb = openpyxl.load_workbook(FILES["kain"], data_only=True, read_only=True)
    for sheet in wb.sheetnames:
        if sheet.strip().lower() == "memo":
            BACKLOG.append(
                "Sheet 'Memo' (rincian per-rol: 12 rol, berat tertulis vs pengecekan, lebar 135cm) "
                "BELUM diimpor — perlu keputusan owner: dijadikan dokumen roll fisik "
                "(wh_fabric_rolls) untuk kain yang mana?")
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        hdr = next((i for i, r in enumerate(rows)
                    if r and any(s(c).upper().startswith("BAHAN") for c in r[:4])), None)
        if hdr is None:
            continue
        sub = rows[hdr + 1] if hdr + 1 < len(rows) else ()
        akhir_cols = [i for i, v in enumerate(sub) if s(v).lower() == "akhir"]
        last_name = ""
        for r in rows[hdr + 2:]:
            if not r:
                continue
            nm = s(cell(r, 1))
            if nm:
                last_name = nm
            color = s(cell(r, 2))
            if not color or not last_name:
                continue
            unit_raw = s(cell(r, 3)).lower()
            unit = UNIT_MAP.get(unit_raw, "kg")
            stock = 0.0
            for i in reversed(akhir_cols):
                v = cell(r, i)
                if v is not None and s(v) != "":
                    stock = num(v)
                    break
            code = f"KAIN-{slug(last_name, 16)}-{slug(color, 12)}"
            ex = await db.rahaza_materials.find_one({"code": code}, {"_id": 0})
            if ex:
                continue
            doc = {
                "id": uid(), "code": code,
                "name": f"{last_name} — {color}", "type": "fabric", "unit": unit,
                "category": "FABRIC", "category_name": "Kain/Fabric",
                "composition": last_name, "color": color,
                "notes": f"Seed dari STOCK BAHAN KAIN · sheet '{sheet}'",
                "min_stock": 0, "unit_cost": 0.0,
                "pack_unit": "rol", "pack_size": 1, "display_in_packs": False,
                "fabric_group": sheet.strip(),
                "active": True, "created_at": now(), "updated_at": now(),
            }
            await db.rahaza_materials.insert_one(dict(doc))
            bump("material_kain")
            if with_stock and stock > 0:
                await set_stock(db, doc["id"], loc_id, stock, {
                    "material_code": code, "material_name": doc["name"],
                    "material_type": "fabric", "unit": unit,
                    "ownership": "cv_da", "inventory_category": "raw_material"})
                bump("stok_kain_baris")
    wb.close()


# ═════════════════════════════════════════════════ 5. AKSESORIS
ACC_UNIT_PATTERNS = [
    (r"\b1?\s*roll?\b", "rol"), (r"\bmeter\b|\b1\s*met\b|\bm\b", "m"),
    (r"\bpack\b|\bpak\b|\bpck\b", "pak"), (r"\bbks\b|\bbungkus\b", "pak"),
    (r"\blusin\b", "lusin"), (r"\bkg\b", "kg"), (r"\bset\b", "set"),
    (r"\bpcs\b|\bpc\b", "pcs"),
]


def acc_unit(name: str) -> str:
    low = f" {name.lower()} "
    for pat, unit in ACC_UNIT_PATTERNS:
        if re.search(pat, low):
            return unit
    return "pcs"


async def seed_accessories(db, loc_id, with_stock=True):
    price = {}
    for r in rows_of(FILES["aksesoris"], "ALL Juli")[6:]:
        kode = s(cell(r, 0))
        if kode and re.match(r"^A\d+$", kode):
            price[kode] = num(cell(r, 2))

    rows = rows_of(FILES["aksesoris"], "Stok")
    hdr = next((i for i, r in enumerate(rows) if r and s(cell(r, 0)).upper() == "KODE"), None)
    if hdr is None:
        BACKLOG.append("Sheet 'Stok' aksesoris: header KODE tidak ditemukan — dilewati.")
        return
    for r in rows[hdr + 1:]:
        kode = s(cell(r, 0))
        name = s(cell(r, 1))
        if not kode or not name:
            continue
        if await db.rahaza_materials.find_one({"code": kode.upper()}, {"_id": 0}):
            continue
        stock = num(cell(r, 7))
        doc = {
            "id": uid(), "code": kode.upper(), "name": name,
            "type": "accessory", "unit": acc_unit(name),
            "category": "Aksesoris", "description": "",
            "min_stock": 0.0, "unit_cost": price.get(kode, 0.0),
            "supplier": "", "notes": "Seed dari Lap. Pemakaian Aksesoris · sheet 'Stok'",
            "pack_unit": "pack", "pack_size": 1, "display_in_packs": False,
            "active": True, "created_by": "seed",
            "created_at": now().isoformat(), "updated_at": now().isoformat(),
        }
        await db.rahaza_materials.insert_one(dict(doc))
        bump("material_aksesoris")
        if with_stock and stock > 0:
            await set_stock(db, doc["id"], loc_id, stock, {
                "material_code": doc["code"], "material_name": name,
                "material_type": "accessory", "unit": doc["unit"],
                "ownership": "cv_da", "inventory_category": "raw_material"})
            bump("stok_aksesoris_baris")


# ═════════════════════════════════════════════════ 6. VENDOR CMT
async def seed_cmt_partners(db):
    rows = rows_of(FILES["cmt"], "MASTER CMT")
    hdr = next((i for i, r in enumerate(rows) if r and s(cell(r, 1)).upper().startswith("KODE")), None)
    if hdr is None:
        return
    for r in rows[hdr + 1:]:
        code = s(cell(r, 1))
        name = s(cell(r, 2))
        if not code or not name:
            continue
        if await db.vendor_partners.find_one({"code": code.upper()}, {"_id": 0}):
            continue
        await db.vendor_partners.insert_one({
            "id": uid(), "name": name, "code": code.upper(),
            "contact_name": "", "contact_phone": s(cell(r, 4)),
            "address": s(cell(r, 3)), "notes": "Seed dari Sistem CMT · MASTER CMT",
            "is_active": s(cell(r, 6)).lower() != "nonaktif",
            "capacity_pcs": int(num(cell(r, 5))), "capacity_note": "",
            "created_at": now(), "created_by": "seed",
        })
        bump("vendor_cmt")


# ═════════════════════════════════════════════════ 7. PRODUK (MODEL + SPEK)
async def seed_products(db):
    rows = rows_of(FILES["cmt"], "GALERI PRODUK")
    hdr = next((i for i, r in enumerate(rows) if r and s(cell(r, 1)).upper() == "SKU"), None)
    models = {}
    if hdr is not None:
        for r in rows[hdr + 1:]:
            sku = s(cell(r, 1)).replace(".0", "")
            name = s(cell(r, 2))
            if not sku or not name:
                continue
            code = sku.upper()
            ex = await db.rahaza_models.find_one({"code": code}, {"_id": 0})
            if ex:
                models[code] = ex
                continue
            doc = {
                "id": uid(), "code": code, "name": name,
                "category": s(cell(r, 3)) or "Umum",
                "material_kg_per_pcs": 0, "yarn_kg_per_pcs": 0,
                "bundle_size": 30, "description": "",
                "cmt_cost_per_pcs": num(cell(r, 4)),
                "sop_steps": [], "reference_videos": [], "reference_images": [],
                "source": "Seed dari Sistem CMT · GALERI PRODUK",
                "active": True, "created_at": now(), "updated_at": now(),
            }
            await db.rahaza_models.insert_one(dict(doc))
            models[code] = doc
            bump("model_produk")

    # SPEK PRODUK → lampirkan spesifikasi (bahan + aksesoris) ke model
    rows = rows_of(FILES["cmt"], "SPEK PRODUK")
    hdr = next((i for i, r in enumerate(rows) if r and s(cell(r, 1)).upper() == "SKU"), None)
    if hdr is None:
        return
    linked = 0
    created_from_spec = 0
    for r in rows[hdr + 1:]:
        sku = s(cell(r, 1)).replace(".0", "").upper()
        if not sku:
            continue
        fabrics = [s(cell(r, i)) for i in (6, 7, 8) if s(cell(r, i))]
        accs = []
        for base in range(9, 30, 3):
            nm = s(cell(r, base))
            if not nm:
                continue
            accs.append({"name": nm, "qty_per_pcs": num(cell(r, base + 1)),
                         "unit": s(cell(r, base + 2)) or "pcs"})
        if not (fabrics or accs):
            continue
        # SPEK PRODUK memuat 57 SKU sementara GALERI PRODUK hanya 13 —
        # SKU yang belum punya master model DIBUATKAN di sini supaya tidak ada
        # produk nyata yang hilang dari master (nama & kategori diambil dari SPEK).
        if not await db.rahaza_models.find_one({"code": sku}, {"_id": 0}):
            await db.rahaza_models.insert_one({
                "id": uid(), "code": sku, "name": s(cell(r, 2)) or sku,
                "category": s(cell(r, 5)) or "Umum",
                "material_kg_per_pcs": 0, "yarn_kg_per_pcs": 0,
                "bundle_size": 30, "description": s(cell(r, 3)),
                "cmt_cost_per_pcs": 0.0,
                "sop_steps": [], "reference_videos": [], "reference_images": [],
                "source": "Seed dari Sistem CMT · SPEK PRODUK",
                "active": True, "created_at": now(), "updated_at": now(),
            })
            created_from_spec += 1
            bump("model_produk")
        res = await db.rahaza_models.update_one({"code": sku}, {"$set": {
            "spec": {
                "construction": s(cell(r, 3)),
                "component_count": num(cell(r, 4)),
                "fabrics": fabrics,
                "accessories": accs,
                "source": "Sistem CMT · SPEK PRODUK",
            },
            "updated_at": now(),
        }})
        if res.matched_count:
            linked += 1
    bump("spek_produk", linked)
    if linked:
        BACKLOG.append(
            f"SPEK PRODUK ({linked} SKU) tersimpan sebagai `spec` di master model "
            "(bahan + aksesoris + qty/pcs). BELUM diubah jadi BOM resmi (rahaza_boms) "
            "karena nama aksesoris di Excel adalah teks bebas (mis. 'Kancing 18L') dan "
            "harus dipetakan manual ke kode master aksesoris (A1, A47, …) + ke size. "
            "Perlu keputusan owner: buat layar pemetaan nama→kode, atau isi BOM manual?")


# ═════════════════════════════════════════════════ 8. TECHPACK → STYLE
async def seed_styles(db):
    rows = rows_of(FILES["techpack"], "Data Techpack Ringkasan Produk ")
    hdr = 0
    cur = None
    buf = []

    async def flush(item, extra):
        if not item:
            return
        name = s(item[1])
        if not name:
            return
        code = (s(item[0]) or slug(name, 20)).upper()
        if await db.rahaza_styles.find_one({"style_code": code}, {"_id": 0}):
            return
        colors = [c.strip() for c in re.split(r"[;,\n]", s(item[6])) if c.strip()]
        size_raw = s(item[8])
        variants = []
        for c in colors or [""]:
            variants.append({"id": uid(), "color": c, "size": size_raw,
                             "sku": f"{code}-{slug(c, 8)}" if c else code, "notes": ""})
        await db.rahaza_styles.insert_one({
            "id": uid(), "style_code": code, "style_name": name,
            "category": s(item[2]), "buyer": s(item[3]),
            "fabric_type": s(item[9]), "season": s(item[4]),
            "description": s(item[5]),
            "status": "active",
            "design_images": [], "techpack_url": None, "techpack_name": None,
            "variants": variants,
            "construction_notes": [x for x in extra["construction"] if x],
            "fabric_consumption": [x for x in extra["consumption"] if x],
            "measurements": [x for x in extra["measure"] if x],
            "source": "Seed dari Data Techpack Ringkasan Produk V5",
            "created_by": "seed", "created_by_name": "Seed Master",
            "created_at": now(), "updated_at": now(),
        })
        bump("style_techpack")

    extra = {"construction": [], "consumption": [], "measure": []}
    for r in rows[hdr + 1:]:
        if not r:
            continue
        if s(cell(r, 1)):
            await flush(cur, extra)
            cur = r
            extra = {"construction": [s(cell(r, 7))], "consumption": [s(cell(r, 10))],
                     "measure": [s(cell(r, 11))]}
            buf = []
        elif cur:
            extra["construction"].append(s(cell(r, 7)))
            extra["consumption"].append(s(cell(r, 10)))
            extra["measure"].append(s(cell(r, 11)))
    await flush(cur, extra)
    _ = buf


# ═════════════════════════════════════════════════ 9. BARANG JADI (FG)
async def seed_fg(db, loc_id, with_stock=True):
    rows = rows_of(FILES["fg"], "DATA TERBARU")
    hdr = next((i for i, r in enumerate(rows) if r and s(cell(r, 0)).upper() == "NAMA BARANG"), None)
    if hdr is None:
        BACKLOG.append("Gudang · Master Data Stok Produk: header 'NAMA BARANG' tidak ditemukan.")
        return
    last = ""
    for r in rows[hdr + 1:]:
        if not r:
            continue
        nm = s(cell(r, 0))
        if nm and nm.upper() not in ("NAMA BARANG",):
            last = nm
        color = s(cell(r, 1))
        if not color or not last:
            continue
        if color.upper() in ("WARNA",):
            continue
        code = f"FG-{slug(last, 16)}-{slug(color, 12)}"
        if await db.rahaza_materials.find_one({"code": code}, {"_id": 0}):
            continue
        rak = num(cell(r, 10))
        doc = {
            "id": uid(), "code": code, "name": f"{last} — {color}",
            "type": "fg", "unit": "pcs",
            "category": "FG", "category_name": "Barang Jadi",
            "color": color, "composition": "",
            "notes": "Seed dari Gudang · Master Data Stok Produk (DATA TERBARU)",
            "min_stock": num(cell(r, 14)), "unit_cost": 0.0,
            "pack_unit": "pack", "pack_size": 1, "display_in_packs": False,
            "product_name": last,
            "active": True, "created_at": now(), "updated_at": now(),
        }
        await db.rahaza_materials.insert_one(dict(doc))
        bump("material_fg")
        if with_stock and rak > 0:
            await set_stock(db, doc["id"], loc_id, rak, {
                "material_code": code, "material_name": doc["name"],
                "material_type": "fg", "unit": "pcs",
                "ownership": "cv_da", "inventory_category": "fg_internal"})
            bump("stok_fg_baris")


# ═════════════════════════════════════════════════ 10. MARKETING
PLATFORM_SHEETS = {
    "PLATFORM SHOPEE GHS": ("shopee", "GHS", "SHOPEE-GHS"),
    "PLATFORM SHOPEE MOEN": ("shopee", "MOEN", "SHOPEE-MOEN"),
    "PLATFORM SHOPEE DALUNA": ("shopee", "DALUNA", "SHOPEE-DALUNA"),
    "PLATFORM TIKTOK DALUNA": ("tiktokshop", "DALUNA", "TIKTOK-DALUNA"),
    "PLATFORM TIKTOK OB": ("tiktokshop", "OB", "TIKTOK-OB"),
    "PLATFORM TIKTOK MOEN": ("tiktokshop", "MOEN", "TIKTOK-MOEN"),
    "PLATFORM TIKTOK FATIMAHIJAB": ("tiktokshop", "FATIMAHIJAB", "TIKTOK-FATIMAHIJAB"),
    "PLATFORM TIKTOK DEZZA": ("tiktokshop", "DEZZA", "TIKTOK-DEZZA"),
}
PIC_BY_ACCOUNT = {
    "SHOPEE-DALUNA": "Ulfa", "TIKTOK-DALUNA": "Ulfa",
    "SHOPEE-GHS": "Niken", "TIKTOK-OB": "Niken",
    "SHOPEE-MOEN": "Dewi", "TIKTOK-MOEN": "Dewi",
}


async def seed_marketing(db):
    wb = openpyxl.load_workbook(FILES["marketing"], data_only=True, read_only=True)
    for sheet, (platform, brand, code) in PLATFORM_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))[:8]
        target = 0.0
        for r in rows:
            if r and r[0] and "target bulan" in s(r[0]).lower():
                target = num(cell(r, 1))
        ex = await db.marketing_platform_accounts.find_one({"account_code": code}, {"_id": 0})
        if not ex:
            acc = {
                "id": uid(), "account_code": code,
                "account_name": f"{brand} ({'Shopee' if platform == 'shopee' else 'TikTok Shop'})",
                "platform": platform, "username": brand.lower(),
                "status": "active", "group": brand.lower(),
                "credentials": {"api_key": "", "api_secret": "", "has_api_integration": False},
                "import_config": {"saved_templates": []},
                "assigned_staff": [PIC_BY_ACCOUNT[code]] if code in PIC_BY_ACCOUNT else [],
                "pic_id": "seed", "health_score": None,
                "monthly_target": target,
                "notes": "Seed dari DASHBOARD MARKETING JULI",
                "created_at": now(), "created_by": "seed", "updated_at": now(),
            }
            await db.marketing_platform_accounts.insert_one(dict(acc))
            bump("akun_marketplace")
        else:
            acc = ex
        if target > 0 and not await db.marketing_account_targets.find_one(
                {"account_id": acc["id"], "year": 2026, "month": 7}, {"_id": 0}):
            await db.marketing_account_targets.insert_one({
                "id": uid(), "account_id": acc["id"],
                "account_name": acc["account_name"], "platform": platform,
                "year": 2026, "month": 7,
                "revenue_target": target, "orders_target": 0,
                "health_score_target": 0,
                "notes": "Target bulanan dari DASHBOARD MARKETING JULI",
                "created_by": "seed", "created_at": now(), "updated_at": now(),
            })
            bump("target_marketing")
    wb.close()
    BACKLOG.append(
        "Dashboard Marketing: sheet DAILY ACTIVITY, KPI TIM, CAMPAIGN, ADS, BUDGET IKLAN, "
        "PENCAIRAN MARKETPLACE, RETUR, REVIEW NEGATIF, KOL & ENDORSEMENT, PENGIRIMAN SAMPEL, "
        "KOMPETITOR, MEETING berisi TRANSAKSI/aktivitas harian — sengaja TIDAK diimpor "
        "sesuai keputusan owner (mulai bersih). Bila nanti ingin dipakai sebagai baseline "
        "historis, perlu konfirmasi.")


# ═════════════════════════════════════════════════ MAIN
async def main(args):
    from database import get_db
    db = get_db()

    emp_rows = [r for r in rows_of(FILES["hr"], "Data Karyawan DA Grosir")[2:] if r and s(cell(r, 0))]
    thp_rows = rows_of(FILES["hr"], "Data THP")[2:]

    print("\n[1/8] Lokasi kerja + unit organisasi + posisi…")
    loc_map = await seed_locations_and_org(db, emp_rows)

    # Lokasi stok DIPILIH SETELAH lokasi kerja nyata di-seed, supaya saldo awal
    # mendarat di gudang sungguhan ("Gudang Lantai 1/Area Gudang"), bukan lokasi
    # bawaan sistem. Kalau tidak ada, jatuh ke lokasi aktif pertama.
    gudang = await db.rahaza_locations.find_one(
        {"name": {"$regex": "gudang", "$options": "i"}}, {"_id": 0}, sort=[("code", 1)])
    loc = gudang or await db.rahaza_locations.find_one({"active": True}, {"_id": 0}, sort=[("code", 1)])
    if not loc:
        raise SystemExit("Lokasi gudang belum ter-seed. Restart backend dulu.")
    loc_id = loc["id"]
    print(f"  · lokasi saldo awal stok: {loc.get('name')} ({loc.get('code')})")

    print("[2/8] Karyawan + akun login + profil payroll + tunjangan…")
    creds = await seed_employees(db, emp_rows, thp_rows, loc_map)
    print("[3/8] Master kain (11 sheet)…")
    await seed_fabrics(db, loc_id, with_stock=not args.no_stock)
    print("[4/8] Master aksesoris…")
    await seed_accessories(db, loc_id, with_stock=not args.no_stock)
    print("[5/8] Vendor CMT…")
    await seed_cmt_partners(db)
    print("[6/8] Model produk + spek…")
    await seed_products(db)
    print("[7/8] Style techpack…")
    await seed_styles(db)
    print("[8/8] Barang jadi + akun marketplace…")
    await seed_fg(db, loc_id, with_stock=not args.no_stock)
    await seed_marketing(db)

    # Pastikan koleksi modul baru SELALU ADA setelah wipe.
    # (mongodump hanya menyalin koleksi yang eksis; kalau `cutting_*` baru lahir saat
    #  transaksi pertama, backup yang diambil lebih dulu tidak memuatnya.)
    existing = await db.list_collection_names()
    for c in ("cutting_orders", "cutting_progress"):
        if c not in existing:
            await db.create_collection(c)
            bump("koleksi_dipastikan_ada")

    print("\n" + "=" * 78)
    print("RINGKASAN SEED MASTER DATA")
    print("=" * 78)
    for k in sorted(SUMMARY):
        print(f"  {k:26s} : {SUMMARY[k]}")
    if creds:
        lines = ["# Kredensial Login Hasil Seed Master (data nyata DA)", "",
                 f"Password default semua karyawan: `{DEFAULT_PASSWORD}`", "",
                 "| Kode | Nama | Jabatan | Email (login) | Role |",
                 "|---|---|---|---|---|"]
        for c in creds:
            lines.append(f"| {c[0]} | {c[1]} | {c[2]} | {c[3]} | {c[4]} |")
        lines += ["", "Superadmin: `admin@garment.com` / `Admin@123`"]
        with open("/app/memory/SEED_CREDENTIALS.md", "w") as f:
            f.write("\n".join(lines) + "\n")
        print(f"\n  Kredensial {len(creds)} akun ditulis ke /app/memory/SEED_CREDENTIALS.md")
    if BACKLOG:
        print("\n" + "-" * 78)
        print("BACKLOG — data Excel yang BELUM bisa diakomodasi sistem (perlu keputusan owner):")
        for i, b in enumerate(BACKLOG, 1):
            print(f"  {i}. {b}")
        with open("/app/docs/BACKLOG_FROM_EXCEL.md", "w") as f:
            f.write("# Backlog — Data Excel yang Belum Terakomodasi\n\n")
            f.write("Dihasilkan otomatis oleh `scripts/seed_da_master_from_excel.py`.\n\n")
            for i, b in enumerate(BACKLOG, 1):
                f.write(f"{i}. {b}\n\n")
    print("=" * 78)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--wipe", action="store_true", help="hapus seluruh koleksi dulu + restart backend")
    ap.add_argument("--no-stock", action="store_true", help="jangan isi saldo awal stok")
    a = ap.parse_args()
    if a.wipe:
        print("[0/8] Menghapus database lama…")
        wipe_and_restart()
    asyncio.run(main(a))
