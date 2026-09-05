#!/usr/bin/env python3
"""uom_rebase_worklist.py — daftar kerja "Ubah Satuan Dasar" massal (A3).

MASALAH
-------
Sebagian material dicatat dengan satuan KEMASAN sebagai satuan dasar
(mis. `rol`, `pak`, `lusin`). Akibatnya stok & HPP tersimpan per kemasan,
padahal pemakaian nyata di produksi memakai satuan eceran (`m`, `pcs`).

CARA PAKAI (3 langkah)
----------------------
1. Ekspor daftar kerja ke Excel:
       python3 scripts/uom_rebase_worklist.py --export /app/data_import/rebase_uom.xlsx

2. Owner mengisi 2 kolom:
       satuan_baru          → satuan dasar yang benar (mis. "m")
       isi_per_satuan_lama  → 1 <satuan_sekarang> = ? <satuan_baru>  (mis. 50)
   Baris yang dikosongkan akan DILEWATI.

3. Pratinjau lalu terapkan (memanggil endpoint resmi
   POST /api/rahaza/materials/{id}/rebase-uom — TIDAK ada logika kedua):
       python3 scripts/uom_rebase_worklist.py --preview /app/data_import/rebase_uom.xlsx
       python3 scripts/uom_rebase_worklist.py --apply   /app/data_import/rebase_uom.xlsx

Catatan: nilai persediaan (qty × HPP) TIDAK berubah — qty dikali faktor, HPP
dibagi faktor. Satuan lama tetap tersedia sebagai kemasan (keep_old_as_pack).
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "backend"))

try:
    from dotenv import load_dotenv
    load_dotenv(ROOT / "backend" / ".env")
except Exception:  # noqa: BLE001
    pass

import openpyxl  # noqa: E402
import requests  # noqa: E402
from motor.motor_asyncio import AsyncIOMotorClient  # noqa: E402

from core import uom as U  # noqa: E402

API = os.environ.get("INTERNAL_API_URL", "http://localhost:8001")
EMAIL = os.environ.get("SEED_ADMIN_EMAIL", "admin@garment.com")
PASSWORD = os.environ.get("SEED_ADMIN_PASSWORD", "Admin@123")

# Satuan yang secara fisik adalah KEMASAN — kandidat rebase.
PACK_LIKE = {"rol", "gulung", "pak", "pack", "lusin", "kodi", "gross",
             "bal", "karton", "dus", "sak", "set", "box", "ball"}

HEADERS = ["material_id", "code", "name", "type", "satuan_sekarang", "stok",
           "hpp_per_satuan", "nilai_persediaan", "satuan_baru", "isi_per_satuan_lama"]


def _db():
    client = AsyncIOMotorClient(os.environ["MONGO_URL"])
    return client[os.environ["DB_NAME"]]


async def collect() -> list[dict]:
    db = _db()
    mats = await db.rahaza_materials.find({}, {"_id": 0}).to_list(20000)
    ids = [m["id"] for m in mats]
    onhand: dict[str, float] = {}
    async for r in db.rahaza_material_stock.aggregate([
        {"$match": {"material_id": {"$in": ids}}},
        {"$group": {"_id": "$material_id", "q": {"$sum": {"$ifNull": ["$qty", 0]}}}},
    ]):
        onhand[r["_id"]] = float(r["q"] or 0)

    rows = []
    for m in mats:
        base = U.base_uom_of(m)
        if base not in PACK_LIKE:
            continue
        qty = round(onhand.get(m["id"], 0.0), 4)
        cost = float(m.get("unit_cost") or 0)
        rows.append({
            "material_id": m["id"], "code": m.get("code", ""), "name": m.get("name", ""),
            "type": m.get("type", ""), "satuan_sekarang": base, "stok": qty,
            "hpp_per_satuan": cost, "nilai_persediaan": round(qty * cost, 2),
            "satuan_baru": "", "isi_per_satuan_lama": "",
        })
    rows.sort(key=lambda r: (r["satuan_sekarang"], r["code"]))
    return rows


def write_xlsx(rows: list[dict], path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "rebase_uom"
    ws.append(HEADERS)
    for j, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=j).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = max(len(h) + 2, 14)
    for r in rows:
        ws.append([r[h] for h in HEADERS])
    ws.freeze_panes = "A2"
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def read_xlsx(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    data = list(ws.iter_rows(values_only=True))
    head = [str(h or "").strip() for h in data[0]]
    out = []
    for raw in data[1:]:
        row = {head[i]: raw[i] for i in range(min(len(head), len(raw)))}
        if not row.get("material_id"):
            continue
        new_uom = str(row.get("satuan_baru") or "").strip().lower()
        try:
            factor = float(row.get("isi_per_satuan_lama") or 0)
        except (TypeError, ValueError):
            factor = 0.0
        if not new_uom or factor <= 0:
            continue
        out.append({**row, "satuan_baru": new_uom, "isi_per_satuan_lama": factor})
    return out


def login() -> str:
    r = requests.post(f"{API}/api/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=30)
    r.raise_for_status()
    body = r.json()
    return body.get("access_token") or body.get("token") or body["data"]["access_token"]


def run_rebase(rows: list[dict], *, preview: bool) -> None:
    token = login()
    hdr = {"Authorization": f"Bearer {token}"}
    ok = fail = 0
    for r in rows:
        payload = {"new_base_uom": r["satuan_baru"], "factor": r["isi_per_satuan_lama"],
                   "keep_old_as_pack": True, "preview": preview}
        resp = requests.post(f"{API}/api/rahaza/materials/{r['material_id']}/rebase-uom",
                             json=payload, headers=hdr, timeout=60)
        if resp.status_code != 200:
            fail += 1
            print(f"  ✗ {r['code']:<18} {resp.status_code} {resp.text[:140]}")
            continue
        d = resp.json()
        b, a = d.get("before", {}), d.get("after", {})
        flag = "OK " if d.get("nilai_persediaan_tetap") else "!! "
        print(f"  {flag}{r['code']:<18} {d['from_uom']}→{d['to_uom']} ×{d['factor']:g} | "
              f"stok {b.get('total_qty'):g}→{a.get('total_qty'):g} | "
              f"HPP {b.get('unit_cost'):,.2f}→{a.get('unit_cost'):,.2f}")
        ok += 1
    print(f"\n{'PRATINJAU' if preview else 'DITERAPKAN'}: {ok} berhasil, {fail} gagal.")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--export", metavar="XLSX", help="ekspor daftar kerja ke Excel")
    g.add_argument("--preview", metavar="XLSX", help="pratinjau rebase dari Excel terisi")
    g.add_argument("--apply", metavar="XLSX", help="terapkan rebase dari Excel terisi")
    args = ap.parse_args()

    if args.export:
        rows = asyncio.run(collect())
        write_xlsx(rows, args.export)
        by_uom: dict[str, int] = {}
        for r in rows:
            by_uom[r["satuan_sekarang"]] = by_uom.get(r["satuan_sekarang"], 0) + 1
        print(f"{len(rows)} item bersatuan kemasan → {args.export}")
        for u, n in sorted(by_uom.items(), key=lambda x: -x[1]):
            print(f"  {u:<10} {n}")
        return

    path = args.preview or args.apply
    rows = read_xlsx(path)
    if not rows:
        print("Tidak ada baris terisi (butuh kolom satuan_baru + isi_per_satuan_lama).")
        return
    print(f"{len(rows)} baris siap diproses dari {path}\n")
    run_rebase(rows, preview=bool(args.preview))


if __name__ == "__main__":
    main()
