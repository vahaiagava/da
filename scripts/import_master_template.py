#!/usr/bin/env python3
"""scripts/import_master_template.py — IMPOR MASTER dari TEMPLATE_MASTER_DA.xlsx.

    python3 scripts/import_master_template.py berkas.xlsx            # PERIKSA saja (dry-run)
    python3 scripts/import_master_template.py berkas.xlsx --apply    # simpan
    python3 scripts/import_master_template.py berkas.xlsx --only 10_BOM,09_BARANG_JADI

PRINSIP
-------
* **Dry-run adalah bawaan.** Menyimpan harus diminta secara sadar (`--apply`).
* **Satu baris salah tidak membatalkan seluruh berkas**, tetapi seluruh KESALAHAN
  dilaporkan lebih dulu dan penyimpanan DITOLAK selama masih ada kesalahan. Impor
  separuh-separuh adalah cara paling cepat melahirkan SKU hantu (kejadian nyata di
  sistem ini: 3 baris SPK dengan SKU yang tidak punya master, Rp 3,6 jt menggantung).
* **Idempoten**: kunci alami (kode/nik/sku/kode_akun) dipakai upsert ⇒ impor ulang
  MEMPERBARUI, tidak menduplikasi.
* **TIDAK menghapus apa pun.** Data lama (termasuk demo) dibiarkan; sesuai keputusan
  pemilik, pembersihan dilakukan di lingkungan produksi masing-masing.
* Yang **tidak** ditangani di sini (dan disebutkan di laporan): password portal
  kreator/livehost, saldo awal stok/piutang/kas, dan varian SSOT RnD.
"""
from __future__ import annotations

import argparse
import asyncio
import os
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "backend"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from master_template_spec import SHEETS, URUTAN  # noqa: E402

G, R, Y, C, B, X = "\033[92m", "\033[91m", "\033[93m", "\033[96m", "\033[1m", "\033[0m"
BATCH = f"master_template_{datetime.now(timezone.utc):%Y%m%d_%H%M%S}"
ERRORS: list[tuple[str, int, str]] = []
STATS: dict[str, dict] = {}


def err(sheet: str, row: int, msg: str) -> None:
    ERRORS.append((sheet, row, msg))


def now():
    return datetime.now(timezone.utc)


def s(v) -> str:
    return "" if v is None else str(v).strip()


def num(v, default=0.0) -> float:
    t = s(v).replace("Rp", "").replace(".", "").replace(",", ".").replace(" ", "")
    if not t:
        return default
    try:
        return float(t)
    except ValueError:
        return default


def is_num(v) -> bool:
    t = s(v)
    if not t:
        return False
    try:
        float(t.replace("Rp", "").replace(".", "").replace(",", ".").replace(" ", ""))
        return True
    except ValueError:
        return False


def yes(v, default=True) -> bool:
    t = s(v).lower()
    if not t:
        return default
    return t in ("ya", "y", "yes", "true", "1", "aktif", "active")


def read_sheet(wb, name: str) -> list[dict]:
    """Baris data sheet sebagai dict {kolom: nilai} + nomor baris asli Excel."""
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [s(h).lower().lstrip("#").strip().rstrip("*") for h in rows[0]]
    out = []
    for i, raw in enumerate(rows[1:], start=2):
        if raw is None or all(s(c) == "" for c in raw):
            continue
        if s(raw[0]).startswith("#"):          # baris contoh / komentar
            continue
        rec = {header[j]: raw[j] for j in range(min(len(header), len(raw))) if header[j]}
        rec["__row"] = i
        out.append(rec)
    return out


def check_required(sheet: str, recs: list[dict]) -> list[dict]:
    req = [k for k, r, *_ in SHEETS[sheet]["kolom"] if r]
    good = []
    for rec in recs:
        miss = [k for k in req if s(rec.get(k)) == ""]
        if miss:
            err(sheet, rec["__row"], f"kolom wajib kosong: {', '.join(miss)}")
            continue
        good.append(rec)
    return good


def check_enum(sheet: str, rec: dict, col: str, allowed: tuple, default: str = "") -> str:
    v = s(rec.get(col)).lower()
    if not v:
        return default
    if v not in allowed:
        err(sheet, rec["__row"], f"'{col}' = '{v}' tidak sah — pilih: {' | '.join(allowed)}")
        return default
    return v


def dupes(sheet: str, recs: list[dict], key: str) -> None:
    seen: dict[str, int] = {}
    for rec in recs:
        k = s(rec.get(key)).upper()
        if k in seen:
            err(sheet, rec["__row"], f"{key} '{k}' kembar dengan baris {seen[k]} "
                                     "di berkas yang sama")
        else:
            seen[k] = rec["__row"]


async def upsert(db, coll: str, find: dict, doc: dict, apply: bool, sheet: str) -> str:
    """Upsert idempoten + hitung statistik. Mengembalikan id dokumen."""
    st = STATS.setdefault(sheet, {"baru": 0, "diperbarui": 0})
    old = await db[coll].find_one(find, {"_id": 0, "id": 1})
    if old:
        st["diperbarui"] += 1
        if apply:
            await db[coll].update_one({"id": old["id"]},
                                      {"$set": {**doc, "updated_at": now(),
                                                "import_batch": BATCH}})
        return old["id"]
    st["baru"] += 1
    new_id = str(uuid.uuid4())
    if apply:
        await db[coll].insert_one({**doc, "id": new_id, "created_at": now(),
                                   "updated_at": now(), "import_batch": BATCH,
                                   "import_source": "master_template_v1"})
    return new_id


# ═══════════════════════════════════════════════════════════════════════════════
async def run(path: Path, apply: bool, only: set[str]) -> int:  # noqa: C901
    from dotenv import load_dotenv
    load_dotenv(ROOT / "backend" / ".env")
    from motor.motor_asyncio import AsyncIOMotorClient
    from openpyxl import load_workbook

    db = AsyncIOMotorClient(os.environ["MONGO_URL"])[os.environ["DB_NAME"]]
    wb = load_workbook(path, data_only=True)
    print(f"{B}Berkas   :{X} {path}")
    print(f"{B}Mode     :{X} " + (f"{R}SIMPAN (--apply){X}" if apply
                                  else f"{G}PERIKSA saja (dry-run){X}"))
    missing = [n for n in URUTAN if n not in wb.sheetnames]
    if missing:
        print(f"{Y}Sheet tidak ada (dilewati): {', '.join(missing)}{X}")

    def want(name: str) -> bool:
        return not only or name in only

    # ── peta kode → id (gabungan yang sudah ada di basis data + yang baru dibuat) ──
    async def code_map(coll: str, field: str = "code", extra: dict | None = None) -> dict:
        q = extra or {}
        docs = await db[coll].find(q, {"_id": 0, "id": 1, field: 1}).to_list(20000)
        return {s(d.get(field)).upper(): d["id"] for d in docs if s(d.get(field))}

    loc = await code_map("rahaza_locations")
    col = await code_map("rahaza_colors")
    siz = await code_map("rahaza_sizes")
    mdl = await code_map("rahaza_models")
    mat = await code_map("rahaza_materials")
    emp = await code_map("rahaza_employees", "employee_code")
    acc = await code_map("marketing_platform_accounts", "account_code")
    # Catatan bahan/SKU yang LAHIR DI BERKAS INI. Tanpa ini, pemeriksaan (dry-run)
    # pada basis data kosong selalu melaporkan "material tidak ada" untuk BOM yang
    # materialnya ada di sheet 06/07 berkas yang sama — dry-run jadi tidak pernah bisa
    # bersih, dan pemakai belajar mengabaikan laporannya.
    pending: dict[str, dict] = {}

    # ── 01 LOKASI ─────────────────────────────────────────────────────────────
    if want("01_LOKASI"):
        recs = check_required("01_LOKASI", read_sheet(wb, "01_LOKASI"))
        dupes("01_LOKASI", recs, "kode")
        for r in recs:
            tipe = check_enum("01_LOKASI", r, "tipe",
                              ("gudang", "kantor", "produksi", "toko"), "gudang")
            code = s(r["kode"]).upper()
            loc[code] = await upsert(db, "rahaza_locations", {"code": code}, {
                "code": code, "name": s(r["nama"]), "type": tipe,
                "parent_id": loc.get(s(r.get("kode_induk")).upper()),
                "active": yes(r.get("aktif")),
            }, apply, "01_LOKASI")

    # ── 02 KARYAWAN (+ profil payroll) ────────────────────────────────────────
    if want("02_KARYAWAN"):
        recs = check_required("02_KARYAWAN", read_sheet(wb, "02_KARYAWAN"))
        dupes("02_KARYAWAN", recs, "nik")
        for r in recs:
            skema = check_enum("02_KARYAWAN", r, "skema_upah",
                               ("bulanan", "borongan", "harian"), "bulanan")
            lk = s(r.get("kode_lokasi")).upper()
            if lk and lk not in loc:
                err("02_KARYAWAN", r["__row"], f"kode_lokasi '{lk}' tidak ada di 01_LOKASI "
                                               "maupun di sistem")
            nik = s(r["nik"]).upper()
            eid = await upsert(db, "rahaza_employees", {"employee_code": nik}, {
                "employee_code": nik, "name": s(r["nama"]),
                "role_hint": s(r.get("jabatan")).lower(), "phone": s(r.get("telepon")),
                "join_date": s(r.get("tanggal_masuk"))[:10] or None,
                "location_id": loc.get(lk), "active": yes(r.get("aktif")),
            }, apply, "02_KARYAWAN")
            emp[nik] = eid
            scheme = {"bulanan": "monthly", "borongan": "piece_rate", "harian": "daily"}[skema]
            await upsert(db, "rahaza_payroll_profiles", {"employee_id": eid}, {
                "employee_id": eid, "pay_scheme": scheme,
                "period_type": "monthly", "base_rate": num(r.get("gaji_pokok")),
                "overtime_rate": num(r.get("tarif_lembur_per_jam")),
                "pcs_process_rates": [], "active": True,
                "notes": f"Impor master {BATCH}",
            }, apply, "02_KARYAWAN")

    # ── 03 WARNA / 04 UKURAN / 05 PROSES ──────────────────────────────────────
    if want("03_WARNA"):
        recs = check_required("03_WARNA", read_sheet(wb, "03_WARNA"))
        dupes("03_WARNA", recs, "kode")
        for r in recs:
            code = s(r["kode"]).upper()
            col[code] = await upsert(db, "rahaza_colors", {"code": code}, {
                "code": code, "name": s(r["nama"]), "hex": s(r.get("hex")),
                "order_seq": int(num(r.get("urutan"))), "active": True,
            }, apply, "03_WARNA")

    if want("04_UKURAN"):
        recs = check_required("04_UKURAN", read_sheet(wb, "04_UKURAN"))
        dupes("04_UKURAN", recs, "kode")
        for r in recs:
            code = s(r["kode"]).upper()
            if not code.replace("-", "").replace(".", "").isalnum():
                err("04_UKURAN", r["__row"], f"kode ukuran '{code}' memuat karakter yang "
                                             "tidak boleh masuk SKU (spasi/garis miring) — "
                                             "pakai mis. ALLSIZE, 2XL, 28-30")
                continue
            siz[code] = await upsert(db, "rahaza_sizes", {"code": code}, {
                "code": code, "name": s(r["nama"]),
                "order_seq": int(num(r.get("urutan"))), "active": True,
            }, apply, "04_UKURAN")

    if want("05_PROSES"):
        recs = check_required("05_PROSES", read_sheet(wb, "05_PROSES"))
        dupes("05_PROSES", recs, "kode")
        for r in recs:
            code = s(r["kode"]).upper()
            await upsert(db, "rahaza_processes", {"code": code}, {
                "code": code, "name": s(r["nama"]),
                "order_seq": int(num(r.get("urutan"))),
                "is_rework": yes(r.get("permak"), False),
                "description": s(r.get("keterangan")), "active": True,
            }, apply, "05_PROSES")

    # ── 06 KAIN & BENANG / 07 AKSESORIS ───────────────────────────────────────
    UNITS = ("pcs", "kg", "gram", "m", "yard", "roll", "pack", "gross", "lusin")
    if want("06_MATERIAL_KAIN"):
        recs = check_required("06_MATERIAL_KAIN", read_sheet(wb, "06_MATERIAL_KAIN"))
        dupes("06_MATERIAL_KAIN", recs, "kode")
        for r in recs:
            jenis = check_enum("06_MATERIAL_KAIN", r, "jenis", ("fabric", "yarn"), "")
            unit = check_enum("06_MATERIAL_KAIN", r, "satuan_dasar", UNITS, "")
            if not jenis or not unit:
                continue
            code = s(r["kode"]).upper()
            mat[code] = await upsert(db, "rahaza_materials", {"code": code}, {
                "code": code, "name": s(r["nama"]), "type": jenis, "unit": unit,
                "base_uom": unit, "purchase_uom": unit, "issue_uom": unit, "display_uom": unit,
                "uoms": [{"code": unit, "name": unit.upper(), "factor": 1.0,
                          "is_base": True, "level": 0}],
                "composition": s(r.get("komposisi")), "color": s(r.get("warna")),
                "gsm": num(r.get("gramasi_gsm")) or None,
                "width_cm": num(r.get("lebar_cm")) or None,
                "unit_cost": num(r.get("harga_per_satuan")),
                "min_stock": num(r.get("stok_minimum")),
                "cost_method": "moving_average", "active": True,
            }, apply, "06_MATERIAL_KAIN")
            pending[code] = {"id": mat[code], "code": code, "name": s(r["nama"]),
                             "type": jenis, "unit": unit, "category_name": ""}

    if want("07_AKSESORIS"):
        recs = check_required("07_AKSESORIS", read_sheet(wb, "07_AKSESORIS"))
        dupes("07_AKSESORIS", recs, "kode")
        for r in recs:
            unit = check_enum("07_AKSESORIS", r, "satuan_dasar", UNITS, "")
            if not unit:
                continue
            pack_size = num(r.get("isi_per_kemasan"), 1.0) or 1.0
            code = s(r["kode"]).upper()
            mat[code] = await upsert(db, "rahaza_materials", {"code": code}, {
                "code": code, "name": s(r["nama"]), "type": "accessory", "unit": unit,
                "base_uom": unit, "purchase_uom": unit, "issue_uom": unit, "display_uom": unit,
                "uoms": [{"code": unit, "name": unit.upper(), "factor": 1.0,
                          "is_base": True, "level": 0}],
                "category_name": s(r.get("kategori")),
                "unit_cost": num(r.get("harga_per_satuan")),
                "min_stock": num(r.get("stok_minimum")),
                "pack_unit": s(r.get("satuan_kemasan")) or "pack", "pack_size": pack_size,
                "cost_method": "moving_average", "active": True,
            }, apply, "07_AKSESORIS")
            pending[code] = {"id": mat[code], "code": code, "name": s(r["nama"]),
                             "type": "accessory", "unit": unit,
                             "category_name": s(r.get("kategori"))}

    # ── 08 MODEL ──────────────────────────────────────────────────────────────
    if want("08_MODEL"):
        recs = check_required("08_MODEL", read_sheet(wb, "08_MODEL"))
        dupes("08_MODEL", recs, "kode")
        for r in recs:
            code = s(r["kode"]).upper()
            mdl[code] = await upsert(db, "rahaza_models", {"code": code}, {
                "code": code, "name": s(r["nama"]),
                "category": s(r.get("kategori")),
                "description": s(r.get("keterangan")),
                "retail_price": num(r.get("harga_jual_dasar")), "active": True,
            }, apply, "08_MODEL")

    # ── 09 BARANG JADI (SKU) ──────────────────────────────────────────────────
    if want("09_BARANG_JADI"):
        recs = check_required("09_BARANG_JADI", read_sheet(wb, "09_BARANG_JADI"))
        dupes("09_BARANG_JADI", recs, "sku")
        colnames = {v: k for k, v in [(c, i) for i, c in enumerate([])]}  # noqa: F841
        cmap = {s(d.get("code")).upper(): d for d in await db.rahaza_colors.find(
            {}, {"_id": 0, "id": 1, "code": 1, "name": 1, "hex": 1}).to_list(5000)}
        smap = {s(d.get("code")).upper(): d for d in await db.rahaza_sizes.find(
            {}, {"_id": 0, "id": 1, "code": 1}).to_list(5000)}
        for r in recs:
            mk, ck, sk = (s(r["kode_model"]).upper(), s(r["kode_warna"]).upper(),
                          s(r["kode_ukuran"]).upper())
            bad = [f"{lbl} '{v}'" for lbl, v, ok in
                   (("kode_model", mk, mk in mdl), ("kode_warna", ck, ck in col),
                    ("kode_ukuran", sk, sk in siz)) if not ok]
            if bad:
                err("09_BARANG_JADI", r["__row"],
                    f"{', '.join(bad)} belum ada — isi sheet masternya lebih dulu")
                continue
            sku = s(r["sku"]).upper()
            c_doc, s_doc = cmap.get(ck, {}), smap.get(sk, {})
            mat[sku] = await upsert(db, "rahaza_materials", {"code": sku}, {
                "code": sku, "sku": sku, "name": s(r["nama"]), "type": "fg",
                "unit": s(r.get("satuan")) or "pcs",
                "model_id": mdl[mk], "model_code": mk,
                "size_id": siz[sk], "size_code": sk,
                "color_id": col[ck], "color_code": ck,
                "color": c_doc.get("name") or ck, "color_name": c_doc.get("name") or ck,
                "color_hex": c_doc.get("hex") or "",
                "weight_gram": num(r.get("berat_gram")),
                "retail_price_master": num(r.get("harga_jual")),
                "min_stock_qty": num(r.get("stok_minimum")), "active": True,
            }, apply, "09_BARANG_JADI")
            pending[sku] = {"id": mat[sku], "code": sku, "name": s(r["nama"]), "type": "fg",
                            "unit": s(r.get("satuan")) or "pcs",
                            "color": c_doc.get("name") or ck, "model_id": mdl[mk]}

    # ── 10 BOM (kain + benang + AKSESORIS) ────────────────────────────────────
    if want("10_BOM"):
        recs = check_required("10_BOM", read_sheet(wb, "10_BOM"))
        mdocs = {s(d.get("code")).upper(): d for d in await db.rahaza_materials.find(
            {}, {"_id": 0, "id": 1, "code": 1, "name": 1, "type": 1, "unit": 1,
                 "category_name": 1}).to_list(50000)}
        mdocs.update(pending)          # bahan yang baru lahir di berkas ini
        groups: dict[tuple, list] = {}
        for r in recs:
            mk, sk, mc = (s(r["kode_model"]).upper(), s(r["kode_ukuran"]).upper(),
                          s(r["kode_material"]).upper())
            if mk not in mdl:
                err("10_BOM", r["__row"], f"kode_model '{mk}' belum ada di 08_MODEL")
                continue
            if sk not in siz:
                err("10_BOM", r["__row"], f"kode_ukuran '{sk}' belum ada di 04_UKURAN")
                continue
            m = mdocs.get(mc)
            if not m:
                err("10_BOM", r["__row"], f"kode_material '{mc}' tidak ada di master bahan/"
                                          "aksesoris — isi 06/07 lebih dulu")
                continue
            if (m.get("type") or "") == "fg":
                err("10_BOM", r["__row"], f"'{mc}' adalah BARANG JADI, tidak boleh menjadi "
                                          "komponen BOM")
                continue
            if not is_num(r.get("qty_per_pcs")) or num(r.get("qty_per_pcs")) <= 0:
                err("10_BOM", r["__row"], f"qty_per_pcs '{s(r.get('qty_per_pcs'))}' harus "
                                          "angka lebih besar dari 0")
                continue
            groups.setdefault((mk, sk), []).append({
                "material_id": m["id"], "code": m.get("code"), "name": m.get("name"),
                "material_type": m.get("type"), "category_name": m.get("category_name") or "",
                "qty": num(r["qty_per_pcs"]),
                "unit": s(r.get("satuan")) or m.get("unit") or "",
                "notes": s(r.get("keterangan")),
            })
        for (mk, sk), lines in groups.items():
            seen = set()
            for ln in lines:
                if ln["material_id"] in seen:
                    err("10_BOM", 0, f"material '{ln['code']}' muncul dua kali pada BOM "
                                     f"{mk}/{sk} — gabungkan qty-nya")
                seen.add(ln["material_id"])
            await upsert(db, "rahaza_boms",
                         {"model_id": mdl[mk], "size_id": siz[sk], "active": True}, {
                             "model_id": mdl[mk], "size_id": siz[sk], "color": "",
                             "version": 1, "is_active": True, "active": True,
                             "materials": lines,
                             "notes": f"Impor master {BATCH} — {mk}/{sk}, {len(lines)} baris",
                         }, apply, "10_BOM")

    # ── 11 VENDOR CMT / 12 KLIEN MAKLON ───────────────────────────────────────
    if want("11_VENDOR_CMT"):
        recs = check_required("11_VENDOR_CMT", read_sheet(wb, "11_VENDOR_CMT"))
        dupes("11_VENDOR_CMT", recs, "kode")
        for r in recs:
            code = s(r["kode"]).upper()
            await upsert(db, "vendor_partners", {"code": code}, {
                "code": code, "name": s(r["nama"]),
                "contact_name": s(r.get("nama_kontak")), "contact_phone": s(r.get("telepon")),
                "address": s(r.get("alamat")), "capacity_pcs": num(r.get("kapasitas_pcs")),
                "notes": s(r.get("keterangan")), "active": True, "is_active": True,
            }, apply, "11_VENDOR_CMT")

    if want("12_KLIEN_MAKLON"):
        recs = check_required("12_KLIEN_MAKLON", read_sheet(wb, "12_KLIEN_MAKLON"))
        dupes("12_KLIEN_MAKLON", recs, "kode")
        for r in recs:
            code = s(r["kode"]).upper()
            await upsert(db, "dewi_maklon_clients", {"code": code}, {
                "code": code, "name": s(r["nama"]),
                "contact_name": s(r.get("nama_kontak")), "contact_phone": s(r.get("telepon")),
                "address": s(r.get("alamat")), "notes": s(r.get("keterangan")),
                "active": True,
            }, apply, "12_KLIEN_MAKLON")

    # ── 13 AKUN TOKO ──────────────────────────────────────────────────────────
    PLAT = ("shopee", "tiktok", "tokopedia", "lazada", "instagram", "facebook")
    if want("13_AKUN_TOKO"):
        recs = check_required("13_AKUN_TOKO", read_sheet(wb, "13_AKUN_TOKO"))
        dupes("13_AKUN_TOKO", recs, "kode_akun")
        for r in recs:
            plat = check_enum("13_AKUN_TOKO", r, "platform", PLAT, "")
            if not plat:
                continue
            code = s(r["kode_akun"]).upper()
            acc[code] = await upsert(db, "marketing_platform_accounts",
                                     {"account_code": code}, {
                                         "account_code": code, "account_name": s(r["nama_akun"]),
                                         "platform": plat, "username": s(r.get("username")),
                                         "group": s(r.get("grup")),
                                         "status": s(r.get("status")).lower() or "active",
                                     }, apply, "13_AKUN_TOKO")

    # ── 14 KATALOG JUAL ───────────────────────────────────────────────────────
    if want("14_KATALOG_JUAL"):
        recs = check_required("14_KATALOG_JUAL", read_sheet(wb, "14_KATALOG_JUAL"))
        fg = {s(d.get("code")).upper(): d for d in await db.rahaza_materials.find(
            {"type": "fg"}, {"_id": 0, "id": 1, "code": 1, "name": 1, "color": 1,
                             "model_id": 1}).to_list(50000)}
        fg.update({k: v for k, v in pending.items() if v.get("type") == "fg"})
        adocs = {d["id"]: d for d in await db.marketing_platform_accounts.find(
            {}, {"_id": 0, "id": 1, "platform": 1, "account_name": 1}).to_list(500)}
        for r in recs:
            ak, sku = s(r["kode_akun"]).upper(), s(r["sku"]).upper()
            if ak not in acc:
                err("14_KATALOG_JUAL", r["__row"], f"kode_akun '{ak}' belum ada di 13_AKUN_TOKO")
                continue
            if sku not in fg:
                err("14_KATALOG_JUAL", r["__row"], f"sku '{sku}' bukan barang jadi yang "
                                                   "terdaftar — isi 09_BARANG_JADI lebih dulu")
                continue
            harga = num(r["harga_jual"])
            if harga <= 0:
                err("14_KATALOG_JUAL", r["__row"], "harga_jual harus lebih besar dari 0 — "
                                                   "margin tidak bisa dihitung dari harga 0")
                continue
            f, a = fg[sku], adocs.get(acc[ak], {})
            await upsert(db, "marketing_catalog_items",
                         {"account_id": acc[ak], "sku": sku}, {
                             "account_id": acc[ak], "platform": a.get("platform") or "",
                             "sku": sku, "name": f.get("name") or sku,
                             "fg_material_id": f["id"], "material_id": f["id"],
                             "fg_code": sku, "fg_name": f.get("name") or sku,
                             "fg_color": f.get("color") or "", "model_id": f.get("model_id"),
                             "unit": "pcs", "source": "master_import",
                             "platform_price": harga, "harga_jual": harga, "price": harga,
                             "harga_coret": num(r.get("harga_coret")),
                             "original_price": num(r.get("harga_coret")),
                             "platform_url": s(r.get("tautan_produk")),
                             "is_active": yes(r.get("aktif")),
                         }, apply, "14_KATALOG_JUAL")

    # ── 15 KOL / KREATOR ──────────────────────────────────────────────────────
    if want("15_KOL_KREATOR"):
        recs = check_required("15_KOL_KREATOR", read_sheet(wb, "15_KOL_KREATOR"))
        dupes("15_KOL_KREATOR", recs, "kode_kreator")
        for r in recs:
            tipe = check_enum("15_KOL_KREATOR", r, "tipe", ("new", "kontrak", "continue"), "")
            mode = check_enum("15_KOL_KREATOR", r, "insentif_mode",
                              ("none", "per_pcs", "target_bonus", "both"), "none")
            if not tipe:
                continue
            if tipe == "new" and mode != "none":
                err("15_KOL_KREATOR", r["__row"], "kreator tipe 'new' tidak berhak insentif — "
                                                  "ubah tipe ke kontrak/continue atau "
                                                  "insentif_mode ke none")
                continue
            ak_codes = [x.strip().upper() for x in s(r.get("kode_akun_toko")).split(",") if x.strip()]
            unknown = [x for x in ak_codes if x not in acc]
            if unknown:
                err("15_KOL_KREATOR", r["__row"], f"kode_akun_toko {unknown} belum ada di "
                                                  "13_AKUN_TOKO")
                continue
            code = s(r["kode_kreator"]).upper()
            await upsert(db, "marketing_kol_creators", {"creator_code": code}, {
                "creator_code": code, "name": s(r["nama"]), "creator_type": tipe,
                "domicile": s(r.get("domisili")), "phone": s(r.get("telepon")),
                "login_email": s(r.get("email_portal")).lower(),
                "assigned_account_ids": [acc[x] for x in ak_codes],
                "status": "active",
                "incentive": {"mode": mode, "rate_per_pcs": num(r.get("insentif_per_pcs")),
                              "target_pcs": int(num(r.get("target_pcs"))),
                              "bonus_amount": num(r.get("bonus_target")),
                              "period_months": int(num(r.get("periode_bulan"), 3)) or 3,
                              "period_start": "", "notes": f"Impor master {BATCH}"},
            }, apply, "15_KOL_KREATOR")

    # ── 16 LIVEHOST ───────────────────────────────────────────────────────────
    if want("16_LIVEHOST"):
        recs = check_required("16_LIVEHOST", read_sheet(wb, "16_LIVEHOST"))
        dupes("16_LIVEHOST", recs, "email")
        for r in recs:
            nik = s(r["nik_karyawan"]).upper()
            if nik not in emp:
                err("16_LIVEHOST", r["__row"], f"nik_karyawan '{nik}' belum ada di "
                                               "02_KARYAWAN — gaji host dibaca dari payroll HR, "
                                               "jadi tautan ini wajib")
                continue
            ak_codes = [x.strip().upper() for x in s(r.get("kode_akun_toko")).split(",") if x.strip()]
            unknown = [x for x in ak_codes if x not in acc]
            if unknown:
                err("16_LIVEHOST", r["__row"], f"kode_akun_toko {unknown} belum ada di "
                                               "13_AKUN_TOKO")
                continue
            email = s(r["email"]).lower()
            await upsert(db, "marketing_livehosts", {"email": email}, {
                "email": email, "name": s(r["nama"]), "phone": s(r.get("telepon")),
                "employee_id": emp[nik], "employee_code": nik,
                "employment_type": "employee", "pay_mode": "monthly_hr",
                "hourly_rate": 0.0,
                "assigned_account_ids": [acc[x] for x in ak_codes],
                "status": s(r.get("status")).lower() or "active",
            }, apply, "16_LIVEHOST")

    # ── LAPORAN ───────────────────────────────────────────────────────────────
    return 0


def print_report(apply: bool) -> int:
    print(f"\n{B}{'─' * 74}{X}")
    print(f"{B}RENCANA PER SHEET{X}")
    if not STATS:
        print("  (tidak ada baris data — hanya baris contoh/komentar?)")
    for name in URUTAN:
        st = STATS.get(name)
        if st:
            print(f"  {name:18s} baru {st['baru']:5d} · diperbarui {st['diperbarui']:5d}")

    if ERRORS:
        print(f"\n{R}{B}{len(ERRORS)} BARIS BERMASALAH — tidak ada yang disimpan{X}")
        by_sheet: dict[str, list] = {}
        for sh, row, msg in ERRORS:
            by_sheet.setdefault(sh, []).append((row, msg))
        for sh, items in by_sheet.items():
            print(f"\n  {Y}{sh}{X} ({len(items)} baris)")
            for row, msg in items[:25]:
                print(f"    baris {row or '-':>4}  {msg}")
            if len(items) > 25:
                print(f"    … {len(items) - 25} baris lain sejenis")
        print(f"\n{R}Perbaiki di berkas Excel lalu jalankan lagi.{X}")
        return 1

    if apply:
        print(f"\n{G}{B}TERSIMPAN.{X} Penanda batch: {C}{BATCH}{X}")
        print(f"{Y}Belum termasuk (memang di luar master):{X}")
        print("  · Saldo awal stok / piutang / hutang / kas — impor terpisah")
        print("  · Password portal kreator & livehost — dibuat dari layar Marketing")
        print("  · HPP: jalankan hitung ulang HPP dari layar Costing setelah BOM masuk")
    else:
        print(f"\n{G}{B}PEMERIKSAAN BERSIH — tidak ada kesalahan.{X}")
        print(f"Jalankan lagi dengan {C}--apply{X} untuk menyimpan.")
    return 0


async def main(path: Path, apply: bool, only: set[str]) -> int:
    """DUA TAHAP. Tahap-1 memeriksa seluruh berkas TANPA menulis; penyimpanan baru
    dijalankan bila tahap-1 bersih. Kalau divalidasi sambil menulis, berkas yang cacat
    di baris ke-500 sudah meninggalkan 499 dokumen setengah jadi — dan itu jauh lebih
    sulit diperbaiki daripada sekadar gagal."""
    await run(path, False, only)
    if ERRORS:
        return print_report(False)
    if not apply:
        return print_report(False)
    ERRORS.clear()
    STATS.clear()
    await run(path, True, only)
    return print_report(True)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("file")
    ap.add_argument("--apply", action="store_true", help="simpan (bawaan: periksa saja)")
    ap.add_argument("--only", default="", help="hanya sheet tertentu, pisahkan koma")
    a = ap.parse_args()
    only = {x.strip() for x in a.only.split(",") if x.strip()}
    sys.exit(asyncio.run(main(Path(a.file), a.apply, only)))
