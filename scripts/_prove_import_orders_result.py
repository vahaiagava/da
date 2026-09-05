#!/usr/bin/env python3
"""BUKTI ANALISIS (READ-ONLY) — S6: apa yang DIHASILKAN kalau berkas asli
diimpor lewat jenis impor 'orders' yang ADA hari ini.

build_rows() adalah fungsi yang sama yang dipakai wizard sebelum commit.
Tidak ada penulisan ke MongoDB.
"""
from __future__ import annotations

import os
import sys
from collections import Counter

sys.path.insert(0, "/app/backend")
os.chdir("/app/backend")

import openpyxl  # noqa: E402
from core import marketing_import_engine as eng  # noqa: E402
from core.marketing_import_schema import get_source_type  # noqa: E402

XLSX = "/app/samples/TikTok_UntukDikirim_2026-07-19.xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb[wb.sheetnames[0]]
raw = [list(r) for r in ws.iter_rows(values_only=True)]
headers = [str(h).strip() if h is not None else "" for h in raw[0]]

# tiru cara wizard membaca: dict per baris pakai header baris 1
all_rows = [{headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
            for r in raw[1:] if any(c not in (None, "") for c in r)]

st = get_source_type("orders")
mapping = eng.auto_map(headers, st)

print("=" * 104)
print("S6 — HASIL NYATA build_rows() PADA BERKAS ASLI, JENIS IMPOR 'orders' (tanpa tulis DB)")
print("=" * 104)
print(f"  baris dikirim ke mesin : {len(all_rows)}  (termasuk baris ke-2 = deskripsi kolom)")

built = eng.build_rows(all_rows, mapping, st, limit=len(all_rows))
ok = [b for b in built if not b.get("errors")]
bad = [b for b in built if b.get("errors")]
print(f"  baris LOLOS validasi   : {len(ok)}")
print(f"  baris BERGALAT         : {len(bad)}")

errc = Counter()
for b in bad:
    for e in b.get("errors") or []:
        errc[str(e)[:80]] += 1
print("\n  Galat terbanyak:")
for k, v in errc.most_common(8):
    print(f"    {v:>4}x  {k}")

warnc = Counter()
for b in built:
    for w in b.get("warnings") or []:
        warnc[str(w)[:80]] += 1
if warnc:
    print("\n  Peringatan terbanyak:")
    for k, v in warnc.most_common(8):
        print(f"    {v:>4}x  {k}")

print("\n  BARIS PERTAMA yang dihasilkan (baris ke-2 berkas = DESKRIPSI KOLOM):")
first = built[0]
print(f"    errors  : {first.get('errors')}")
print(f"    data    : { {k: v for k, v in (first.get('data') or {}).items()} }")
print("    ==> kalau baris deskripsi tidak dilewati, ia MASUK sebagai satu 'pesanan' palsu")
print("        (atau menghasilkan 1 galat yang menutupi galat asli).")

print("\n  UANG — field yang benar-benar terisi di hasil impor:")
money_fields = ["revenue", "total_payment", "price_final", "price_original",
                "discount_seller", "shipping_cost"]
tot = {m: 0.0 for m in money_fields}
filled = {m: 0 for m in money_fields}
for b in ok:
    d = b.get("data") or {}
    for m in money_fields:
        v = d.get(m)
        if v not in (None, ""):
            filled[m] += 1
            try:
                tot[m] += float(v)
            except (TypeError, ValueError):
                pass
for m in money_fields:
    print(f"    {m:<18} terisi {filled[m]:>4}/{len(ok)}   jumlah = Rp {tot[m]:>14,.0f}")
print(f"\n    Nilai sebenarnya di berkas: Order Amount = Rp 62.805.113 · "
      f"SKU Subtotal After Discount = Rp 59.783.811")
print(f"    ==> jenis impor 'orders' hari ini menghasilkan omzet Rp {tot['revenue']:,.0f} "
      f"dari berkas bernilai Rp 62.805.113.")

print("\n  DEDUPE — berapa pesanan yang saling menimpa:")
oids = [(b.get("data") or {}).get("order_id") for b in ok]
c = Counter(o for o in oids if o)
dup = {k: v for k, v in c.items() if v > 1}
print(f"    order_id unik {len(c)} dari {len(oids)} baris · order_id muncul >1x: {len(dup)}")
print(f"    dedupe jenis 'orders' = {st.dedupe} ⇒ {sum(dup.values()) - len(dup)} baris SKU "
      f"akan ditimpa/hilang.")

print("\n  KOLOM YANG DIBUANG (tidak ada field tujuannya di jenis 'orders'):")
lost = [m["column"] for m in mapping if not m.get("field") or m.get("method") == "suggest"]
penting = [c for c in lost if c in (
    "Order Amount", "SKU Subtotal After Discount", "SKU Subtotal Before Discount",
    "Shipping Fee After Discount", "Order Channel", "Creator Handle", "Tracking ID",
    "Normal or Pre-order", "Order Substatus", "Warehouse Name", "Purchase Channel",
    "Shipped Time", "Delivered Time", "Cancelled Time", "Cancelation/Return Type",
    "Order Refund Amount", "Sku Quantity of return", "Package ID", "Province",
    "Regency and City", "SKU Unit Original Price", "Paid Time", "RTS Time")]
print(f"    total dibuang {len(lost)}/65 · di antaranya yang PENTING untuk rencana Anda:")
for c in penting:
    print(f"      - {c}")
print("=" * 104)
