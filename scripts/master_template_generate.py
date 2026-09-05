#!/usr/bin/env python3
"""scripts/master_template_generate.py — buat berkas Excel TEMPLATE MASTER.

    python3 scripts/master_template_generate.py [tujuan.xlsx]

Isi: sheet 00_PETUNJUK (urutan wajib + arti kolom) · 16 sheet data · 99_DAFTAR_PILIHAN.
Baris contoh diberi awalan `#` sehingga importir MELEWATINYA — jadi template boleh
langsung diimpor tanpa memasukkan data karangan.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from master_template_spec import ENUMS, SHEETS, URUTAN  # noqa: E402

HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")
REQ_FILL = PatternFill("solid", fgColor="FDE68A")
NOTE_FONT = Font(color="6B7280", italic=True, size=9)


def build(dest: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_PETUNJUK"
    ws.append(["TEMPLATE MASTER DATA — CV. Dewi Aditya"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for line in [
        "CARA PAKAI",
        "1. Isi sheet SESUAI URUTAN NOMOR. Sheet bernomor kecil menjadi acuan sheet berikutnya",
        "   (mis. BOM butuh model, ukuran, dan material yang sudah ada).",
        "2. JANGAN mengubah nama kolom pada baris pertama. Urutan kolom boleh berbeda —",
        "   importir mencari berdasarkan NAMA kolom.",
        "3. Baris yang dimulai dengan tanda # adalah CONTOH dan akan dilewati. Hapus atau biarkan.",
        "4. Kolom bertanda * WAJIB diisi. Baris tanpa kolom wajib akan DILAPORKAN, bukan diam-diam dibuang.",
        "5. Kode (kode/nik/sku/kode_akun) adalah KUNCI: impor ulang dengan kode sama = MEMPERBARUI,",
        "   bukan menduplikasi. Jadi berkas ini boleh diimpor berkali-kali.",
        "6. Angka ditulis angka saja (95000), bukan 'Rp 95.000'. Tanggal: YYYY-MM-DD.",
        "7. Jalankan pemeriksaan dulu (tanpa menyimpan):",
        "       python3 scripts/import_master_template.py <berkas.xlsx>",
        "   lalu simpan bila laporan bersih:",
        "       python3 scripts/import_master_template.py <berkas.xlsx> --apply",
        "",
        "YANG TIDAK ADA DI SINI (memang bukan master):",
        "· Saldo awal stok, piutang/hutang, saldo kas/bank — diimpor terpisah setelah master siap.",
        "· Password portal kreator/livehost — dibuat dari layar Marketing (tidak pernah lewat Excel).",
        "· CoA & posting profile akuntansi — sudah terpasang di sistem (353 akun · 33 profil).",
        "",
        "URUTAN SHEET & TUJUANNYA",
    ]:
        ws.append([line])
    ws.append([])
    ws.append(["Sheet", "Isi", "Kunci (penentu perbarui/baru)", "Kolom wajib"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
    for name in URUTAN:
        spec = SHEETS[name]
        wajib = ", ".join(k for k, req, *_ in spec["kolom"] if req)
        ws.append([name, spec["judul"], spec["kunci"], wajib])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 60

    for name in URUTAN:
        spec = SHEETS[name]
        s = wb.create_sheet(name)
        head = [k + ("*" if req else "") for k, req, *_ in spec["kolom"]]
        s.append([k for k, *_ in spec["kolom"]])
        for i, c in enumerate(s[1]):
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = HEAD_FILL
            c.alignment = Alignment(horizontal="center")
            if spec["kolom"][i][1]:
                c.fill = REQ_FILL
                c.font = Font(bold=True, color="7C2D12")
            s.column_dimensions[c.column_letter].width = max(14, min(34, len(head[i]) + 8))
        s.append(["# " + spec["kolom"][0][3]] + [k[3] for k in spec["kolom"][1:]])
        for c in s[2]:
            c.font = NOTE_FONT
        s.append(["# arti kolom →"] + [k[2] for k in spec["kolom"][1:]])
        for c in s[3]:
            c.font = NOTE_FONT
        s.freeze_panes = "A2"

    d = wb.create_sheet("99_DAFTAR_PILIHAN")
    d.append(["Kolom", "Nilai yang sah"])
    for c in d[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
    for k, v in ENUMS.items():
        d.append([k, v])
    d.column_dimensions["A"].width = 24
    d.column_dimensions["B"].width = 78

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    print(f"Template dibuat: {dest}")
    print(f"Sheet: 00_PETUNJUK + {len(URUTAN)} sheet data + 99_DAFTAR_PILIHAN")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("/app/data_import/TEMPLATE_MASTER_DA.xlsx")
    build(out)
