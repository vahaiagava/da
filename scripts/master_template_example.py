#!/usr/bin/env python3
"""scripts/master_template_example.py — CONTOH TERISI template master.

    python3 scripts/master_template_example.py [tujuan.xlsx]

Isinya data kecil tetapi **utuh dan saling tertaut**: 1 model kaos + 1 model kemeja,
2 warna × 3 ukuran, BOM memakai kain DAN aksesoris, 2 toko, katalog, KOL, live host.
Semua kode diberi awalan `CTH-` supaya jelas ini contoh — kalau tidak sengaja terimpor,
mudah dikenali dan dihapus.

Berkas ini WAJIB lolos `import_master_template.py` tanpa satu pun kesalahan (dijaga
gate INV-F41 C2): contoh yang tidak bisa diimpor lebih buruk daripada tidak ada contoh.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from master_template_spec import SHEETS, URUTAN  # noqa: E402

P = "CTH"          # awalan semua kode contoh
HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")

WARNA = [("NVY", "Navy", "#1B2A4A"), ("WHT", "Putih", "#F8FAFC")]
UKURAN = [("S", "S", 1), ("M", "M", 2), ("L", "L", 3), ("XL", "XL", 4), ("ALLSIZE", "All Size", 9)]
MODEL = [(f"{P}-TS01", "Kaos Basic Daluna", "Kaos", 89_000),
         (f"{P}-KM01", "Kemeja Katun Daluna", "Kemeja", 189_000)]
# (kode_model, kode_ukuran) → [(kode_material, qty_per_pcs, satuan, catatan)]
BOM = {
    (f"{P}-TS01", "M"): [(f"{P}-KN-CTN", 0.22, "kg", "badan + lengan"),
                         (f"{P}-ACC-LBL", 1, "pcs", "label merek"),
                         (f"{P}-ACC-HTG", 1, "pcs", "hangtag"),
                         (f"{P}-ACC-BNG", 0.02, "kg", "benang jahit")],
    (f"{P}-TS01", "L"): [(f"{P}-KN-CTN", 0.24, "kg", "badan + lengan"),
                         (f"{P}-ACC-LBL", 1, "pcs", "label merek"),
                         (f"{P}-ACC-HTG", 1, "pcs", "hangtag"),
                         (f"{P}-ACC-BNG", 0.02, "kg", "benang jahit")],
    (f"{P}-KM01", "M"): [(f"{P}-KN-OXF", 1.35, "m", "badan + lengan + kerah"),
                         (f"{P}-ACC-BTN", 9, "pcs", "kancing depan + manset"),
                         (f"{P}-ACC-LBL", 1, "pcs", "label merek"),
                         (f"{P}-ACC-BNG", 0.03, "kg", "benang jahit")],
}
SKU = [(f"{P}-TS01", "NVY", "M"), (f"{P}-TS01", "NVY", "L"), (f"{P}-TS01", "WHT", "M"),
       (f"{P}-KM01", "NVY", "M"), (f"{P}-KM01", "WHT", "M")]
HARGA_JUAL = {f"{P}-TS01": 89_000, f"{P}-KM01": 189_000}
TOKO = [(f"{P}-SHP", "Shopee Daluna", "shopee", "daluna.official"),
        (f"{P}-TTK", "TikTok Daluna", "tiktok", "@daluna.id")]


def sku_code(m, c, s) -> str:
    return f"{m}-{c}-{s}"


def rows_for(sheet: str) -> list[list]:
    if sheet == "01_LOKASI":
        return [[f"{P}-GD01", "Gudang Utama", "gudang", "", "ya"],
                [f"{P}-PRD01", "Ruang Produksi", "produksi", f"{P}-GD01", "ya"],
                [f"{P}-KTR", "Kantor Pusat", "kantor", "", "ya"]]
    if sheet == "02_KARYAWAN":
        return [[f"{P}-E001", "Siti Aminah", "penjahit", "081200000001", "2024-03-01",
                 f"{P}-PRD01", "borongan", 0, 0, "ya"],
                [f"{P}-E002", "Dewi Lestari", "qc", "081200000002", "2024-05-12",
                 f"{P}-PRD01", "bulanan", 3_200_000, 20_000, "ya"],
                [f"{P}-E003", "Ayu Pratiwi", "livehost", "081200000003", "2025-01-06",
                 f"{P}-KTR", "bulanan", 4_500_000, 25_000, "ya"]]
    if sheet == "03_WARNA":
        return [[k, n, h, i + 1] for i, (k, n, h) in enumerate(WARNA)]
    if sheet == "04_UKURAN":
        return [[k, n, u] for k, n, u in UKURAN]
    if sheet == "05_PROSES":
        return [[f"{P}-POT", "Potong", 1, "tidak", "cutting kain"],
                [f"{P}-JHT", "Jahit", 2, "tidak", "jahit utama"],
                [f"{P}-QC", "QC", 3, "tidak", "periksa mutu"],
                [f"{P}-PMK", "Permak", 4, "ya", "perbaikan hasil QC"],
                [f"{P}-PCK", "Packing", 5, "tidak", "kemas kirim"]]
    if sheet == "06_MATERIAL_KAIN":
        return [[f"{P}-KN-CTN", "Kain Cotton Combed 30s", "fabric", "kg", "cotton 100%",
                 "Navy", 180, 160, 95_000, 50],
                [f"{P}-KN-OXF", "Kain Oxford Katun", "fabric", "m", "cotton 80% poly 20%",
                 "Putih", 140, 150, 38_000, 100],
                [f"{P}-BNG-POL", "Benang Poly 40/2", "yarn", "kg", "polyester 100%",
                 "Putih", "", "", 62_000, 20]]
    if sheet == "07_AKSESORIS":
        return [[f"{P}-ACC-BTN", "Kancing kemeja 12mm", "pcs", "Kancing", 150, 2000, "pack", 144],
                [f"{P}-ACC-LBL", "Label merek jahit", "pcs", "Label", 350, 1000, "pack", 100],
                [f"{P}-ACC-HTG", "Hangtag karton", "pcs", "Hangtag", 500, 500, "pack", 100],
                [f"{P}-ACC-BNG", "Benang jahit jadi", "kg", "Benang Jahit", 68_000, 10, "", 1]]
    if sheet == "08_MODEL":
        return [[k, n, kat, "", h] for k, n, kat, h in MODEL]
    if sheet == "09_BARANG_JADI":
        out = []
        for m, c, s in SKU:
            nm = dict((k, n) for k, n, *_ in MODEL)[m]
            wn = dict((k, n) for k, n, _ in WARNA)[c]
            out.append([sku_code(m, c, s), f"{nm} [{wn} · {s}]", m, c, s, "pcs",
                        220 if m.endswith("TS01") else 320, HARGA_JUAL[m], 10])
        return out
    if sheet == "10_BOM":
        return [[m, s, mat, qty, unit, note]
                for (m, s), lines in BOM.items() for mat, qty, unit, note in lines]
    if sheet == "11_VENDOR_CMT":
        return [[f"{P}-CMT01", "CMT Pak Aan", "Aan", "081300000001", "Cimahi", 3000,
                 "spesialis kaos"],
                [f"{P}-CMT02", "CMT Bu Iis", "Iis", "081300000002", "Bandung", 1500,
                 "spesialis kemeja"]]
    if sheet == "12_KLIEN_MAKLON":
        return [[f"{P}-MK01", "Brand SnBM", "Koh Tri", "081400000001", "Jakarta",
                 "order bulanan"]]
    if sheet == "13_AKUN_TOKO":
        return [[k, n, p, u, "Daluna", "active"] for k, n, p, u in TOKO]
    if sheet == "14_KATALOG_JUAL":
        out = []
        for ak, _, _, _ in TOKO:
            for m, c, s in SKU[:3]:
                harga = HARGA_JUAL[m] + (10_000 if ak.endswith("TTK") else 0)
                out.append([ak, sku_code(m, c, s), harga, harga + 40_000, "", "ya"])
        return out
    if sheet == "15_KOL_KREATOR":
        return [[f"{P}-KOL01", "Rina Andira", "kontrak", "Bandung", "081500000001",
                 "rina.contoh@kreator.id", f"{P}-SHP,{P}-TTK", "both", 2000, 500, 500_000, 3],
                [f"{P}-KOL02", "Bagas Putra", "continue", "Jakarta", "081500000002",
                 "bagas.contoh@kreator.id", f"{P}-TTK", "per_pcs", 1500, 0, 0, 3],
                [f"{P}-KOL03", "Nia Safira", "new", "Bandung", "081500000003",
                 "nia.contoh@kreator.id", f"{P}-SHP", "none", 0, 0, 0, 3]]
    if sheet == "16_LIVEHOST":
        return [["Ayu Pratiwi", "ayu.contoh@dewiaditya.id", f"{P}-E003", "081200000003",
                 f"{P}-SHP,{P}-TTK", "active"]]
    return []


def build(dest: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "00_BACA_DULU"
    for line in [
        "CONTOH TERISI — TEMPLATE MASTER DATA CV. Dewi Aditya",
        "",
        "Berkas ini BUKAN data Anda. Semua kode diawali 'CTH-' supaya jelas contoh.",
        "Gunakan sebagai acuan bentuk pengisian, lalu ganti isinya dengan data asli",
        "(atau salin polanya ke TEMPLATE_MASTER_DA.xlsx yang masih kosong).",
        "",
        "Yang bisa Anda tiru dari contoh ini:",
        "· 09_BARANG_JADI: pola SKU = KODEMODEL-WARNA-UKURAN (mis. CTH-TS01-NVY-M)",
        "· 10_BOM: satu model+ukuran memakai BEBERAPA baris — kain, benang, DAN aksesoris",
        "  (kancing, label, hangtag). Aksesoris memang bagian HPP, jangan dilewat.",
        "· BOM per ukuran boleh berbeda qty-nya (M 0,22 kg vs L 0,24 kg).",
        "· 02_KARYAWAN: 'borongan' untuk penjahit, 'bulanan' untuk QC/live host.",
        "· 16_LIVEHOST wajib menunjuk nik_karyawan — gaji host dibaca dari payroll HR.",
        "· 15_KOL_KREATOR: tipe 'new' TIDAK boleh punya insentif (mode none).",
        "· 14_KATALOG_JUAL: satu SKU bisa berharga beda di tiap toko.",
        "",
        "Uji dulu sebelum menyimpan:",
        "  python3 scripts/import_master_template.py <berkas.xlsx>",
        "  python3 scripts/import_master_template.py <berkas.xlsx> --apply",
    ]:
        ws.append([line])
    ws["A1"].font = Font(bold=True, size=13)
    ws.column_dimensions["A"].width = 96

    for name in URUTAN:
        s = wb.create_sheet(name)
        s.append([k for k, *_ in SHEETS[name]["kolom"]])
        for c in s[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = HEAD_FILL
            s.column_dimensions[c.column_letter].width = 20
        for row in rows_for(name):
            s.append(row)
        s.freeze_panes = "A2"

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)
    total = sum(len(rows_for(n)) for n in URUTAN)
    print(f"Contoh terisi dibuat: {dest}  ({total} baris data pada {len(URUTAN)} sheet)")


if __name__ == "__main__":
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "/app/data_import/CONTOH_TERISI_MASTER_DA.xlsx")
    build(out)
