#!/usr/bin/env python3
"""ANALISIS SAJA (read-only) — bongkar isi ekspor Seller Center TikTok.

Tidak menulis apa pun ke database. Dipakai untuk dokumen analisis
memory/ANALISIS_IMPOR_SELLER_CENTER_KEPUTUSAN_2026-08-11.md
"""
import sys
from collections import Counter, defaultdict

import openpyxl

PATH = sys.argv[1] if len(sys.argv) > 1 else "/app/samples/TikTok_UntukDikirim_2026-07-19.xlsx"


def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    return wb.sheetnames, rows


def main():
    sheets, rows = load(PATH)
    print(f"FILE      : {PATH}")
    print(f"SHEETS    : {sheets}")
    print(f"RAW ROWS  : {len(rows)}")
    header = [str(h) if h is not None else "" for h in rows[0]]
    print(f"COLUMNS   : {len(header)}")
    print(f"ROW2 (desc?): {str(rows[1][0])[:80]!r} | {str(rows[1][1])[:60]!r}")
    data = rows[2:]
    data = [r for r in data if any(c not in (None, "") for c in r)]
    print(f"DATA ROWS : {len(data)}")
    print()

    idx = {h: i for i, h in enumerate(header)}

    def col(name):
        i = idx.get(name)
        if i is None:
            return []
        return [r[i] for r in data]

    print("=" * 100)
    print("PER-KOLOM: nama | terisi/total | nilai unik | 3 contoh")
    print("=" * 100)
    for i, h in enumerate(header):
        vals = [r[i] for r in data]
        filled = [v for v in vals if v not in (None, "")]
        uniq = len(set(str(v) for v in filled))
        ex = " | ".join(str(v)[:26] for v in filled[:3])
        flag = "  <== KOSONG TOTAL" if not filled else ""
        print(f"{i:>3} {h[:38]:<38} {len(filled):>4}/{len(vals):<4} uniq={uniq:<5} {ex[:60]}{flag}")
    print()

    print("=" * 100)
    print("DISTRIBUSI KOLOM KUNCI")
    print("=" * 100)
    for name in [
        "Order Status", "Order Substatus", "Cancelation/Return Type",
        "Normal or Pre-order", "Payment Method", "Order Channel",
        "Warehouse Name", "Shipping Provider Name", "Fulfillment Type",
        "Shipping Information", "Buyer Message", "Cancel By", "Cancel Reason",
    ]:
        if name in idx:
            c = Counter(str(v) if v not in (None, "") else "<kosong>" for v in col(name))
            print(f"\n{name}:")
            for k, v in c.most_common(8):
                print(f"    {k[:60]:<60} {v}")
        else:
            print(f"\n{name}: <KOLOM TIDAK ADA>")
    print()

    print("=" * 100)
    print("MATEMATIKA DOBEL: Order Amount vs SKU Subtotal")
    print("=" * 100)
    oid_i = idx.get("Order ID")
    if oid_i is None:
        for h in header:
            if "order id" in h.lower():
                oid_i = idx[h]
                break
    oids = [str(r[oid_i]).strip() for r in data]
    c = Counter(oids)
    multi = {k: v for k, v in c.items() if v > 1}
    print(f"Order ID unik      : {len(c)}")
    print(f"Order multi-baris  : {len(multi)} (total baris ekstra {sum(multi.values()) - len(multi)})")
    print(f"Baris terbanyak    : {max(c.values())}")

    def num(v):
        if v is None or v == "":
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace("Rp", "").replace(" ", "").strip()
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(".", "")
        try:
            return float(s)
        except Exception:
            return 0.0

    for money in ["Order Amount", "SKU Subtotal After Discount", "SKU Subtotal Before Discount",
                  "Shipping Fee After Discount", "Original Shipping Fee", "Taxes",
                  "SKU Platform Discount", "SKU Seller Discount"]:
        if money not in idx:
            print(f"{money:<32} <KOLOM TIDAK ADA>")
            continue
        vals = col(money)
        per_row = sum(num(v) for v in vals)
        seen = {}
        for o, v in zip(oids, vals):
            if o not in seen:
                seen[o] = num(v)
        per_order = sum(seen.values())
        print(f"{money:<32} jumlah per-BARIS = {per_row:>16,.0f} | jumlah per-ORDER (1x) = {per_order:>16,.0f} | selisih = {per_row - per_order:>13,.0f}")
    print()

    print("=" * 100)
    print("KUNCI PRODUK: Seller SKU vs SKU ID")
    print("=" * 100)
    for name in ["Seller SKU", "SKU ID", "Product Name", "Variation", "Product Category", "Quantity"]:
        if name not in idx:
            print(f"{name}: <TIDAK ADA>")
            continue
        vals = col(name)
        filled = [v for v in vals if v not in (None, "")]
        print(f"{name:<20} terisi {len(filled)}/{len(vals)} | unik {len(set(str(v) for v in filled))}")
    if "SKU ID" in idx and "Product Name" in idx:
        pair = defaultdict(set)
        for r in data:
            pair[str(r[idx['SKU ID']])].add(f"{r[idx['Product Name']]} / {r[idx.get('Variation', 0)]}")
        amb = {k: v for k, v in pair.items() if len(v) > 1}
        print(f"\nSKU ID unik = {len(pair)} | SKU ID yang punya >1 nama/variasi = {len(amb)}")
        for k, v in list(amb.items())[:3]:
            print(f"   {k}: {list(v)[:2]}")
    print()

    print("=" * 100)
    print("KUANTITAS (untuk skenario potong stok)")
    print("=" * 100)
    if "Quantity" in idx:
        qs = [int(num(v)) for v in col("Quantity")]
        print(f"total pcs = {sum(qs)} | min={min(qs)} max={max(qs)}")
        preo = idx.get("Normal or Pre-order")
        if preo is not None:
            byp = defaultdict(int)
            for r in data:
                byp[str(r[preo])] += int(num(r[idx['Quantity']]))
            for k, v in byp.items():
                print(f"   {k:<20} {v} pcs")
        # top SKU by qty
        top = defaultdict(int)
        for r in data:
            top[f"{r[idx['SKU ID']]} | {str(r[idx['Product Name']])[:34]} | {r[idx.get('Variation',0)]}"] += int(num(r[idx['Quantity']]))
        print("\n   TOP 10 SKU by pcs (kandidat pemetaan pertama):")
        for k, v in sorted(top.items(), key=lambda x: -x[1])[:10]:
            print(f"      {v:>4} pcs  {k[:88]}")
    print()

    print("=" * 100)
    print("FORMAT WAKTU (uji parse)")
    print("=" * 100)
    for name in ["Created Time", "Paid Time", "RTS Time", "Shipped Time", "Delivered Time", "Cancelled Time"]:
        if name not in idx:
            print(f"{name:<18} <TIDAK ADA>")
            continue
        vals = [v for v in col(name) if v not in (None, "")]
        print(f"{name:<18} terisi {len(vals):>4} | contoh {str(vals[0])[:30] if vals else '-'} | tipe {type(vals[0]).__name__ if vals else '-'}")


if __name__ == "__main__":
    main()
